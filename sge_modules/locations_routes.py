"""Domínio locations_routes extraído do app.py consolidado.

Carregado no arranque pelo mecanismo transitório de compatibilidade.
"""

@app.route('/locais')
def listar_locais():
    q = (request.args.get('q') or '').strip()
    incluir_inativos = (request.args.get('inativos') == '1')
    sort = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')
    status = (request.args.get('status') or 'todos').strip().lower()
    tipo = (request.args.get('tipo') or 'todos').strip()
    qualidade = (request.args.get('qualidade') or 'todos').strip().lower()
    estado_tecnico = (request.args.get('estado_tecnico') or 'todos').strip()
    prioridade = (request.args.get('prioridade') or 'todas').strip()

    locais = get_locais_with_cfg(q if q else None, incluir_inativos=incluir_inativos, sort=sort, order=order,
                                 tipo=tipo, qualidade=qualidade, estado_tecnico=estado_tecnico, prioridade=prioridade)
    if status == 'ativos':
        locais = [r for r in locais if int(r.get('ativo', 1)) == 1]
    elif status == 'arquivados':
        locais = [r for r in locais if int(r.get('ativo', 1)) != 1]
    elif status == 'sem_config':
        locais = [r for r in locais if not r.get('config_ok')]

    resumo = get_locais_module_summary(locais)
    ranking_atencao = sorted(locais, key=lambda x: (x.get('maturidade', 0), x.get('pot_contratada', 0) + x.get('pot_instalada', 0)))[:5]
    tipos_disponiveis = ['todos'] + sorted({r.get('tipo','Outro') for r in get_locais_with_cfg(None, incluir_inativos=True)})
    estados_tecnicos = ['todos', 'Normal', 'Atenção', 'Crítico']
    prioridades = ['todas', 'Baixa', 'Média', 'Alta']
    return render_template('locais.html',
                           locais=locais, q=q,
                           incluir_inativos=incluir_inativos,
                           sort=sort, order=order,
                           status=status, tipo=tipo, qualidade=qualidade,
                           estado_tecnico=estado_tecnico, prioridade=prioridade,
                           tipos_disponiveis=tipos_disponiveis,
                           estados_tecnicos=estados_tecnicos,
                           prioridades=prioridades,
                           ranking_atencao=ranking_atencao,
                           resumo=resumo)


@app.route('/locais/template.csv')
def export_locais_template_csv():
    si = StringIO()
    w = csv.writer(si, delimiter=';')
    w.writerow(['nome','codigo','endereco','contato_nome','contato_tel','email','responsavel_alt','tipo_local','categoria_operacional','estado_tecnico','prioridade','ativo','fator_mult','pot_contratada','pot_instalada','tarifa_ativa','tarifa_reativa','tarifa_ponta','tarifa_perdas','taxa_fixa','taxa_radio','taxa_lixo','iva','notas'])
    w.writerow(['Ex.: ETA Umbeluzi','ETA-UMB','Umbeluzi, Maputo','Supervisor Local','84xxxxxxx','supervisor@exemplo.co.mz','Chefe de turno','ETA','Produção','Normal','Alta',1,1.0,6000,11750,4.780,1.430,497.03,4.780,207.28,297.00,150.00,16,'Local de referência'])
    output = si.getvalue()
    return Response(output.encode('utf-8'), mimetype='text/csv; charset=utf-8', headers={"Content-Disposition": "attachment; filename=template_locais.csv"})


@app.route('/locais/export.csv')
def export_locais_csv():
    q = (request.args.get('q') or '').strip()
    incluir_inativos = (request.args.get('inativos') == '1')
    sort = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')
    data = get_locais_with_cfg(q if q else None, incluir_inativos=incluir_inativos, sort=sort, order=order)

    si = StringIO()
    w = csv.writer(si, delimiter=';')
    w.writerow([
        'id','nome','codigo','endereco','contato_nome','contato_tel','email','responsavel_alt','tipo_local','categoria_operacional','estado_tecnico','prioridade','ativo',
        'fator_mult','pot_contratada_kW','pot_instalada_kW','notas'
    ])
    for r in data:
        w.writerow([
            r['id'], r['nome'], r.get('codigo','') or '', r.get('endereco','') or '',
            r.get('contato_nome','') or '', r.get('contato_tel','') or '', r.get('email','') or '', r.get('responsavel_alt','') or '',
            r.get('tipo_local','') or '', r.get('categoria_operacional','') or '', r.get('estado_tecnico','') or '', r.get('prioridade','') or '',
            r.get('ativo',1),
            f"{r['fator_mult']:.4f}",
            f"{r['pot_contratada']:.2f}",
            f"{r['pot_instalada']:.2f}",
            (r.get('notas','') or '').replace('\n',' ').strip()
        ])
    output = si.getvalue()
    return Response(output.encode('utf-8'),
                    mimetype='text/csv; charset=utf-8',
                    headers={"Content-Disposition": "attachment; filename=locais.csv"})

@app.route('/locais/export.xlsx')
def export_locais_xlsx():
    q = (request.args.get('q') or '').strip()
    incluir_inativos = (request.args.get('inativos') == '1')
    sort = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')
    data = get_locais_with_cfg(q if q else None, incluir_inativos=incluir_inativos, sort=sort, order=order)

    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    _set_xlsx_identity(wb, 'Centros e Locais')
    ws = wb.add_worksheet('Locais')
    hdr = wb.add_format({'bold': True, 'bg_color': '#EAF4FF', 'font_color': '#174983', 'border': 1})
    txt = wb.add_format({'border': 1})
    num = wb.add_format({'border': 1, 'num_format': '0.00'})
    headers = ['ID','Nome','Tipo','Categoria','Sector','Código','Província','Município','Distrito','Bairro','Latitude','Longitude','Endereço','Contacto','Telefone','Email','Responsável alt.','Estado técnico','Prioridade','Ativo','Fator','Pot. contratada (kW)','Pot. instalada (kW)','Maturidade (%)','Fonte','Referência externa','Última sincronização']
    for col,h in enumerate(headers): ws.write(0,col,h,hdr)
    for i,r in enumerate(data, start=1):
        vals = [r['id'], r['nome'], r.get('tipo'), r.get('categoria_operacional') or '', r.get('sector_operacional') or '',
                r.get('codigo') or '', r.get('provincia') or '', r.get('municipio') or '', r.get('distrito') or '', r.get('bairro') or '',
                r.get('latitude'), r.get('longitude'), r.get('endereco') or '', r.get('contato_nome') or '', r.get('contato_tel') or '',
                r.get('email') or '', r.get('responsavel_alt') or '', r.get('estado_tecnico') or '', r.get('prioridade') or '',
                'Sim' if int(r.get('ativo',1))==1 else 'Não', float(r.get('fator_mult',1) or 1),
                float(r.get('pot_contratada',0) or 0), float(r.get('pot_instalada',0) or 0), int(r.get('maturidade',0) or 0),
                r.get('fonte_cadastro') or '', r.get('referencia_externa') or '', r.get('ultima_sincronizacao') or '']
        for col,v in enumerate(vals):
            fmt = num if isinstance(v,(int,float)) and col in [10,11,20,21,22,23] else txt
            ws.write(i,col,v,fmt)
    ws.autofilter(0,0,max(len(data),1),len(headers)-1)
    ws.freeze_panes(1,0)
    for idx,w in enumerate([8,28,18,22,15,14,16,18,18,18,12,12,30,20,16,24,22,16,12,10,10,18,18,14,24,28,22]): ws.set_column(idx,idx,w)
    wb.close()
    output.seek(0)
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment; filename=locais.xlsx'})


@app.route('/locais/<int:local_id>')
def detalhes_local(local_id):
    local = get_local_full(local_id)
    if not local:
        flash('Local não encontrado.', 'warning')
        return redirect(url_for('listar_locais'))

    cfg_full = get_local_cfg_full(local_id)
    equipamentos = get_equipamentos_por_local(local_id)
    overview = get_local_overview(local_id)
    local['tipo'] = local.get('tipo_local') or infer_local_tipo(local.get('nome'), local.get('endereco'))
    local['maturidade'] = calcular_maturidade_local({**local, **cfg_full})
    alertas = get_local_alertas(local, cfg_full, overview)
    history = get_local_history(local_id, limit=12)
    sublocais = get_local_children(local_id, include_inactive=True)
    return render_template('detalhes_local.html', local=local, cfg=cfg_full, equipamentos=equipamentos, overview=overview, alertas=alertas, history=history, sublocais=sublocais)


@app.route('/locais/adicionar', methods=['GET', 'POST'])
def adicionar_local():
    parent_options = get_local_choices(include_inactive=True)
    if request.method == 'POST':
        nome = (request.form.get('nome') or '').strip()
        codigo = (request.form.get('codigo') or '').strip() or None
        endereco = (request.form.get('endereco') or '').strip() or None
        contato_nome = (request.form.get('contato_nome') or '').strip() or None
        contato_tel = (request.form.get('contato_tel') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        responsavel_alt = (request.form.get('responsavel_alt') or '').strip() or None
        tipo_local = (request.form.get('tipo_local') or '').strip() or None
        categoria_operacional = (request.form.get('categoria_operacional') or '').strip() or None
        estado_tecnico = (request.form.get('estado_tecnico') or 'Normal').strip() or 'Normal'
        prioridade = (request.form.get('prioridade') or 'Média').strip() or 'Média'
        notas = (request.form.get('notas') or '').strip() or None
        ativo = 1 if (request.form.get('ativo', '1') == '1') else 0
        parent_raw = (request.form.get('parent_id') or '').strip()
        parent_id = int(parent_raw) if parent_raw.isdigit() else None
        provincia = (request.form.get('provincia') or '').strip() or None
        municipio = (request.form.get('municipio') or '').strip() or None
        distrito = (request.form.get('distrito') or '').strip() or None
        bairro = (request.form.get('bairro') or '').strip() or None
        sector_operacional = (request.form.get('sector_operacional') or '').strip() or None
        try: latitude = float((request.form.get('latitude') or '').replace(',','.'))
        except Exception: latitude = None
        try: longitude = float((request.form.get('longitude') or '').replace(',','.'))
        except Exception: longitude = None
        if not nome:
            flash('O nome do local é obrigatório.', 'warning')
            return render_template('adicionar_local.html', form=request.form, parent_options=parent_options)
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        try:
            c.execute('''INSERT INTO locais (nome, codigo, endereco, contato_nome, contato_tel, email, responsavel_alt, tipo_local, categoria_operacional, estado_tecnico, prioridade, notas, ativo, parent_id,
                                             provincia, municipio, distrito, bairro, latitude, longitude, sector_operacional, classificacao_confirmada)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)''',
                      (nome, codigo, endereco, contato_nome, contato_tel, email, responsavel_alt, tipo_local, categoria_operacional, estado_tecnico, prioridade, notas, ativo, parent_id,
                       provincia, municipio, distrito, bairro, latitude, longitude, sector_operacional))
            lid = c.lastrowid
            c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (lid,))
            conn.commit()
            log_local_history(lid, 'Local criado', f'Cadastro inicial do local {nome}', actor='locais_fase4')
            flash(f'Local "{nome}" criado com sucesso.', 'success')
            return redirect(url_for('detalhes_local', local_id=lid))
        except sqlite3.IntegrityError:
            flash('Já existe um local com esse nome.', 'danger')
        except Exception as e:
            flash(f'Não foi possível criar o local: {e}', 'danger')
        finally:
            conn.close()
    return render_template('adicionar_local.html', form=request.form, parent_options=parent_options)


@app.route('/locais/editar/<int:local_id>', methods=['GET', 'POST'])
def editar_local(local_id):
    local = get_local_full(local_id)
    parent_options = get_local_choices(include_inactive=True, exclude_id=local_id)
    if not local:
        flash('Local não encontrado.', 'warning')
        return redirect(url_for('listar_locais'))
    if request.method == 'POST':
        novo_nome = (request.form.get('nome') or '').strip()
        codigo = (request.form.get('codigo') or '').strip() or None
        endereco = (request.form.get('endereco') or '').strip() or None
        contato_nome = (request.form.get('contato_nome') or '').strip() or None
        contato_tel = (request.form.get('contato_tel') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        responsavel_alt = (request.form.get('responsavel_alt') or '').strip() or None
        tipo_local = (request.form.get('tipo_local') or '').strip() or None
        categoria_operacional = (request.form.get('categoria_operacional') or '').strip() or None
        estado_tecnico = (request.form.get('estado_tecnico') or 'Normal').strip() or 'Normal'
        prioridade = (request.form.get('prioridade') or 'Média').strip() or 'Média'
        notas = (request.form.get('notas') or '').strip() or None
        ativo = 1 if (request.form.get('ativo', '1') == '1') else 0
        parent_raw = (request.form.get('parent_id') or '').strip()
        parent_id = int(parent_raw) if parent_raw.isdigit() else None
        if parent_id == local_id:
            parent_id = None
        provincia = (request.form.get('provincia') or '').strip() or None
        municipio = (request.form.get('municipio') or '').strip() or None
        distrito = (request.form.get('distrito') or '').strip() or None
        bairro = (request.form.get('bairro') or '').strip() or None
        sector_operacional = (request.form.get('sector_operacional') or '').strip() or None
        try: latitude = float((request.form.get('latitude') or '').replace(',','.'))
        except Exception: latitude = None
        try: longitude = float((request.form.get('longitude') or '').replace(',','.'))
        except Exception: longitude = None
        if not novo_nome:
            flash('O nome do local é obrigatório.', 'warning')
            local.update({'nome': novo_nome, 'codigo': codigo, 'endereco': endereco, 'contato_nome': contato_nome, 'contato_tel': contato_tel, 'email': email, 'responsavel_alt': responsavel_alt, 'tipo_local': tipo_local, 'categoria_operacional': categoria_operacional, 'estado_tecnico': estado_tecnico, 'prioridade': prioridade, 'notas': notas, 'ativo': ativo})
            return render_template('editar_local.html', local=local, parent_options=parent_options)
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        try:
            c.execute('''UPDATE locais
                            SET nome=?, codigo=?, endereco=?, contato_nome=?, contato_tel=?, email=?, responsavel_alt=?, tipo_local=?, categoria_operacional=?, estado_tecnico=?, prioridade=?, notas=?, ativo=?, parent_id=?,
                                provincia=?, municipio=?, distrito=?, bairro=?, latitude=?, longitude=?, sector_operacional=?, classificacao_confirmada=1
                          WHERE id=?''',
                      (novo_nome, codigo, endereco, contato_nome, contato_tel, email, responsavel_alt, tipo_local, categoria_operacional, estado_tecnico, prioridade, notas, ativo, parent_id,
                       provincia, municipio, distrito, bairro, latitude, longitude, sector_operacional, local_id))
            conn.commit()
            log_local_history(local_id, 'Perfil atualizado', f'Nome: {novo_nome}; prioridade: {prioridade}; estado técnico: {estado_tecnico}', actor='locais_fase4')
            flash('Local atualizado com sucesso.', 'success')
            return redirect(url_for('detalhes_local', local_id=local_id))
        except sqlite3.IntegrityError:
            flash('Já existe outro local com esse nome.', 'danger')
        except Exception as e:
            flash(f'Não foi possível atualizar o local: {e}', 'danger')
        finally:
            conn.close()
        local.update({'nome': novo_nome, 'codigo': codigo, 'endereco': endereco, 'contato_nome': contato_nome, 'contato_tel': contato_tel, 'notas': notas, 'ativo': ativo})
    return render_template('editar_local.html', local=local, parent_options=parent_options)

# === NOVO: Importar Locais
# === NOVO: Importar Locais (CSV)
@app.route('/locais/import', methods=['GET','POST'])
def locais_import():
    if request.method == 'POST':
        f = request.files.get('arquivo')
        if not f or f.filename == '':
            return redirect(url_for('listar_locais', msg='selecione um arquivo CSV'))

        from asset_registry_service import header_key
        extension = os.path.splitext(f.filename or '')[1].lower()
        if extension == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
            matrix = list(ws.iter_rows(values_only=True))
            header_index = next((i for i,row in enumerate(matrix[:30]) if 'nome' in [header_key(v) for v in row]), None)
            if header_index is None:
                wb.close()
                return redirect(url_for('listar_locais', msg='Excel precisa conter a coluna "nome"'))
            headers = [header_key(v) for v in matrix[header_index]]
            reader = [
                {headers[i]: value for i,value in enumerate(row) if i < len(headers) and headers[i]}
                for row in matrix[header_index+1:]
            ]
            wb.close()
        else:
            content = f.read().decode('utf-8-sig', errors='ignore')
            delimiter = ';' if content.count(';') > content.count(',') else ','
            raw_reader = csv.DictReader(StringIO(content), delimiter=delimiter)
            reader = [{header_key(k): v for k,v in row.items()} for row in raw_reader]
            headers = [header_key(h) for h in (raw_reader.fieldnames or [])]
        if 'nome' not in headers:
            return redirect(url_for('listar_locais', msg='O ficheiro precisa conter a coluna "nome"'))

        add_count = 0; upd_count = 0; err_count = 0
        history_updates = []
        pending_parents = []
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()

        def ftext(value):
            return '' if value is None else str(value).strip()

        def fbool(value, default=1):
            text = header_key(value)
            if text in ('1', 'sim', 's', 'yes', 'true', 'activo', 'ativo'):
                return 1
            if text in ('0', 'nao', 'n', 'no', 'false', 'inactivo', 'inativo'):
                return 0
            return default

        def ffloat(v, dflt):
            try:
                value = float(str(v).replace(',','.'))
                return value if math.isfinite(value) and value >= 0 else dflt
            except Exception:
                return dflt

        def fcoord(v):
            try:
                value = float(str(v).replace(',','.'))
                return value if math.isfinite(value) else None
            except Exception:
                return None

        for row in reader:
            try:
                nome = ftext(row.get('nome'))
                if not nome:
                    err_count += 1; continue

                # Upsert em locais (com colunas novas)
                campos_locais = {
                    "codigo": row.get('codigo'),
                    "endereco": row.get('endereco'),
                    "contato_nome": row.get('contato_nome'),
                    "contato_tel": row.get('contato_tel'),
                    "email": row.get('email'),
                    "responsavel_alt": row.get('responsavel_alt'),
                    "tipo_local": row.get('tipo_local'),
                    "categoria_operacional": row.get('categoria_operacional'),
                    "estado_tecnico": row.get('estado_tecnico') or 'Normal',
                    "prioridade": row.get('prioridade') or 'Média',
                    "notas": row.get('notas'),
                    "ativo": fbool(row.get('ativo'), 1),
                    "provincia": row.get('provincia'), "municipio": row.get('municipio'),
                    "distrito": row.get('distrito'), "bairro": row.get('bairro'),
                    "latitude": fcoord(row.get('latitude')), "longitude": fcoord(row.get('longitude')),
                    "sector_operacional": row.get('sector_operacional') or row.get('sector'),
                    "parent_nome": row.get('parent_nome') or row.get('local_pai'),
                }
                parent_id = None
                if ftext(campos_locais['parent_nome']):
                    c.execute('SELECT id FROM locais WHERE lower(trim(nome))=lower(trim(?))', (ftext(campos_locais['parent_nome']),))
                    parent_found = c.fetchone(); parent_id = parent_found[0] if parent_found else None
                c.execute('SELECT id FROM locais WHERE lower(trim(nome))=lower(trim(?))', (nome,))
                found = c.fetchone()
                if found:
                    lid = found[0]
                    c.execute('''UPDATE locais
                                    SET codigo=?, endereco=?, contato_nome=?, contato_tel=?, email=?, responsavel_alt=?, tipo_local=?, categoria_operacional=?, estado_tecnico=?, prioridade=?, notas=?, ativo=?,
                                        provincia=?, municipio=?, distrito=?, bairro=?, latitude=?, longitude=?, sector_operacional=?,
                                        parent_id=COALESCE(?,parent_id), classificacao_confirmada=1
                                  WHERE id=?''',
                              (campos_locais["codigo"], campos_locais["endereco"],
                               campos_locais["contato_nome"], campos_locais["contato_tel"], campos_locais["email"], campos_locais["responsavel_alt"], campos_locais["tipo_local"], campos_locais["categoria_operacional"], campos_locais["estado_tecnico"], campos_locais["prioridade"],
                               campos_locais["notas"], campos_locais["ativo"], campos_locais['provincia'], campos_locais['municipio'],
                               campos_locais['distrito'], campos_locais['bairro'], campos_locais['latitude'], campos_locais['longitude'],
                               campos_locais['sector_operacional'], parent_id, lid))
                    upd_count += 1
                else:
                    c.execute('''INSERT INTO locais (nome, codigo, endereco, contato_nome, contato_tel, email, responsavel_alt, tipo_local, categoria_operacional, estado_tecnico, prioridade, notas, ativo,
                                                       provincia, municipio, distrito, bairro, latitude, longitude, sector_operacional, parent_id, classificacao_confirmada)
                                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)''',
                              (nome, campos_locais["codigo"], campos_locais["endereco"],
                               campos_locais["contato_nome"], campos_locais["contato_tel"], campos_locais["email"], campos_locais["responsavel_alt"], campos_locais["tipo_local"], campos_locais["categoria_operacional"], campos_locais["estado_tecnico"], campos_locais["prioridade"],
                               campos_locais["notas"], campos_locais["ativo"], campos_locais['provincia'], campos_locais['municipio'],
                               campos_locais['distrito'], campos_locais['bairro'], campos_locais['latitude'], campos_locais['longitude'],
                               campos_locais['sector_operacional'], parent_id))
                    lid = c.lastrowid
                    add_count += 1
                if ftext(campos_locais['parent_nome']):
                    pending_parents.append((lid, ftext(campos_locais['parent_nome'])))

                # Upsert em locais_cfg (se vierem colunas)
                cfg = {
                    "fator_mult": ffloat(row.get('fator_mult'), 1.0),
                    "pot_contratada": ffloat(row.get('pot_contratada'), 0.0),
                    "pot_instalada": ffloat(row.get('pot_instalada'), 0.0),
                    "tarifa_ativa": ffloat(row.get('tarifa_ativa'), 4.780),
                    "tarifa_reativa": ffloat(row.get('tarifa_reativa'), 1.430),
                    "tarifa_ponta": ffloat(row.get('tarifa_ponta'), 497.03),
                    "tarifa_perdas": ffloat(row.get('tarifa_perdas'), 4.780),
                    "taxa_fixa": ffloat(row.get('taxa_fixa'), 207.28),
                    "taxa_radio": ffloat(row.get('taxa_radio'), 297.00),
                    "taxa_lixo": ffloat(row.get('taxa_lixo'), 150.00),
                    "iva": 16.0,
                }
                c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (lid,))
                c.execute('''UPDATE locais_cfg
                                SET fator_mult=?, pot_contratada=?, pot_instalada=?,
                                    tarifa_ativa=?, tarifa_reativa=?, tarifa_ponta=?, tarifa_perdas=?,
                                    taxa_fixa=?, taxa_radio=?, taxa_lixo=?, iva=?
                              WHERE local_id=?''',
                          (cfg["fator_mult"], cfg["pot_contratada"], cfg["pot_instalada"],
                           cfg["tarifa_ativa"], cfg["tarifa_reativa"], cfg["tarifa_ponta"], cfg["tarifa_perdas"],
                           cfg["taxa_fixa"], cfg["taxa_radio"], cfg["taxa_lixo"], cfg["iva"], lid))
                history_updates.append((lid, cfg))
            except Exception:
                err_count += 1

        # Segunda passagem: permite referenciar um local-pai que aparece mais
        # abaixo na mesma folha Excel/CSV.
        for child_id, parent_name in pending_parents:
            try:
                parent = c.execute(
                    'SELECT id FROM locais WHERE lower(trim(nome))=lower(trim(?))', (parent_name,)
                ).fetchone()
                if parent and int(parent[0]) != int(child_id):
                    c.execute('UPDATE locais SET parent_id=? WHERE id=?', (parent[0], child_id))
            except Exception:
                err_count += 1

        conn.commit(); conn.close()
        for lid, cfg in history_updates:
            try:
                _guardar_tarifa_historica(
                    lid, datetime.now().strftime('%Y-%m-%d'), cfg,
                    actor=_actor_name(), notes='Atualização por importação CSV de locais',
                )
            except Exception:
                err_count += 1
        if add_count or upd_count:
            log_local_history(0, 'Importação CSV', f'{add_count} adicionados, {upd_count} atualizados, {err_count} com erro', actor='locais_fase4')
        return redirect(url_for('listar_locais', msg=f'import:{add_count} add, {upd_count} upd, {err_count} err'))
    colunas = ['nome','codigo','tipo_local','categoria_operacional','sector_operacional','parent_nome','provincia','municipio','distrito','bairro','endereco','latitude','longitude','contato_nome','contato_tel','email','responsavel_alt','estado_tecnico','prioridade','ativo','fator_mult','pot_contratada','pot_instalada','tarifa_ativa','tarifa_reativa','tarifa_ponta','tarifa_perdas','taxa_fixa','taxa_radio','taxa_lixo','iva','notas']
    return render_template('locais_import.html', colunas=colunas)

# === NOVO: Duplicar e Ativar/Arquivar Local
@app.route('/locais/duplicar/<int:local_id>', methods=['POST'])
def locais_duplicar(local_id):
    info = get_local_full(local_id)
    if not info:
        return redirect(url_for('listar_locais', msg='local_nao_encontrado'))
    base = info['nome'] + " (Cópia)"
    novo_nome = base
    i = 2
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    # evita nome duplicado
    while True:
        c.execute('SELECT 1 FROM locais WHERE nome=?', (novo_nome,))
        if not c.fetchone():
            break
        novo_nome = f"{base} {i}"; i += 1
    # dup local
    c.execute('''INSERT INTO locais (nome, codigo, endereco, contato_nome, contato_tel, notas, ativo, parent_id)
                 VALUES (?, ?, ?, ?, ?, ?, 1, ?)''',
              (novo_nome,
               (info['codigo'] or '') + '-copy' if info['codigo'] else None,
               info['endereco'], info['contato_nome'], info['contato_tel'], info['notas'], info.get('parent_id')))
    novo_id = c.lastrowid
    # dup cfg
    cfg = get_local_cfg_full(local_id)
    c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (novo_id,))
    c.execute('''UPDATE locais_cfg SET fator_mult=?, pot_contratada=?, pot_instalada=?, tarifa_ativa=?, tarifa_reativa=?, tarifa_ponta=?,
                 tarifa_perdas=?, taxa_fixa=?, taxa_radio=?, taxa_lixo=?, iva=16.0 WHERE local_id=?''',
              (cfg['fator_mult'], cfg['pot_contratada'], cfg['pot_instalada'], cfg['tarifa_ativa'], cfg['tarifa_reativa'],
               cfg['tarifa_ponta'], cfg['tarifa_perdas'], cfg['taxa_fixa'], cfg['taxa_radio'], cfg['taxa_lixo'], novo_id))
    conn.commit(); conn.close()
    _guardar_tarifa_historica(
        novo_id, datetime.now().strftime('%Y-%m-%d'), cfg,
        actor=_actor_name(), notes=f'Copiado de {info["nome"]}',
    )
    log_local_history(novo_id, 'Local duplicado', f'Criado a partir de {info["nome"]}', actor='locais_fase4')
    return redirect(url_for('listar_locais', msg=f'duplicado:{novo_nome}'))

@app.route('/locais/arquivar/<int:local_id>', methods=['POST'])
def arquivar_local(local_id):
    return locais_toggle(local_id)

@app.route('/locais/toggle/<int:local_id>', methods=['POST'])
def locais_toggle(local_id):
    info = get_local_full(local_id)
    if not info:
        return redirect(url_for('listar_locais', msg='local_nao_encontrado'))
    novo = 0 if info['ativo'] == 1 else 1
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('UPDATE locais SET ativo=? WHERE id=?', (novo, local_id))
    conn.commit(); conn.close()
    log_local_history(local_id, 'Estado alterado', 'Ativado' if novo==1 else 'Arquivado', actor='locais_fase4')
    return redirect(url_for('listar_locais', msg=('ativado' if novo==1 else 'arquivado')))

# === CONFIG LOCAL ===
@app.route('/locais/config/<int:local_id>', methods=['GET', 'POST'])
def configurar_local(local_id):
    local = get_local_full(local_id)
    if not local:
        flash('Local não encontrado.', 'warning')
        return redirect(url_for('listar_locais'))

    if request.method == 'POST':
        try:
            fator_mult       = float(request.form.get('fator_mult', 1) or 1)
            pot_contratada   = float(request.form.get('pot_contratada', 0) or 0)
            pot_instalada    = float(request.form.get('pot_instalada', 0) or 0)
            tarifa_ativa     = float(request.form.get('tarifa_ativa', 4.780) or 4.780)
            tarifa_reativa   = float(request.form.get('tarifa_reativa', 1.430) or 1.430)
            tarifa_ponta     = float(request.form.get('tarifa_ponta', 497.03) or 497.03)
            tarifa_perdas    = float(request.form.get('tarifa_perdas', 4.780) or 4.780)
            taxa_fixa        = float(request.form.get('taxa_fixa', 207.28) or 207.28)
            taxa_radio       = float(request.form.get('taxa_radio', 297.00) or 297.00)
            taxa_lixo        = float(request.form.get('taxa_lixo', 150.00) or 150.00)
            tarifa_valid_from = request.form.get('tarifa_valid_from') or datetime.now().strftime('%Y-%m-%d')
            datetime.strptime(tarifa_valid_from, '%Y-%m-%d')
            numbers = (fator_mult, pot_contratada, pot_instalada, tarifa_ativa, tarifa_reativa,
                       tarifa_ponta, tarifa_perdas, taxa_fixa, taxa_radio, taxa_lixo)
            if fator_mult <= 0 or any(not math.isfinite(value) or value < 0 for value in numbers):
                raise ValueError('valores fora do intervalo permitido')
        except (TypeError, ValueError):
            flash('Os parâmetros técnicos, tarifas ou data de vigência são inválidos.', 'danger')
            return redirect(url_for('configurar_local', local_id=local_id))

        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (local_id,))
        c.execute('''
            UPDATE locais_cfg
               SET fator_mult=?, pot_instalada=?, iva=16.0
             WHERE local_id=?
        ''', (fator_mult, pot_instalada, local_id))
        conn.commit(); conn.close()
        _guardar_tarifa_historica(
            local_id,
            tarifa_valid_from,
            {
                'pot_contratada': pot_contratada,
                'tarifa_ativa': tarifa_ativa,
                'tarifa_reativa': tarifa_reativa,
                'tarifa_ponta': tarifa_ponta,
                'tarifa_perdas': tarifa_perdas,
                'taxa_fixa': taxa_fixa,
                'taxa_radio': taxa_radio,
                'taxa_lixo': taxa_lixo,
            },
            actor=_actor_name(),
            notes='Atualização pela configuração do local',
        )
        log_local_history(local_id, 'Configuração atualizada', f'Fator {fator_mult:.4f}; Pot. contratada {pot_contratada:.2f} kW; Pot. instalada {pot_instalada:.2f} kW', actor='locais_fase4')
        flash('Configuração do local guardada com sucesso.', 'success')
        return redirect(url_for('configurar_local', local_id=local_id))

    cfg = get_local_cfg_full(local_id)
    overview = get_local_overview(local_id)
    return render_template('local_config.html', local=local, cfg=cfg, overview=overview,
                           tarifa_valid_from=datetime.now().strftime('%Y-%m-%d'))


@app.route('/locais/<int:local_id>/tarifas', methods=['GET', 'POST'])
def tarifas_historico_local(local_id):
    local = get_local_full(local_id)
    if not local:
        flash('Local não encontrado.', 'warning')
        return redirect(url_for('listar_locais'))
    if request.method == 'POST':
        try:
            _guardar_tarifa_historica(
                local_id,
                request.form.get('valid_from'),
                {
                    'pot_contratada': request.form.get('pot_contratada'),
                    'tarifa_ativa': request.form.get('tarifa_ativa'),
                    'tarifa_reativa': request.form.get('tarifa_reativa'),
                    'tarifa_ponta': request.form.get('tarifa_ponta'),
                    'tarifa_perdas': request.form.get('tarifa_perdas'),
                    'taxa_fixa': request.form.get('taxa_fixa'),
                    'taxa_radio': request.form.get('taxa_radio'),
                    'taxa_lixo': request.form.get('taxa_lixo'),
                },
                actor=_actor_name(),
                notes=(request.form.get('notes') or '').strip(),
            )
            log_local_history(local_id, 'Tarifário histórico atualizado', request.form.get('valid_from') or '', actor=_actor_name())
            flash('Nova vigência tarifária guardada sem alterar faturas históricas anteriores.', 'success')
            return redirect(url_for('tarifas_historico_local', local_id=local_id))
        except (ValueError, TypeError):
            flash('Data ou valores tarifários inválidos.', 'danger')
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        'SELECT * FROM tarifas_historico WHERE local_id=? ORDER BY valid_from DESC, id DESC',
        (local_id,),
    ).fetchall()
    conn.close()
    cfg = get_local_cfg_full(local_id)
    return render_template('tarifas_historico.html', local=local, tarifas=rows, cfg=cfg,
                           hoje=datetime.now().strftime('%Y-%m-%d'), iva_rate=16.0,
                           iva_base_percent=62.0)

# === EQUIPAMENTOS ===
# === EQUIPAMENTOS ===
