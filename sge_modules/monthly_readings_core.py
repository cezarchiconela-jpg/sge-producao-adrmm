"""Domínio monthly_readings_core extraído do app.py consolidado.

Carregado no arranque pelo mecanismo transitório de compatibilidade.
"""

def ensure_leituras_mensais_phase2_schema():
    """Garante pequenos índices de performance sem alterar dados existentes.
    Mantém compatibilidade total com o sge.db atual e evita leituras lentas
    em meses com muitos registos.
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_mensais_local_mes_ano_data ON leituras_mensais(local, mes, ano, data)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_mensais_local_data ON leituras_mensais(local, data)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_locais_cfg_local_id ON locais_cfg(local_id)")
        conn.commit()
        conn.close()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
    return True


def _local_id_nome_from_request(raw_local, locais_db):
    raw = (str(raw_local or '').strip())
    if not raw and locais_db:
        raw = str(locais_db[0][0])
    selected_id = None
    local_nome = raw
    if raw.isdigit():
        selected_id = int(raw)
        nm = get_local_by_id(selected_id)
        local_nome = nm[1] if nm else raw
    else:
        for lid, lname in locais_db:
            if lname == raw:
                selected_id = lid
                local_nome = lname
                break
    return selected_id, local_nome


def _cfg_map_locais(locais_db):
    """Mapa de configurações dos locais em consulta única.
    Evita abrir uma ligação SQLite para cada local, o que deixava a página
    de leituras mensais lenta quando havia muitos locais cadastrados.
    """
    mapa = {}
    if not locais_db:
        return mapa
    nomes_por_id = {int(lid): lname for lid, lname in locais_db}
    ids = list(nomes_por_id.keys())
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        qmarks = ','.join(['?'] * len(ids))
        c.execute(f'''
            SELECT local_id, fator_mult, pot_contratada, tarifa_ativa, tarifa_reativa, tarifa_ponta,
                   tarifa_perdas, taxa_fixa, taxa_radio, taxa_lixo, iva, COALESCE(pot_instalada, 0.0)
            FROM locais_cfg
            WHERE local_id IN ({qmarks})
        ''', ids)
        rows = c.fetchall()
        conn.close()
        for lid, fator, potc, ta, tr, tp, perdas, fixa, radio, lixo, iva, poti in rows:
            lname = nomes_por_id.get(int(lid))
            if not lname:
                continue
            mapa[lname] = {
                'fator_mult': fator, 'pot_contratada': potc, 'tarifa_ativa': ta,
                'tarifa_reativa': tr, 'tarifa_ponta': tp, 'tarifa_perdas': perdas,
                'taxa_fixa': fixa, 'taxa_radio': radio, 'taxa_lixo': lixo,
                'iva': iva, 'pot_instalada': poti
            }
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
    # Garante defaults sem chamadas adicionais ao banco
    for lid, lname in locais_db:
        mapa.setdefault(lname, {
            'fator_mult': 1.0, 'pot_contratada': 0.0, 'tarifa_ativa': 0.0,
            'tarifa_reativa': 0.0, 'tarifa_ponta': 0.0, 'tarifa_perdas': 0.0,
            'taxa_fixa': 0.0, 'taxa_radio': 0.0, 'taxa_lixo': 0.0,
            'iva': 16.0, 'pot_instalada': 0.0
        })
    return mapa


def _safe_float(v, default=None):
    if v is None or v == '':
        return default
    try:
        return float(str(v).replace(',', '.'))
    except Exception:
        return default




def _fmt_mil(v, casas=2):
    """Formata números com separador de milhares no padrão PT: 1.234.567,89."""
    try:
        n = float(v or 0)
    except Exception:
        n = 0.0
    txt = f"{n:,.{int(casas)}f}"
    return txt.replace(',', 'X').replace('.', ',').replace('X', '.')


def _mzn_extenso(valor):
    """Valor monetário por extenso em português: meticais e centavos."""
    try:
        total_cent = int(round(float(valor or 0) * 100))
    except Exception:
        total_cent = 0
    meticais = total_cent // 100
    centavos = total_cent % 100

    unidades = ['', 'um', 'dois', 'três', 'quatro', 'cinco', 'seis', 'sete', 'oito', 'nove']
    especiais = {
        10:'dez', 11:'onze', 12:'doze', 13:'treze', 14:'catorze', 15:'quinze', 16:'dezasseis',
        17:'dezassete', 18:'dezoito', 19:'dezanove'
    }
    dezenas = ['', '', 'vinte', 'trinta', 'quarenta', 'cinquenta', 'sessenta', 'setenta', 'oitenta', 'noventa']
    centenas = ['', 'cento', 'duzentos', 'trezentos', 'quatrocentos', 'quinhentos', 'seiscentos', 'setecentos', 'oitocentos', 'novecentos']

    def ate_999(n):
        n = int(n)
        if n == 0:
            return ''
        if n == 100:
            return 'cem'
        parts = []
        c, r = divmod(n, 100)
        if c:
            parts.append(centenas[c])
        if r:
            if r < 10:
                parts.append(unidades[r])
            elif r < 20:
                parts.append(especiais[r])
            else:
                d, u = divmod(r, 10)
                if u:
                    parts.append(dezenas[d] + ' e ' + unidades[u])
                else:
                    parts.append(dezenas[d])
        return ' e '.join([x for x in parts if x])

    def inteiro_ext(n):
        n = int(n)
        if n == 0:
            return 'zero'
        grupos = []
        escala = [('', ''), ('mil', 'mil'), ('milhão', 'milhões'), ('mil milhões', 'mil milhões'), ('bilião', 'biliões')]
        i = 0
        while n > 0:
            grupos.append(n % 1000)
            n //= 1000
            i += 1
        partes = []
        for idx in range(len(grupos)-1, -1, -1):
            g = grupos[idx]
            if not g:
                continue
            if idx == 1 and g == 1:
                partes.append('mil')
            elif idx > 0:
                sing, plur = escala[idx] if idx < len(escala) else ('', '')
                partes.append(ate_999(g) + ' ' + (sing if g == 1 else plur))
            else:
                partes.append(ate_999(g))
        return ' e '.join(partes)

    texto = inteiro_ext(meticais) + (' metical' if meticais == 1 else ' meticais')
    if centavos:
        texto += ' e ' + inteiro_ext(centavos) + (' centavo' if centavos == 1 else ' centavos')
    return texto.capitalize()


try:
    app.jinja_env.filters['fmt_mil'] = _fmt_mil
    app.jinja_env.filters['mzn_extenso'] = _mzn_extenso
except Exception:
    pass


def _ponta_faturavel_edm(pot_contratada, ponta_lida_corrigida):
    """
    Regra operacional da ponta faturável EDM MT usada no SGE:
    20% da potência contratada + 80% da ponta lida corrigida pelo fator multiplicativo.
    A potência contratada vem diretamente da configuração do Local e NÃO leva fator multiplicativo.
    A ponta lida já deve chegar aqui corrigida pelo fator multiplicativo.
    """
    return billing_demand(pot_contratada, ponta_lida_corrigida)



def _quantidades_fatura_mensal(local_nome: str, mes_str: str, ano_int: int):
    """
    Calcula as quantidades físicas de faturação a partir da planilha mensal.

    Regras aplicadas:
    - Leituras de ativa e reativa gravadas em leituras_mensais já estão corrigidas pelo fator multiplicativo.
    - Energia ativa do mês = última leitura ativa válida - leitura base inicial.
    - Energia reativa do mês = última leitura reativa válida - leitura base inicial.
    - Se existir leitura do mês anterior, ela é a base inicial.
    - Se não existir leitura do mês anterior, a primeira leitura válida do mês é apenas referência/base.
    - Ponta considerada = maior ponta corrigida registada no mês.
    - Diferenças negativas não são faturadas automaticamente; ficam sinalizadas para auditoria.
    """
    ensure_leituras_mensais_phase2_schema()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute("""
        SELECT data, ativa, reativa, ponta, agua, anterior
        FROM leituras_mensais
        WHERE local=? AND mes=? AND ano=?
          AND (ativa IS NOT NULL OR reativa IS NOT NULL OR ponta IS NOT NULL OR agua IS NOT NULL)
        ORDER BY data
    """, (local_nome, str(mes_str).zfill(2), int(ano_int))).fetchall()
    conn.close()

    prev_ativa, prev_reativa = get_prev_month_last_readings(local_nome, str(mes_str).zfill(2), int(ano_int))

    ativas = []
    reativas = []
    ponta_max = 0.0
    agua_total = 0.0
    avisos = []

    for r in rows:
        data = r['data']
        ativa = _safe_float(r['ativa'], None)
        reativa = _safe_float(r['reativa'], None)
        ponta = _safe_float(r['ponta'], None)
        agua = _safe_float(r['agua'], 0.0) or 0.0
        if ativa is not None and ativa > 0:
            ativas.append((data, ativa))
        if reativa is not None and reativa > 0:
            reativas.append((data, reativa))
        if ponta is not None and ponta > 0:
            ponta_max = max(ponta_max, ponta)
        agua_total += agua

    leitura_base_ativa = float(prev_ativa or 0.0) if prev_ativa and prev_ativa > 0 else (ativas[0][1] if ativas else 0.0)
    leitura_final_ativa = ativas[-1][1] if ativas else leitura_base_ativa
    kwh_ativa = leitura_final_ativa - leitura_base_ativa
    if kwh_ativa < 0:
        avisos.append('Leitura ativa final inferior à leitura base; consumo ativo faturável foi tratado como zero. Verificar leitura anterior, troca/reinício de contador ou fator multiplicativo.')
        kwh_ativa = 0.0

    leitura_base_reativa = float(prev_reativa or 0.0) if prev_reativa and prev_reativa > 0 else (reativas[0][1] if reativas else 0.0)
    leitura_final_reativa = reativas[-1][1] if reativas else leitura_base_reativa
    kvarh_reativa = leitura_final_reativa - leitura_base_reativa
    if kvarh_reativa < 0:
        avisos.append('Leitura reativa final inferior à leitura base; consumo reativo faturável foi tratado como zero. Verificar leitura anterior, troca/reinício de contador ou fator multiplicativo.')
        kvarh_reativa = 0.0

    limite_reativa = 0.75 * kwh_ativa
    kvarh_excedente = max(kvarh_reativa - limite_reativa, 0.0)
    consumo_especifico = (kwh_ativa / agua_total) if agua_total > 0 else None

    return {
        'rows': rows,
        'kwh_ativa': kwh_ativa,
        'kvarh_reativa': kvarh_reativa,
        'limite_reativa': limite_reativa,
        'kvarh_excedente': kvarh_excedente,
        'kw_ponta_lida': ponta_max,
        'agua_total': agua_total,
        'consumo_especifico': consumo_especifico,
        'leitura_base_ativa': leitura_base_ativa,
        'leitura_final_ativa': leitura_final_ativa,
        'leitura_base_reativa': leitura_base_reativa,
        'leitura_final_reativa': leitura_final_reativa,
        'tem_base_mes_anterior_ativa': bool(prev_ativa and prev_ativa > 0),
        'tem_base_mes_anterior_reativa': bool(prev_reativa and prev_reativa > 0),
        'avisos': avisos,
    }



# === Arquivo interno de faturas mensais e PDF em 1 página ===
def ensure_faturas_mensais_archive_schema():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS faturas_mensais_arquivo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            local TEXT NOT NULL,
            mes TEXT NOT NULL,
            ano INTEGER NOT NULL,
            periodo TEXT,
            total REAL DEFAULT 0,
            subtotal REAL DEFAULT 0,
            kwh_ativa REAL DEFAULT 0,
            kvarh_excedente REAL DEFAULT 0,
            demanda_ponta_kw REAL DEFAULT 0,
            agua_total REAL DEFAULT 0,
            consumo_especifico REAL,
            snapshot_json TEXT,
            criado_em TEXT DEFAULT (datetime('now','localtime')),
            atualizado_em TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(local, mes, ano)
        )
    ''')
    conn.commit(); conn.close()


def ensure_leituras_mensais_status_schema():
    """Tabela leve para controlo operacional do mês: aberto/fechado.
    Não altera leituras existentes. Serve para evitar que uma fatura já conferida
    seja modificada por engano sem reabertura consciente do período.
    """
    ensure_leituras_mensais_phase2_schema()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS leituras_mensais_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            local TEXT NOT NULL,
            mes TEXT NOT NULL,
            ano INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'aberto',
            fechado_em TEXT,
            fechado_por TEXT,
            reaberto_em TEXT,
            reaberto_por TEXT,
            observacao TEXT,
            atualizado_em TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(local, mes, ano)
        )
    ''')
    conn.commit(); conn.close()


def _get_periodo_status(local, mes, ano):
    ensure_leituras_mensais_status_schema()
    mes = str(mes).zfill(2); ano = int(ano)
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    row = c.execute('SELECT * FROM leituras_mensais_status WHERE local=? AND mes=? AND ano=?', (local, mes, ano)).fetchone()
    conn.close()
    if not row:
        return {'status':'aberto', 'fechado_em':'', 'fechado_por':'', 'reaberto_em':'', 'reaberto_por':'', 'observacao':'', 'fechado':False}
    d = dict(row); d['fechado'] = str(d.get('status') or '').lower() == 'fechado'
    return d


def _set_periodo_status(local, mes, ano, status='aberto', actor='operador', observacao=''):
    ensure_leituras_mensais_status_schema()
    mes = str(mes).zfill(2); ano = int(ano); status = (status or 'aberto').lower()
    now_sql = "datetime('now','localtime')"
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    if status == 'fechado':
        c.execute(f'''
            INSERT INTO leituras_mensais_status(local, mes, ano, status, fechado_em, fechado_por, observacao, atualizado_em)
            VALUES(?, ?, ?, 'fechado', {now_sql}, ?, ?, {now_sql})
            ON CONFLICT(local, mes, ano) DO UPDATE SET
                status='fechado', fechado_em={now_sql}, fechado_por=excluded.fechado_por,
                observacao=excluded.observacao, atualizado_em={now_sql}
        ''', (local, mes, ano, actor, observacao))
    else:
        c.execute(f'''
            INSERT INTO leituras_mensais_status(local, mes, ano, status, reaberto_em, reaberto_por, observacao, atualizado_em)
            VALUES(?, ?, ?, 'aberto', {now_sql}, ?, ?, {now_sql})
            ON CONFLICT(local, mes, ano) DO UPDATE SET
                status='aberto', reaberto_em={now_sql}, reaberto_por=excluded.reaberto_por,
                observacao=excluded.observacao, atualizado_em={now_sql}
        ''', (local, mes, ano, actor, observacao))
    conn.commit(); conn.close()


def _validar_periodo_mensal_operacional(local, mes, ano, pot_contratada=0, fator_mult=1):
    """Valida rapidamente o mês para orientar operador antes de gerar/fechar fatura."""
    mes = str(mes).zfill(2); ano = int(ano)
    qfat = _quantidades_fatura_mensal(local, mes, ano) if local else {}
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute('''SELECT data, ativa, reativa, ponta, fp, diferenca, agua
                        FROM leituras_mensais WHERE local=? AND mes=? AND ano=? ORDER BY data''', (local, mes, ano)).fetchall()
    conn.close()
    total_dias = calendar.monthrange(ano, int(mes))[1]
    dias_preenchidos = 0
    dias_fp_baixo = 0
    quedas_ativa = 0
    quedas_reativa = 0
    dias_sem_agua = 0
    ultima_ativa = None; ultima_reativa = None
    for r in rows:
        ativa = _safe_float(r['ativa'], None); reativa = _safe_float(r['reativa'], None)
        ponta = _safe_float(r['ponta'], None); agua = _safe_float(r['agua'], 0) or 0
        fp = _safe_float(r['fp'], None)
        tem_linha = any(v is not None and v != 0 for v in [ativa, reativa, ponta, agua])
        if tem_linha:
            dias_preenchidos += 1
            if agua <= 0:
                dias_sem_agua += 1
        if fp is not None and fp > 0 and fp < 0.85:
            dias_fp_baixo += 1
        if ativa is not None and ativa > 0:
            if ultima_ativa is not None and ativa < ultima_ativa:
                quedas_ativa += 1
            ultima_ativa = max(ultima_ativa or ativa, ativa)
        if reativa is not None and reativa > 0:
            if ultima_reativa is not None and reativa < ultima_reativa:
                quedas_reativa += 1
            ultima_reativa = max(ultima_reativa or reativa, reativa)

    avisos = list(qfat.get('avisos', []))
    criticos = []
    if dias_preenchidos == 0:
        criticos.append('Ainda não existem leituras preenchidas para este período.')
    if qfat.get('kwh_ativa', 0) <= 0 and dias_preenchidos > 1:
        criticos.append('Consumo ativo faturável igual a zero. Verificar leitura base, leitura final ou fator multiplicativo.')
    if not qfat.get('tem_base_mes_anterior_ativa'):
        avisos.append('Não foi encontrada base ativa do mês anterior. A primeira leitura válida do mês atual será tratada como base inicial.')
    if not qfat.get('tem_base_mes_anterior_reativa'):
        avisos.append('Não foi encontrada base reativa do mês anterior. A primeira leitura válida do mês atual será tratada como base inicial.')
    if qfat.get('kw_ponta_lida', 0) <= 0:
        avisos.append('Ponta máxima do mês ainda não foi registada.')
    if qfat.get('agua_total', 0) <= 0:
        avisos.append('Água elevada total está zerada; o consumo específico mensal não será representativo.')
    if quedas_ativa:
        criticos.append(f'Foram encontradas {quedas_ativa} ocorrência(s) de leitura ativa inferior à leitura válida anterior.')
    if quedas_reativa:
        criticos.append(f'Foram encontradas {quedas_reativa} ocorrência(s) de leitura reativa inferior à leitura válida anterior.')
    if dias_fp_baixo:
        avisos.append(f'{dias_fp_baixo} dia(s) com fator de potência abaixo de 0,85.')
    if dias_sem_agua and qfat.get('agua_total', 0) > 0:
        avisos.append(f'{dias_sem_agua} dia(s) preenchido(s) sem volume de água registado.')

    return {
        'ok_para_faturar': len(criticos) == 0,
        'criticos': criticos,
        'avisos': avisos,
        'dias_preenchidos': dias_preenchidos,
        'total_dias': total_dias,
        'progresso_pct': round((dias_preenchidos / total_dias) * 100, 1) if total_dias else 0,
        'qfat': qfat,
        'dias_fp_baixo': dias_fp_baixo,
        'quedas_ativa': quedas_ativa,
        'quedas_reativa': quedas_reativa,
    }


def _arquivar_fatura_mensal_snapshot(ctx):
    try:
        ensure_faturas_mensais_archive_schema()
        local = str(ctx.get('local') or '').strip()
        periodo = str(ctx.get('periodo') or '')
        mes = periodo.split('/')[0].zfill(2) if '/' in periodo else str(ctx.get('mes') or '').zfill(2)
        ano = int(periodo.split('/')[1]) if '/' in periodo else int(ctx.get('ano') or 0)
        if not local or not mes or not ano:
            return None
        serializable = {}
        for k, v in ctx.items():
            if k in ('request', 'qfat'):
                continue
            try:
                json.dumps(v)
                serializable[k] = v
            except Exception:
                serializable[k] = str(v)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            INSERT INTO faturas_mensais_arquivo
            (local, mes, ano, periodo, total, subtotal, kwh_ativa, kvarh_excedente, demanda_ponta_kw, agua_total, consumo_especifico, snapshot_json)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(local, mes, ano) DO UPDATE SET
                periodo=excluded.periodo,
                total=excluded.total,
                subtotal=excluded.subtotal,
                kwh_ativa=excluded.kwh_ativa,
                kvarh_excedente=excluded.kvarh_excedente,
                demanda_ponta_kw=excluded.demanda_ponta_kw,
                agua_total=excluded.agua_total,
                consumo_especifico=excluded.consumo_especifico,
                snapshot_json=excluded.snapshot_json,
                atualizado_em=datetime('now','localtime')
        ''', (
            local, mes, ano, periodo,
            float(ctx.get('total') or 0), float(ctx.get('subtotal') or 0),
            float(ctx.get('kwh_ativa') or 0), float(ctx.get('kvarh_excedente') or 0),
            float(ctx.get('demanda_ponta_kw') or 0), float(ctx.get('agua_total') or 0),
            ctx.get('consumo_especifico_medio'),
            json.dumps(serializable, ensure_ascii=False)
        ))
        conn.commit()
        rid = c.execute('SELECT id FROM faturas_mensais_arquivo WHERE local=? AND mes=? AND ano=?', (local, mes, ano)).fetchone()[0]
        conn.close()
        return rid
    except Exception as e:
        print('Falha ao arquivar fatura mensal:', e)
        return None


def _montar_contexto_fatura_mensal(local, mes, ano):
    mes_str = str(mes).zfill(2)
    ano_int = int(ano)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT id, nome FROM locais WHERE nome = ?', (local,))
    row_loc = c.fetchone()
    local_id = row_loc['id'] if row_loc else None
    fator_mult = 1.0
    if local_id is not None:
        try:
            row_cfg = c.execute('SELECT fator_mult FROM locais_cfg WHERE local_id=?', (local_id,)).fetchone()
            fator_mult = float(row_cfg[0] or 1.0) if row_cfg else 1.0
        except Exception:
            fator_mult = 1.0
    tarifas = resolve_tariffs(conn, local_id, f"{ano_int:04d}-{int(mes_str):02d}-01")
    conn.close()
    qfat = _quantidades_fatura_mensal(local, mes_str, ano_int)
    kwh_ativa = qfat['kwh_ativa']
    kvarh_reativa = qfat['kvarh_reativa']
    limite_reativa = qfat['limite_reativa']
    kvarh_excedente = qfat['kvarh_excedente']
    kw_ponta_lida = qfat['kw_ponta_lida']
    agua_total = qfat['agua_total']
    fatura = calculate_invoice(
        active_kwh=kwh_ativa,
        reactive_kvarh=kvarh_reativa,
        measured_peak_kw=kw_ponta_lida,
        contracted_power_kw=tarifas.get('pot_contratada'),
        tariffs=tarifas,
        bill_losses=False,
    )
    consumo_especifico_medio = (kwh_ativa / agua_total) if agua_total and kwh_ativa > 0 else None
    return dict(
        local=local, mes=mes_str, ano=ano_int, periodo=f"{mes_str}/{ano_int}",
        fator_mult=fator_mult, pot_contratada=fatura['contracted_power_kw'],
        tarifa_ativa=tarifas['tarifa_ativa'], tarifa_reativa=tarifas['tarifa_reativa'],
        tarifa_ponta=tarifas['tarifa_ponta'], tarifa_perdas=tarifas['tarifa_perdas'],
        taxa_fixa=tarifas['taxa_fixa'], taxa_radio=tarifas['taxa_radio'], taxa_lixo=tarifas['taxa_lixo'],
        tarifa_fonte=tarifas.get('source'), tarifa_valid_from=tarifas.get('valid_from'),
        kwh_ativa=kwh_ativa, kvarh_reativa=kvarh_reativa, limite_reativa=limite_reativa,
        kvarh_excedente=fatura['reactive_excess_kvarh'], kw_ponta_lida=kw_ponta_lida,
        demanda_ponta_kw=fatura['billing_demand_kw'],
        valor_ativa=fatura['active_cost_mzn'], valor_reativa=fatura['reactive_cost_mzn'],
        valor_ponta=fatura['demand_cost_mzn'], valor_perdas=fatura['losses_cost_mzn'],
        subtotal_energia=fatura['energy_subtotal_mzn'], subtotal_taxas=fatura['fees_subtotal_mzn'],
        subtotal=fatura['subtotal_mzn'], base_iva=fatura['vat_base_mzn'],
        valor_iva=fatura['vat_mzn'], total=fatura['total_mzn'],
        total_extenso=_mzn_extenso(fatura['total_mzn']),
        consumo_especifico_medio=consumo_especifico_medio, agua_total=agua_total,
        iva_percent=fatura['vat_rate_percent'], base_iva_percent=fatura['vat_base_percent'],
        leitura_base_ativa=qfat.get('leitura_base_ativa', 0),
        leitura_final_ativa=qfat.get('leitura_final_ativa', 0),
        leitura_base_reativa=qfat.get('leitura_base_reativa', 0),
        leitura_final_reativa=qfat.get('leitura_final_reativa', 0),
        avisos_fatura=qfat.get('avisos', []), qfat=qfat, fatura=fatura
    )

@app.route('/leituras_mensal', methods=['GET', 'POST'])
def leituras_mensal():
    ensure_leituras_mensais_phase2_schema()
    hoje = datetime.now()
    locais_db = get_locais()
    meses = [(str(i).zfill(2), calendar.month_name[i]) for i in range(1, 13)]

    if not locais_db:
        flash("Nenhum local configurado. Primeiro cadastre pelo menos um Local.", "warning")
        return render_template('leituras_mensal.html', locais_db=[], selected_local_id=None, local='', meses=meses,
                               mes=hoje.strftime('%m'), ano=hoje.year, dias=[], leituras={}, fator_mult=1.0,
                               pot_contratada=0.0, pot_instalada=0.0, fp_medio=0.0, pot_max_ponta=0.0,
                               cfg_selected={}, cfg_map={}, first_prev_ativa=0.0, first_prev_reativa=0.0,
                               resumo={})

    # Leitura dos filtros: em POST, prioriza o formulário; em GET, usa a URL.
    # Isto evita o problema de a página manter o mês/ano antigo quando a URL
    # ainda trazia parâmetros anteriores.
    if request.method == 'POST':
        raw_local = request.form.get('local', str(locais_db[0][0]))
        mes_req = request.form.get('mes')
        ano_req = request.form.get('ano')
    else:
        raw_local = request.args.get('local', str(locais_db[0][0]))
        mes_req = request.args.get('mes')
        ano_req = request.args.get('ano')

    selected_local_id, local_nome = _local_id_nome_from_request(raw_local, locais_db)
    mes = (mes_req or hoje.strftime('%m')).zfill(2)
    ano = int(ano_req or hoje.year)

    cfg_selected = get_local_cfg_full(selected_local_id) if selected_local_id else {}
    fator_mult = float(cfg_selected.get('fator_mult') or 1.0)
    pot_contratada = float(cfg_selected.get('pot_contratada') or 0.0)
    pot_instalada = float(cfg_selected.get('pot_instalada') or 0.0)

    num_dias = calendar.monthrange(ano, int(mes))[1]
    dias = [f"{ano}-{mes}-{str(dia).zfill(2)}" for dia in range(1, num_dias + 1)]

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    rows_db = c.execute("""
        SELECT data,hora,ativa,reativa,ponta,fp,potc,anterior,atual,
               diferenca,agua,esp,acum,valor
        FROM leituras_mensais
        WHERE local=? AND mes=? AND ano=?
        ORDER BY data
    """, (local_nome, mes, ano)).fetchall()
    conn.close()

    leituras_map = {}
    for r in rows_db:
        fator = fator_mult if fator_mult else 1.0
        ativa_lida = ((r[2] or 0) / fator if r[2] not in (None, '') else '')
        reativa_lida = ((r[3] or 0) / fator if r[3] not in (None, '') else '')
        ponta_lida = ((r[4] or 0) / fator if r[4] not in (None, '') else '')
        leituras_map[r[0]] = {
            'hora': r[1] or '',
            'ativa': r[2] if r[2] is not None else '',
            'reativa': r[3] if r[3] is not None else '',
            'ponta': r[4] if r[4] is not None else '',
            'fp': r[5] if r[5] is not None else '',
            'potc': r[6] if r[6] is not None else pot_contratada,
            'anterior': r[7] if r[7] is not None else '',
            'atual': r[8] if r[8] is not None else '',
            'diferenca': r[9] if r[9] is not None else '',
            'agua': r[10] if r[10] is not None else '',
            'esp': r[11] if r[11] is not None else '',
            'acum': r[12] if r[12] is not None else '',
            'valor': r[13] if r[13] is not None else '',
            'ativa_lida': ativa_lida,
            'reativa_lida': reativa_lida,
            'ponta_lida': ponta_lida,
            'reativa_excedente': 0,
            'valor_ativa': 0,
            'valor_reativa': 0,
            'valor_ponta': 0,
            'valor_total_dia': r[13] or 0,
        }

    fp_vals = []
    pot_max_ponta = 0.0
    resumo = {'kwh_total':0.0, 'kvarh_total':0.0, 'ponta_max':0.0, 'agua_total':0.0, 'consumo_especifico':0.0,
              'valor_total':0.0, 'reativa_excedente':0.0, 'dias_preenchidos':0}
    for d in dias:
        row = leituras_map.get(d)
        if not row:
            continue
        resumo['dias_preenchidos'] += 1
        dif = _safe_float(row.get('diferenca'), 0.0) or 0.0
        agua = _safe_float(row.get('agua'), 0.0) or 0.0
        val = _safe_float(row.get('valor_total_dia'), _safe_float(row.get('valor'), 0.0)) or 0.0
        rea_exc = _safe_float(row.get('reativa_excedente'), 0.0) or 0.0
        ponta = _safe_float(row.get('ponta'), 0.0) or 0.0
        resumo['kwh_total'] += dif
        resumo['agua_total'] += agua
        resumo['valor_total'] += val
        resumo['reativa_excedente'] += rea_exc
        resumo['ponta_max'] = max(resumo['ponta_max'], ponta)
        try:
            if row['fp'] not in ('', None):
                fp_vals.append(float(row['fp']))
        except Exception:
            pass
    resumo['consumo_especifico'] = (resumo['kwh_total'] / resumo['agua_total']) if resumo['agua_total'] else 0.0
    fp_medio = round(sum(fp_vals) / len(fp_vals), 3) if fp_vals else 0.0
    pot_max_ponta = resumo['ponta_max']
    first_prev_ativa, first_prev_reativa = get_prev_month_last_readings(local_nome, mes, ano)
    periodo_status = _get_periodo_status(local_nome, mes, ano)
    validacao_periodo = _validar_periodo_mensal_operacional(local_nome, mes, ano, pot_contratada, fator_mult)

    return render_template('leituras_mensal.html', locais_db=locais_db, selected_local_id=selected_local_id,
                           local=local_nome, meses=meses, mes=mes, ano=ano, dias=dias, leituras=leituras_map,
                           fator_mult=fator_mult, pot_contratada=pot_contratada, pot_instalada=pot_instalada,
                           fp_medio=fp_medio, pot_max_ponta=pot_max_ponta, cfg_selected=cfg_selected,
                           cfg_map={}, first_prev_ativa=first_prev_ativa,
                           first_prev_reativa=first_prev_reativa, resumo=resumo,
                           periodo_status=periodo_status, validacao_periodo=validacao_periodo)


def get_prev_month_last_readings(local_nome: str, mes: str, ano: int):
    """Última leitura faturada de ativa/reativa do mês anterior."""
    ensure_leituras_mensais_phase2_schema()
    mes_int = int(mes); ano_int = int(ano)
    prev_mes = mes_int - 1; prev_ano = ano_int
    if prev_mes == 0:
        prev_mes = 12; prev_ano = ano_int - 1
    prev_mes_str = str(prev_mes).zfill(2)
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    row = c.execute("""
        SELECT ativa, reativa
        FROM leituras_mensais
        WHERE local=? AND mes=? AND ano=? AND ativa IS NOT NULL
        ORDER BY data DESC
        LIMIT 1
    """, (local_nome, prev_mes_str, prev_ano)).fetchone()
    conn.close()
    if row:
        return float(row[0] or 0), float(row[1] or 0)
    return 0.0, 0.0


@app.route('/leituras_mensal_salvar', methods=['POST'])
def leituras_mensal_salvar():
    ensure_leituras_mensais_phase2_schema()
    hoje = datetime.now()
    locais_db = get_locais()
    raw_local = (request.form.get('local') or '').strip()
    selected_local_id, local_nome = _local_id_nome_from_request(raw_local, locais_db)
    mes = (request.form.get('mes') or hoje.strftime('%m')).zfill(2)
    ano = int(request.form.get('ano') or hoje.year)
    cfg = get_local_cfg_full(selected_local_id) if selected_local_id else {}
    fator_mult = _safe_float(request.form.get('fator_mult'), cfg.get('fator_mult', 1.0)) or 1.0
    pot_contratada = _safe_float(cfg.get('pot_contratada'), 0.0) or 0.0
    t_ativa = _safe_float(cfg.get('tarifa_ativa'), 0.0) or 0.0
    t_reativa = _safe_float(cfg.get('tarifa_reativa'), 0.0) or 0.0
    t_ponta = _safe_float(cfg.get('tarifa_ponta'), 0.0) or 0.0
    acao = (request.form.get('acao') or '').strip()
    if _get_periodo_status(local_nome, mes, ano).get('fechado'):
        flash('Este mês está FECHADO. Para alterar leituras, reabra o período de forma controlada.', 'warning')
        return redirect(url_for('leituras_mensal', local=local_nome, mes=mes, ano=ano))
    num_dias = calendar.monthrange(ano, int(mes))[1]

    prev_ativa, prev_reativa = get_prev_month_last_readings(local_nome, mes, ano)
    # Se não existir leitura do mês anterior, a primeira leitura preenchida do mês
    # passa a ser a linha de base. Isto evita faturar indevidamente o valor
    # acumulado histórico do contador como se fosse consumo do mês.
    has_prev_ativa = bool(prev_ativa and prev_ativa > 0)
    has_prev_reativa = bool(prev_reativa and prev_reativa > 0)
    prev_ponta_corrigida = 0.0
    acum_mes = 0.0
    linhas_processadas = 0

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    for i in range(num_dias):
        data_str = request.form.get(f"data_{i}")
        if not data_str:
            continue
        hora = request.form.get(f"hora_{i}") or ""
        ativa_lida = _safe_float(request.form.get(f"ativa_lida_{i}"), None)
        reativa_lida = _safe_float(request.form.get(f"reativa_lida_{i}"), None)
        ponta_lida = _safe_float(request.form.get(f"ponta_lida_{i}"), None)
        agua_val = _safe_float(request.form.get(f"agua_{i}"), None)
        # compatibilidade com template antigo
        if ativa_lida is None:
            ativa_lida = _safe_float(request.form.get(f"ativa_{i}"), None)
        if reativa_lida is None:
            reativa_lida = _safe_float(request.form.get(f"reativa_{i}"), None)
        if ponta_lida is None:
            ponta_lida = _safe_float(request.form.get(f"ponta_{i}"), None)

        if ativa_lida is None and reativa_lida is None and ponta_lida is None and agua_val is None:
            continue

        ativa_fat = (ativa_lida * fator_mult) if ativa_lida is not None else None
        reativa_fat = (reativa_lida * fator_mult) if reativa_lida is not None else None
        ponta_fat = (ponta_lida * fator_mult) if ponta_lida is not None else None
        if ponta_fat is not None:
            # A ponta lida é tratada como registo máximo mensal: pode manter-se ou aumentar, nunca reduzir.
            if ponta_fat < prev_ponta_corrigida:
                ponta_fat = prev_ponta_corrigida
            else:
                prev_ponta_corrigida = ponta_fat
        # Energia ativa: contador acumulativo. Consumo do dia = leitura atual - leitura anterior.
        # Se não houver leitura anterior do mês anterior, a primeira leitura do mês é apenas referência inicial.
        anterior_val = prev_ativa if has_prev_ativa else (ativa_fat if ativa_fat is not None else prev_ativa)
        atual_val = ativa_fat if ativa_fat is not None else prev_ativa
        if ativa_fat is not None and has_prev_ativa:
            dif_val = atual_val - anterior_val
        else:
            dif_val = 0.0
        if dif_val < 0:
            # Leitura menor que a anterior: não fatura consumo negativo e NÃO atualiza
            # a referência anterior. Assim, o próximo dia válido continua a comparar
            # contra a última leitura correta do mês anterior ou do dia anterior válido.
            dif_operacional = 0.0
        else:
            dif_operacional = dif_val
        if ativa_fat is not None:
            if not has_prev_ativa:
                prev_ativa = atual_val
                has_prev_ativa = True
            elif dif_val >= 0:
                prev_ativa = atual_val

        # Energia reativa: mesma lógica da ativa. Nunca se deve usar a leitura acumulada total
        # como reativa excedente; usa-se apenas a diferença mensal/diária. Se a leitura recuar,
        # mantém-se a última referência válida para não contaminar os dias seguintes.
        if reativa_fat is not None:
            if has_prev_reativa:
                delta_reativa_real = reativa_fat - prev_reativa
            else:
                delta_reativa_real = 0.0
            if delta_reativa_real < 0:
                delta_reativa = 0.0
            else:
                delta_reativa = delta_reativa_real
            if not has_prev_reativa:
                prev_reativa = reativa_fat
                has_prev_reativa = True
            elif delta_reativa_real >= 0:
                prev_reativa = reativa_fat
        else:
            delta_reativa = 0.0

        if dif_operacional > 0 or delta_reativa > 0:
            fp_val = dif_operacional / math.sqrt((dif_operacional ** 2) + (delta_reativa ** 2)) if (dif_operacional or delta_reativa) else None
        else:
            fp_val = None
        reativa_excedente = max(delta_reativa - (0.75 * dif_operacional), 0.0)
        esp_val = (dif_operacional / agua_val) if agua_val else None
        acum_mes += dif_operacional
        valor_ativa = dif_operacional * t_ativa
        valor_reativa = reativa_excedente * t_reativa
        valor_ponta = 0.0  # A ponta é uma cobrança mensal: 20% PC + 80% ponta máxima corrigida. Não é somada por dia.
        valor_total_dia = valor_ativa + valor_reativa

        c.execute("""
            INSERT INTO leituras_mensais
            (local, data, hora, ativa, reativa, ponta, fp, potc, anterior, atual, diferenca,
             agua, esp, acum, valor, mes, ano)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(local, data) DO UPDATE SET
                hora=excluded.hora,
                ativa=excluded.ativa,
                reativa=excluded.reativa,
                ponta=excluded.ponta,
                fp=excluded.fp,
                potc=excluded.potc,
                anterior=excluded.anterior,
                atual=excluded.atual,
                diferenca=excluded.diferenca,
                agua=excluded.agua,
                esp=excluded.esp,
                acum=excluded.acum,
                valor=excluded.valor,
                mes=excluded.mes,
                ano=excluded.ano
        """, (local_nome, data_str, hora, ativa_fat, reativa_fat, ponta_fat, fp_val, pot_contratada,
              anterior_val, atual_val, dif_operacional, agua_val, esp_val, acum_mes, valor_total_dia,
              mes, ano))
        linhas_processadas += 1

    conn.commit(); conn.close()
    flash(f"Leituras mensais salvas e calculadas automaticamente ({linhas_processadas} linhas processadas).", "success")
    if acao == 'fatura_edm':
        return redirect(url_for('leituras_mensal_fatura_edm', local=local_nome, mes=mes, ano=str(ano)))
    return redirect(url_for('leituras_mensal', local=local_nome, mes=mes, ano=ano))


@app.route('/leituras_mensal/status', methods=['POST'])
def leituras_mensal_status_periodo():
    ensure_leituras_mensais_status_schema()
    local = (request.form.get('local') or request.args.get('local') or '').strip()
    mes = (request.form.get('mes') or request.args.get('mes') or datetime.now().strftime('%m')).zfill(2)
    ano = int(request.form.get('ano') or request.args.get('ano') or datetime.now().year)
    operacao = (request.form.get('operacao') or '').strip().lower()
    obs = (request.form.get('observacao') or '').strip()
    actor = session.get('username') if 'session' in globals() else None
    actor = actor or 'operador'
    if not local:
        flash('Selecione um local antes de alterar o estado do período.', 'warning')
        return redirect(url_for('leituras_mensal'))
    if operacao == 'fechar':
        validacao = _validar_periodo_mensal_operacional(local, mes, ano)
        if not validacao.get('ok_para_faturar'):
            flash('O mês não foi fechado porque existem inconsistências críticas. Abra a Auditoria Inteligente para corrigir.', 'danger')
            return redirect(url_for('leituras_mensal_audit', local=local, mes=mes, ano=ano))
        ctx = _montar_contexto_fatura_mensal(local, mes, ano)
        _arquivar_fatura_mensal_snapshot(ctx)
        _set_periodo_status(local, mes, ano, 'fechado', actor=actor, observacao=obs)
        flash('Período fechado com sucesso. A fatura foi arquivada e as leituras ficaram protegidas contra alterações acidentais.', 'success')
    elif operacao == 'reabrir':
        _set_periodo_status(local, mes, ano, 'aberto', actor=actor, observacao=obs)
        flash('Período reaberto. Pode alterar e salvar as leituras novamente.', 'info')
    else:
        flash('Operação de estado inválida.', 'warning')
    return redirect(url_for('leituras_mensal', local=local, mes=mes, ano=ano))


@app.route('/leituras_mensal/visualizar')
def visualizar_mensal():
    locais = [l[1] for l in get_locais()]
    mes = request.args.get('mes') or datetime.now().strftime('%Y-%m')
    # aceitar formatos YYYY-MM ou MM
    if len(mes)==2:
        mes_val = mes
        ano_val = int(request.args.get('ano') or datetime.now().year)
        mes_str = f"{ano_val}-{mes_val}"
    else:
        ano_val, mes_val = mes.split('-')
        ano_val = int(ano_val)
    local = request.args.get('local', locais[0] if locais else '')
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute('''SELECT data, diferenca, ativa, reativa, ponta, fp FROM leituras_mensais
                        WHERE local=? AND mes=? AND ano=? ORDER BY data''',
                     (local, mes_val, ano_val)).fetchall()
    conn.close()
    # converter para listas de plot
    dias = [r[0] for r in rows]
    difs = [r[1] or 0 for r in rows]
    ativa = [r[2] or 0 for r in rows]
    reativa = [r[3] or 0 for r in rows]
    ponta = [r[4] or 0 for r in rows]
    fp = [r[5] or 0 for r in rows]
    return render_template('visualizar_mensal.html', locais=locais, local=local, mes=f"{ano_val}-{mes_val}", dias=dias,
                           difs=difs, ativas=ativa, reativas=reativa, pontas=ponta, fps=fp)

# === Importar Leituras Mensais Históricas (Excel/CSV inteligente) ===

def _import_norm_txt(v):
    import unicodedata, re
    s = '' if v is None else str(v).strip().lower()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = s.replace('\xa0', ' ')
    s = re.sub(r'[^a-z0-9]+', '_', s).strip('_')
    return s


def _import_float(v, default=None):
    if v is None or v == '':
        return default
    try:
        if isinstance(v, str):
            s = v.strip().replace('\xa0', '').replace(' ', '')
            if not s or s.upper() in ('#DIV/0!', '#VALUE!', '#REF!', '#N/A', 'N/A'):
                return default
            if ',' in s and '.' not in s:
                s = s.replace(',', '.')
            return float(s)
        return float(v)
    except Exception:
        return default


def _import_parse_data(v):
    from datetime import datetime, date, timedelta
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
        except Exception:
            return None
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _import_month_year_from_sheet(sheet_name):
    import re
    key = _import_norm_txt(sheet_name).replace('_', '')
    mapa = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
            'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}
    mes = None
    for pref, num in mapa.items():
        if key.startswith(pref):
            mes = num
            break
    m = re.search(r'(20\d{2}|\d{2})', key)
    ano = None
    if m:
        ano_raw = int(m.group(1))
        ano = ano_raw if ano_raw > 100 else 2000 + ano_raw
    return mes, ano


def _import_find_umbeluzi_local(locais_db):
    """Localiza o registo oficial da ETA de Umbeluzi na tabela de locais.
    Isto evita importar folhas históricas ETAU para um centro errado.
    """
    for lid, lname in (locais_db or []):
        if 'umbeluzi' in _import_norm_txt(lname):
            return lid, lname
    for lid, lname in (locais_db or []):
        n = _import_norm_txt(lname)
        if n in ('etau', 'eta_u', 'eta_umbeluzi', 'eta_de_umbeluzi'):
            return lid, lname
    return None, None


def _import_is_eta_umbeluzi_file(filename, linhas=None):
    """Detecta ficheiros/linhas do histórico ETAU/ETA Umbeluzi."""
    markers = ('etau', 'umbeluzi', 'eta_umbeluzi', 'eta_de_umbeluzi')
    filename_key = _import_norm_txt(filename)
    if any(m in filename_key for m in markers):
        return True
    for r in (linhas or [])[:120]:
        campos = ' '.join(str(r.get(k, '') or '') for k in ('local', 'origem', 'folha', 'origem_ficheiro'))
        key = _import_norm_txt(campos)
        if any(m in key for m in markers):
            return True
    return False


def _import_find_col(headers, aliases):
    norm_aliases = {_import_norm_txt(a) for a in aliases}
    for idx, h in enumerate(headers):
        if h in norm_aliases:
            return idx
    for idx, h in enumerate(headers):
        for a in norm_aliases:
            if a and (h == a or h.endswith('_' + a) or a in h):
                return idx
    return None


def _import_read_uploaded_workbook(file_bytes, filename, default_local):
    import csv as _csv
    from collections import Counter
    rows, issues, resumo = [], [], []
    filename_l = (filename or '').lower()

    def add_row(local, data_obj, hora, ativa_lida, reativa_lida, ponta_lida, agua, origem, folha,
                ativa_anterior=None, status='OK', observacao=''):
        if not data_obj:
            issues.append({'nivel': 'erro', 'folha': folha, 'data': '', 'mensagem': 'Linha sem data válida.'})
            return
        ano, mes = int(data_obj.year), int(data_obj.month)
        row = {
            'local': (str(local or default_local or '').strip()),
            'data': data_obj.strftime('%Y-%m-%d'),
            'ano': ano,
            'mes': str(mes).zfill(2),
            'hora': str(hora or '').strip(),
            'ativa_lida': _import_float(ativa_lida, None),
            'reativa_lida': _import_float(reativa_lida, None),
            'ponta_lida': _import_float(ponta_lida, None),
            'agua': _import_float(agua, None),
            'ativa_anterior': _import_float(ativa_anterior, None),
            'status': str(status or 'OK').strip().upper(),
            'observacao': str(observacao or ''),
            'origem': origem or filename,
            'folha': folha or '',
        }
        if row['status'] in ('NAO_IMPORTAR', 'NÃO_IMPORTAR', 'IGNORAR'):
            issues.append({'nivel': 'ignorado', 'folha': folha, 'data': row['data'], 'mensagem': 'Linha marcada para não importar.'})
            return
        if row['ativa_lida'] is None:
            issues.append({'nivel': 'erro', 'folha': folha, 'data': row['data'], 'mensagem': 'Sem leitura ativa lida; linha não importada.'})
            return
        if row['reativa_lida'] is None:
            issues.append({'nivel': 'aviso', 'folha': folha, 'data': row['data'], 'mensagem': 'Sem leitura reativa; será gravada vazia/zero no cálculo.'})
        if row['ponta_lida'] is None:
            issues.append({'nivel': 'aviso', 'folha': folha, 'data': row['data'], 'mensagem': 'Sem ponta lida.'})
        if row['agua'] is None:
            issues.append({'nivel': 'aviso', 'folha': folha, 'data': row['data'], 'mensagem': 'Sem água elevada.'})
        if row['ponta_lida'] is not None and row['ponta_lida'] > 1:
            issues.append({'nivel': 'aviso', 'folha': folha, 'data': row['data'], 'mensagem': f"Ponta elevada ({row['ponta_lida']}). Confirmar antes de faturar."})
        rows.append(row)

    if filename_l.endswith('.csv'):
        raw = file_bytes.decode('utf-8-sig', errors='ignore')
        dialect = _csv.Sniffer().sniff(raw[:2048], delimiters=';,') if raw.strip() else None
        reader = _csv.DictReader(io.StringIO(raw), dialect=dialect) if dialect else _csv.DictReader(io.StringIO(raw), delimiter=';')
        for line_no, r in enumerate(reader, start=2):
            rm = {_import_norm_txt(k): v for k, v in (r or {}).items()}
            data_obj = _import_parse_data(rm.get('data'))
            add_row(rm.get('local') or rm.get('centro') or default_local, data_obj, rm.get('hora'),
                    rm.get('ativa_lida') or rm.get('ativa'), rm.get('reativa_lida') or rm.get('reativa'),
                    rm.get('ponta_lida') or rm.get('ponta'), rm.get('agua') or rm.get('agua_m3'),
                    filename, f'CSV linha {line_no}', rm.get('ativa_anterior'), rm.get('status') or 'OK', rm.get('observacao') or '')
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        for ws in wb.worksheets:
            ws_key = _import_norm_txt(ws.title)
            if ws_key.startswith(('02_resumo', '03_a_rever', '04_ignorados', '05_modelo')):
                # Folhas auxiliares do modelo padronizado; a importação real fica em 01_IMPORTAR_NO_SGE.
                continue
            # Leitura robusta e rápida: alguns XLSX gerados por sistemas externos ficam
            # "unsized" no openpyxl, deixando ws.max_row/ws.max_column vazios. Além disso,
            # usar ws.cell() repetidamente em modo read_only pode ser muito lento. Por isso,
            # calculamos a dimensão quando necessário e usamos iter_rows(values_only=True).
            try:
                ws.calculate_dimension(force=True)
            except Exception:
                pass
            max_row = min(ws.max_row or 5000, 5000)
            max_col = min(ws.max_column or 60, 60)
            sheet_rows = []
            for idx, row_vals in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True), start=1):
                sheet_rows.append(list(row_vals or []))
                if idx >= 5000:
                    break
            if not sheet_rows:
                issues.append({'nivel': 'ignorado', 'folha': ws.title, 'data': '', 'mensagem': 'Folha vazia; ignorada.'})
                continue
            max_row = len(sheet_rows)
            max_col = max((len(r) for r in sheet_rows), default=0)
            header_row = None
            headers = []
            for ridx, vals in enumerate(sheet_rows[:20], start=1):
                norm = [_import_norm_txt(v) for v in vals]
                if 'data' in norm and any(x in norm for x in ('ativa_lida', 'activa_lida', 'ativa', 'energia_activa', 'energia_ativa')):
                    header_row = ridx
                    headers = norm
                    break
            if header_row:
                c_local = _import_find_col(headers, ['local', 'centro', 'instalacao', 'instalação'])
                c_data = _import_find_col(headers, ['data'])
                c_hora = _import_find_col(headers, ['hora'])
                c_ativa = _import_find_col(headers, ['ativa_lida', 'activa_lida', 'energia_activa', 'energia_ativa', 'ativa'])
                c_reativa = _import_find_col(headers, ['reativa_lida', 'reactiva_lida', 'reativa', 'reactiva'])
                c_ponta = _import_find_col(headers, ['ponta_lida', 'ponta'])
                c_agua = _import_find_col(headers, ['agua', 'agua_m3', 'volume_produzido', 'volume_elevado'])
                c_status = _import_find_col(headers, ['status', 'estado'])
                c_obs = _import_find_col(headers, ['observacao', 'observacoes', 'obs'])
                c_at_ant = _import_find_col(headers, ['ativa_anterior', 'activa_anterior', 'leitura_anterior'])
                for ridx, vals in enumerate(sheet_rows[header_row:], start=header_row + 1):
                    if not any(v not in (None, '') for v in vals):
                        continue
                    data_obj = _import_parse_data(vals[c_data] if c_data is not None and c_data < len(vals) else None)
                    local_file = vals[c_local] if c_local is not None and c_local < len(vals) else default_local
                    status = vals[c_status] if c_status is not None and c_status < len(vals) else 'OK'
                    obs = vals[c_obs] if c_obs is not None and c_obs < len(vals) else ''
                    add_row(local_file, data_obj,
                            vals[c_hora] if c_hora is not None and c_hora < len(vals) else '',
                            vals[c_ativa] if c_ativa is not None and c_ativa < len(vals) else None,
                            vals[c_reativa] if c_reativa is not None and c_reativa < len(vals) else None,
                            vals[c_ponta] if c_ponta is not None and c_ponta < len(vals) else None,
                            vals[c_agua] if c_agua is not None and c_agua < len(vals) else None,
                            filename, ws.title,
                            vals[c_at_ant] if c_at_ant is not None and c_at_ant < len(vals) else None,
                            str(status or 'OK'), str(obs or ''))
            else:
                mes, ano = _import_month_year_from_sheet(ws.title)
                if not mes or not ano:
                    issues.append({'nivel': 'ignorado', 'folha': ws.title, 'data': '', 'mensagem': 'Folha sem mês/ano reconhecido; ignorada.'})
                    continue
                dias_mes = calendar.monthrange(ano, mes)[1]
                for ridx in range(6, min(max_row, 36) + 1):
                    vals = sheet_rows[ridx - 1] if ridx - 1 < len(sheet_rows) else []
                    dia = vals[0] if len(vals) > 0 else None
                    if not isinstance(dia, (int, float)):
                        continue
                    dia = int(dia)
                    if dia < 1 or dia > 31:
                        continue
                    if dia > dias_mes:
                        issues.append({'nivel': 'ignorado', 'folha': ws.title, 'data': f'{ano}-{mes:02d}-{dia:02d}', 'mensagem': 'Dia inexistente no mês; ignorado.'})
                        continue
                    data_obj = datetime(ano, mes, dia).date()
                    add_row(default_local, data_obj,
                            vals[1] if len(vals) > 1 else '',
                            vals[3] if len(vals) > 3 else None,
                            vals[4] if len(vals) > 4 else None,
                            vals[6] if len(vals) > 6 else None,
                            vals[5] if len(vals) > 5 else None,
                            filename, ws.title,
                            ativa_anterior=vals[2] if len(vals) > 2 else None,
                            status='OK', observacao='Importado de folha histórica ETAU')
    unique = {}
    for r in rows:
        unique[(r['local'], r['data'])] = r
    rows = sorted(unique.values(), key=lambda x: (x['ano'], x['mes'], x['data']))
    cnt = Counter((r['ano'], r['mes']) for r in rows)
    for (ano, mes), n in sorted(cnt.items()):
        resumo.append({'ano': ano, 'mes': mes, 'linhas': n})
    return rows, issues, resumo


def _gravar_importacao_leituras_mensais(local_nome, selected_local_id, linhas, substituir_periodos=False):
    from collections import defaultdict
    ensure_leituras_mensais_phase2_schema()
    cfg = get_local_cfg_full(selected_local_id) if selected_local_id else {}
    fator_mult = _import_float(cfg.get('fator_mult'), 1.0) or 1.0
    pot_contratada = _import_float(cfg.get('pot_contratada'), 0.0) or 0.0
    t_ativa = _import_float(cfg.get('tarifa_ativa'), 0.0) or 0.0
    t_reativa = _import_float(cfg.get('tarifa_reativa'), 0.0) or 0.0
    grupos = defaultdict(list)
    for r in linhas:
        grupos[(int(r['ano']), str(r['mes']).zfill(2))].append(r)

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    inseridos = atualizados = ignorados = erros = 0
    periodos = []
    avisos = []
    try:
        for (ano, mes), rows_mes in sorted(grupos.items()):
            if _get_periodo_status(local_nome, mes, ano).get('fechado'):
                ignorados += len(rows_mes)
                avisos.append(f'{mes}/{ano}: período fechado; {len(rows_mes)} linha(s) ignorada(s).')
                continue
            if substituir_periodos:
                c.execute('DELETE FROM leituras_mensais WHERE local=? AND mes=? AND ano=?', (local_nome, mes, ano))
            rows_mes = sorted(rows_mes, key=lambda x: x['data'])
            prev_ativa, prev_reativa = get_prev_month_last_readings(local_nome, mes, ano)
            has_prev_ativa = bool(prev_ativa and prev_ativa > 0)
            has_prev_reativa = bool(prev_reativa and prev_reativa > 0)
            prev_ponta_corrigida = 0.0
            acum_mes = 0.0
            gravadas_mes = 0
            for r in rows_mes:
                try:
                    data_str = r['data']
                    hora = r.get('hora') or ''
                    ativa_lida = _import_float(r.get('ativa_lida'), None)
                    reativa_lida = _import_float(r.get('reativa_lida'), None)
                    ponta_lida = _import_float(r.get('ponta_lida'), None)
                    agua_val = _import_float(r.get('agua'), None)
                    at_ant = _import_float(r.get('ativa_anterior'), None)
                    if ativa_lida is None and reativa_lida is None and ponta_lida is None and agua_val is None:
                        ignorados += 1
                        continue
                    ativa_fat = (ativa_lida * fator_mult) if ativa_lida is not None else None
                    reativa_fat = (reativa_lida * fator_mult) if reativa_lida is not None else None
                    ponta_fat = (ponta_lida * fator_mult) if ponta_lida is not None else None
                    if not has_prev_ativa and at_ant is not None:
                        prev_ativa = at_ant * fator_mult
                        has_prev_ativa = True
                    if ponta_fat is not None:
                        if ponta_fat < prev_ponta_corrigida:
                            ponta_fat = prev_ponta_corrigida
                        else:
                            prev_ponta_corrigida = ponta_fat
                    anterior_val = prev_ativa if has_prev_ativa else (ativa_fat if ativa_fat is not None else prev_ativa)
                    atual_val = ativa_fat if ativa_fat is not None else prev_ativa
                    if ativa_fat is not None and has_prev_ativa:
                        dif_val = atual_val - anterior_val
                    else:
                        dif_val = 0.0
                    dif_operacional = dif_val if dif_val >= 0 else 0.0
                    if ativa_fat is not None:
                        if not has_prev_ativa:
                            prev_ativa = atual_val
                            has_prev_ativa = True
                        elif dif_val >= 0:
                            prev_ativa = atual_val
                    if reativa_fat is not None:
                        if has_prev_reativa:
                            delta_reativa_real = reativa_fat - prev_reativa
                        else:
                            delta_reativa_real = 0.0
                        delta_reativa = delta_reativa_real if delta_reativa_real >= 0 else 0.0
                        if not has_prev_reativa:
                            prev_reativa = reativa_fat
                            has_prev_reativa = True
                        elif delta_reativa_real >= 0:
                            prev_reativa = reativa_fat
                    else:
                        delta_reativa = 0.0
                    if dif_operacional > 0 or delta_reativa > 0:
                        fp_val = dif_operacional / math.sqrt((dif_operacional ** 2) + (delta_reativa ** 2))
                    else:
                        fp_val = None
                    reativa_excedente = max(delta_reativa - (0.75 * dif_operacional), 0.0)
                    esp_val = (dif_operacional / agua_val) if agua_val else None
                    acum_mes += dif_operacional
                    valor_total_dia = (dif_operacional * t_ativa) + (reativa_excedente * t_reativa)
                    exists = c.execute('SELECT 1 FROM leituras_mensais WHERE local=? AND data=?', (local_nome, data_str)).fetchone()
                    c.execute('''
                        INSERT INTO leituras_mensais
                        (local, data, hora, ativa, reativa, ponta, fp, potc, anterior, atual, diferenca,
                         agua, esp, acum, valor, mes, ano)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(local, data) DO UPDATE SET
                            hora=excluded.hora,
                            ativa=excluded.ativa,
                            reativa=excluded.reativa,
                            ponta=excluded.ponta,
                            fp=excluded.fp,
                            potc=excluded.potc,
                            anterior=excluded.anterior,
                            atual=excluded.atual,
                            diferenca=excluded.diferenca,
                            agua=excluded.agua,
                            esp=excluded.esp,
                            acum=excluded.acum,
                            valor=excluded.valor,
                            mes=excluded.mes,
                            ano=excluded.ano
                    ''', (local_nome, data_str, hora, ativa_fat, reativa_fat, ponta_fat, fp_val, pot_contratada,
                          anterior_val, atual_val, dif_operacional, agua_val, esp_val, acum_mes, valor_total_dia,
                          mes, ano))
                    if exists:
                        atualizados += 1
                    else:
                        inseridos += 1
                    gravadas_mes += 1
                except Exception as e:
                    erros += 1
                    avisos.append(f"{r.get('data','')}: erro ao gravar ({e}).")
            periodos.append({'ano': ano, 'mes': mes, 'linhas': gravadas_mes})
        conn.commit()
    finally:
        conn.close()
    return {
        'inseridos': inseridos,
        'atualizados': atualizados,
        'ignorados': ignorados,
        'erros': erros,
        'periodos': periodos,
        'avisos': avisos[:30],
        'fator_mult': fator_mult,
    }


@app.route('/leituras_mensal/template_importacao_xlsx')
def leituras_mensal_template_importacao_xlsx():
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    _set_xlsx_identity(workbook, 'Modelo de Importação de Leituras MT')
    ws = workbook.add_worksheet('MODELO_IMPORTACAO')
    fmt_title = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#073B78', 'align': 'center', 'valign': 'vcenter', 'font_size': 14})
    fmt_head = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#0F6FC6', 'align': 'center', 'valign': 'vcenter', 'border': 1})
    fmt_date = workbook.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})
    fmt_num = workbook.add_format({'num_format': '0.00', 'border': 1})
    fmt_txt = workbook.add_format({'border': 1})
    ws.merge_range('A1:K1', 'MODELO PADRÃO PARA IMPORTAÇÃO DE LEITURAS MT NO SGE', fmt_title)
    headers = ['local', 'data', 'hora', 'ativa_lida', 'reativa_lida', 'ponta_lida', 'agua', 'ativa_anterior', 'status', 'observacao', 'origem']
    for col, h in enumerate(headers):
        ws.write(2, col, h, fmt_head)
    exemplo = ['ETA DE UMBELUZI', datetime.now().date(), '00:00', 4806.55, 2260.97, 0.206, 196400, 4802.24, 'OK', 'Exemplo; substituir pelos dados reais', 'modelo']
    for col, val in enumerate(exemplo):
        if col == 1:
            ws.write_datetime(3, col, datetime.combine(val, datetime.min.time()), fmt_date)
        elif col in (3,4,5,6,7):
            ws.write_number(3, col, float(val), fmt_num)
        else:
            ws.write(3, col, val, fmt_txt)
    ws.set_column('A:A', 18); ws.set_column('B:B', 13); ws.set_column('C:C', 10); ws.set_column('D:H', 16); ws.set_column('I:I', 14); ws.set_column('J:K', 36)
    ws.freeze_panes(3, 0)
    workbook.close()
    output.seek(0)
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=modelo_importacao_leituras_mt_sge.xlsx'})


@app.route('/leituras_mensal/import', methods=['GET', 'POST'])
def leituras_mensal_import():
    locais_db = get_locais()
    locais = [{'id': lid, 'nome': lname} for lid, lname in locais_db]
    umbeluzi_id, umbeluzi_nome = _import_find_umbeluzi_local(locais_db)
    preview = []
    report = None
    msg = None

    def render_import_page():
        return render_template(
            'leituras_mensal_import.html',
            locais=locais,
            msg=msg,
            report=report,
            preview=preview,
            local_recomendado_id=umbeluzi_id,
            local_recomendado_nome=umbeluzi_nome
        )

    if request.method == 'POST':
        raw_local = request.form.get('local', '')
        selected_local_id, local_nome = _local_id_nome_from_request(raw_local, locais_db)
        if not local_nome or not selected_local_id:
            msg = ('danger', 'Selecione um local válido para importar as leituras.')
            return render_import_page()
        file = request.files.get('arquivo')
        if not file or not file.filename:
            msg = ('warning', 'Selecione um ficheiro Excel ou CSV.')
            return render_import_page()
        filename = secure_filename(file.filename)
        data_bytes = file.read()
        try:
            linhas, issues, resumo = _import_read_uploaded_workbook(data_bytes, filename, local_nome)
        except Exception as e:
            msg = ('danger', f'Erro ao ler o ficheiro: {e}')
            return render_import_page()

        local_forcado = False
        if _import_is_eta_umbeluzi_file(filename, linhas):
            if umbeluzi_id and umbeluzi_nome:
                if selected_local_id != umbeluzi_id:
                    issues.insert(0, {
                        'nivel': 'aviso',
                        'folha': '',
                        'data': '',
                        'mensagem': f'Ficheiro reconhecido como ETAU/ETA Umbeluzi. O destino foi corrigido automaticamente de "{local_nome}" para "{umbeluzi_nome}".'
                    })
                    selected_local_id, local_nome = umbeluzi_id, umbeluzi_nome
                    local_forcado = True
            elif 'umbeluzi' not in _import_norm_txt(local_nome):
                msg = ('danger', 'Este ficheiro foi reconhecido como ETAU/ETA Umbeluzi, mas não encontrei o local "ETA DE UMBELUZI" cadastrado no SGE. Cadastre ou selecione o local correto antes de importar.')
                return render_import_page()

        substituir = request.form.get('substituir_periodos') == 'on'
        acao = request.form.get('acao') or 'previsualizar'
        preview = linhas[:80]
        report = {
            'filename': filename,
            'local': local_nome,
            'linhas_validas': len(linhas),
            'issues': issues[:80],
            'resumo': resumo,
            'substituir': substituir,
            'local_forcado': local_forcado,
            'gravacao': None,
        }
        if acao == 'importar':
            if not linhas:
                msg = ('warning', 'O ficheiro foi lido, mas nenhuma linha válida foi encontrada para importar.')
            else:
                grav = _gravar_importacao_leituras_mensais(local_nome, selected_local_id, linhas, substituir_periodos=substituir)
                report['gravacao'] = grav
                if grav.get('erros'):
                    msg = ('warning', f"Importação concluída com alertas. Local destino: {local_nome}. Inseridos={grav['inseridos']}, Atualizados={grav['atualizados']}, Ignorados={grav['ignorados']}, Erros={grav['erros']}.")
                else:
                    msg = ('success', f"Importação concluída para {local_nome}. Inseridos={grav['inseridos']}, Atualizados={grav['atualizados']}, Ignorados={grav['ignorados']}. Fator usado: {grav.get('fator_mult', 1)}.")
        else:
            msg = ('info', f'Pré-visualização pronta para {local_nome}: {len(linhas)} linha(s) válidas encontradas. Revise o resumo e clique em Importar agora se estiver correto.')
    return render_import_page()


# === FATURA (manual) ===

@app.route('/calcular_fatura', methods=['GET', 'POST'])
def calcular_fatura():
    locais = get_locais()
    if request.method == 'POST':
        local_nome = request.form['local']
        periodo = request.form['periodo']
        fator_mult = float(request.form['fator_mult'])
        demanda_max = float(request.form['demanda_max'])
        pot_contratada = float(request.form['pot_contratada'])

        ativa_ant = float(request.form['ativa_ant'])
        ativa_atu = float(request.form['ativa_atu'])
        reativa_ant = float(request.form['reativa_ant'])
        reativa_atu = float(request.form['reativa_atu'])
        perdas_ant = float(request.form['perdas_ant'])
        perdas_atu = float(request.form['perdas_atu'])
        ponta_ant = float(request.form['ponta_ant'])
        ponta_atu = float(request.form['ponta_atu'])

        tarifa_ativa = float(request.form['tarifa_ativa'])
        tarifa_reativa = float(request.form['tarifa_reativa'])
        tarifa_ponta = float(request.form['tarifa_ponta'])
        tarifa_perdas = float(request.form['tarifa_perdas'])
        taxa_fixa = float(request.form['taxa_fixa'])
        taxa_radio = float(request.form['taxa_radio'])
        taxa_lixo = float(request.form['taxa_lixo'])
        saldo_ant = float(request.form['saldo_ant'])

        ativa = (ativa_atu - ativa_ant) * fator_mult
        reativa = (reativa_atu - reativa_ant) * fator_mult
        perdas = (perdas_atu - perdas_ant) * fator_mult
        ponta_lida_corrigida = max(demanda_max, (ponta_atu - ponta_ant) * fator_mult)
        fatura = calculate_invoice(
            active_kwh=ativa,
            reactive_kvarh=reativa,
            measured_peak_kw=ponta_lida_corrigida,
            contracted_power_kw=pot_contratada,
            losses_kwh=perdas,
            tariffs={
                'tarifa_ativa': tarifa_ativa, 'tarifa_reativa': tarifa_reativa,
                'tarifa_ponta': tarifa_ponta, 'tarifa_perdas': tarifa_perdas,
                'taxa_fixa': taxa_fixa, 'taxa_radio': taxa_radio, 'taxa_lixo': taxa_lixo,
            },
            previous_balance_mzn=saldo_ant,
            bill_losses=False,
        )
        reativa_faturavel = fatura['reactive_excess_kvarh']
        ponta = fatura['billing_demand_kw']

        return render_template('fatura_resultado.html',
                               local=local_nome, periodo=periodo,
                               ativa=ativa, reativa=reativa,
                               reativa_faturavel=reativa_faturavel,
                               perdas=perdas, ponta=ponta,
                               demanda_max=ponta_lida_corrigida, pot_contratada=pot_contratada,
                               demanda_faturavel=ponta,
                               valor_ativa=fatura['active_cost_mzn'],
                               valor_reativa=fatura['reactive_cost_mzn'],
                               valor_perdas=fatura['losses_cost_mzn'],
                               valor_ponta=fatura['demand_cost_mzn'],
                               taxa_fixa=fatura['fixed_fee_mzn'], taxa_radio=fatura['radio_fee_mzn'],
                               taxa_lixo=fatura['waste_fee_mzn'], subtotal=fatura['subtotal_mzn'],
                               valor_iva=fatura['vat_mzn'], total=fatura['total_mzn'],
                               iva_percent=16.0, base_iva_percent=62.0)
    # GET -> cfg_map para auto-preencher
    cfg_map = {}
    for lid, lname in locais:
        cfg = get_local_cfg(lid)
        cfg_map[str(lid)] = {
            "fator_mult": cfg[0],
            "pot_contratada": cfg[1],
            "tarifa_ativa": cfg[2],
            "tarifa_reativa": cfg[3],
            "tarifa_ponta": cfg[4],
            "tarifa_perdas": cfg[5],
            "taxa_fixa": cfg[6],
            "taxa_radio": cfg[7],
            "taxa_lixo": cfg[8],
            "iva": cfg[9],
        }
    return render_template('fatura.html', locais=locais, cfg_map=cfg_map)

# === FATURA (a partir do mês) ===
@app.route('/fatura/mes', methods=['GET', 'POST'])
def fatura_mes():
    locais = get_locais()
    hoje = datetime.now()
    if request.method == 'POST':
        local_id = int(request.form['local_id'])
        mes = request.form['mes']  # "YYYY-MM"
        ano_int = int(mes.split('-')[0]); mes_int = int(mes.split('-')[1])
        local_nome = [l[1] for l in locais if l[0]==local_id][0]

        ctx = _montar_contexto_fatura_mensal(local_nome, str(mes_int).zfill(2), ano_int)

        periodo_leg = f"{mes}-01 a {mes}-{calendar.monthrange(ano_int, mes_int)[1]}"

        return render_template('fatura_resultado.html',
                               local=local_nome, periodo=periodo_leg,
                               ativa=ctx['kwh_ativa'], reativa=ctx['kvarh_reativa'],
                               reativa_faturavel=ctx['kvarh_excedente'], perdas=0.0,
                               ponta=ctx['demanda_ponta_kw'], demanda_max=ctx['kw_ponta_lida'],
                               pot_contratada=ctx['pot_contratada'], demanda_faturavel=ctx['demanda_ponta_kw'],
                               valor_ativa=ctx['valor_ativa'], valor_reativa=ctx['valor_reativa'],
                               valor_perdas=ctx['valor_perdas'], valor_ponta=ctx['valor_ponta'],
                               taxa_fixa=ctx['taxa_fixa'], taxa_radio=ctx['taxa_radio'], taxa_lixo=ctx['taxa_lixo'],
                               subtotal=ctx['subtotal'], valor_iva=ctx['valor_iva'], total=ctx['total'],
                               iva_percent=ctx['iva_percent'], base_iva_percent=ctx['base_iva_percent'])

    return render_template('fatura_mes.html', locais=locais, hoje=hoje.strftime('%Y-%m'))

# =========================
# === MÓDULO "MOTORES" ===
# =========================
# A — o módulo de Motores deixa de ser uma página isolada e passa a ser
# uma análise técnica especializada, integrada com Equipamentos, Monitoria e Alertas.

