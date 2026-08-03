"""Domínio equipment_extended extraído do app.py consolidado.

Carregado no arranque pelo mecanismo transitório de compatibilidade.
"""

@app.route('/api/local_cfg/<int:local_id>')
def api_local_cfg(local_id):
    local = get_local_by_id(local_id)
    if not local:
        return {"ok": False, "error": "Local não encontrado"}, 404
    cfg = get_local_cfg_full(local_id)
    return {"ok": True, "local_id": local_id, "local_nome": local[1], "cfg": cfg}, 200

# === NOVO: API Locais JSON (mantida) ===
@app.route('/api/locais.json')
def api_locais_json():
    incluir_inativos = (request.args.get('inativos') == '1')
    q = (request.args.get('q') or '').strip()
    data = get_locais_with_cfg(q if q else None, incluir_inativos=incluir_inativos)
    return {"ok": True, "locais": data}

# ==============================

# ==============================

@app.route('/equipamentos/remover/<int:equipamento_id>', methods=['POST'])
def remover_equipamento(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET deleted_at=datetime('now','localtime'), updated_at=datetime('now','localtime') WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "arquivar", "")
    flash("Equipamento removido da lista ativa.", "warning")
    return redirect(url_for('listar_equipamentos'))



@app.route('/equipamentos/export/csv')
def exportar_equipamentos_csv():
    q = request.args.get('q','').strip()
    local_id = request.args.get('local_id','').strip()
    where = []; params=[]
    if q:
        like = f"%{q}%"
        where.append("(e.nome LIKE ? OR e.tag LIKE ? OR e.especificacao LIKE ?)")
        params += [like, like, like]
    if local_id and local_id.isdigit():
        where.append("e.local_id=?"); params.append(int(local_id))
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.id, e.nome, COALESCE(l.nome,''), e.tag, e.especificacao,
               COALESCE(e.ano_instalacao,''), COALESCE(e.quantidade,0)
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    rows = c.fetchall(); conn.close()

    output = io.StringIO()
    writer = csv.writer(output, lineterminator='\n')
    writer.writerow(["ID","Nome","Local","TAG","Especificação","Ano instalação","Quantidade"])
    for r in rows:
        writer.writerow(r)
    csv_data = output.getvalue().encode('utf-8')
    return Response(csv_data, mimetype='text/csv',
                    headers={"Content-Disposition":"attachment; filename=equipamentos.csv"})



@app.route('/equipamentos/export/pdf')
def exportar_equipamentos_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm

    q = request.args.get('q','').strip()
    local_id = request.args.get('local_id','').strip()
    where = []; params=[]
    if q:
        like = f"%{q}%"
        where.append("(e.nome LIKE ? OR e.tag LIKE ? OR e.especificacao LIKE ?)")
        params += [like, like, like]
    if local_id and local_id.isdigit():
        where.append("e.local_id=?"); params.append(int(local_id))
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.id, e.nome, COALESCE(l.nome,''), e.tag, e.especificacao,
               COALESCE(e.ano_instalacao,''), COALESCE(e.quantidade,0)
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    rows = c.fetchall(); conn.close()

    buffer = io.BytesIO()
    page_size = landscape(A4)
    cpdf = canvas.Canvas(buffer, pagesize=page_size)
    _set_pdf_identity(cpdf, 'Relatório de Equipamentos')
    width, height = page_size

    title = f"Equipamentos - Relatório"
    cpdf.setFont("Helvetica-Bold", 16)
    cpdf.drawString(2*cm, height-1.5*cm, title)
    cpdf.setFont("Helvetica", 9)
    cpdf.drawString(2*cm, height-2.2*cm, f"Filtro: q='{q}'  local_id='{local_id}'")

    # Table header
    headers = ["ID","Nome","Local","TAG","Especificação","Ano","Qtd"]
    col_x = [1.0*cm, 2.5*cm, 8.5*cm, 12.5*cm, 15.0*cm, 24.0*cm, 26.0*cm]
    y = height - 3.0*cm
    cpdf.setFont("Helvetica-Bold", 8)
    for i, h in enumerate(headers):
        cpdf.drawString(col_x[i], y, h)
    cpdf.line(1*cm, y-0.2*cm, width-1*cm, y-0.2*cm)
    y -= 0.5*cm
    cpdf.setFont("Helvetica", 8)

    for r in rows:
        if y < 1.5*cm:
            cpdf.showPage()
            cpdf.setFont("Helvetica-Bold", 8)
            y = height - 1.5*cm
            for i, h in enumerate(headers):
                cpdf.drawString(col_x[i], y, h)
            cpdf.line(1*cm, y-0.2*cm, width-1*cm, y-0.2*cm)
            y -= 0.5*cm
            cpdf.setFont("Helvetica", 8)

        values = [str(r[0]), r[1] or "", r[2] or "", r[3] or "", (r[4] or "")[:70], str(r[5] or ""), str(r[6] or "")]
        for i, val in enumerate(values):
            cpdf.drawString(col_x[i], y, val)
        y -= 0.45*cm

    cpdf.showPage()
    cpdf.save()
    pdf_data = buffer.getvalue()
    buffer.close()
    return Response(pdf_data, mimetype='application/pdf',
                    headers={"Content-Disposition":"attachment; filename=equipamentos.pdf"})



def log_equip_audit(equipamento_id, acao, detalhes=""):
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("INSERT INTO mecanismos_dummy (x) VALUES (1)") if False else None  # no-op to keep pattern
        c.execute("INSERT INTO equipamentos_audit (equipamento_id, acao, detalhes, actor) VALUES (?, ?, ?, ?)",
                  (equipamento_id, acao, detalhes, _actor_name()))
        conn.commit(); conn.close()
    except Exception:
        pass


@app.route('/equipamentos/desativar/<int:equipamento_id>', methods=['POST'])
def desativar_equipamento(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET ativo=0, updated_at=datetime('now','localtime') WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "desativar", "")
    flash("Equipamento desativado.", "warning")
    return redirect(url_for('listar_equipamentos'))

@app.route('/equipamentos/ativar/<int:equipamento_id>', methods=['POST'])
def ativar_equipamento(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET ativo=1, updated_at=datetime('now','localtime') WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "ativar", "")
    flash("Equipamento reativado.", "success")
    return redirect(url_for('listar_equipamentos'))


@app.route('/equipamentos/<int:equipamento_id>', methods=['GET', 'POST'])
def equipamento_detalhe(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''
        SELECT e.id, e.nome, COALESCE(l.nome,''), COALESCE(e.tag,''), COALESCE(e.especificacao,''),
               COALESCE(e.ano_instalacao,''), COALESCE(e.quantidade,0), COALESCE(e.ativo,1),
               e.created_at, e.updated_at, COALESCE(e.categoria,''), COALESCE(e.fabricante,''),
               COALESCE(e.modelo,''), COALESCE(e.numero_serie,''), COALESCE(e.custo_aquisicao,0.0),
               COALESCE(e.vida_util_anos,''), COALESCE(e.criticidade,''), COALESCE(e.potencia_kw,''),
               COALESCE(e.tensao_v,''), COALESCE(e.corrente_a,''), COALESCE(e.fornecedor,''),
               COALESCE(e.contrato_num,''), COALESCE(e.garantia_fim,'')
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        WHERE e.id=?
    ''', (equipamento_id,))
    eq = c.fetchone()

    if not eq:
        conn.close()
        flash("Equipamento não encontrado.", "warning")
        return redirect(url_for('listar_equipamentos'))

    c.execute('SELECT id, original_name, filename, mime, size, uploaded_at FROM equipamentos_files WHERE equipamento_id=? ORDER BY uploaded_at DESC', (equipamento_id,))
    files = c.fetchall()
    c.execute('SELECT id, filename, thumb_filename, caption FROM equipamentos_photos WHERE equipamento_id=? ORDER BY uploaded_at DESC', (equipamento_id,))
    photos = c.fetchall()

    try:
        c.execute('SELECT tensao_nominal, corrente_nominal, potencia_nominal_kw, fp_nominal, eficiencia_nominal FROM equipamentos_cfg WHERE equipamento_id=?', (equipamento_id,))
        cfg = c.fetchone()
    except Exception:
        cfg = None

    conn.close()

    detalhe = {
        'id': eq[0],
        'nome': _equip_clean_text(eq[1]),
        'local': _equip_clean_text(eq[2]),
        'tag': _equip_clean_text(eq[3]),
        'especificacao': _equip_clean_text(eq[4]),
        'ano': _equip_clean_text(eq[5]),
        'quantidade': int(eq[6] or 0),
        'ativo': _equip_bool(eq[7]),
        'categoria': _equip_clean_text(eq[10]),
        'fabricante': _equip_clean_text(eq[11]),
        'modelo': _equip_clean_text(eq[12]),
        'numero_serie': _equip_clean_text(eq[13]),
        'custo': _equip_clean_number(eq[14], decimals=2, fallback='0.00', zero_as_value=True),
        'vida_util': _equip_clean_text(eq[15]),
        'criticidade': _equip_clean_text(eq[16]),
        'potencia_kw': _equip_clean_number(eq[17], decimals=2),
        'tensao_v': _equip_clean_number(eq[18], decimals=0),
        'corrente_a': _equip_clean_number(eq[19], decimals=2),
        'fornecedor': _equip_clean_text(eq[20]),
        'contrato_num': _equip_clean_text(eq[21]),
        'garantia_fim': _equip_clean_text(eq[22]),
        'created_at': _equip_clean_text(eq[8]),
        'updated_at': _equip_clean_text(eq[9]),
    }
    return render_template('equipamento_detalhe.html', eq=eq, detalhe=detalhe, files=files, photos=photos, cfg=cfg)

# Upload endpoint
@app.route('/equipamentos/<int:equipamento_id>/upload', methods=['POST'])
def equipamento_upload(equipamento_id):
    file = request.files.get('file')
    if not file or file.filename == '':
        flash("Nenhum ficheiro selecionado.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    filename = secure_filename(file.filename)
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    file.save(save_path)
    size = os.path.getsize(save_path)
    mime = file.mimetype or ""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('INSERT INTO equipamentos_files (equipamento_id, filename, original_name, mime, size) VALUES (?,?,?,?,?)',
              (equipamento_id, filename, file.filename, mime, size))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "upload", file.filename)
    flash("Ficheiro carregado.", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/files/<path:filename>')
def equipamento_download(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)


@app.route('/equipamentos/import', methods=['GET', 'POST'])
def importar_equipamentos():
    if request.method == 'POST':
        file = request.files.get('csv')
        if not file or file.filename == '':
            flash("Selecione um ficheiro CSV.", "danger")
            return redirect(url_for('importar_equipamentos'))
        import csv, io
        content = file.read().decode('utf-8', errors='ignore')
        reader = csv.DictReader(io.StringIO(content))
        inserted = 0; skipped = 0
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        for row in reader:
            nome = (row.get('nome') or row.get('Nome') or '').strip()
            if not nome:
                skipped += 1; continue
            local_nome = (row.get('local') or row.get('Local') or '').strip()
            # resolve local_id by name if provided
            local_id = None
            if local_nome:
                c.execute("SELECT id FROM locais WHERE nome=?", (local_nome,))
                r = c.fetchone()
                if r: local_id = r[0]
            tag = (row.get('tag') or row.get('TAG') or '').strip()
            especificacao = (row.get('especificacao') or row.get('Especificacao') or row.get('Especificação') or '').strip()
            ano = (row.get('ano_instalacao') or row.get('ano') or '').strip()
            qtd = (row.get('quantidade') or row.get('qtd') or '1').strip()
            try:
                qtd_i = int(qtd)
            except Exception:
                qtd_i = 1

            c.execute('''
                INSERT INTO equipamentos (nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'), datetime('now','localtime'))
            ''', (nome, local_id, tag, especificacao, ano or None, qtd_i))
            inserted += 1
        conn.commit(); conn.close()
        flash(f"Importação concluída: {inserted} inseridos, {skipped} ignorados.", "success")
        return redirect(url_for('listar_equipamentos'))
    return render_template('importar_equipamentos.html')





@app.route('/equipamentos/<int:equipamento_id>/upload_photo', methods=['POST'])
def equipamento_upload_photo(equipamento_id):
    file = request.files.get('photo')
    if not file or file.filename == '':
        flash("Selecione uma imagem.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    if not (file.mimetype or "").lower().startswith("image/"):
        flash("O ficheiro deve ser uma imagem.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    filename, thumb_name, w, h = _save_image_and_thumb(file, dest_dir=app.config['UPLOAD_FOLDER']+'/photos', thumb_dir=app.config['UPLOAD_FOLDER']+'/thumbs')
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('INSERT INTO equipamentos_photos (equipamento_id, filename, thumb_filename, width, height) VALUES (?,?,?,?,?)',
              (equipamento_id, 'photos/'+filename, ('thumbs/'+thumb_name if thumb_name else None), w, h))
    photo_id = c.lastrowid
    # set as cover if not set
    c.execute('SELECT cover_photo_id FROM equipamentos WHERE id=?', (equipamento_id,))
    r = c.fetchone()
    if not r or not r[0]:
        c.execute('UPDATE equipamentos SET cover_photo_id=? WHERE id=?', (photo_id, equipamento_id))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "upload_foto", filename)
    flash("Foto adicionada.", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/<int:equipamento_id>/upload_photos', methods=['POST'])
def equipamento_upload_photos(equipamento_id):
    files = request.files.getlist('photos[]')
    if not files:
        flash("Selecione uma ou mais imagens.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    added = 0
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    for file in files:
        if not file or file.filename == '': continue
        if not (file.mimetype or "").lower().startswith("image/"): continue
        filename, thumb_name, w, h = _save_image_and_thumb(file, dest_dir=app.config['UPLOAD_FOLDER']+'/photos', thumb_dir=app.config['UPLOAD_FOLDER']+'/thumbs')
        c.execute('INSERT INTO equipamentos_photos (equipamento_id, filename, thumb_filename, width, height) VALUES (?,?,?,?,?)',
                  (equipamento_id, 'photos/'+filename, ('thumbs/'+thumb_name if thumb_name else None), w, h))
        if added == 0:
            # set cover if empty
            c.execute('UPDATE equipamentos SET cover_photo_id=COALESCE(cover_photo_id, last_insert_rowid()) WHERE id=?', (equipamento_id,))
        added += 1
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "upload_fotos_multi", f"{added} fotos")
    flash(f"{added} foto(s) adicionada(s).", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/<int:equipamento_id>/photo/<int:photo_id>/delete', methods=['POST'])
def equipamento_photo_delete(equipamento_id, photo_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT filename, thumb_filename FROM equipamentos_photos WHERE id=? AND equipamento_id=?', (photo_id, equipamento_id))
    r = c.fetchone()
    if r:
        fn, tfn = r[0], r[1]
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], fn))
        except Exception: pass
        try:
            if tfn: os.remove(os.path.join(app.config['UPLOAD_FOLDER'], tfn))
        except Exception: pass
    c.execute('DELETE FROM equipamentos_photos WHERE id=? AND equipamento_id=?', (photo_id, equipamento_id))
    # if it was cover, unset
    c.execute('UPDATE equipamentos SET cover_photo_id=NULL WHERE id=? AND cover_photo_id=?', (equipamento_id, photo_id))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "apagar_foto", str(photo_id))
    flash("Foto removida.", "warning")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/<int:equipamento_id>/photo/<int:photo_id>/cover', methods=['POST'])
def equipamento_photo_cover(equipamento_id, photo_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('UPDATE equipamentos SET cover_photo_id=? WHERE id=?', (photo_id, equipamento_id))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "definir_capa", str(photo_id))
    flash("Foto definida como capa.", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))



@app.route('/equipamentos/<int:equipamento_id>/label')
def equipamento_label(equipamento_id):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A7, landscape
    from reportlab.lib.units import mm
    from reportlab.graphics.barcode import qr
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id, nome, tag FROM equipamentos WHERE id=?', (equipamento_id,))
    r = c.fetchone(); conn.close()
    if not r:
        flash("Equipamento não encontrado.", "warning")
        return redirect(url_for('listar_equipamentos'))
    _id, nome, tag = r
    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=landscape(A7))
    _set_pdf_identity(cpdf, 'Etiqueta de Equipamento')
    w, h = landscape(A7)
    cpdf.setFont("Helvetica-Bold", 10)
    cpdf.drawString(10*mm, h-10*mm, f"EQ #{_id}")
    cpdf.setFont("Helvetica", 9)
    cpdf.drawString(10*mm, h-16*mm, (nome or "")[:40])
    if tag:
        cpdf.drawString(10*mm, h-22*mm, f"TAG: {tag[:30]}")
    # QR code content
    code_val = f"SGE:EQ:{_id}"
    qr_code = qr.QrCodeWidget(code_val)
    bounds = qr_code.getBounds()
    size = 30*mm
    width = bounds[2] - bounds[0]
    height = bounds[3] - bounds[1]
    d = size / max(width, height)
    from reportlab.graphics.shapes import Drawing
    drawing = Drawing(size, size, transform=[d,0,0,d,0,0])
    drawing.add(qr_code)
    from reportlab.graphics import renderPDF
    renderPDF.draw(drawing, cpdf, w-40*mm, h-40*mm)
    cpdf.showPage(); cpdf.save()
    pdf_data = buf.getvalue(); buf.close()
    return Response(pdf_data, mimetype='application/pdf',
                    headers={"Content-Disposition": f"attachment; filename=label_eq_{_id}.pdf"})

import openpyxl


@app.route('/equipamentos/bulk', methods=['POST'])
def equipamentos_bulk():
    action = request.form.get('action')
    ids = request.form.getlist('ids')
    ids = [int(x) for x in ids if x.isdigit()]
    if not ids:
        flash("Selecione pelo menos um item.", "warning")
        return redirect(url_for('listar_equipamentos'))
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()

    if action == 'ativar':
        c.executemany("UPDATE equipamentos SET ativo=1, updated_at=datetime('now','localtime') WHERE id=?", [(i,) for i in ids])
        conn.commit(); conn.close()
        flash(f"{len(ids)} ativado(s).", "success")
        return redirect(url_for('listar_equipamentos'))
    if action == 'desativar':
        c.executemany("UPDATE equipamentos SET ativo=0, updated_at=datetime('now','localtime') WHERE id=?", [(i,) for i in ids])
        conn.commit(); conn.close()
        flash(f"{len(ids)} desativado(s).", "warning")
        return redirect(url_for('listar_equipamentos'))
    if action == 'remover':
        c.executemany("UPDATE equipamentos SET deleted_at=datetime('now','localtime'), updated_at=datetime('now','localtime') WHERE id=?", [(i,) for i in ids])
        conn.commit(); conn.close()
        flash(f"{len(ids)} removido(s) da lista ativa.", "warning")
        return redirect(url_for('listar_equipamentos'))
    if action == 'labels':
        conn.close()
        ids_str = ",".join([str(i) for i in ids])
        return redirect(url_for('equipamentos_labels_pdf', ids=ids_str))
    if action == 'export_csv':
        conn.close()
        ids_str = ",".join([str(i) for i in ids])
        return redirect(url_for('exportar_equipamentos_csv') + f"?ids={ids_str}")
    if action == 'export_pdf':
        conn.close()
        ids_str = ",".join([str(i) for i in ids])
        return redirect(url_for('exportar_equipamentos_pdf') + f"?ids={ids_str}")

    conn.close()
    flash("Ação desconhecida.", "danger")
    return redirect(url_for('listar_equipamentos'))


@app.route('/equipamentos/duplicados')
def equipamentos_duplicados():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    # Duplicados por (nome, local)
    c.execute('''
        SELECT e.nome, e.local_id, COUNT(*)
        FROM equipamentos e
        GROUP BY e.nome, e.local_id
        HAVING COUNT(*) > 1
        ORDER BY COUNT(*) DESC, e.nome
    ''')
    by_nome_local = c.fetchall()

    # Duplicados por TAG
    c.execute('''
        SELECT e.tag, COUNT(*)
        FROM equipamentos e
        WHERE e.tag IS NOT NULL AND e.tag <> ''
        GROUP BY e.tag
        HAVING COUNT(*) > 1
        ORDER BY COUNT(*) DESC, e.tag
    ''')
    by_tag = c.fetchall()

    # Duplicados por número de série
    try:
        c.execute('''
            SELECT e.numero_serie, COUNT(*)
            FROM equipamentos e
            WHERE e.numero_serie IS NOT NULL AND e.numero_serie <> ''
            GROUP BY e.numero_serie
            HAVING COUNT(*) > 1
            ORDER BY COUNT(*) DESC, e.numero_serie
        ''')
        by_serial = c.fetchall()
    except Exception:
        by_serial = []

    conn.close()
    return render_template('equipamentos_duplicados.html',
                           by_nome_local=by_nome_local, by_tag=by_tag, by_serial=by_serial)


@app.route('/equipamentos/export/detalhe/<int:equipamento_id>.pdf')
def equipamento_detalhe_pdf(equipamento_id):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''
        SELECT e.id, e.nome, COALESCE(l.nome,''), e.tag, e.especificacao, e.ano_instalacao,
               e.quantidade, e.categoria, e.fabricante, e.modelo, e.numero_serie, e.custo_aquisicao,
               e.vida_util_anos, e.criticidade, e.created_at, e.updated_at
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        WHERE e.id=?
    ''', (equipamento_id,))
    eq = c.fetchone()

    cover = None
    c.execute('SELECT filename FROM equipamentos_photos WHERE id=(SELECT cover_photo_id FROM equipamentos WHERE id=?)', (equipamento_id,))
    r = c.fetchone()
    if r: cover = r[0]
    conn.close()

    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=A4)
    _set_pdf_identity(cpdf, 'Ficha de Equipamento')
    w, h = A4
    cpdf.setFont("Helvetica-Bold", 14)
    cpdf.drawString(2*cm, h-2*cm, f"Ficha do Equipamento #{equipamento_id}")
    cpdf.setFont("Helvetica", 10)
    y = h-3*cm

    labels = ["Nome","Local","TAG","Especificação","Ano","Qtd","Categoria","Fabricante","Modelo","Nº Série","Custo (MZN)","Vida Útil (anos)","Criticidade","Criado","Atualizado"]
    vals = [eq[1],eq[2],eq[3],eq[4],eq[5],eq[6],eq[7],eq[8],eq[9],eq[10],eq[11],eq[12],eq[13],eq[14],eq[15]]
    for L,V in zip(labels,vals):
        cpdf.drawString(2*cm, y, f"{L}: {V if V is not None else '-'}")
        y -= 0.6*cm
        if y < 4*cm:
            cpdf.showPage(); y = h-2*cm

    if cover:
        try:
            from reportlab.lib.utils import ImageReader
            img = ImageReader(os.path.join(app.config['UPLOAD_FOLDER'], cover))
            cpdf.drawImage(img, 2*cm, 2*cm, width=12*cm, height=8*cm, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    cpdf.showPage(); cpdf.save()
    pdf = buf.getvalue(); buf.close()
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=equip_{equipamento_id}.pdf"})


@app.route('/equipamentos/labels.pdf')
def equipamentos_labels_pdf():
    ids_str = request.args.get('ids','').strip()
    if not ids_str:
        flash("Nenhum ID selecionado.", "warning")
        return redirect(url_for('listar_equipamentos'))
    try:
        ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
    except Exception:
        ids = []
    if not ids:
        flash("IDs inválidos.", "danger")
        return redirect(url_for('listar_equipamentos'))

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF
    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=A4)
    _set_pdf_identity(cpdf, 'Etiquetas de Equipamentos')
    w, h = A4

    # grid 3x8 (aprox) de etiquetas
    cols, rows = 3, 8
    margin_x, margin_y = 10*mm, 10*mm
    cell_w = (w - 2*margin_x) / cols
    cell_h = (h - 2*margin_y) / rows

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f"SELECT id, nome, tag FROM equipamentos WHERE id IN ({','.join(['?']*len(ids))}) ORDER BY id", ids)
    rows_data = c.fetchall(); conn.close()

    i = 0
    for r in rows_data:
        x = i % cols
        y = i // cols
        if y >= rows:
            cpdf.showPage()
            y = 0; i = 0
        pos_x = margin_x + x*cell_w
        pos_y = h - margin_y - (y+1)*cell_h
        cpdf.rect(pos_x, pos_y, cell_w, cell_h)
        cpdf.setFont("Helvetica-Bold", 10)
        cpdf.drawString(pos_x+5*mm, pos_y+cell_h-10*mm, f"EQ #{r[0]}")
        cpdf.setFont("Helvetica", 9)
        cpdf.drawString(pos_x+5*mm, pos_y+cell_h-16*mm, (r[1] or "")[:32])
        if r[2]:
            cpdf.drawString(pos_x+5*mm, pos_y+cell_h-22*mm, f"TAG: {r[2][:24]}")
        qr_code = qr.QrCodeWidget(f"SGE:EQ:{r[0]}")
        bounds = qr_code.getBounds()
        size = min(cell_w, cell_h) * 0.5
        width = bounds[2] - bounds[0]; height = bounds[3] - bounds[1]
        scale = size / max(width, height)
        drawing = Drawing(size, size, transform=[scale,0,0,scale,0,0])
        drawing.add(qr_code)
        renderPDF.draw(drawing, cpdf, pos_x + cell_w - size - 5*mm, pos_y + 5*mm)
        i += 1

    cpdf.showPage(); cpdf.save()
    data = buf.getvalue(); buf.close()
    return Response(data, mimetype="application/pdf",
                    headers={"Content-Disposition":"attachment; filename=labels_equipamentos.pdf"})


@app.route('/equipamentos/import/xlsx', methods=['GET','POST'])
def importar_equipamentos_xlsx():
    if request.method == 'POST':
        file = request.files.get('xlsx')
        if not file or file.filename == '':
            flash("Selecione um ficheiro XLSX.", "danger")
            return redirect(url_for('importar_equipamentos_xlsx'))
        from openpyxl import load_workbook
        data = io.BytesIO(file.read())
        wb = load_workbook(filename=data, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h:i for i,h in enumerate(headers)}
        def get(row, key):
            i = idx.get(key)
            return (str(row[i].value).strip() if (i is not None and row[i].value is not None) else '')
        inserted=0; skipped=0
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        for row in ws.iter_rows(min_row=2):
            nome = get(row,'nome')
            if not nome: skipped+=1; continue
            local_nome = get(row,'local')
            local_id = None
            if local_nome:
                c.execute("SELECT id FROM locais WHERE nome=?", (local_nome,))
                r = c.fetchone()
                if r: local_id = r[0]
            tag = get(row,'tag')
            especificacao = get(row,'especificacao') or get(row,'especificação')
            ano = get(row,'ano') or get(row,'ano_instalacao')
            quantidade = get(row,'quantidade') or '1'
            categoria = get(row,'categoria')
            fabricante = get(row,'fabricante')
            modelo = get(row,'modelo')
            numero_serie = get(row,'numero_serie') or get(row,'nº série')
            custo_aquisicao = get(row,'custo_aquisicao')
            vida_util_anos = get(row,'vida_util_anos')
            criticidade = get(row,'criticidade')

            try: qtd_i = int(quantidade)
            except: qtd_i = 1
            try: custo_val = float(custo_aquisicao) if custo_aquisicao else None
            except: custo_val = None
            try: vida_val = int(vida_util_anos) if vida_util_anos else None
            except: vida_val = None

            c.execute('''
                INSERT INTO equipamentos (nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo, created_at, updated_at,
                                          categoria, fabricante, modelo, numero_serie, custo_aquisicao, vida_util_anos, criticidade)
                VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'), datetime('now','localtime'),
                        ?, ?, ?, ?, ?, ?, ?)
            ''', (nome, local_id, tag, especificacao, ano or None, qtd_i, categoria, fabricante, modelo, numero_serie, custo_val, vida_val, criticidade))
            inserted += 1
        conn.commit(); conn.close()
        flash(f"Importação XLSX concluída: {inserted} inseridos, {skipped} ignorados.", "success")
        return redirect(url_for('listar_equipamentos'))
    return render_template('importar_equipamentos_xlsx.html')


@app.route('/equipamentos/<int:equipamento_id>/upload_zip', methods=['POST'])
def equipamento_upload_zip(equipamento_id):
    file = request.files.get('zip')
    if not file or file.filename == '':
        flash("Selecione um ZIP.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    zdata = io.BytesIO(file.read())
    try:
        with zipfile.ZipFile(zdata) as z:
            added = 0
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            for name in z.namelist():
                if name.endswith('/'):
                    continue
                data = z.read(name)
                # create a FileStorage-like wrapper
                class _FS:
                    def __init__(self, filename, data):
                        self.filename = filename
                        self.data = data
                        self.mimetype = "image/jpeg"
                    def save(self, path):
                        with open(path, 'wb') as f: f.write(self.data)
                fs = _FS(os.path.basename(name), data)
                filename, thumb_name, w, h = _save_image_and_thumb(fs, dest_dir=app.config['UPLOAD_FOLDER']+'/photos', thumb_dir=app.config['UPLOAD_FOLDER']+'/thumbs')
                c.execute('INSERT INTO equipamentos_photos (equipamento_id, filename, thumb_filename, width, height) VALUES (?,?,?,?,?)',
                          (equipamento_id, 'photos/'+filename, ('thumbs/'+thumb_name if thumb_name else None), w, h))
                added += 1
            conn.commit(); conn.close()
            flash(f"{added} foto(s) importadas do ZIP.", "success")
    except Exception as e:
        flash(f"ZIP inválido: {e}", "danger")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))


# === SETTINGS HELPER ===
def get_setting(key, default='0'):
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("SELECT value FROM settings WHERE key=?", (key,))
        r = c.fetchone(); conn.close()
        return r[0] if r and r[0] is not None else default
    except Exception:
        return default


def _apply_advanced_query(q, where_clauses, params):
    """Operadores suportados:
    fabricante:, categoria:, local:, tag:, modelo:, serie:/nserie:,
    crit:/criticidade:, ano>=, ano<=, além de termos livres.
    """
    import shlex
    try:
        parts = shlex.split(q)
    except Exception:
        parts = q.split()
    rest = []
    for p in parts:
        pl = p.lower()
        if pl.startswith('fabricante:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('categoria:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('local:'):
            v = p.split(':',1)[1]; where_clauses.append("EXISTS (SELECT 1 FROM locais lx WHERE lx.id=e.local_id AND lx.nome LIKE ?)"); params.append(f"%{v}%"); continue
        if pl.startswith('tag:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.tag,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('modelo:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('serie:') or pl.startswith('nserie:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.numero_serie,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('crit:') or pl.startswith('criticidade:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.criticidade,'') = ?"); params.append(v); continue
        if pl.startswith('ano>='):
            try:
                v = int(p.split('>=',1)[1]); where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(v); continue
            except: pass
        if pl.startswith('ano<='):
            try:
                v = int(p.split('<=',1)[1]); where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(v); continue
            except: pass
        rest.append(p)
    if rest:
        like_q = "%" + " ".join(rest) + "%"
        where_clauses.append("(e.nome LIKE ? OR e.tag LIKE ? OR e.especificacao LIKE ? OR e.modelo LIKE ? OR e.fabricante LIKE ?)")
        params.extend([like_q, like_q, like_q, like_q, like_q])


# === EQUIPAMENTOS: Export JSON ===
@app.route('/equipamentos/export/json')
def equipamentos_export_json():
    ids = request.args.get('ids','').strip()
    where = []; params = []
    if ids:
        try:
            arr = [int(x) for x in ids.split(',') if x.strip().isdigit()]
            if arr:
                where.append("e.id IN (" + ",".join(["?"]*len(arr)) + ")")
                params += arr
        except Exception:
            pass
    if not where:
        q = request.args.get('q','').strip()
        categoria = request.args.get('categoria','').strip()
        fabricante = request.args.get('fabricante','').strip()
        modelo = request.args.get('modelo','').strip()
        criticidade = request.args.get('criticidade','').strip()
        local_id = request.args.get('local_id','').strip()
        ano_min = request.args.get('ano_min','').strip()
        ano_max = request.args.get('ano_max','').strip()
        incluir_inativos = request.args.get('incluir_inativos','0')=='1'
        if q: _apply_advanced_query(q, where, params)
        if local_id and local_id.isdigit(): where.append("e.local_id=?"); params.append(int(local_id))
        if categoria: where.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
        if fabricante: where.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
        if modelo: where.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
        if criticidade: where.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
        if ano_min and ano_min.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
        if ano_max and ano_max.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
        if not incluir_inativos: where.append("COALESCE(e.ativo,1)=1")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.*, COALESCE(l.nome,'') AS local_nome
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    cols = [d[0] for d in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    conn.close()
    return Response(json.dumps(rows, ensure_ascii=False, indent=2), mimetype="application/json")


# === EQUIPAMENTOS: Import JSON ===
@app.route('/equipamentos/import/json', methods=['GET','POST'])
def equipamentos_import_json():
    if request.method == 'POST':
        file = request.files.get('json')
        if not file or file.filename == '':
            flash("Selecione um ficheiro JSON.", "danger")
            return redirect(url_for('equipamentos_import_json'))
        try:
            data = json.loads(file.read().decode('utf-8', errors='ignore'))
            if not isinstance(data, list):
                raise ValueError("JSON deve ser uma lista de objetos.")
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            inserted=0; skipped=0
            for obj in data:
                nome = (obj.get('nome') or obj.get('NOME') or '').strip()
                if not nome: skipped+=1; continue
                local_id = obj.get('local_id')
                tag = obj.get('tag') or ''
                especificacao = obj.get('especificacao') or obj.get('especificação') or ''
                ano = obj.get('ano_instalacao') or obj.get('ano') or None
                qtd = obj.get('quantidade') or 1
                categoria = obj.get('categoria') or ''
                fabricante = obj.get('fabricante') or ''
                modelo = obj.get('modelo') or ''
                numero_serie = obj.get('numero_serie') or ''
                custo = obj.get('custo_aquisicao')
                vida = obj.get('vida_util_anos')
                criticidade = obj.get('criticidade') or ''
                try: qtd_i = int(qtd)
                except: qtd_i = 1
                try: custo_val = float(custo) if custo is not None else None
                except: custo_val = None
                try: vida_val = int(vida) if vida is not None else None
                except: vida_val = None
                c.execute('''
                    INSERT INTO equipamentos (nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo, created_at, updated_at,
                                              categoria, fabricante, modelo, numero_serie, custo_aquisicao, vida_util_anos, criticidade)
                    VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'), datetime('now','localtime'),
                            ?, ?, ?, ?, ?, ?, ?)
                ''', (nome, local_id, tag, especificacao, ano, qtd_i, categoria, fabricante, modelo, numero_serie, custo_val, vida_val, criticidade))
                inserted+=1
            conn.commit(); conn.close()
            flash(f"Importação JSON concluída: {inserted} inseridos, {skipped} ignorados.", "success")
            return redirect(url_for('listar_equipamentos'))
        except Exception as e:
            flash(f"JSON inválido: {e}", "danger")
            return redirect(url_for('equipamentos_import_json'))
    return render_template('importar_equipamentos_json.html')


# === XLSX template & export filtrado ===
@app.route('/equipamentos/export/xlsx_template')
def equipamentos_xlsx_template():
    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out, {'in_memory': True})
    _set_xlsx_identity(wb, 'Modelo de Importação de Equipamentos')
    ws = wb.add_worksheet('Equipamentos')
    headers = ["nome","local","tag","especificacao","ano","quantidade","categoria","fabricante","modelo","numero_serie","custo_aquisicao","vida_util_anos","criticidade"]
    for i,hx in enumerate(headers): ws.write(0, i, hx)
    ws.data_validation(1, 12, 10000, 12, {'validate': 'list', 'source': ['Baixa','Média','Alta']})
    wb.close(); out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=template_equipamentos.xlsx"})


@app.route('/equipamentos/export/xlsx')
def equipamentos_export_xlsx():
    q = request.args.get('q','').strip()
    local_id = request.args.get('local_id','').strip()
    incluir_inativos = request.args.get('incluir_inativos','0')=='1'
    categoria = request.args.get('categoria','').strip()
    fabricante = request.args.get('fabricante','').strip()
    modelo = request.args.get('modelo','').strip()
    criticidade = request.args.get('criticidade','').strip()
    ano_min = request.args.get('ano_min','').strip()
    ano_max = request.args.get('ano_max','').strip()

    where=[]; params=[]
    if q: _apply_advanced_query(q, where, params)
    if local_id and local_id.isdigit(): where.append("e.local_id=?"); params.append(int(local_id))
    if categoria: where.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
    if fabricante: where.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
    if modelo: where.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
    if criticidade: where.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
    if ano_min and ano_min.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
    if ano_max and ano_max.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
    if not incluir_inativos: where.append("COALESCE(e.ativo,1)=1")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.nome, COALESCE(l.nome,''), e.tag, e.especificacao, e.ano_instalacao, e.quantidade,
               e.categoria, e.fabricante, e.modelo, e.numero_serie, e.custo_aquisicao, e.vida_util_anos, e.criticidade
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    rows = c.fetchall(); conn.close()

    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out, {'in_memory': True})
    _set_xlsx_identity(wb, 'Exportação de Equipamentos')
    ws = wb.add_worksheet('Equipamentos')
    headers = ["nome","local","tag","especificacao","ano","quantidade","categoria","fabricante","modelo","numero_serie","custo_aquisicao","vida_util_anos","criticidade"]
    for i,hx in enumerate(headers): ws.write(0, i, hx)
    for r_i, r in enumerate(rows, start=1):
        for c_i, v in enumerate(r): ws.write(r_i, c_i, v)
    ws.data_validation(1, 12, 10000, 12, {'validate': 'list', 'source': ['Baixa','Média','Alta']})
    wb.close(); out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=equipamentos_filtrados.xlsx"})


# === Relatório Consolidado (PDF) ===
@app.route('/equipamentos/export/relatorio.pdf')
def equipamentos_relatorio_pdf():
    q = request.args.get('q','').strip()
    local_id = request.args.get('local_id','').strip()
    incluir_inativos = request.args.get('incluir_inativos','0')=='1'
    categoria = request.args.get('categoria','').strip()
    fabricante = request.args.get('fabricante','').strip()
    modelo = request.args.get('modelo','').strip()
    criticidade = request.args.get('criticidade','').strip()
    ano_min = request.args.get('ano_min','').strip()
    ano_max = request.args.get('ano_max','').strip()

    where_clauses=[]; params=[]
    if q: _apply_advanced_query(q, where_clauses, params)
    if local_id and local_id.isdigit(): where_clauses.append("e.local_id=?"); params.append(int(local_id))
    if categoria: where_clauses.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
    if fabricante: where_clauses.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
    if modelo: where_clauses.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
    if criticidade: where_clauses.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
    if ano_min and ano_min.isdigit(): where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
    if ano_max and ano_max.isdigit(): where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
    if not incluir_inativos: where_clauses.append("COALESCE(e.ativo,1)=1")
    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.id, e.nome, COALESCE(l.nome,''), e.tag, e.especificacao, e.ano_instalacao, e.quantidade,
               COALESCE(e.categoria,''), COALESCE(e.fabricante,''), COALESCE(e.modelo,''), COALESCE(e.criticidade,''),
               COALESCE(e.custo_aquisicao,0.0)
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    rows = c.fetchall(); conn.close()

    from collections import Counter
    def summarize(values):
        c = Counter([v or '' for v in values])
        return sorted(c.items(), key=lambda x:(-x[1], x[0]))

    sum_local = summarize([r[2] for r in rows])
    sum_cat = summarize([r[7] for r in rows])
    sum_fab = summarize([r[8] for r in rows])

    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=A4)
    _set_pdf_identity(cpdf, 'Relatório Consolidado de Equipamentos')
    w, h = A4
    from reportlab.lib.units import cm
    cpdf.setFont("Helvetica-Bold", 14)
    cpdf.drawString(2*cm, h-2*cm, "Relatório Consolidado - Equipamentos")
    cpdf.setFont("Helvetica", 9)
    cpdf.drawString(2*cm, h-2.7*cm, f"Filtros: q='{q}' local_id='{local_id}' cat='{categoria}' fab='{fabricante}' mod='{modelo}' crit='{criticidade}' ano[{ano_min},{ano_max}]")

    y = h-3.5*cm
    cpdf.setFont("Helvetica-Bold", 10); cpdf.drawString(2*cm, y, "Sumário por Local"); y -= 0.5*cm
    cpdf.setFont("Helvetica", 9)
    for k,v in sum_local[:22]:
        cpdf.drawString(2*cm, y, f"{k or '-'}: {v}"); y -= 0.4*cm
        if y < 3*cm: cpdf.showPage(); y=h-2*cm

    cpdf.setFont("Helvetica-Bold", 10); cpdf.drawString(10*cm, h-3.5*cm, "Sumário por Categoria")
    y2 = h-4.0*cm; cpdf.setFont("Helvetica", 9)
    for k,v in sum_cat[:22]:
        cpdf.drawString(10*cm, y2, f"{k or '-'}: {v}"); y2 -= 0.4*cm

    cpdf.showPage()
    cpdf.setFont("Helvetica-Bold", 10)
    cpdf.drawString(2*cm, h-2*cm, "Lista Detalhada")
    y = h-2.8*cm; cpdf.setFont("Helvetica", 8)
    headers = ["ID","Nome","Local","TAG","Ano","Qtd","Cat","Fab","Mod","Crit","Custo","Custo Total"]
    col = [1.0, 2.0, 7.0, 11.0, 15.0, 16.5, 18.0, 20.0, 23.0, 26.0, 28.0, 31.0]
    for i,hx in enumerate(headers): cpdf.drawString(col[i]*cm, y, hx)
    y -= 0.4*cm
    total_custo = 0.0
    for r in rows:
        custo = float(r[11] or 0.0); qtd = int(r[6] or 0)
        custo_total = custo * qtd; total_custo += custo_total
        vals = [r[0], r[1], r[2], r[3] or "", r[5] or "", qtd, r[7] or "", r[8] or "", r[9] or "", r[10] or "", f"{custo:.2f}", f"{custo_total:.2f}"]
        for i,v in enumerate(vals):
            cpdf.drawString(col[i]*cm, y, str(v)[:18])
        y -= 0.35*cm
        if y < 2.5*cm:
            cpdf.drawString(28.0*cm, 1.5*cm, f"Total custo: {total_custo:.2f}")
            cpdf.showPage(); cpdf.setFont("Helvetica", 8); y = h-2.5*cm
    cpdf.drawString(28.0*cm, 1.5*cm, f"Total custo: {total_custo:.2f}")
    cpdf.showPage(); cpdf.save()
    data = buf.getvalue(); buf.close()
    return Response(data, mimetype="application/pdf",
                    headers={"Content-Disposition":"attachment; filename=relatorio_equipamentos.pdf"})


# === Links: add/remove ===
@app.route('/equipamentos/<int:equipamento_id>/links/add', methods=['POST'])
def equipamento_add_link(equipamento_id):
    urlv = (request.form.get('url') or '').strip()
    title = (request.form.get('title') or '').strip()
    if not urlv:
        flash("URL é obrigatório.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("INSERT INTO equipamentos_links (equipamento_id, url, title) VALUES (?,?,?)", (equipamento_id, urlv, title or urlv))
    conn.commit(); conn.close()
    flash("Link adicionado.", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/<int:equipamento_id>/links/<int:link_id>/delete', methods=['POST'])
def equipamento_del_link(equipamento_id, link_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("DELETE FROM equipamentos_links WHERE id=? AND equipamento_id=?", (link_id, equipamento_id))
    conn.commit(); conn.close()
    flash("Link removido.", "warning")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))


# === API simples com cover_thumb ===


@app.route('/api/equipamentos')
def api_equipamentos():
    q = request.args.get('q','').strip()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    if q:
        like = f"%{q}%"
        c.execute('''
            SELECT e.id, e.nome, COALESCE(l.nome,''), COALESCE(e.tag,''),
                   COALESCE(cp.thumb_filename,(SELECT thumb_filename FROM equipamentos_photos WHERE equipamento_id=e.id ORDER BY uploaded_at DESC LIMIT 1),'') as cover_thumb
            FROM equipamentos e
            LEFT JOIN locais l ON e.local_id=l.id
            LEFT JOIN equipamentos_photos cp ON cp.id = e.cover_photo_id
            WHERE e.nome LIKE ? OR COALESCE(e.tag,'') LIKE ? OR COALESCE(l.nome,'') LIKE ?
            ORDER BY e.nome ASC
            LIMIT 200
        ''', (like, like, like))
    else:
        c.execute('''
            SELECT e.id, e.nome, COALESCE(l.nome,''), COALESCE(e.tag,''),
                   COALESCE(cp.thumb_filename,(SELECT thumb_filename FROM equipamentos_photos WHERE equipamento_id=e.id ORDER BY uploaded_at DESC LIMIT 1),'') as cover_thumb
            FROM equipamentos e
            LEFT JOIN locais l ON e.local_id=l.id
            LEFT JOIN equipamentos_photos cp ON cp.id = e.cover_photo_id
            ORDER BY e.nome ASC
            LIMIT 200
        ''')
    data = [{"id": r[0], "nome": r[1], "local": r[2], "tag": r[3], "cover_thumb": r[4]} for r in c.fetchall()]
    conn.close()
    return Response(json.dumps(data, ensure_ascii=False), mimetype='application/json')


@app.route('/equipamentos/<int:equipamento_id>/files/<int:file_id>/delete', methods=['POST'])
def delete_equip_file(equipamento_id, file_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT filename, original_name FROM equipamentos_files WHERE id=? AND equipamento_id=?", (file_id, equipamento_id))
    row = c.fetchone()
    if not row:
        conn.close(); flash("Documento não encontrado.", "warning")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], row[0])
        if os.path.exists(filepath):
            os.remove(filepath)
        c.execute("DELETE FROM equipamentos_files WHERE id=?", (file_id,))
        conn.commit()
        flash("Documento removido.", "success")
        log_equip_audit(equipamento_id, "delete_file", row[1] or row[0])
    except Exception as ex:
        flash(f"Erro ao apagar documento: {ex}", "danger")
    finally:
        conn.close()
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))


@app.route('/api/equipamentos/<int:equipamento_id>')
def api_equipamento_detalhe(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT id, nome, local_id, tag, especificacao, ano_instalacao, quantidade, categoria, custo_aquisicao, criticidade, fabricante, modelo, numero_serie, vida_util_anos, ativo, potencia_kw, tensao_v, corrente_a, fornecedor, contrato_num, garantia_fim FROM equipamentos WHERE id=?", (equipamento_id,))
    e = c.fetchone()
    if not e:
        conn.close(); return Response(json.dumps({"error":"not found"}), status=404, mimetype="application/json")
    c.execute("SELECT id, filename, caption FROM equipamentos_photos WHERE equipamento_id=? ORDER BY uploaded_at DESC", (equipamento_id,))
    photos = c.fetchall()
    c.execute("SELECT id, filename, size, original_name FROM equipamentos_files WHERE equipamento_id=? ORDER BY uploaded_at DESC", (equipamento_id,))
    files = c.fetchall()
    c.execute("SELECT id, url, title FROM equipamentos_links WHERE equipamento_id=? ORDER BY added_at DESC", (equipamento_id,))
    links = c.fetchall()
    conn.close()
    return Response(json.dumps({
        "equipamento": e, "photos": photos, "files": files, "links": links
    }, ensure_ascii=False), mimetype="application/json")

def _equip_where_from_request(prefer_ids=True):
    ids = (request.args.get('ids') or '').strip()
    q = (request.args.get('q') or '').strip()
    local_id = (request.args.get('local_id') or '').strip()
    incluir_inativos = (request.args.get('incluir_inativos','0') == '1')
    categoria = (request.args.get('categoria') or '').strip()
    fabricante = (request.args.get('fabricante') or '').strip()
    modelo = (request.args.get('modelo') or '').strip()
    criticidade = (request.args.get('criticidade') or '').strip()
    ano_min = (request.args.get('ano_min') or '').strip()
    ano_max = (request.args.get('ano_max') or '').strip()
    where = []; params = []
    if prefer_ids and ids:
        try:
            arr = [int(x) for x in ids.split(',') if x.strip().isdigit()]
            if arr:
                where.append("e.id IN (" + ",".join(["?"]*len(arr)) + ")")
                params += arr
        except Exception:
            pass
    if not where:
        if q:
            _apply_advanced_query(q, where, params)
        if local_id and local_id.isdigit():
            where.append("e.local_id=?"); params.append(int(local_id))
        if categoria:
            where.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
        if fabricante:
            where.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
        if modelo:
            where.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
        if criticidade:
            where.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
        if ano_min and ano_min.isdigit():
            where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
        if ano_max and ano_max.isdigit():
            where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
        if not incluir_inativos:
            where.append("COALESCE(e.ativo,1)=1")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    return where_sql, params

# ==== SOFT DELETE ====
@app.route('/equipamentos/<int:equipamento_id>/arquivar', methods=['POST'])
def equipamentos_arquivar(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET deleted_at=datetime('now','localtime') WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    flash("Equipamento arquivado.", "warning")
    return redirect(url_for('listar_equipamentos'))

@app.route('/equipamentos/<int:equipamento_id>/restaurar', methods=['POST'])
def equipamentos_restaurar(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET deleted_at=NULL WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    flash("Equipamento restaurado.", "success")
    return redirect(url_for('listar_equipamentos'))
# ==== /SOFT DELETE ====


# ==== HISTÓRICO ====
@app.route('/equipamentos/<int:equipamento_id>/historico')
def equipamentos_historico(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    try:
        c.execute("SELECT acao, detalhes, datetime(ts,'localtime') FROM equipamentos_audit WHERE equipamento_id=? ORDER BY ts DESC", (equipamento_id,))
        rows = c.fetchall()
    except Exception:
        rows = []
    conn.close()
    return render_template('equipamentos_historico.html', equipamento_id=equipamento_id, rows=rows)
# ==== /HISTÓRICO ====


# ==== IMPORT PREVIEW ====
@app.route('/equipamentos/import/preview', methods=['GET','POST'])
def equipamentos_import_preview():
    if request.method == 'GET':
        return render_template('importar_equipamentos_preview.html')
    file = request.files.get('file')
    ftype = (request.form.get('tipo') or '').lower()
    if not file or file.filename == '' or ftype not in ('csv','xlsx','json'):
        flash("Selecione um ficheiro e o tipo correto (csv/xlsx/json).", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    import csv, io, json
    rows, errors = [], []
    try:
        if ftype=='csv':
            content = file.read().decode('utf-8', errors='ignore')
            rd = csv.DictReader(io.StringIO(content))
            for i,row in enumerate(rd, start=2):
                nome = (row.get('nome') or row.get('Nome') or '').strip()
                if not nome: errors.append((i,"Nome vazio")); continue
                rows.append(row)
        elif ftype=='xlsx':
            from openpyxl import load_workbook
            data = io.BytesIO(file.read()); wb = load_workbook(filename=data, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1,max_row=1))]
            for r_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                obj = {headers[i]: (str(cell.value).strip() if cell.value is not None else '') for i,cell in enumerate(row)}
                if not (obj.get('nome') or ''): errors.append((r_i,"Nome vazio")); continue
                rows.append(obj)
        else:
            data = json.loads(file.read().decode('utf-8', errors='ignore'))
            if not isinstance(data, list): raise ValueError("JSON deve ser uma lista")
            for i,obj in enumerate(data, start=1):
                nome = (obj.get('nome') or obj.get('Nome') or '').strip()
                if not nome: errors.append((i,"Nome vazio")); continue
                rows.append(obj)
    except Exception as e:
        flash(f"Erro a ler ficheiro: {e}", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    preview_key = f"equip_preview_{int(datetime.timestamp(datetime.now()))}"
    tmp_path = os.path.join(UPLOAD_DIR, preview_key + '.json')
    try:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump({'ftype': ftype, 'rows': rows}, f, ensure_ascii=False)
    except Exception:
        tmp_path = None
    return render_template('importar_equipamentos_preview.html', rows=rows[:200], errors=errors, preview_key=preview_key)

@app.route('/equipamentos/import/confirm', methods=['POST'])
def equipamentos_import_confirm():
    key = (request.form.get('preview_key') or '').strip()
    if not key:
        flash("Pré-visualização expirada.", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    path = os.path.join(UPLOAD_DIR, key + '.json')
    if not os.path.exists(path):
        flash("Pré-visualização não encontrada.", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        flash(f"Erro ao carregar pré-visualização: {e}", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    rows = data.get('rows') or []
    inserted=0; skipped=0
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    for row in rows:
        nome = (row.get('nome') or row.get('Nome') or '').strip()
        if not nome: skipped+=1; continue
        local_nome = (row.get('local') or row.get('Local') or '').strip()
        local_id = None
        if local_nome:
            c.execute("SELECT id FROM locais WHERE nome=?", (local_nome,)); r = c.fetchone()
            if r: local_id = r[0]
        tag = (row.get('tag') or row.get('TAG') or '').strip()
        especificacao = (row.get('especificacao') or row.get('Especificacao') or row.get('Especificação') or '').strip()
        ano = (row.get('ano_instalacao') or row.get('ano') or '').strip() or None
        try: qtd_i = int((row.get('quantidade') or row.get('qtd') or '1').strip())
        except: qtd_i = 1
        c.execute('''
            INSERT INTO equipamentos (nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'), datetime('now','localtime'))
        ''', (nome, local_id, tag, especificacao, ano, qtd_i))
        inserted+=1
    conn.commit(); conn.close()
    try: os.remove(path)
    except Exception: pass
    flash(f"Importação concluída: {inserted} inseridos, {skipped} ignorados.", "success")
    return redirect(url_for('listar_equipamentos'))
# ==== /IMPORT PREVIEW ====


# ==== DASHBOARD ====
@app.route('/equipamentos/dashboard')
def equipamentos_dashboard():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM equipamentos WHERE COALESCE(deleted_at,'')=''"); total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM equipamentos WHERE COALESCE(ativo,1)=1 AND COALESCE(deleted_at,'')=''"); ativos = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM equipamentos WHERE COALESCE(ativo,1)=0 AND COALESCE(deleted_at,'')=''"); inativos = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM equipamentos WHERE garantia_fim IS NOT NULL AND garantia_fim<>'' AND date(garantia_fim)>=date('now') AND COALESCE(deleted_at,'')=''")
    em_garantia = c.fetchone()[0]
    c.execute("SELECT COALESCE(categoria,''), COUNT(*) FROM equipamentos WHERE COALESCE(deleted_at,'')='' GROUP BY COALESCE(categoria,'') ORDER BY COUNT(*) DESC")
    por_cat = c.fetchall()
    c.execute("SELECT COALESCE(l.nome,''), COUNT(*) FROM equipamentos e LEFT JOIN locais l ON e.local_id=l.id WHERE COALESCE(e.deleted_at,'')='' GROUP BY COALESCE(l.nome,'') ORDER BY COUNT(*) DESC")
    por_local = c.fetchall()
    conn.close()
    return render_template('equipamentos_dashboard.html', total=total, ativos=ativos, inativos=inativos, em_garantia=em_garantia, por_cat=por_cat, por_local=por_local)
# ==== /DASHBOARD ====


# ==== FILTROS SALVOS ====
@app.route('/equipamentos/filtros', methods=['GET'])
def equipamentos_filtros():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT id, user, nome, query_json, datetime(created_at,'localtime') FROM saved_filters WHERE modulo='equipamentos' ORDER BY created_at DESC")
    rows = c.fetchall(); conn.close()
    return render_template('equipamentos_filtros.html', rows=rows)

@app.route('/equipamentos/filtros/salvar', methods=['POST'])
def equipamentos_filtros_salvar():
    user = (request.form.get('user') or 'admin').strip()
    nome = (request.form.get('nome') or '').strip()
    query_json = (request.form.get('query_json') or '').strip()
    if not nome or not query_json:
        flash("Informe nome e filtros.", "danger")
        return redirect(url_for('equipamentos_filtros'))
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("INSERT INTO saved_filters (user, modulo, nome, query_json) VALUES (?, 'equipamentos', ?, ?)", (user, nome, query_json))
    conn.commit(); conn.close()
    flash("Filtro salvo.", "success")
    return redirect(url_for('equipamentos_filtros'))
# ==== /FILTROS SALVOS ====


# ==== ETIQUETAS CUSTOM ====
@app.route('/equipamentos/labels/options')
def equipamentos_labels_options():
    return render_template('equipamentos_labels_options.html')

@app.route('/equipamentos/labels_custom.pdf')
def equipamentos_labels_custom():
    ids_str = (request.args.get('ids') or '').strip()
    if not ids_str:
        flash("IDs obrigatórios", "warning")
        return redirect(url_for('listar_equipamentos'))
    ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
    show_modelo = (request.args.get('modelo','0')=='1')
    show_serie = (request.args.get('serie','0')=='1')
    layout = (request.args.get('layout') or 'A4_3x8')
    include_logo = (request.args.get('logo','0')=='1')

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF
    import math

    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=A4)
    _set_pdf_identity(cpdf, 'Etiquetas Personalizadas de Equipamentos')
    w, h = A4

    cols, rows = (3,8) if layout=='A4_3x8' else (2,6)
    margin_x, margin_y = 10, 10
    cell_w = (w - 2*margin_x) / cols
    cell_h = (h - 2*margin_y) / rows

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    qmarks = ",".join(["?"]*len(ids))
    c.execute(f"SELECT id, nome, COALESCE(tag,''), COALESCE(modelo,''), COALESCE(numero_serie,'') FROM equipamentos WHERE id IN ({qmarks}) ORDER BY id", ids)
    rows_data = c.fetchall(); conn.close()

    i = 0
    for r in rows_data:
        x = i % cols; y = i // cols
        if y >= rows:
            cpdf.showPage(); y = 0; i = 0
        pos_x = margin_x + x*cell_w
        pos_y = h - margin_y - (y+1)*cell_h
        cpdf.rect(pos_x, pos_y, cell_w, cell_h)
        cpdf.setFont("Helvetica-Bold", 10)
        cpdf.drawString(pos_x+6, pos_y+cell_h-14, f"EQ #{r[0]}")
        cpdf.setFont("Helvetica", 9)
        cpdf.drawString(pos_x+6, pos_y+cell_h-26, (r[1] or "")[:36])
        if r[2]:
            cpdf.drawString(pos_x+6, pos_y+cell_h-38, f"TAG: {r[2][:26]}")
        yoff = 50
        if show_modelo and r[3]:
            cpdf.drawString(pos_x+6, pos_y+cell_h-yoff, f"Modelo: {r[3][:26]}"); yoff += 12
        if show_serie and r[4]:
            cpdf.drawString(pos_x+6, pos_y+cell_h-yoff, f"Série: {r[4][:26]}"); yoff += 12
        qr_code = qr.QrCodeWidget(f"SGE:EQ:{r[0]}")
        bounds = qr_code.getBounds()
        size = min(cell_w, cell_h) * 0.40
        width = bounds[2] - bounds[0]; height = bounds[3] - bounds[1]
        scale = size / max(width, height)
        drawing = Drawing(size, size, transform=[scale,0,0,scale,0,0])
        drawing.add(qr_code)
        renderPDF.draw(drawing, cpdf, pos_x + cell_w - size - 8, pos_y + 8)
        if include_logo:
            cpdf.setFont("Helvetica-Bold", 8)
            cpdf.drawString(pos_x+6, pos_y+6, "LOGO")
        i += 1

    cpdf.showPage(); cpdf.save()
    data = buf.getvalue(); buf.close()
    return Response(data, mimetype="application/pdf",
                    headers={"Content-Disposition":"attachment; filename=labels_custom.pdf"})
# ==== /ETIQUETAS CUSTOM ====


# ==== API v2 (GET) ====
@app.route('/api/v2/equipamentos')
def api_v2_equip_list():
    where=[]; params=[]
    q = (request.args.get('q') or '').strip()
    if q:
        _apply_advanced_query(q, where, params)
    local_id = (request.args.get('local_id') or '').strip()
    categoria = (request.args.get('categoria') or '').strip()
    fabricante = (request.args.get('fabricante') or '').strip()
    modelo = (request.args.get('modelo') or '').strip()
    criticidade = (request.args.get('criticidade') or '').strip()
    ano_min = (request.args.get('ano_min') or '').strip()
    ano_max = (request.args.get('ano_max') or '').strip()
    incluir_inativos = (request.args.get('incluir_inativos','0')=='1')
    if local_id.isdigit(): where.append("e.local_id=?"); params.append(int(local_id))
    if categoria: where.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
    if fabricante: where.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
    if modelo: where.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
    if criticidade: where.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
    if ano_min.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
    if ano_max.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
    if not incluir_inativos: where.append("COALESCE(e.ativo,1)=1")
    where.append("COALESCE(e.deleted_at,'')=''")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f"""
        SELECT e.id, e.nome, e.local_id, COALESCE(l.nome,''), e.tag, e.modelo, e.fabricante,
               e.criticidade, e.ativo, e.deleted_at, e.garantia_fim
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id=l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    """, params)
    rows = c.fetchall(); conn.close()
    data = [{
        "id":r[0],"nome":r[1],"local_id":r[2],"local":r[3],"tag":r[4],"modelo":r[5],"fabricante":r[6],
        "criticidade":r[7],"ativo":r[8],"arquivado": bool(r[9]), "garantia_fim": r[10]
    } for r in rows]
    return jsonify(data)

@app.route('/api/v2/equipamentos/<int:equipamento_id>')
def api_v2_equip_detail(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""
        SELECT e.*, COALESCE(l.nome,'') as local_nome
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id=l.id
        WHERE e.id=?
    """, (equipamento_id,))
    e = c.fetchone()
    if not e:
        conn.close(); return jsonify({"error":"not found"}), 404
    c.execute("SELECT id, nome, COALESCE(fabricante,''), COALESCE(modelo,''), COALESCE(qtd,1) FROM equipamentos_componentes WHERE equipamento_id=? ORDER BY nome", (equipamento_id,))
    comps = c.fetchall()
    try:
        c.execute("SELECT id, filename, size FROM equipamentos_files WHERE equipamento_id=? ORDER BY uploaded_at DESC", (equipamento_id,)); files = c.fetchall()
    except Exception:
        files = []
    try:
        c.execute("SELECT id, filename, thumb_filename, width, height FROM equipamentos_photos WHERE equipamento_id=? ORDER BY uploaded_at DESC", (equipamento_id,)); photos = c.fetchall()
    except Exception:
        photos = []
    cols = [d[0] for d in c.description] if c.description else []
    conn.close()
    return jsonify({
        "equipamento": dict(zip(cols, e)) if cols else {},
        "componentes": [{"id":r[0],"nome":r[1],"fabricante":r[2],"modelo":r[3],"qtd":r[4]} for r in comps],
        "files": [{"id":r[0],"filename":r[1],"size":r[2]} for r in files],
        "photos": [{"id":r[0],"filename":r[1],"thumb":r[2],"w":r[3],"h":r[4]} for r in photos]
    })
# ==== /API v2 ====



@app.route('/equipamentos/filtros/delete/<int:fid>', methods=['POST'])
@app.route('/equipamentos/filtros/<int:fid>/delete', methods=['POST'])
def equipamentos_filtros_delete(fid):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("DELETE FROM saved_filters WHERE id=?", (fid,))
    conn.commit(); conn.close()
    flash("Filtro removido.", "warning")
    return redirect(url_for('equipamentos_filtros'))

# === Util: configuração completa por Local ===
def _db_columns(conn, table):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    return {row[1] for row in cur.fetchall()}

def _get_local_id_by_any(conn, any_id_or_name):
    cur = conn.cursor()
    try:
        lid = int(str(any_id_or_name).strip())
        cur.execute("SELECT id, nome FROM locais WHERE id = ?", (lid,))
        row = cur.fetchone()
        if row: return row[0], row[1]
    except Exception:
        pass
    cur.execute("SELECT id, nome FROM locais WHERE nome = ?", (str(any_id_or_name).strip(),))
    row = cur.fetchone()
    if row: return row[0], row[1]
    return None, None

def _get_local_cfg_full(conn, local_id):
    cfg = {
        "fator_mult": 1.0,
        "pot_contratada": 0.0,
        "pot_instalada": 0.0,
        "tarifa_ativa": 0.0,
        "tarifa_reativa": 0.0,
        "tarifa_ponta": 0.0,
        "tarifa_perdas": 0.0,
        "taxa_fixa": 0.0,
        "taxa_radio": 0.0,
        "taxa_lixo": 0.0,
        "iva": 0.0,
    }
    cur = conn.cursor()
    cols_cfg = _db_columns(conn, "locais_cfg")
    select_cols = [c for c in ["fator_mult","pot_contratada","pot_instalada",
                               "tarifa_ativa","tarifa_reativa","tarifa_ponta","tarifa_perdas",
                               "taxa_fixa","taxa_radio","taxa_lixo","iva"] if c in cols_cfg]
    if select_cols:
        sql = "SELECT " + ",".join(select_cols) + " FROM locais_cfg WHERE local_id = ?"
        cur.execute(sql, (local_id,))
        row = cur.fetchone()
        if row:
            for idx, col in enumerate(select_cols):
                cfg[col] = _to_float(row[idx], cfg[col])

    cols_locais = _db_columns(conn, "locais")
    if "pot_instalada" in cols_locais:
        cur.execute("SELECT pot_instalada FROM locais WHERE id = ?", (local_id,))
        r2 = cur.fetchone()
        if r2 and r2[0] is not None:
            cfg["pot_instalada"] = _to_float(r2[0], cfg["pot_instalada"])

    return cfg

