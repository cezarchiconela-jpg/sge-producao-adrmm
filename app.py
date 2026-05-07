
# --- Pack3: Validations per local + stronger audit table ---
def migrate_pack3():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    # Config table for validations per local
    c.execute('''CREATE TABLE IF NOT EXISTS validacoes_locais (
        local TEXT PRIMARY KEY,
        fp_min REAL DEFAULT 0.85,
        kwh_dia_max REAL,
        permitir_regressivo INTEGER DEFAULT 0
    )''')
    # Strengthen audit with actor and period
    c.execute('''CREATE TABLE IF NOT EXISTS leituras_mensais_audit (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        local TEXT, data TEXT, mes TEXT, ano INTEGER,
        field TEXT, old_value TEXT, new_value TEXT,
        acao TEXT,
        actor TEXT,
        ts TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.commit(); conn.close()


def get_validacao_local(local: str):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    row = c.execute("SELECT fp_min, kwh_dia_max, permitir_regressivo FROM validacoes_locais WHERE local=?",(local,)).fetchone()
    conn.close()
    if not row:
        return {'fp_min':0.85, 'kwh_dia_max':None, 'permitir_regressivo':0}
    return {'fp_min': float(row[0] or 0.85), 'kwh_dia_max': (float(row[1]) if row[1] is not None else None), 'permitir_regressivo': int(row[2] or 0)}

def set_validacao_local(local, fp_min, kwh_dia_max, permitir_regressivo):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''INSERT INTO validacoes_locais(local, fp_min, kwh_dia_max, permitir_regressivo)
                 VALUES(?,?,?,?)
                 ON CONFLICT(local) DO UPDATE SET
                   fp_min=excluded.fp_min, kwh_dia_max=excluded.kwh_dia_max, permitir_regressivo=excluded.permitir_regressivo''',
              (local, fp_min, kwh_dia_max, permitir_regressivo))
    conn.commit(); conn.close()

def log_audit(local, data, mes, ano, field, old, new, acao="update", actor="pack3"):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''INSERT INTO leituras_mensais_audit(local,data,mes,ano,field,old_value,new_value,acao,actor)
                 VALUES(?,?,?,?,?,?,?,?,?)''', (local, data, mes, ano, field, str(old), str(new), acao, actor))
    conn.commit(); conn.close()
from flask import Flask, request, render_template, redirect, url_for, Response, flash, jsonify, send_from_directory, g, session
import os
import secrets
print(">> SGE a arrancar a partir do ficheiro:", __file__)
print(">> Pasta atual:", os.getcwd())
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash
from PIL import Image
from PIL import Image, ImageOps
from reportlab.lib.units import cm, mm
import zipfile
import io
import time 
import sqlite3
import calendar
from datetime import datetime, timedelta
import math
from io import StringIO
import json
from jinja2 import TemplateNotFound
from io import BytesIO
import csv
import xlsxwriter
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from werkzeug.utils import secure_filename

app = Flask(__name__)
RATE_LIMIT_UPLOADS = {}

app.secret_key = os.environ.get('SECRET_KEY') or os.environ.get('FLASK_SECRET_KEY') or secrets.token_hex(32)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# === LOGGING ESTRUTURADO ===
import logging, uuid, time
from logging.handlers import RotatingFileHandler

def _setup_logging():
    logger = logging.getLogger('sge')
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        fh = RotatingFileHandler(os.path.join(BASE_DIR, 'sge.log'), maxBytes=1_000_000, backupCount=3)
        fmt = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
        fh.setFormatter(fmt)
        logger.addHandler(fh)
    return logger

_logger = None
try:
    _logger = _setup_logging()
    _logger.info("SGE logging inicializado")
except Exception as _e:
    print("Falha ao iniciar logging:", _e)

@app.before_request
def _add_request_context():
    # Contexto de correlação para cada request
    rid = str(uuid.uuid4())[:8]
    setattr(g, 'rid', rid)
    setattr(g, 't0', time.time())

@app.after_request
def _after_request(resp):
    try:
        dt = time.time() - getattr(g, 't0', time.time())
        _logger and _logger.info("RID=%s %s %s %s %.3fs", getattr(g,'rid','-'), request.remote_addr, request.method, request.path, dt)
    except Exception:
        pass
    # Cabeçalhos defensivos para ambiente online.
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    resp.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    resp.headers.setdefault('Permissions-Policy', 'geolocation=(), microphone=(), camera=()')
    if os.environ.get('SGE_HSTS', '0').lower() in ('1','true','yes'):
        resp.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    return resp
app.config['UPLOAD_FOLDER'] = os.environ.get('SGE_UPLOAD_FOLDER', os.path.join(BASE_DIR, 'uploads'))
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('SGE_MAX_UPLOAD_MB', '25')) * 1024 * 1024
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE=os.environ.get('SESSION_COOKIE_SAMESITE', 'Lax'),
    SESSION_COOKIE_SECURE=os.environ.get('SESSION_COOKIE_SECURE', '0').lower() in ('1','true','yes'),
)

# === AUTENTICAÇÃO OPCIONAL PARA AMBIENTE ONLINE ===
# Em produção recomenda-se activar: SGE_REQUIRE_LOGIN=1
# Credenciais via variáveis de ambiente: SGE_ADMIN_USER e SGE_ADMIN_PASSWORD
# Alternativa mais segura: SGE_ADMIN_PASSWORD_HASH com hash Werkzeug.
AUTH_EXEMPT_PREFIXES = ('/static/', '/uploads/')
AUTH_EXEMPT_PATHS = {'/login', '/logout', '/healthz', '/robots.txt', '/favicon.ico'}

def _truthy_env(name, default='0'):
    return os.environ.get(name, default).lower() in ('1', 'true', 'yes', 'on')

def _login_required_enabled():
    return _truthy_env('SGE_REQUIRE_LOGIN', '0')

def _auth_configured():
    return bool(os.environ.get('SGE_ADMIN_PASSWORD') or os.environ.get('SGE_ADMIN_PASSWORD_HASH'))

def _check_admin_password(password):
    stored_hash = os.environ.get('SGE_ADMIN_PASSWORD_HASH')
    stored_plain = os.environ.get('SGE_ADMIN_PASSWORD')
    if stored_hash:
        try:
            return check_password_hash(stored_hash, password or '')
        except Exception:
            return False
    return bool(stored_plain and secrets.compare_digest(stored_plain, password or ''))

@app.before_request
def _require_login_online():
    if not _login_required_enabled():
        return
    path = request.path or '/'
    if path in AUTH_EXEMPT_PATHS or any(path.startswith(prefix) for prefix in AUTH_EXEMPT_PREFIXES):
        return
    if session.get('sge_logged_in'):
        return
    if request.path.startswith('/api/') or request.headers.get('X-Requested-With','').lower() == 'xmlhttprequest':
        return jsonify(success=False, error='auth_required', message='Autenticação necessária.'), 401
    return redirect(url_for('login', next=request.url))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if not _login_required_enabled():
        flash('Autenticação não está activa neste ambiente.', 'info')
        return redirect(url_for('index'))
    if request.method == 'POST':
        if not _auth_configured():
            flash('Login activo, mas a palavra-passe de administrador ainda não foi configurada nas variáveis de ambiente.', 'error')
        else:
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            expected_user = os.environ.get('SGE_ADMIN_USER', 'admin')
            if secrets.compare_digest(username, expected_user) and _check_admin_password(password):
                session.clear()
                session['sge_logged_in'] = True
                session['username'] = username
                session.permanent = True
                return redirect(request.args.get('next') or url_for('index'))
            flash('Credenciais inválidas.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Sessão terminada com sucesso.', 'success')
    return redirect(url_for('login') if _login_required_enabled() else url_for('index'))

@app.get('/healthz')
def healthz():
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute('SELECT 1')
        conn.close()
        db_ok = True
    except Exception:
        db_ok = False
    status = 200 if db_ok else 503
    return jsonify(status='ok' if db_ok else 'degraded', app='SGE', database=db_ok), status

DB_PATH = os.environ.get('SGE_DB_PATH', os.path.join(BASE_DIR, 'sge.db'))

def _prepare_runtime_paths():
    # Permite deploy com base de dados e uploads em disco persistente (ex.: Render Disk).
    try:
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        seed_db = os.path.join(BASE_DIR, 'sge.db')
        if os.path.abspath(DB_PATH) != os.path.abspath(seed_db) and (not os.path.exists(DB_PATH)) and os.path.exists(seed_db):
            import shutil
            shutil.copy2(seed_db, DB_PATH)
    except Exception as _e:
        print('Aviso: preparação de paths de runtime falhou:', _e)

_prepare_runtime_paths()

@app.context_processor
def _inject_global_template_helpers():
    return {'now': datetime.now}

try:
    migrate_pack3()
except Exception as _e:
    print("Pack3 migration failed:", _e)

# --- Locais premium fase 3: enriquecimento do cadastro ---
def migrate_locais_premium_fase3():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        cols = {row[1] for row in c.execute("PRAGMA table_info(locais)").fetchall()}
        wanted = {
            'tipo_local': "TEXT",
            'categoria_operacional': "TEXT",
            'email': "TEXT",
            'responsavel_alt': "TEXT",
            'estado_tecnico': "TEXT DEFAULT 'Normal'",
            'prioridade': "TEXT DEFAULT 'Média'",
            'parent_id': "INTEGER"
        }
        for col, spec in wanted.items():
            if col not in cols:
                c.execute(f"ALTER TABLE locais ADD COLUMN {col} {spec}")
        conn.commit()
    finally:
        conn.close()

try:
    migrate_locais_premium_fase3()
except Exception as _e:
    print("Locais premium fase 3 migration failed:", _e)


def migrate_locais_hierarquia_online_v3():
    """Migração segura para centros, sublocais e hierarquia de locais."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        cols = {row[1] for row in c.execute("PRAGMA table_info(locais)").fetchall()}
        if 'parent_id' not in cols:
            c.execute("ALTER TABLE locais ADD COLUMN parent_id INTEGER")
        c.execute("CREATE INDEX IF NOT EXISTS idx_locais_parent_id ON locais(parent_id)")
        conn.commit()
    finally:
        conn.close()

try:
    migrate_locais_hierarquia_online_v3()
except Exception as _e:
    print("Migração de hierarquia de locais falhou:", _e)


# --- Locais premium fase 4: histórico, alertas e exportação executiva ---
def migrate_locais_premium_fase4():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("""CREATE TABLE IF NOT EXISTS locais_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            local_id INTEGER NOT NULL,
            evento TEXT NOT NULL,
            detalhe TEXT,
            actor TEXT DEFAULT 'sge',
            ts TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.commit()
    finally:
        conn.close()

def log_local_history(local_id, evento, detalhe='', actor='sge'):
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('INSERT INTO locais_history(local_id, evento, detalhe, actor) VALUES(?,?,?,?)',
                  (local_id, evento, detalhe, actor))
        conn.commit(); conn.close()
    except Exception:
        pass

def get_local_history(local_id, limit=12):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT evento, detalhe, actor, ts FROM locais_history WHERE local_id=? ORDER BY id DESC LIMIT ?', (local_id, int(limit)))
    rows = c.fetchall(); conn.close()
    return [{'evento':r[0], 'detalhe':r[1], 'actor':r[2], 'ts':r[3]} for r in rows]

def get_local_alertas(local, cfg, overview):
    itens = []
    if not (local.get('contato_nome') or '').strip() and not (local.get('contato_tel') or '').strip() and not (local.get('email') or '').strip():
        itens.append(('warning', 'Sem contacto principal registado', 'Definir responsável, telefone ou email para garantir governança do local.'))
    if float(cfg.get('pot_contratada', 0) or 0) <= 0:
        itens.append(('danger', 'Potência contratada não definida', 'Sem este parâmetro a pré-fatura e os controlos de ponta ficam limitados.'))
    if float(cfg.get('pot_instalada', 0) or 0) <= 0:
        itens.append(('info', 'Potência instalada não definida', 'Preenche este campo para relatórios e indicadores executivos.'))
    if overview.get('leituras_mensais_count', 0) == 0:
        itens.append(('info', 'Sem histórico de leituras mensais', 'Abrir Leituras Mensais para começar o histórico do local.'))
    if (local.get('estado_tecnico') or 'Normal').lower() in ['atenção', 'atencao', 'crítico', 'critico']:
        itens.append(('warning', f"Estado técnico: {local.get('estado_tecnico')}", 'Este local merece acompanhamento mais próximo no plano operacional.'))
    if (local.get('prioridade') or 'Média').lower() == 'alta':
        itens.append(('danger', 'Local marcado como prioridade alta', 'Convém validar cadastro, configuração, leituras e equipamentos com maior frequência.'))
    return itens

try:
    migrate_locais_premium_fase4()
except Exception as _e:
    print("Locais premium fase 4 migration failed:", _e)


# --- Auto-migrations for leituras_mensais ---
def migrate_leituras_mensais():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    # Table may already exist; ensure schema columns exist
    c.execute('''CREATE TABLE IF NOT EXISTS leituras_mensais (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        local TEXT,
        data TEXT,
        hora TEXT,
        ativa REAL, reativa REAL, ponta REAL,
        fp REAL, potc REAL,
        anterior REAL, atual REAL, diferenca REAL,
        agua REAL, esp REAL, acum REAL, valor REAL,
        mes TEXT, ano INTEGER
    )''')
    # Add unique index to avoid duplicates per (local, data)
    try:
        c.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_leituras_mensais_unique ON leituras_mensais(local, data)')
    except Exception:
        pass
    # Audit table
    c.execute('''CREATE TABLE IF NOT EXISTS leituras_mensais_audit (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        lm_id INTEGER,
        acao TEXT,
        field TEXT,
        old_value TEXT,
        new_value TEXT,
        actor TEXT,
        ts TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.commit(); conn.close()

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# === BANCO DE DADOS E TABELAS ===


def _migrar_audit_leituras():
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS leituras_audit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                leitura_id INTEGER,
                acao TEXT,
                field TEXT,
                old_value TEXT,
                new_value TEXT,
                actor TEXT,
                ts TEXT DEFAULT (datetime('now','localtime'))
            )
        ''')
        conn.commit(); conn.close()
    except Exception:
        pass
def init_db():
    # Pacote 2: migrações (idempotente)
    try:
        _apply_pacote2_migrations()
    except Exception:
        pass
    _migrar_audit_leitras_safe = _migrar_audit_leituras()


    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # Config de Local (tabela + índices)
    c.execute('''
        CREATE TABLE IF NOT EXISTS locais_cfg (
            local_id INTEGER PRIMARY KEY,
            fator_mult REAL DEFAULT 1.0,
            pot_contratada REAL DEFAULT 0.0,
            tarifa_ativa REAL DEFAULT 4.780,
            tarifa_reativa REAL DEFAULT 1.430,
            tarifa_ponta REAL DEFAULT 4.970,
            tarifa_perdas REAL DEFAULT 4.780,
            taxa_fixa REAL DEFAULT 207.28,
            taxa_radio REAL DEFAULT 297.00,
            taxa_lixo REAL DEFAULT 150.00,
            iva REAL DEFAULT 16.0,
            FOREIGN KEY (local_id) REFERENCES locais(id)
        )
    ''')
    try:
        c.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_leit_mensal_unique ON leituras_mensais(local, data, mes, ano)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_leit_mensal_periodo ON leituras_mensais(mes, ano, local)')
    except Exception as _e:
        pass
    # MIGRATION: add pot_instalada if missing
    try:
        c.execute("PRAGMA table_info(locais_cfg)")
        cols_cfg = {row[1] for row in c.fetchall()}
        if 'pot_instalada' not in cols_cfg:
            c.execute("ALTER TABLE locais_cfg ADD COLUMN pot_instalada REAL DEFAULT 0.0")
    except Exception:
        pass

    # Locais (base)
    c.execute('''
        CREATE TABLE IF NOT EXISTS locais (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE
        )
    ''')
    # --- MIGRAÇÕES EQUIPAMENTOS (Excelência) ---
    try:
        c.execute("PRAGMA table_info(equipamentos)")
        cols = [r[1] for r in c.fetchall()]
        if "ativo" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN ativo INTEGER DEFAULT 1")
        if "created_at" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN created_at TEXT")
        if "updated_at" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN updated_at TEXT")
    except Exception:
        pass


    # --- MIGRAÇÕES EQUIPAMENTOS (Campos avançados + Fotos) ---
    try:
        c.execute("PRAGMA table_info(equipamentos)")
        cols = [r[1] for r in c.fetchall()]
        if "categoria" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN categoria TEXT")
        if "fabricante" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN fabricante TEXT")
        if "modelo" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN modelo TEXT")
        if "numero_serie" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN numero_serie TEXT")
        if "custo_aquisicao" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN custo_aquisicao REAL")
        if "vida_util_anos" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN vida_util_anos INTEGER")
        if "criticidade" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN criticidade TEXT")
        if "cover_photo_id" not in cols:
            c.execute("ALTER TABLE equipamentos ADD COLUMN cover_photo_id INTEGER")
    except Exception:
        pass

    c.execute('''
        CREATE TABLE IF NOT EXISTS equipamentos_photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipamento_id INTEGER,
            filename TEXT,
            thumb_filename TEXT,
            caption TEXT,
            width INTEGER,
            height INTEGER,
            uploaded_at TEXT DEFAULT (datetime('now','localtime'))
        )
    ''')
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_ep_equip ON equipamentos_photos(equipamento_id)")
    except Exception:
        pass

    # Audit log
    c.execute('''
        CREATE TABLE IF NOT EXISTS equipamentos_audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipamento_id INTEGER,
            acao TEXT,
            detalhes TEXT,
            ts TEXT DEFAULT (datetime('now','localtime'))
        )
    ''')

    # Files for equipamentos
    c.execute('''
        CREATE TABLE IF NOT EXISTS equipamentos_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipamento_id INTEGER,
            filename TEXT,
            original_name TEXT,
            mime TEXT,
            size INTEGER,
            uploaded_at TEXT DEFAULT (datetime('now','localtime'))
        )
    ''')

    # Índices
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_equip_local ON equipamentos(local_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_equip_nome ON equipamentos(nome)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_equip_ativo ON equipamentos(ativo)")
    except Exception:
        pass
    # --- MIGRAÇÕES EQUIPAMENTOS (Excelência Pack 3) ---
    try:
        c.execute('''
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        c.execute("INSERT OR IGNORE INTO settings (key,value) VALUES ('unique_tag','0')")
        c.execute("INSERT OR IGNORE INTO settings (key,value) VALUES ('unique_nome_local','0')")
        c.execute("CREATE INDEX IF NOT EXISTS idx_equip_tag ON equipamentos(tag)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_equip_nome_local ON equipamentos(nome, local_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_equip_numero_serie ON equipamentos(numero_serie)")
    except Exception:
        pass

    # --- Links de documentação por equipamento ---
    try:
        c.execute('''
            CREATE TABLE IF NOT EXISTS equipamentos_links (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipamento_id INTEGER,
                url TEXT,
                title TEXT,
                added_at TEXT DEFAULT (datetime('now','localtime'))
            )
        ''')
        # Fecho seguro da ligação principal do init_db.
        # Na versão anterior esta ligação ficava aberta e podia bloquear o sge.db.
        conn.commit()
        conn.close()
    except Exception:
        pass



        # ===

        # === MIGRAÇÃO: colunas extras em locais ===
        try:
            c.execute("PRAGMA table_info(locais)")
            cols = {row[1] for row in c.fetchall()}
            if "codigo" not in cols:
                c.execute("ALTER TABLE locais ADD COLUMN codigo TEXT")
            if "endereco" not in cols:
                c.execute("ALTER TABLE locais ADD COLUMN endereco TEXT")
            if "contato_nome" not in cols:
                c.execute("ALTER TABLE locais ADD COLUMN contato_nome TEXT")
            if "contato_tel" not in cols:
                c.execute("ALTER TABLE locais ADD COLUMN contato_tel TEXT")
            if "notas" not in cols:
                c.execute("ALTER TABLE locais ADD COLUMN notas TEXT")
            if "ativo" not in cols:
                c.execute("ALTER TABLE locais ADD COLUMN ativo INTEGER DEFAULT 1")
        except Exception:
            pass

        # Equipamentos
        c.execute('''
            CREATE TABLE IF NOT EXISTS equipamentos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                local_id INTEGER,
                tag TEXT,
                especificacao TEXT,
                ano_instalacao TEXT,
                quantidade INTEGER,
                FOREIGN KEY (local_id) REFERENCES locais (id)
            )
        ''')

        # Leituras Diárias
        c.execute('''
            CREATE TABLE IF NOT EXISTS leituras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                datahora TEXT,
                local TEXT,
                equipamento TEXT,
                energia_ativa REAL,
                energia_reativa REAL,
                energia_aparente REAL,
                pot_ativa REAL,
                pot_reativa REAL,
                pot_aparente REAL,
                fp REAL,
                ponta REAL,
                caudal_elevada REAL,
                corrente REAL,
                tensao REAL,
                observacoes TEXT
            )
        ''')

        # Leituras Mensais
        c.execute('''
            CREATE TABLE IF NOT EXISTS leituras_mensais (
                local TEXT, data TEXT, hora TEXT, ativa REAL, reativa REAL, ponta REAL, fp REAL, potc REAL,
                anterior REAL, atual REAL, diferenca REAL, agua REAL, esp REAL, acum REAL, valor REAL,
                mes TEXT, ano INTEGER
            )
        ''')


# Config de Local


        # --- MIGRATION: add pot_instalada if missing ---
        try:
            c.execute("PRAGMA table_info(locais_cfg)")
            cols_cfg = {row[1] for row in c.fetchall()}
            if 'pot_instalada' not in cols_cfg:
                c.execute("ALTER TABLE locais_cfg ADD COLUMN pot_instalada REAL DEFAULT 0.0")
        except Exception:
            pass

        # MOTORES – medições
        c.execute('''
            CREATE TABLE IF NOT EXISTS motor_medicoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipamento_id INTEGER NOT NULL,
                datahora TEXT NOT NULL,
                tensao_v REAL,
                corrente_a REAL,
                fator_potencia REAL,
                frequencia_hz REAL,
                fases INTEGER DEFAULT 3,
                pot_ativa_kw REAL,
                pot_reativa_kvar REAL,
                pot_aparente_kva REAL,
                eficiencia REAL,
                energia_kwh REAL,
                observacoes TEXT,
                FOREIGN KEY (equipamento_id) REFERENCES equipamentos(id)
            )
        ''')

        # MOTORES – runs
        c.execute('''
            CREATE TABLE IF NOT EXISTS motor_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipamento_id INTEGER NOT NULL,
                start_time TEXT NOT NULL,
                stop_time TEXT,
                duracao_min REAL,
                FOREIGN KEY (equipamento_id) REFERENCES equipamentos(id)
            )
        ''')

        # Config por equipamento
        c.execute('''
            CREATE TABLE IF NOT EXISTS equipamentos_cfg (
                equipamento_id INTEGER PRIMARY KEY,
                tensao_nominal REAL,
                corrente_nominal REAL,
                potencia_nominal_kw REAL,
                fp_nominal REAL,
                eficiencia_nominal REAL,
                limite_corrente REAL,
                limite_fp REAL DEFAULT 0.80,
                FOREIGN KEY (equipamento_id) REFERENCES equipamentos(id)
            )
        ''')

        # === Tabela de projetos solares ===
        c.execute('''
            CREATE TABLE IF NOT EXISTS solar_projetos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                criado_em TEXT,
                local_id INTEGER,
                local_nome TEXT,
                periodo TEXT,
                modo TEXT,
                tipo_sistema TEXT,
                daily_kwh REAL,
                total_mes_kwh REAL,
                psh REAL,
                derate REAL,
                panel_wp REAL,
                panel_area REAL,
                n_paineis INTEGER,
                kwp_necessario REAL,
                kwp_real REAL,
                area_total REAL,
                inv_dcac REAL,
                inversor_kw REAL,
                tarifa_kwh REAL,
                economia_mensal REAL,
                autonomy_days REAL,
                battery_dod REAL,
                battery_eff REAL,
                system_voltage REAL,
                battery_module_kwh REAL,
                bateria_kwh_util REAL,
                bateria_kwh_bruta REAL,
                n_modulos_bateria INTEGER,
                mes TEXT,
                ano INTEGER,
                dias_utilizados INTEGER,
                fator_mult REAL,
                resultado_json TEXT,
                params_json TEXT
            )
        ''')
        # Migração (campos novos)
        new_cols = [
            ('capex_kwp', 'REAL'), ('capex_total', 'REAL'), ('opex_pct', 'REAL'),
            ('opex_anual', 'REAL'), ('tarifa_esc', 'REAL'), ('desconto', 'REAL'),
            ('anos_analise', 'INTEGER'), ('payback_anos', 'REAL'), ('npv', 'REAL'),
            ('co2_factor', 'REAL'), ('co2_t_ano', 'REAL'),
            ('producao_anual_kwh', 'REAL'), ('producao_mensal_json', 'TEXT'),
            ('perfil_sazonal_json', 'TEXT')
        ]
        c.execute("PRAGMA table_info('solar_projetos')")
        existing = {row[1] for row in c.fetchall()}
        for col, typ in new_cols:
            if col not in existing:
                c.execute(f"ALTER TABLE solar_projetos ADD COLUMN {col} {typ}")

        conn.commit()
        conn.close()

init_db()



def _to_float(val, default=0.0):
    try:
        if val is None: 
            return float(default)
        s = str(val).strip().replace(",", ".")
        if s == "": 
            return float(default)
        return float(s)
    except Exception:
        return float(default)
# === UTILITÁRIOS ===


def _apply_pacote2_migrations():
    """Cria/altera colunas e índices do Pacote 2 (sem PMP). Idempotente."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            c.execute("PRAGMA table_info(equipamentos)")
            cols = [r[1] for r in c.fetchall()]
            to_add = []
            if "deleted_at" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN deleted_at TEXT")
            if "potencia_kw" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN potencia_kw REAL")
            if "tensao_v" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN tensao_v REAL")
            if "corrente_a" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN corrente_a REAL")
            if "ip_class" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN ip_class TEXT")
            if "peso_kg" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN peso_kg REAL")
            if "garantia_fim" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN garantia_fim TEXT")
            if "fornecedor" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN fornecedor TEXT")
            if "contrato_num" not in cols: to_add.append("ALTER TABLE equipamentos ADD COLUMN contrato_num TEXT")
            for sqlx in to_add:
                c.execute(sqlx)
        except Exception:
            pass

        c.execute('''
            CREATE TABLE IF NOT EXISTS equipamentos_componentes (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              equipamento_id INTEGER NOT NULL,
              nome TEXT NOT NULL,
              fabricante TEXT,
              modelo TEXT,
              qtd INTEGER DEFAULT 1,
              created_at TEXT DEFAULT (datetime('now','localtime')),
              FOREIGN KEY(equipamento_id) REFERENCES equipamentos(id) ON DELETE CASCADE
            )
        ''')

        c.execute('''
            CREATE TABLE IF NOT EXISTS saved_filters (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              user TEXT,
              modulo TEXT,
              nome TEXT,
              query_json TEXT,
              created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        ''')

        try:
            c.execute("CREATE INDEX IF NOT EXISTS idx_equip_local ON equipamentos(local_id)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_equip_categoria ON equipamentos(categoria)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_equip_fabricante ON equipamentos(fabricante)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_equip_modelo ON equipamentos(modelo)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_equip_crit ON equipamentos(criticidade)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_equip_ativo ON equipamentos(ativo)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_equip_deleted ON equipamentos(deleted_at)")
        except Exception:
            pass

        conn.commit()
    except Exception:
        pass
    finally:
        try:
            conn.close()
        except Exception:
            pass

# ==== /MIGRAÇÕES PACOTE 2 ====



def ensure_locais_parent_id_column():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        cols = {row[1] for row in c.execute("PRAGMA table_info(locais)").fetchall()}
        if 'parent_id' not in cols:
            c.execute("ALTER TABLE locais ADD COLUMN parent_id INTEGER")
        c.execute("CREATE INDEX IF NOT EXISTS idx_locais_parent_id ON locais(parent_id)")
        conn.commit()
    finally:
        conn.close()

def _get_locais_rows(include_inactive=True):
    ensure_locais_parent_id_column()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    sql = "SELECT id, nome, COALESCE(parent_id, NULL) AS parent_id, COALESCE(ativo,1) AS ativo FROM locais"
    params = []
    if not include_inactive:
        sql += " WHERE COALESCE(ativo,1)=1"
    sql += " ORDER BY nome COLLATE NOCASE"
    rows = c.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_locais_hierarchy(include_inactive=True, exclude_id=None):
    rows = _get_locais_rows(include_inactive=include_inactive)
    if exclude_id is not None:
        rows = [r for r in rows if int(r['id']) != int(exclude_id)]
    by_id = {int(r['id']): r for r in rows}
    children = {}
    roots = []
    for r in rows:
        rid = int(r['id'])
        pid = r.get('parent_id')
        if pid is not None:
            try:
                pid = int(pid)
            except Exception:
                pid = None
        if pid and pid in by_id and pid != rid:
            children.setdefault(pid, []).append(rid)
        else:
            roots.append(rid)

    for key in children:
        children[key].sort(key=lambda cid: (by_id[cid].get('nome') or '').lower())
    roots = sorted(set(roots), key=lambda rid: (by_id[rid].get('nome') or '').lower())

    ordered = []
    visited = set()

    def walk(rid, depth=0, trail=None):
        if rid in visited:
            return
        visited.add(rid)
        row = dict(by_id[rid])
        trail = list(trail or [])
        trail.append(row.get('nome') or '')
        row['depth'] = depth
        row['display_name'] = (('— ' * depth) + (row.get('nome') or '')).strip()
        row['full_name'] = ' › '.join([p for p in trail if p])
        ordered.append(row)
        for child_id in children.get(rid, []):
            walk(child_id, depth + 1, trail)

    for rid in roots:
        walk(rid, 0, [])
    for rid in sorted(by_id.keys()):
        if rid not in visited:
            walk(rid, 0, [])
    return ordered


def get_local_choices(include_inactive=True, exclude_id=None):
    return [(r['id'], r['display_name']) for r in get_locais_hierarchy(include_inactive=include_inactive, exclude_id=exclude_id)]


def get_local_children(parent_id, include_inactive=True):
    parent_id = int(parent_id)
    rows = [r for r in get_locais_hierarchy(include_inactive=include_inactive) if (r.get('parent_id') is not None and int(r.get('parent_id')) == parent_id)]
    if not rows:
        return []
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for row in rows:
        cnt = c.execute("SELECT COUNT(*) FROM equipamentos WHERE local_id=? AND COALESCE(deleted_at,'')=''", (row['id'],)).fetchone()[0]
        row['equipamentos_count'] = int(cnt or 0)
    conn.close()
    return rows


def get_descendant_local_ids(local_id, include_self=True):
    """Devolve o local e todos os sublocais abaixo dele. Usado para filtros hierárquicos."""
    try:
        root = int(local_id)
    except Exception:
        return []
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        rows = c.execute("SELECT id, parent_id FROM locais").fetchall()
    except Exception:
        conn.close()
        return [root] if include_self else []
    conn.close()
    children = {}
    for rid, pid in rows:
        try:
            rid = int(rid)
            pid = int(pid) if pid is not None else None
        except Exception:
            continue
        if pid:
            children.setdefault(pid, []).append(rid)
    found = []
    def walk(pid):
        for cid in children.get(pid, []):
            if cid not in found:
                found.append(cid)
                walk(cid)
    if include_self:
        found.append(root)
    walk(root)
    return found


def get_local_names_for_ids(local_ids):
    ids = [int(x) for x in (local_ids or []) if str(x).isdigit()]
    if not ids:
        return []
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    placeholders = ','.join('?' for _ in ids)
    rows = c.execute(f"SELECT nome FROM locais WHERE id IN ({placeholders})", ids).fetchall()
    conn.close()
    return [r[0] for r in rows if r and (r[0] or '').strip()]

def get_locais():
    return get_local_choices(include_inactive=True)

def get_local_by_id(local_id):
    """Compatibilidade: devolve o local no formato antigo (id, nome).
    Algumas rotas antigas ainda chamam esta função; sem ela o módulo de leituras mensais gera erro 500.
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT id, nome FROM locais WHERE id=?', (int(local_id),))
        row = c.fetchone()
        conn.close()
        return row
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
        return None

def get_local_full(local_id: int):
    """Dados completos do local (inclui colunas premium e hierarquia)."""
    ensure_locais_parent_id_column()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''
        SELECT l.id, l.nome, l.codigo, l.endereco, l.contato_nome, l.contato_tel, l.notas, COALESCE(l.ativo,1),
               COALESCE(l.tipo_local,''), COALESCE(l.categoria_operacional,''), COALESCE(l.email,''),
               COALESCE(l.responsavel_alt,''), COALESCE(l.estado_tecnico,'Normal'), COALESCE(l.prioridade,'Média'),
               l.parent_id, COALESCE(p.nome,'')
        FROM locais l
        LEFT JOIN locais p ON p.id = l.parent_id
        WHERE l.id=?
    ''', (local_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return None
    return {
        "id": row[0], "nome": row[1], "codigo": row[2], "endereco": row[3],
        "contato_nome": row[4], "contato_tel": row[5], "notas": row[6],
        "ativo": int(row[7] or 1),
        "tipo_local": row[8], "categoria_operacional": row[9], "email": row[10],
        "responsavel_alt": row[11], "estado_tecnico": row[12], "prioridade": row[13],
        "parent_id": row[14], "parent_nome": row[15]
    }

def infer_local_tipo(nome: str, endereco: str = '') -> str:
    txt = f"{nome or ''} {endereco or ''}".lower()
    checks = [
        ('ETA', ['eta', 'estacao de tratamento', 'estação de tratamento']),
        ('CD', ['cd ', 'centro distrib', 'centro distribuidor']),
        ('Furo', ['furo', 'furos', 'poco', 'poço']),
        ('Reservatório', ['reservatorio', 'reservatório', 'tanque']),
        ('Estação', ['estacao', 'estação', 'psaa', 'elevatoria', 'elevatória']),
        ('Escritório', ['escritorio', 'escritório', 'administracao', 'administração']),
    ]
    for tipo, keys in checks:
        if any(k in txt for k in keys):
            return tipo
    return 'Outro'


def calcular_maturidade_local(local: dict) -> int:
    score = 0
    if (local.get('codigo') or '').strip():
        score += 10
    if (local.get('endereco') or '').strip():
        score += 10
    if (local.get('contato_nome') or '').strip():
        score += 10
    if (local.get('contato_tel') or '').strip():
        score += 10
    if (local.get('email') or '').strip():
        score += 8
    if (local.get('responsavel_alt') or '').strip():
        score += 7
    if (local.get('tipo_local') or '').strip():
        score += 10
    if (local.get('categoria_operacional') or '').strip():
        score += 5
    if float(local.get('pot_contratada', 0) or 0) > 0:
        score += 12
    if float(local.get('pot_instalada', 0) or 0) > 0:
        score += 12
    if float(local.get('fator_mult', 1) or 1) != 1:
        score += 6
    if (local.get('notas') or '').strip():
        score += 5
    if (local.get('estado_tecnico') or '').strip() and (local.get('estado_tecnico') or '').strip().lower() != 'normal':
        score += 3
    if (local.get('prioridade') or '').strip():
        score += 2
    return min(score, 100)


def get_locais_with_cfg(search=None, incluir_inativos=False, sort='nome', order='asc', tipo=None, qualidade=None, estado_tecnico=None, prioridade=None):
    """Locais + config com filtros premium."""
    ensure_locais_parent_id_column()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    where = []
    params = []
    if not incluir_inativos:
        where.append("COALESCE(l.ativo,1)=1")
    if search:
        where.append("(l.nome LIKE ? OR COALESCE(l.codigo,'') LIKE ? OR COALESCE(l.endereco,'') LIKE ? OR COALESCE(l.contato_nome,'') LIKE ? OR COALESCE(l.email,'') LIKE ?)")
        like = f"%{search}%"
        params += [like, like, like, like, like]
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    sort_map = {
        "nome": "l.nome COLLATE NOCASE",
        "codigo": "COALESCE(l.codigo,'') COLLATE NOCASE",
        "fator": "COALESCE(cfg.fator_mult,1.0)",
        "pot_contratada": "COALESCE(cfg.pot_contratada,0.0)",
        "pot_instalada": "COALESCE(cfg.pot_instalada,0.0)",
        "maturidade": "l.nome COLLATE NOCASE",
    }
    order_by = sort_map.get(sort, "l.nome COLLATE NOCASE")
    direction = "DESC" if str(order).lower() == "desc" else "ASC"

    c.execute(f"""
        SELECT l.id, l.nome, l.codigo, l.endereco, l.contato_nome, l.contato_tel,
               COALESCE(l.ativo,1),
               COALESCE(cfg.fator_mult,1.0),
               COALESCE(cfg.pot_contratada,0.0),
               COALESCE(cfg.pot_instalada,0.0),
               COALESCE(l.notas,''), COALESCE(l.tipo_local,''), COALESCE(l.categoria_operacional,''),
               COALESCE(l.email,''), COALESCE(l.responsavel_alt,''), COALESCE(l.estado_tecnico,'Normal'),
               COALESCE(l.prioridade,'Média'), l.parent_id
        FROM locais l
        LEFT JOIN locais_cfg cfg ON cfg.local_id = l.id
        {where_sql}
        ORDER BY {order_by} {direction}
    """, tuple(params))
    rows = c.fetchall()
    conn.close()
    data = []
    for r in rows:
        item = {
            "id": r[0], "nome": r[1], "codigo": r[2], "endereco": r[3],
            "contato_nome": r[4], "contato_tel": r[5], "ativo": int(r[6] or 1),
            "fator_mult": float(r[7]), "pot_contratada": float(r[8]),
            "pot_instalada": float(r[9]), "notas": r[10],
            "tipo_local": r[11], "categoria_operacional": r[12], "email": r[13],
            "responsavel_alt": r[14], "estado_tecnico": r[15], "prioridade": r[16], "parent_id": r[17]
        }
        item['tipo'] = item['tipo_local'] or infer_local_tipo(item['nome'], item['endereco'])
        item['maturidade'] = calcular_maturidade_local(item)
        item['config_ok'] = (item['pot_contratada'] > 0 or item['pot_instalada'] > 0)
        data.append(item)

    hierarchy_map = {int(r['id']): r for r in get_locais_hierarchy(include_inactive=True)}
    for item in data:
        href = hierarchy_map.get(int(item['id']))
        item['display_name'] = (href.get('full_name') if href else item['nome']) or item['nome']
        item['depth'] = int(href.get('depth', 0)) if href else 0

    if tipo and tipo != 'todos':
        data = [r for r in data if (r['tipo'] or '').lower() == tipo.lower()]
    if estado_tecnico and estado_tecnico != 'todos':
        data = [r for r in data if (r.get('estado_tecnico') or 'Normal').lower() == estado_tecnico.lower()]
    if prioridade and prioridade != 'todas':
        data = [r for r in data if (r.get('prioridade') or 'Média').lower() == prioridade.lower()]
    if qualidade == 'completo':
        data = [r for r in data if r['maturidade'] >= 70]
    elif qualidade == 'incompleto':
        data = [r for r in data if r['maturidade'] < 70]
    elif qualidade == 'sem_contato':
        data = [r for r in data if not (r.get('contato_nome') or '').strip() and not (r.get('contato_tel') or '').strip()]
    elif qualidade == 'alta_prontidao':
        data = [r for r in data if r['config_ok'] and r['maturidade'] >= 70]

    if sort == 'maturidade':
        data = sorted(data, key=lambda x: x.get('maturidade', 0), reverse=(direction=='DESC'))
    return data

def get_local_overview(local_id: int):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    out = {
        'equipamentos_count': 0,
        'equipamentos_qtd_total': 0,
        'leituras_mensais_count': 0,
        'ultima_leitura': None,
        'primeira_leitura': None,
        'leituras_dias_com_dados': 0,
        'equipamentos_principais': [],
        'prontidao': 'Base',
        'prontidao_score': 0,
    }
    try:
        local_ids_scope = get_descendant_local_ids(local_id, include_self=True)
        placeholders_scope = ','.join('?' for _ in local_ids_scope) if local_ids_scope else '?'
        params_scope = local_ids_scope if local_ids_scope else [local_id]
        c.execute(f'''SELECT COUNT(*), COALESCE(SUM(COALESCE(quantidade,1)),0)
                     FROM equipamentos WHERE local_id IN ({placeholders_scope})''', params_scope)
        row = c.fetchone() or (0, 0)
        out['equipamentos_count'] = int(row[0] or 0)
        out['equipamentos_qtd_total'] = int(row[1] or 0)

        c.execute('SELECT nome FROM locais WHERE id=?', (local_id,))
        row = c.fetchone()
        local_nome = row[0] if row else ''
        local_names_scope = get_local_names_for_ids(local_ids_scope)
        if local_names_scope:
            placeholders_names = ','.join('?' for _ in local_names_scope)
            c.execute(f'''SELECT COUNT(*), MIN(data), MAX(data),
                                SUM(CASE WHEN COALESCE(ativa,0)<>0 OR COALESCE(reativa,0)<>0 OR COALESCE(ponta,0)<>0 OR COALESCE(agua,0)<>0 THEN 1 ELSE 0 END)
                         FROM leituras_mensais
                         WHERE local IN ({placeholders_names})''', local_names_scope)
            row = c.fetchone() or (0, None, None, 0)
            out['leituras_mensais_count'] = int(row[0] or 0)
            out['primeira_leitura'] = row[1]
            out['ultima_leitura'] = row[2]
            out['leituras_dias_com_dados'] = int(row[3] or 0)

        c.execute(f'''SELECT nome, COALESCE(quantidade,1)
                     FROM equipamentos WHERE local_id IN ({placeholders_scope})
                     ORDER BY nome LIMIT 5''', params_scope)
        out['equipamentos_principais'] = c.fetchall()
        score = 0
        if out['equipamentos_count'] > 0:
            score += 25
        if out['leituras_mensais_count'] > 0:
            score += 25
        if out['leituras_dias_com_dados'] > 0:
            score += 25
        if out['ultima_leitura']:
            score += 25
        out['prontidao_score'] = score
        out['prontidao'] = 'Alta' if score >= 75 else ('Média' if score >= 40 else 'Base')
    finally:
        conn.close()
    return out

def get_locais_module_summary(locais):
    resumo = {
        'total': len(locais),
        'ativos': 0,
        'arquivados': 0,
        'pot_contratada_total': 0.0,
        'pot_instalada_total': 0.0,
        'com_config': 0,
        'sem_config': 0,
        'fator_medio': 0.0,
        'maturidade_media': 0.0,
        'sem_contato': 0,
        'tipos': {},
        'tipo_dominante': '—',
        'alta_prioridade': 0,
        'criticos': 0,
    }
    if not locais:
        return resumo
    soma_fator = 0.0
    soma_maturidade = 0.0
    for r in locais:
        ativo = int(r.get('ativo', 1) or 1)
        resumo['ativos' if ativo == 1 else 'arquivados'] += 1
        pot_c = float(r.get('pot_contratada', 0) or 0)
        pot_i = float(r.get('pot_instalada', 0) or 0)
        resumo['pot_contratada_total'] += pot_c
        resumo['pot_instalada_total'] += pot_i
        if pot_c > 0 or pot_i > 0:
            resumo['com_config'] += 1
        else:
            resumo['sem_config'] += 1
        soma_fator += float(r.get('fator_mult', 1) or 1)
        soma_maturidade += float(r.get('maturidade', 0) or 0)
        if not (r.get('contato_nome') or '').strip() and not (r.get('contato_tel') or '').strip():
            resumo['sem_contato'] += 1
        if (r.get('prioridade') or '').lower() == 'alta':
            resumo['alta_prioridade'] += 1
        if (r.get('estado_tecnico') or '').lower() in ('crítico', 'critico'):
            resumo['criticos'] += 1
        tipo = r.get('tipo') or infer_local_tipo(r.get('nome'), r.get('endereco'))
        resumo['tipos'][tipo] = resumo['tipos'].get(tipo, 0) + 1
    resumo['fator_medio'] = soma_fator / max(len(locais), 1)
    resumo['maturidade_media'] = soma_maturidade / max(len(locais), 1)
    if resumo['tipos']:
        resumo['tipo_dominante'] = max(resumo['tipos'].items(), key=lambda kv: kv[1])[0]
    return resumo

def get_equipamentos():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT id, nome FROM equipamentos')
    equipamentos = c.fetchall()
    conn.close()
    return equipamentos

def get_equipamentos_por_local(local_id):
    local_ids_scope = get_descendant_local_ids(local_id, include_self=True)
    if not local_ids_scope:
        local_ids_scope = [local_id]
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    placeholders = ','.join('?' for _ in local_ids_scope)
    c.execute(f'''
        SELECT e.id, e.nome, e.tag, e.especificacao, e.ano_instalacao, e.quantidade
        FROM equipamentos e
        WHERE e.local_id IN ({placeholders})
        ORDER BY e.nome
    ''', local_ids_scope)
    equipamentos = c.fetchall()
    conn.close()
    return equipamentos

# Config por local (dict completo, inclui pot_instalada)
def get_local_cfg_full(local_id: int):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''
        SELECT fator_mult, pot_contratada, tarifa_ativa, tarifa_reativa, tarifa_ponta,
               tarifa_perdas, taxa_fixa, taxa_radio, taxa_lixo, iva,
               COALESCE(pot_instalada, 0.0)
          FROM locais_cfg WHERE local_id=?
    ''', (local_id,))
    row = c.fetchone()
    if not row:
        c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (local_id,))
        conn.commit()
        row = (1.0, 0.0, 4.780, 1.430, 4.970, 4.780, 207.28, 297.00, 150.00, 16.0, 0.0)
    conn.close()
    return {
        "fator_mult": row[0],
        "pot_contratada": row[1],
        "tarifa_ativa": row[2],
        "tarifa_reativa": row[3],
        "tarifa_ponta": row[4],
        "tarifa_perdas": row[5],
        "taxa_fixa": row[6],
        "taxa_radio": row[7],
        "taxa_lixo": row[8],
        "iva": row[9],
        "pot_instalada": row[10],
    }

def get_local_cfg(local_id: int):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT fator_mult, pot_contratada, tarifa_ativa, tarifa_reativa, tarifa_ponta,
               tarifa_perdas, taxa_fixa, taxa_radio, taxa_lixo, iva
        FROM locais_cfg WHERE local_id=?
    ''', (local_id,))
    row = c.fetchone()
    if not row:
        c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (local_id,))
        conn.commit()
        row = (1.0, 0.0, 4.780, 1.430, 4.970, 4.780, 207.28, 297.00, 150.00, 16.0)
    conn.close()
    return row

# Consumo mensal (kWh) a partir de leituras_mensais
def consumo_mensal_kwh(local_nome, mes, ano, fator_mult):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT diferenca, ativa FROM leituras_mensais
        WHERE local=? AND mes=? AND ano=?
    ''', (local_nome, mes, ano))
    rows = c.fetchall()
    conn.close()

    total = 0.0
    dias = 0
    for dif, ativa in rows:
        if dif is not None and dif != '':
            try:
                total += float(dif) * float(fator_mult)
                dias += 1
                continue
            except:
                pass
        if ativa is not None and ativa != '':
            try:
                total += float(ativa)
                dias += 1
            except:
                pass
    return (total, dias)

# === ROTAS PRINCIPAIS ===

@app.route('/')
def index():
    return render_template('index.html')



# === MÓDULO UNIFICADO: GESTÃO DE LEITURAS, CONSUMO E FATURAÇÃO ===
@app.route('/energia')
@app.route('/gestao_leituras')
def gestao_leituras():
    """Centro único para leitura diária, consulta, planilha mensal e fatura."""
    from datetime import datetime
    mes = request.args.get('mes', default=datetime.now().month, type=int)
    ano = request.args.get('ano', default=datetime.now().year, type=int)
    local_id = request.args.get('local_id', type=int)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    locais = c.execute("""
        SELECT id, nome
        FROM locais
        ORDER BY nome
    """).fetchall()

    selected_local = None
    if locais:
        if local_id is None:
            local_id = locais[0]['id']
        selected_local = c.execute("SELECT * FROM locais WHERE id=?", (local_id,)).fetchone()
        if selected_local is None:
            selected_local = locais[0]
            local_id = selected_local['id']

    cfg = {}
    if selected_local:
        cfg_row = c.execute("SELECT * FROM locais_cfg WHERE local_id=?", (local_id,)).fetchone()
        if cfg_row:
            cfg = dict(cfg_row)

    fator_mult = float(cfg.get('fator_mult') or selected_local['fator_multiplicativo'] if selected_local and 'fator_multiplicativo' in selected_local.keys() else 1) if selected_local else 1.0
    tarifa_ativa = float(cfg.get('tarifa_ativa') or 0)
    tarifa_reativa = float(cfg.get('tarifa_reativa') or 0)
    tarifa_ponta = float(cfg.get('tarifa_ponta') or 0)
    taxa_fixa = float(cfg.get('taxa_fixa') or 0)
    taxa_radio = float(cfg.get('taxa_radio') or 0)
    taxa_lixo = float(cfg.get('taxa_lixo') or 0)
    iva_pct = float(cfg.get('iva') or 16)

    totais = dict(dias=0, ativa=0, reativa=0, ponta=0, agua=0, fp_medio=0, diferenca=0)
    if selected_local:
        r = c.execute("""
            SELECT COUNT(*) dias,
                   COALESCE(SUM(ativa),0) ativa,
                   COALESCE(SUM(reativa),0) reativa,
                   COALESCE(MAX(ponta),0) ponta,
                   COALESCE(SUM(agua),0) agua,
                   COALESCE(AVG(NULLIF(fp,0)),0) fp_medio,
                   COALESCE(SUM(CASE WHEN diferenca IS NOT NULL AND diferenca != '' THEN diferenca ELSE ativa END),0) diferenca
            FROM leituras_mensais
            WHERE local=? AND mes=? AND ano=?
        """, (selected_local['nome'], str(mes), int(ano))).fetchone()
        if r:
            totais.update(dict(r))

    ativa_faturavel = float(totais.get('diferenca') or totais.get('ativa') or 0) * fator_mult
    reativa_faturavel = float(totais.get('reativa') or 0) * fator_mult
    ponta_faturavel = _ponta_faturavel_edm(float(cfg.get('pot_contratada') or (selected_local['potencia_contratada'] if selected_local and 'potencia_contratada' in selected_local.keys() else 0) or 0), float(totais.get('ponta') or 0) * fator_mult)
    agua = float(totais.get('agua') or 0)
    consumo_especifico = (ativa_faturavel / agua) if agua > 0 else 0
    reativa_excedente = max(0.0, reativa_faturavel - (0.75 * ativa_faturavel))
    custo_ativa = ativa_faturavel * tarifa_ativa
    custo_reativa = reativa_excedente * tarifa_reativa
    custo_ponta = ponta_faturavel * tarifa_ponta
    subtotal = custo_ativa + custo_reativa + custo_ponta + taxa_fixa + taxa_radio + taxa_lixo
    iva = subtotal * (iva_pct / 100.0) * 0.62
    total_estimado = subtotal + iva

    resumo = {
        'fator_mult': fator_mult,
        'pot_contratada': float(cfg.get('pot_contratada') or (selected_local['potencia_contratada'] if selected_local and 'potencia_contratada' in selected_local.keys() else 0) or 0),
        'pot_instalada': float(cfg.get('pot_instalada') or 0),
        'tarifa_ativa': tarifa_ativa,
        'tarifa_reativa': tarifa_reativa,
        'tarifa_ponta': tarifa_ponta,
        'ativa_faturavel': ativa_faturavel,
        'reativa_faturavel': reativa_faturavel,
        'ponta_faturavel': ponta_faturavel,
        'consumo_especifico': consumo_especifico,
        'fp_medio': float(totais.get('fp_medio') or 0),
        'total_estimado': total_estimado,
    }
    conn.close()
    return render_template('gestao_leituras.html', locais=locais, selected_local=selected_local, mes=mes, ano=ano, resumo=resumo)

# === DASHBOARD ===
# === DASHBOARD (versão excelência v2) ===
from collections import defaultdict

def _ensure_idx_dashboard():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_mensais_local_mes_ano ON leituras_mensais(local, mes, ano)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_mensais_data ON leituras_mensais(data)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_motor_runs_start ON motor_runs(start_time)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_equip_local ON equipamentos(local_id)")
        conn.commit()
    finally:
        conn.close()

    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_local_datahora ON leituras(local, datahora)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_datahora ON leituras(datahora)")
    except Exception:
        pass


def _detect_tariff_column(c):
    # tenta descobrir a coluna de tarifa em locais_cfg
    try:
        cols = [r[1] for r in c.execute("PRAGMA table_info(locais_cfg)").fetchall()]
    except Exception:
        return None
    for name in ("tarifa_kwh","tarifa_ativa","tarifa"):
        if name in cols:
            return name
    return None

def _prev_month(mes, ano):
    m = int(mes)
    if m == 1:
        return "12", ano - 1
    return f"{m-1:02d}", ano

def _dias_no_mes(mes, ano):
    import calendar
    return calendar.monthrange(int(ano), int(mes))[1]

def _agg_dashboard(mes, ano, local_id=None):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    try:
        # detetar coluna de tarifa (se existir)
        tarifa_col = _detect_tariff_column(c)
        where_local = ""
        where_local_runs = ""
        params_main = [mes, ano]
        params_runs = [mes, str(ano)]
        params_daily = [mes, ano]
        if local_id:
            where_local = " AND l.id = ? "
            params_main.append(int(local_id))
            where_local_runs = " AND e.local_id = ? "
            params_runs.append(int(local_id))
            params_daily.append(int(local_id))

        # agregados por local no mês
        rows = c.execute(f"""
            SELECT l.id AS local_id, l.nome AS local,
                   ROUND(COALESCE(SUM(m.diferenca),0),2) AS energia_mes,
                   SUM(CASE WHEN m.fp IS NOT NULL AND m.fp < 0.80 THEN 1 ELSE 0 END) AS fp_baixo,
                   MAX(m.ponta) AS ponta_max,
                   COUNT(DISTINCT m.data) AS dias_com_dados
            FROM locais l
            LEFT JOIN leituras_mensais m
                   ON m.local = l.nome AND m.mes = ? AND m.ano = ?
            WHERE 1=1 {where_local}
            GROUP BY l.id, l.nome
            ORDER BY energia_mes DESC, l.nome ASC
        """, params_main).fetchall()

        # horas de motor por local no mês
        hrs = dict(c.execute(f"""
            SELECT e.local_id, ROUND(COALESCE(SUM(r.duracao_min),0)/60.0, 2) as horas
            FROM motor_runs r
            JOIN equipamentos e ON e.id = r.equipamento_id
            WHERE strftime('%m', r.start_time)=? AND strftime('%Y', r.start_time)=? {where_local_runs}
            GROUP BY e.local_id
        """, params_runs).fetchall())

        # tarifas por local (se houver)
        tarifas = {}
        if tarifa_col:
            try:
                tarifas = dict(c.execute(f"""
                    SELECT lc.local_id, ROUND(COALESCE(lc.{tarifa_col},0),4) as tarifa
                    FROM locais_cfg lc
                """).fetchall())
            except Exception:
                tarifas = {}

        # KPI globais e cartões
        cards, energia_total, ponta_max_global, locais_fp_baixo = [], 0.0, 0.0, 0
        dias_mes = _dias_no_mes(mes, ano)
        locais_cobertura_baixa = 0
        custo_total = 0.0
        custos_habilitados = bool(tarifa_col)

        for lid, lname, energia_mes, fp_baixo, ponta_max, dias_com_dados in rows:
            energia_mes = float(energia_mes or 0)
            fp_baixo = int(fp_baixo or 0)
            ponta_max = float(ponta_max or 0)
            dias_com_dados = int(dias_com_dados or 0)
            horas_motores = float(hrs.get(lid, 0.0))

            cobertura_pct = round( (dias_com_dados * 100.0 / dias_mes) if dias_mes else 0.0 , 1)
            if cobertura_pct < 80.0:
                locais_cobertura_baixa += 1

            tarifa_kwh = tarifas.get(lid) if custos_habilitados else None
            custo_estimado = round(energia_mes * float(tarifa_kwh), 2) if (custos_habilitados and tarifa_kwh is not None) else None
            if custo_estimado is not None:
                custo_total += custo_estimado

            cards.append({
                "local_id": lid,
                "local": lname,
                "energia_mes": energia_mes,
                "fp_baixo": fp_baixo,
                "ponta_max": ponta_max,
                "horas_motores": horas_motores,
                "dias_com_dados": dias_com_dados,
                "dias_mes": dias_mes,
                "cobertura_pct": cobertura_pct,
                "tarifa_kwh": tarifa_kwh if tarifa_kwh is not None else None,
                "custo_estimado": custo_estimado if custo_estimado is not None else None
            })

            energia_total += energia_mes
            ponta_max_global = max(ponta_max_global, ponta_max)
            if fp_baixo > 0:
                locais_fp_baixo += 1

        # Ranking (Top 8)
        top = sorted(cards, key=lambda x: x["energia_mes"], reverse=True)[:8]
        rank = {"labels": [x["local"] for x in top], "data": [x["energia_mes"] for x in top]}

        # Tendência diária (todas as leituras do mês, somadas por dia)
        daily = c.execute(f"""
            SELECT m.data, ROUND(COALESCE(SUM(m.diferenca),0),2) as kwh
            FROM leituras_mensais m
            JOIN locais l ON l.nome = m.local
            WHERE m.mes = ? AND m.ano = ? {(" AND l.id = ?" if local_id else "")}
            GROUP BY m.data
            ORDER BY m.data ASC
        """, params_daily).fetchall()
        trend = {"labels": [r[0] for r in daily], "data": [float(r[1] or 0) for r in daily]}

        # M-1 comparação
        mes_prev, ano_prev = _prev_month(mes, ano)
        energia_prev = c.execute("""
            SELECT ROUND(COALESCE(SUM(m.diferenca),0),2)
            FROM leituras_mensais m
            JOIN locais l ON l.nome = m.local
            WHERE m.mes = ? AND m.ano = ?
            """ + ( " AND l.id = ?" if local_id else "" ),
            ([mes_prev, ano_prev] + ([int(local_id)] if local_id else []))
        ).fetchone()[0] or 0.0
        energia_prev = float(energia_prev)

        # custo M-1
        custo_prev = None
        if custos_habilitados:
            # aproximação: custo_prev = sum(energia_prev_por_local * tarifa_local)
            # para simplificar, usa a mesma tarifa atual por local
            if local_id:
                lid = int(local_id)
                t = tarifas.get(lid, 0.0)
                # energia_prev por local selecionado
                eprev_local = c.execute("""
                    SELECT ROUND(COALESCE(SUM(m.diferenca),0),2)
                    FROM leituras_mensais m
                    JOIN locais l ON l.nome = m.local
                    WHERE m.mes = ? AND m.ano = ? AND l.id = ?
                """, [mes_prev, ano_prev, lid]).fetchone()[0] or 0.0
                custo_prev = round(float(eprev_local) * float(t), 2)
            else:
                # calcula energia_prev por local
                eprev_rows = c.execute("""
                    SELECT l.id, ROUND(COALESCE(SUM(m.diferenca),0),2)
                    FROM leituras_mensais m
                    JOIN locais l ON l.nome = m.local
                    WHERE m.mes = ? AND m.ano = ?
                    GROUP BY l.id
                """, [mes_prev, ano_prev]).fetchall()
                cprev = 0.0
                for lid, eprev in eprev_rows:
                    cprev += float(eprev or 0) * float(tarifas.get(lid, 0.0))
                custo_prev = round(cprev, 2)

        # deltas
        def _pct_delta(cur, prev):
            try:
                if prev == 0:
                    return None if cur == 0 else 100.0
                return (float(cur) - float(prev)) * 100.0 / float(prev)
            except Exception:
                return None

        delta_energia_pct = _pct_delta(energia_total, energia_prev)
        delta_custo_pct = _pct_delta(custo_total if custos_habilitados else None, custo_prev) if custos_habilitados else None

        kpis = {
            "energia_total": round(energia_total, 2),
            "ponta_max_global": round(ponta_max_global, 2),
            "locais_fp_baixo": locais_fp_baixo,
            "horas_motores_total": round(sum(x["horas_motores"] for x in cards), 2),
            "locais_cobertura_baixa": int(locais_cobertura_baixa),
            "custo_total": round(custo_total, 2) if custos_habilitados else 0,
            "custos_habilitados": custos_habilitados,
            "delta_energia_pct": delta_energia_pct,
            "delta_custo_pct": delta_custo_pct
        }

        return cards, kpis, rank, trend
    finally:
        conn.close()

# === DASHBOARD (versão excelência v2) ===
from collections import defaultdict

def _ensure_idx_dashboard():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_mensais_local_mes_ano ON leituras_mensais(local, mes, ano)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_mensais_data ON leituras_mensais(data)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_motor_runs_start ON motor_runs(start_time)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_equip_local ON equipamentos(local_id)")
        conn.commit()
    finally:
        conn.close()

def _detect_tariff_column(c):
    # tenta descobrir a coluna de tarifa em locais_cfg
    try:
        cols = [r[1] for r in c.execute("PRAGMA table_info(locais_cfg)").fetchall()]
    except Exception:
        return None
    for name in ("tarifa_kwh","tarifa_ativa","tarifa"):
        if name in cols:
            return name
    return None

def _prev_month(mes, ano):
    m = int(mes)
    if m == 1:
        return "12", ano - 1
    return f"{m-1:02d}", ano

def _dias_no_mes(mes, ano):
    import calendar
    return calendar.monthrange(int(ano), int(mes))[1]

def _agg_dashboard(mes, ano, local_id=None):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    try:
        # detetar coluna de tarifa (se existir)
        tarifa_col = _detect_tariff_column(c)
        where_local = ""
        where_local_runs = ""
        params_main = [mes, ano]
        params_runs = [mes, str(ano)]
        params_daily = [mes, ano]
        if local_id:
            where_local = " AND l.id = ? "
            params_main.append(int(local_id))
            where_local_runs = " AND e.local_id = ? "
            params_runs.append(int(local_id))
            params_daily.append(int(local_id))

        # agregados por local no mês
        rows = c.execute(f"""
            SELECT l.id AS local_id, l.nome AS local,
                   ROUND(COALESCE(SUM(m.diferenca),0),2) AS energia_mes,
                   SUM(CASE WHEN m.fp IS NOT NULL AND m.fp < 0.80 THEN 1 ELSE 0 END) AS fp_baixo,
                   MAX(m.ponta) AS ponta_max,
                   COUNT(DISTINCT m.data) AS dias_com_dados
            FROM locais l
            LEFT JOIN leituras_mensais m
                   ON m.local = l.nome AND m.mes = ? AND m.ano = ?
            WHERE 1=1 {where_local}
            GROUP BY l.id, l.nome
            ORDER BY energia_mes DESC, l.nome ASC
        """, params_main).fetchall()

        # horas de motor por local no mês
        hrs = dict(c.execute(f"""
            SELECT e.local_id, ROUND(COALESCE(SUM(r.duracao_min),0)/60.0, 2) as horas
            FROM motor_runs r
            JOIN equipamentos e ON e.id = r.equipamento_id
            WHERE strftime('%m', r.start_time)=? AND strftime('%Y', r.start_time)=? {where_local_runs}
            GROUP BY e.local_id
        """, params_runs).fetchall())

        # tarifas por local (se houver)
        tarifas = {}
        if tarifa_col:
            try:
                tarifas = dict(c.execute(f"""
                    SELECT lc.local_id, ROUND(COALESCE(lc.{tarifa_col},0),4) as tarifa
                    FROM locais_cfg lc
                """).fetchall())
            except Exception:
                tarifas = {}

        # KPI globais e cartões
        cards, energia_total, ponta_max_global, locais_fp_baixo = [], 0.0, 0.0, 0
        dias_mes = _dias_no_mes(mes, ano)
        locais_cobertura_baixa = 0
        custo_total = 0.0
        custos_habilitados = bool(tarifa_col)

        for lid, lname, energia_mes, fp_baixo, ponta_max, dias_com_dados in rows:
            energia_mes = float(energia_mes or 0)
            fp_baixo = int(fp_baixo or 0)
            ponta_max = float(ponta_max or 0)
            dias_com_dados = int(dias_com_dados or 0)
            horas_motores = float(hrs.get(lid, 0.0))

            cobertura_pct = round( (dias_com_dados * 100.0 / dias_mes) if dias_mes else 0.0 , 1)
            if cobertura_pct < 80.0:
                locais_cobertura_baixa += 1

            tarifa_kwh = tarifas.get(lid) if custos_habilitados else None
            custo_estimado = round(energia_mes * float(tarifa_kwh), 2) if (custos_habilitados and tarifa_kwh is not None) else None
            if custo_estimado is not None:
                custo_total += custo_estimado

            cards.append({
                "local_id": lid,
                "local": lname,
                "energia_mes": energia_mes,
                "fp_baixo": fp_baixo,
                "ponta_max": ponta_max,
                "horas_motores": horas_motores,
                "dias_com_dados": dias_com_dados,
                "dias_mes": dias_mes,
                "cobertura_pct": cobertura_pct,
                "tarifa_kwh": tarifa_kwh if tarifa_kwh is not None else None,
                "custo_estimado": custo_estimado if custo_estimado is not None else None
            })

            energia_total += energia_mes
            ponta_max_global = max(ponta_max_global, ponta_max)
            if fp_baixo > 0:
                locais_fp_baixo += 1

        # Ranking (Top 8)
        top = sorted(cards, key=lambda x: x["energia_mes"], reverse=True)[:8]
        rank = {"labels": [x["local"] for x in top], "data": [x["energia_mes"] for x in top]}

        # Tendência diária (todas as leituras do mês, somadas por dia)
        daily = c.execute(f"""
            SELECT m.data, ROUND(COALESCE(SUM(m.diferenca),0),2) as kwh
            FROM leituras_mensais m
            JOIN locais l ON l.nome = m.local
            WHERE m.mes = ? AND m.ano = ? {(" AND l.id = ?" if local_id else "")}
            GROUP BY m.data
            ORDER BY m.data ASC
        """, params_daily).fetchall()
        trend = {"labels": [r[0] for r in daily], "data": [float(r[1] or 0) for r in daily]}

        # M-1 comparação
        mes_prev, ano_prev = _prev_month(mes, ano)
        energia_prev = c.execute("""
            SELECT ROUND(COALESCE(SUM(m.diferenca),0),2)
            FROM leituras_mensais m
            JOIN locais l ON l.nome = m.local
            WHERE m.mes = ? AND m.ano = ?
            """ + ( " AND l.id = ?" if local_id else "" ),
            ([mes_prev, ano_prev] + ([int(local_id)] if local_id else []))
        ).fetchone()[0] or 0.0
        energia_prev = float(energia_prev)

        # custo M-1
        custo_prev = None
        if custos_habilitados:
            # aproximação: custo_prev = sum(energia_prev_por_local * tarifa_local)
            # para simplificar, usa a mesma tarifa atual por local
            if local_id:
                lid = int(local_id)
                t = tarifas.get(lid, 0.0)
                # energia_prev por local selecionado
                eprev_local = c.execute("""
                    SELECT ROUND(COALESCE(SUM(m.diferenca),0),2)
                    FROM leituras_mensais m
                    JOIN locais l ON l.nome = m.local
                    WHERE m.mes = ? AND m.ano = ? AND l.id = ?
                """, [mes_prev, ano_prev, lid]).fetchone()[0] or 0.0
                custo_prev = round(float(eprev_local) * float(t), 2)
            else:
                # calcula energia_prev por local
                eprev_rows = c.execute("""
                    SELECT l.id, ROUND(COALESCE(SUM(m.diferenca),0),2)
                    FROM leituras_mensais m
                    JOIN locais l ON l.nome = m.local
                    WHERE m.mes = ? AND m.ano = ?
                    GROUP BY l.id
                """, [mes_prev, ano_prev]).fetchall()
                cprev = 0.0
                for lid, eprev in eprev_rows:
                    cprev += float(eprev or 0) * float(tarifas.get(lid, 0.0))
                custo_prev = round(cprev, 2)

        # deltas
        def _pct_delta(cur, prev):
            try:
                if prev == 0:
                    return None if cur == 0 else 100.0
                return (float(cur) - float(prev)) * 100.0 / float(prev)
            except Exception:
                return None

        delta_energia_pct = _pct_delta(energia_total, energia_prev)
        delta_custo_pct = _pct_delta(custo_total if custos_habilitados else None, custo_prev) if custos_habilitados else None

        kpis = {
            "energia_total": round(energia_total, 2),
            "ponta_max_global": round(ponta_max_global, 2),
            "locais_fp_baixo": locais_fp_baixo,
            "horas_motores_total": round(sum(x["horas_motores"] for x in cards), 2),
            "locais_cobertura_baixa": int(locais_cobertura_baixa),
            "custo_total": round(custo_total, 2) if custos_habilitados else 0,
            "custos_habilitados": custos_habilitados,
            "delta_energia_pct": delta_energia_pct,
            "delta_custo_pct": delta_custo_pct
        }

        return cards, kpis, rank, trend
    finally:
        conn.close()

@app.route('/dashboard')
def dashboard():
    _ensure_idx_dashboard()
    hoje = datetime.now()
    mes = (request.args.get('mes') or hoje.strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or hoje.year)
    local_id = request.args.get('local_id')

    cards, kpis, rank, trend = _agg_dashboard(mes, ano, local_id=local_id)
    locais = get_locais()
    return render_template('dashboard.html', cards=cards, kpis=kpis, rank=rank, trend=trend, locais=locais, mes=mes, ano=ano, local_id=local_id)

@app.route('/dashboard/export')
def dashboard_export():
    # export CSV com os agregados mostrados na tela
    hoje = datetime.now()
    mes = (request.args.get('mes') or hoje.strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or hoje.year)
    local_id = request.args.get('local_id')

    cards, kpis, rank, trend = _agg_dashboard(mes, ano, local_id=local_id)

    import csv
    from io import StringIO
    si = StringIO()
    w = csv.writer(si, delimiter=';')
    header = ["local_id","local","energia_kwh","fp_baixo_dias","ponta_max_kw","horas_motores_h","dias_com_dados","dias_mes","cobertura_pct","tarifa_kwh","custo_estimado_mzn"]
    w.writerow(header)
    for c in cards:
        w.writerow([
            c["local_id"], c["local"], c["energia_mes"], c["fp_baixo"], c["ponta_max"],
            c["horas_motores"], c["dias_com_dados"], c["dias_mes"], c["cobertura_pct"],
            ("" if c["tarifa_kwh"] is None else c["tarifa_kwh"]),
            ("" if c["custo_estimado"] is None else c["custo_estimado"]),
        ])
    output = si.getvalue().encode("utf-8-sig")
    return Response(output, mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename=dashboard_{ano}-{mes}.csv"})# === LOCAIS ===

@app.route('/locais')
def listar_locais():
    q = (request.args.get('q') or '').strip()
    incluir_inativos = (request.args.get('inativos') == '1')
    sort = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')
    status = (request.args.get('status') or 'todos').strip().lower()
    tipo = (request.args.get('tipo') or 'todos').strip()
    qualidade = (request.args.get('qualidade') or 'todos').strip().lower()
    estado_tecnico = (request.args.get('estado_tecnico') or 'todos').strip()
    prioridade = (request.args.get('prioridade') or 'todas').strip()

    locais = get_locais_with_cfg(q if q else None, incluir_inativos=incluir_inativos, sort=sort, order=order,
                                 tipo=tipo, qualidade=qualidade, estado_tecnico=estado_tecnico, prioridade=prioridade)
    if status == 'ativos':
        locais = [r for r in locais if int(r.get('ativo', 1)) == 1]
    elif status == 'arquivados':
        locais = [r for r in locais if int(r.get('ativo', 1)) != 1]
    elif status == 'sem_config':
        locais = [r for r in locais if not r.get('config_ok')]

    resumo = get_locais_module_summary(locais)
    ranking_atencao = sorted(locais, key=lambda x: (x.get('maturidade', 0), x.get('pot_contratada', 0) + x.get('pot_instalada', 0)))[:5]
    tipos_disponiveis = ['todos'] + sorted({r.get('tipo','Outro') for r in get_locais_with_cfg(None, incluir_inativos=True)})
    estados_tecnicos = ['todos', 'Normal', 'Atenção', 'Crítico']
    prioridades = ['todas', 'Baixa', 'Média', 'Alta']
    return render_template('locais.html',
                           locais=locais, q=q,
                           incluir_inativos=incluir_inativos,
                           sort=sort, order=order,
                           status=status, tipo=tipo, qualidade=qualidade,
                           estado_tecnico=estado_tecnico, prioridade=prioridade,
                           tipos_disponiveis=tipos_disponiveis,
                           estados_tecnicos=estados_tecnicos,
                           prioridades=prioridades,
                           ranking_atencao=ranking_atencao,
                           resumo=resumo)


@app.route('/locais/template.csv')
def export_locais_template_csv():
    si = StringIO()
    w = csv.writer(si, delimiter=';')
    w.writerow(['nome','codigo','endereco','contato_nome','contato_tel','email','responsavel_alt','tipo_local','categoria_operacional','estado_tecnico','prioridade','ativo','fator_mult','pot_contratada','pot_instalada','tarifa_ativa','tarifa_reativa','tarifa_ponta','tarifa_perdas','taxa_fixa','taxa_radio','taxa_lixo','iva','notas'])
    w.writerow(['Ex.: ETA Umbeluzi','ETA-UMB','Umbeluzi, Maputo','Supervisor Local','84xxxxxxx','supervisor@adrmm.co.mz','Chefe de turno','ETA','Produção','Normal','Alta',1,1.0,6000,11750,4.780,1.430,4.970,4.780,207.28,297.00,150.00,16,'Local de referência'])
    output = si.getvalue()
    return Response(output.encode('utf-8'), mimetype='text/csv; charset=utf-8', headers={"Content-Disposition": "attachment; filename=template_locais.csv"})


@app.route('/locais/export.csv')
def export_locais_csv():
    q = (request.args.get('q') or '').strip()
    incluir_inativos = (request.args.get('inativos') == '1')
    sort = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')
    data = get_locais_with_cfg(q if q else None, incluir_inativos=incluir_inativos, sort=sort, order=order)

    si = StringIO()
    w = csv.writer(si, delimiter=';')
    w.writerow([
        'id','nome','codigo','endereco','contato_nome','contato_tel','email','responsavel_alt','tipo_local','categoria_operacional','estado_tecnico','prioridade','ativo',
        'fator_mult','pot_contratada_kW','pot_instalada_kW','notas'
    ])
    for r in data:
        w.writerow([
            r['id'], r['nome'], r.get('codigo','') or '', r.get('endereco','') or '',
            r.get('contato_nome','') or '', r.get('contato_tel','') or '', r.get('email','') or '', r.get('responsavel_alt','') or '',
            r.get('tipo_local','') or '', r.get('categoria_operacional','') or '', r.get('estado_tecnico','') or '', r.get('prioridade','') or '',
            r.get('ativo',1),
            f"{r['fator_mult']:.4f}",
            f"{r['pot_contratada']:.2f}",
            f"{r['pot_instalada']:.2f}",
            (r.get('notas','') or '').replace('\n',' ').strip()
        ])
    output = si.getvalue()
    return Response(output.encode('utf-8'),
                    mimetype='text/csv; charset=utf-8',
                    headers={"Content-Disposition": "attachment; filename=locais.csv"})

@app.route('/locais/export.xlsx')
def export_locais_xlsx():
    q = (request.args.get('q') or '').strip()
    incluir_inativos = (request.args.get('inativos') == '1')
    sort = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')
    data = get_locais_with_cfg(q if q else None, incluir_inativos=incluir_inativos, sort=sort, order=order)

    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet('Locais')
    hdr = wb.add_format({'bold': True, 'bg_color': '#EAF4FF', 'font_color': '#174983', 'border': 1})
    txt = wb.add_format({'border': 1})
    num = wb.add_format({'border': 1, 'num_format': '0.00'})
    headers = ['ID','Nome','Tipo','Categoria','Código','Endereço','Contacto','Telefone','Email','Responsável alt.','Estado técnico','Prioridade','Ativo','Fator','Pot. contratada (kW)','Pot. instalada (kW)','Maturidade (%)']
    for col,h in enumerate(headers): ws.write(0,col,h,hdr)
    for i,r in enumerate(data, start=1):
        vals = [r['id'], r['nome'], r.get('tipo'), r.get('categoria_operacional') or '', r.get('codigo') or '', r.get('endereco') or '',
                r.get('contato_nome') or '', r.get('contato_tel') or '', r.get('email') or '', r.get('responsavel_alt') or '',
                r.get('estado_tecnico') or '', r.get('prioridade') or '', 'Sim' if int(r.get('ativo',1))==1 else 'Não',
                float(r.get('fator_mult',1) or 1), float(r.get('pot_contratada',0) or 0), float(r.get('pot_instalada',0) or 0), int(r.get('maturidade',0) or 0)]
        for col,v in enumerate(vals):
            fmt = num if isinstance(v,(int,float)) and col in [13,14,15,16] else txt
            ws.write(i,col,v,fmt)
    ws.autofilter(0,0,max(len(data),1),len(headers)-1)
    ws.freeze_panes(1,0)
    for idx,w in enumerate([8,28,12,18,12,24,18,15,22,20,16,12,10,10,18,18,14]): ws.set_column(idx,idx,w)
    wb.close()
    output.seek(0)
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment; filename=locais.xlsx'})


@app.route('/locais/<int:local_id>')
def detalhes_local(local_id):
    local = get_local_full(local_id)
    if not local:
        flash('Local não encontrado.', 'warning')
        return redirect(url_for('listar_locais'))

    cfg_full = get_local_cfg_full(local_id)
    equipamentos = get_equipamentos_por_local(local_id)
    overview = get_local_overview(local_id)
    local['tipo'] = local.get('tipo_local') or infer_local_tipo(local.get('nome'), local.get('endereco'))
    local['maturidade'] = calcular_maturidade_local({**local, **cfg_full})
    alertas = get_local_alertas(local, cfg_full, overview)
    history = get_local_history(local_id, limit=12)
    sublocais = get_local_children(local_id, include_inactive=True)
    return render_template('detalhes_local.html', local=local, cfg=cfg_full, equipamentos=equipamentos, overview=overview, alertas=alertas, history=history, sublocais=sublocais)


@app.route('/locais/adicionar', methods=['GET', 'POST'])
def adicionar_local():
    parent_options = get_local_choices(include_inactive=True)
    if request.method == 'POST':
        nome = (request.form.get('nome') or '').strip()
        codigo = (request.form.get('codigo') or '').strip() or None
        endereco = (request.form.get('endereco') or '').strip() or None
        contato_nome = (request.form.get('contato_nome') or '').strip() or None
        contato_tel = (request.form.get('contato_tel') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        responsavel_alt = (request.form.get('responsavel_alt') or '').strip() or None
        tipo_local = (request.form.get('tipo_local') or '').strip() or None
        categoria_operacional = (request.form.get('categoria_operacional') or '').strip() or None
        estado_tecnico = (request.form.get('estado_tecnico') or 'Normal').strip() or 'Normal'
        prioridade = (request.form.get('prioridade') or 'Média').strip() or 'Média'
        notas = (request.form.get('notas') or '').strip() or None
        ativo = 1 if (request.form.get('ativo', '1') == '1') else 0
        parent_raw = (request.form.get('parent_id') or '').strip()
        parent_id = int(parent_raw) if parent_raw.isdigit() else None
        if not nome:
            flash('O nome do local é obrigatório.', 'warning')
            return render_template('adicionar_local.html', form=request.form, parent_options=parent_options)
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        try:
            c.execute('''INSERT INTO locais (nome, codigo, endereco, contato_nome, contato_tel, email, responsavel_alt, tipo_local, categoria_operacional, estado_tecnico, prioridade, notas, ativo, parent_id)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                      (nome, codigo, endereco, contato_nome, contato_tel, email, responsavel_alt, tipo_local, categoria_operacional, estado_tecnico, prioridade, notas, ativo, parent_id))
            lid = c.lastrowid
            c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (lid,))
            conn.commit()
            log_local_history(lid, 'Local criado', f'Cadastro inicial do local {nome}', actor='locais_fase4')
            flash(f'Local "{nome}" criado com sucesso.', 'success')
            return redirect(url_for('detalhes_local', local_id=lid))
        except sqlite3.IntegrityError:
            flash('Já existe um local com esse nome.', 'danger')
        except Exception as e:
            flash(f'Não foi possível criar o local: {e}', 'danger')
        finally:
            conn.close()
    return render_template('adicionar_local.html', form=request.form, parent_options=parent_options)


@app.route('/locais/editar/<int:local_id>', methods=['GET', 'POST'])
def editar_local(local_id):
    local = get_local_full(local_id)
    parent_options = get_local_choices(include_inactive=True, exclude_id=local_id)
    if not local:
        flash('Local não encontrado.', 'warning')
        return redirect(url_for('listar_locais'))
    if request.method == 'POST':
        novo_nome = (request.form.get('nome') or '').strip()
        codigo = (request.form.get('codigo') or '').strip() or None
        endereco = (request.form.get('endereco') or '').strip() or None
        contato_nome = (request.form.get('contato_nome') or '').strip() or None
        contato_tel = (request.form.get('contato_tel') or '').strip() or None
        email = (request.form.get('email') or '').strip() or None
        responsavel_alt = (request.form.get('responsavel_alt') or '').strip() or None
        tipo_local = (request.form.get('tipo_local') or '').strip() or None
        categoria_operacional = (request.form.get('categoria_operacional') or '').strip() or None
        estado_tecnico = (request.form.get('estado_tecnico') or 'Normal').strip() or 'Normal'
        prioridade = (request.form.get('prioridade') or 'Média').strip() or 'Média'
        notas = (request.form.get('notas') or '').strip() or None
        ativo = 1 if (request.form.get('ativo', '1') == '1') else 0
        parent_raw = (request.form.get('parent_id') or '').strip()
        parent_id = int(parent_raw) if parent_raw.isdigit() else None
        if parent_id == local_id:
            parent_id = None
        if not novo_nome:
            flash('O nome do local é obrigatório.', 'warning')
            local.update({'nome': novo_nome, 'codigo': codigo, 'endereco': endereco, 'contato_nome': contato_nome, 'contato_tel': contato_tel, 'email': email, 'responsavel_alt': responsavel_alt, 'tipo_local': tipo_local, 'categoria_operacional': categoria_operacional, 'estado_tecnico': estado_tecnico, 'prioridade': prioridade, 'notas': notas, 'ativo': ativo})
            return render_template('editar_local.html', local=local, parent_options=parent_options)
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        try:
            c.execute('''UPDATE locais
                            SET nome=?, codigo=?, endereco=?, contato_nome=?, contato_tel=?, email=?, responsavel_alt=?, tipo_local=?, categoria_operacional=?, estado_tecnico=?, prioridade=?, notas=?, ativo=?, parent_id=?
                          WHERE id=?''',
                      (novo_nome, codigo, endereco, contato_nome, contato_tel, email, responsavel_alt, tipo_local, categoria_operacional, estado_tecnico, prioridade, notas, ativo, parent_id, local_id))
            conn.commit()
            log_local_history(local_id, 'Perfil atualizado', f'Nome: {novo_nome}; prioridade: {prioridade}; estado técnico: {estado_tecnico}', actor='locais_fase4')
            flash('Local atualizado com sucesso.', 'success')
            return redirect(url_for('detalhes_local', local_id=local_id))
        except sqlite3.IntegrityError:
            flash('Já existe outro local com esse nome.', 'danger')
        except Exception as e:
            flash(f'Não foi possível atualizar o local: {e}', 'danger')
        finally:
            conn.close()
        local.update({'nome': novo_nome, 'codigo': codigo, 'endereco': endereco, 'contato_nome': contato_nome, 'contato_tel': contato_tel, 'notas': notas, 'ativo': ativo})
    return render_template('editar_local.html', local=local, parent_options=parent_options)

# === NOVO: Importar Locais
# === NOVO: Importar Locais (CSV)
@app.route('/locais/import', methods=['GET','POST'])
def locais_import():
    if request.method == 'POST':
        f = request.files.get('arquivo')
        if not f or f.filename == '':
            return redirect(url_for('listar_locais', msg='selecione um arquivo CSV'))

        content = f.read().decode('utf-8', errors='ignore')
        delimiter = ';' if content.count(';') > content.count(',') else ','
        reader = csv.DictReader(StringIO(content), delimiter=delimiter)

        # Checa coluna obrigatória
        headers = [h.strip().lower() for h in (reader.fieldnames or [])]
        if 'nome' not in headers:
            return redirect(url_for('listar_locais', msg='CSV precisa conter a coluna "nome"'))

        add_count = 0; upd_count = 0; err_count = 0
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()

        def ffloat(v, dflt):
            try: return float(str(v).replace(',','.'))
            except: return dflt

        for row in reader:
            try:
                nome = (row.get('nome') or '').strip()
                if not nome:
                    err_count += 1; continue

                # Upsert em locais (com colunas novas)
                campos_locais = {
                    "codigo": row.get('codigo'),
                    "endereco": row.get('endereco'),
                    "contato_nome": row.get('contato_nome'),
                    "contato_tel": row.get('contato_tel'),
                    "email": row.get('email'),
                    "responsavel_alt": row.get('responsavel_alt'),
                    "tipo_local": row.get('tipo_local'),
                    "categoria_operacional": row.get('categoria_operacional'),
                    "estado_tecnico": row.get('estado_tecnico') or 'Normal',
                    "prioridade": row.get('prioridade') or 'Média',
                    "notas": row.get('notas'),
                    "ativo": int(row.get('ativo', '1')) if str(row.get('ativo','1')).strip() != '' else 1
                }
                c.execute('SELECT id FROM locais WHERE nome=?', (nome,))
                found = c.fetchone()
                if found:
                    lid = found[0]
                    c.execute('''UPDATE locais
                                    SET codigo=?, endereco=?, contato_nome=?, contato_tel=?, email=?, responsavel_alt=?, tipo_local=?, categoria_operacional=?, estado_tecnico=?, prioridade=?, notas=?, ativo=?
                                  WHERE id=?''',
                              (campos_locais["codigo"], campos_locais["endereco"],
                               campos_locais["contato_nome"], campos_locais["contato_tel"], campos_locais["email"], campos_locais["responsavel_alt"], campos_locais["tipo_local"], campos_locais["categoria_operacional"], campos_locais["estado_tecnico"], campos_locais["prioridade"],
                               campos_locais["notas"], campos_locais["ativo"], lid))
                    upd_count += 1
                else:
                    c.execute('''INSERT INTO locais (nome, codigo, endereco, contato_nome, contato_tel, email, responsavel_alt, tipo_local, categoria_operacional, estado_tecnico, prioridade, notas, ativo)
                                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                              (nome, campos_locais["codigo"], campos_locais["endereco"],
                               campos_locais["contato_nome"], campos_locais["contato_tel"], campos_locais["email"], campos_locais["responsavel_alt"], campos_locais["tipo_local"], campos_locais["categoria_operacional"], campos_locais["estado_tecnico"], campos_locais["prioridade"],
                               campos_locais["notas"], campos_locais["ativo"]))
                    lid = c.lastrowid
                    add_count += 1

                # Upsert em locais_cfg (se vierem colunas)
                cfg = {
                    "fator_mult": ffloat(row.get('fator_mult'), 1.0),
                    "pot_contratada": ffloat(row.get('pot_contratada'), 0.0),
                    "pot_instalada": ffloat(row.get('pot_instalada'), 0.0),
                    "tarifa_ativa": ffloat(row.get('tarifa_ativa'), 4.780),
                    "tarifa_reativa": ffloat(row.get('tarifa_reativa'), 1.430),
                    "tarifa_ponta": ffloat(row.get('tarifa_ponta'), 4.970),
                    "tarifa_perdas": ffloat(row.get('tarifa_perdas'), 4.780),
                    "taxa_fixa": ffloat(row.get('taxa_fixa'), 207.28),
                    "taxa_radio": ffloat(row.get('taxa_radio'), 297.00),
                    "taxa_lixo": ffloat(row.get('taxa_lixo'), 150.00),
                    "iva": ffloat(row.get('iva'), 16.0),
                }
                c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (lid,))
                c.execute('''UPDATE locais_cfg
                                SET fator_mult=?, pot_contratada=?, pot_instalada=?,
                                    tarifa_ativa=?, tarifa_reativa=?, tarifa_ponta=?, tarifa_perdas=?,
                                    taxa_fixa=?, taxa_radio=?, taxa_lixo=?, iva=?
                              WHERE local_id=?''',
                          (cfg["fator_mult"], cfg["pot_contratada"], cfg["pot_instalada"],
                           cfg["tarifa_ativa"], cfg["tarifa_reativa"], cfg["tarifa_ponta"], cfg["tarifa_perdas"],
                           cfg["taxa_fixa"], cfg["taxa_radio"], cfg["taxa_lixo"], cfg["iva"], lid))
            except Exception:
                err_count += 1

        conn.commit(); conn.close()
        if add_count or upd_count:
            log_local_history(0, 'Importação CSV', f'{add_count} adicionados, {upd_count} atualizados, {err_count} com erro', actor='locais_fase4')
        return redirect(url_for('listar_locais', msg=f'import:{add_count} add, {upd_count} upd, {err_count} err'))
    colunas = ['nome','codigo','endereco','contato_nome','contato_tel','ativo','fator_mult','pot_contratada','pot_instalada','tarifa_ativa','tarifa_reativa','tarifa_ponta','tarifa_perdas','taxa_fixa','taxa_radio','taxa_lixo','iva','notas']
    return render_template('locais_import.html', colunas=colunas)

# === NOVO: Duplicar e Ativar/Arquivar Local
@app.route('/locais/duplicar/<int:local_id>')
def locais_duplicar(local_id):
    info = get_local_full(local_id)
    if not info:
        return redirect(url_for('listar_locais', msg='local_nao_encontrado'))
    base = info['nome'] + " (Cópia)"
    novo_nome = base
    i = 2
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    # evita nome duplicado
    while True:
        c.execute('SELECT 1 FROM locais WHERE nome=?', (novo_nome,))
        if not c.fetchone():
            break
        novo_nome = f"{base} {i}"; i += 1
    # dup local
    c.execute('''INSERT INTO locais (nome, codigo, endereco, contato_nome, contato_tel, notas, ativo, parent_id)
                 VALUES (?, ?, ?, ?, ?, ?, 1, ?)''',
              (novo_nome,
               (info['codigo'] or '') + '-copy' if info['codigo'] else None,
               info['endereco'], info['contato_nome'], info['contato_tel'], info['notas'], info.get('parent_id')))
    novo_id = c.lastrowid
    # dup cfg
    cfg = get_local_cfg_full(local_id)
    c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (novo_id,))
    c.execute('''UPDATE locais_cfg SET fator_mult=?, pot_contratada=?, pot_instalada=?, tarifa_ativa=?, tarifa_reativa=?, tarifa_ponta=?,
                 tarifa_perdas=?, taxa_fixa=?, taxa_radio=?, taxa_lixo=?, iva=? WHERE local_id=?''',
              (cfg['fator_mult'], cfg['pot_contratada'], cfg['pot_instalada'], cfg['tarifa_ativa'], cfg['tarifa_reativa'],
               cfg['tarifa_ponta'], cfg['tarifa_perdas'], cfg['taxa_fixa'], cfg['taxa_radio'], cfg['taxa_lixo'], cfg['iva'], novo_id))
    conn.commit(); conn.close()
    log_local_history(novo_id, 'Local duplicado', f'Criado a partir de {info["nome"]}', actor='locais_fase4')
    return redirect(url_for('listar_locais', msg=f'duplicado:{novo_nome}'))

@app.route('/locais/arquivar/<int:local_id>')
def arquivar_local(local_id):
    return locais_toggle(local_id)

@app.route('/locais/toggle/<int:local_id>')
def locais_toggle(local_id):
    info = get_local_full(local_id)
    if not info:
        return redirect(url_for('listar_locais', msg='local_nao_encontrado'))
    novo = 0 if info['ativo'] == 1 else 1
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('UPDATE locais SET ativo=? WHERE id=?', (novo, local_id))
    conn.commit(); conn.close()
    log_local_history(local_id, 'Estado alterado', 'Ativado' if novo==1 else 'Arquivado', actor='locais_fase4')
    return redirect(url_for('listar_locais', msg=('ativado' if novo==1 else 'arquivado')))

# === CONFIG LOCAL ===
@app.route('/locais/config/<int:local_id>', methods=['GET', 'POST'])
def configurar_local(local_id):
    local = get_local_full(local_id)
    if not local:
        flash('Local não encontrado.', 'warning')
        return redirect(url_for('listar_locais'))

    if request.method == 'POST':
        fator_mult       = float(request.form.get('fator_mult', 1) or 1)
        pot_contratada   = float(request.form.get('pot_contratada', 0) or 0)
        pot_instalada    = float(request.form.get('pot_instalada', 0) or 0)
        tarifa_ativa     = float(request.form.get('tarifa_ativa', 4.780) or 4.780)
        tarifa_reativa   = float(request.form.get('tarifa_reativa', 1.430) or 1.430)
        tarifa_ponta     = float(request.form.get('tarifa_ponta', 4.970) or 4.970)
        tarifa_perdas    = float(request.form.get('tarifa_perdas', 4.780) or 4.780)
        taxa_fixa        = float(request.form.get('taxa_fixa', 207.28) or 207.28)
        taxa_radio       = float(request.form.get('taxa_radio', 297.00) or 297.00)
        taxa_lixo        = float(request.form.get('taxa_lixo', 150.00) or 150.00)
        iva              = float(request.form.get('iva', 16.0) or 16.0)

        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('INSERT OR IGNORE INTO locais_cfg (local_id) VALUES (?)', (local_id,))
        c.execute('''
            UPDATE locais_cfg
               SET fator_mult=?, pot_contratada=?, pot_instalada=?,
                   tarifa_ativa=?, tarifa_reativa=?, tarifa_ponta=?, tarifa_perdas=?,
                   taxa_fixa=?, taxa_radio=?, taxa_lixo=?, iva=?
             WHERE local_id=?
        ''', (fator_mult, pot_contratada, pot_instalada,
              tarifa_ativa, tarifa_reativa, tarifa_ponta, tarifa_perdas,
              taxa_fixa, taxa_radio, taxa_lixo, iva, local_id))
        conn.commit(); conn.close()
        log_local_history(local_id, 'Configuração atualizada', f'Fator {fator_mult:.4f}; Pot. contratada {pot_contratada:.2f} kW; Pot. instalada {pot_instalada:.2f} kW', actor='locais_fase4')
        flash('Configuração do local guardada com sucesso.', 'success')
        return redirect(url_for('configurar_local', local_id=local_id))

    cfg = get_local_cfg_full(local_id)
    overview = get_local_overview(local_id)
    return render_template('local_config.html', local=local, cfg=cfg, overview=overview)

# === EQUIPAMENTOS ===
# === EQUIPAMENTOS ===
def _form_str(name):
    return (request.form.get(name) or '').strip()

def _form_int(name):
    v = _form_str(name)
    if not v:
        return None
    return int(v)

def _form_float(name):
    v = _form_str(name).replace(',', '.')
    if not v:
        return None
    return float(v)

def _equip_form_payload():
    return {
        'nome': _form_str('nome'),
        'local_id': _form_int('local_id'),
        'tag': _form_str('tag'),
        'especificacao': _form_str('especificacao'),
        'ano_instalacao': _form_str('ano_instalacao') or None,
        'quantidade': _form_int('quantidade') or 1,
        'categoria': _form_str('categoria'),
        'fabricante': _form_str('fabricante'),
        'modelo': _form_str('modelo'),
        'numero_serie': _form_str('numero_serie'),
        'custo_aquisicao': _form_float('custo'),
        'vida_util_anos': _form_int('vida_util'),
        'criticidade': _form_str('criticidade'),
        'ativo': 1 if request.form.get('ativo') else 0,
        'potencia_kw': _form_float('potencia_kw'),
        'tensao_v': _form_float('tensao_v'),
        'corrente_a': _form_float('corrente_a'),
        'fornecedor': _form_str('fornecedor'),
        'contrato_num': _form_str('contrato_num'),
        'garantia_fim': _form_str('garantia_fim') or None,
    }

def _equip_validate_payload(payload):
    errors = []
    if not payload['nome']:
        errors.append('O nome do equipamento é obrigatório.')
    if payload['quantidade'] is not None and payload['quantidade'] <= 0:
        errors.append('Quantidade deve ser maior que zero.')
    ano = payload['ano_instalacao']
    if ano and (not str(ano).isdigit() or len(str(ano)) != 4):
        errors.append('Ano de instalação deve estar no formato AAAA.')
    if payload['criticidade'] and payload['criticidade'] not in ('Baixa', 'Média', 'Alta'):
        errors.append('Criticidade inválida.')
    return errors


def _equip_clean_text(value, fallback='—'):
    if value is None:
        return fallback
    if isinstance(value, float):
        try:
            if math.isnan(value):
                return fallback
        except Exception:
            pass
    text = str(value).strip()
    if not text or text.lower() in ('nan', 'none', 'null'):
        return fallback
    return text


def _equip_clean_number(value, decimals=2, fallback='—', zero_as_value=True):
    if value is None:
        return fallback
    try:
        number = float(value)
        if math.isnan(number):
            return fallback
        if not zero_as_value and abs(number) < 1e-12:
            return fallback
        if decimals == 0:
            return str(int(round(number)))
        return f"{number:.{decimals}f}"
    except Exception:
        return fallback


def _equip_bool(value):
    return 1 if str(value).strip() in ('1', 'True', 'true', 'on') else 0


def _equip_form_defaults(equipamento):
    if not equipamento:
        return {}
    return {
        'id': equipamento[0],
        'nome': equipamento[1] or '',
        'local_id': equipamento[2] or '',
        'tag': equipamento[3] or '',
        'especificacao': equipamento[4] or '',
        'ano_instalacao': equipamento[5] or '',
        'quantidade': equipamento[6] or 1,
        'ativo': _equip_bool(equipamento[7]),
        'categoria': equipamento[10] or '',
        'fabricante': equipamento[11] or '',
        'modelo': equipamento[12] or '',
        'numero_serie': equipamento[13] or '',
        'custo': '' if equipamento[14] in (None, '') else equipamento[14],
        'vida_util': '' if equipamento[15] in (None, '') else equipamento[15],
        'criticidade': equipamento[16] or '',
        'potencia_kw': '' if equipamento[19] in (None, '') else equipamento[19],
        'tensao_v': '' if equipamento[20] in (None, '') else equipamento[20],
        'corrente_a': '' if equipamento[21] in (None, '') else equipamento[21],
        'garantia_fim': equipamento[24] or '',
        'fornecedor': equipamento[25] or '',
        'contrato_num': equipamento[26] or '',
    }

@app.route('/equipamentos')
def listar_equipamentos():
    q = request.args.get('q', '').strip()
    local_id = request.args.get('local_id', '').strip()
    incluir_inativos = request.args.get('incluir_inativos', '0') == '1'
    categoria = request.args.get('categoria', '').strip()
    fabricante = request.args.get('fabricante', '').strip()
    modelo = request.args.get('modelo', '').strip()
    criticidade = request.args.get('criticidade', '').strip()
    ano_min = request.args.get('ano_min', '').strip()
    ano_max = request.args.get('ano_max', '').strip()
    sort = request.args.get('sort', 'local_nome')
    order = request.args.get('order', 'asc').lower()
    page = max(1, int(request.args.get('page', 1) or 1))
    per_page = int(request.args.get('per_page', 20) or 20)
    per_page = max(10, min(per_page, 200))
    offset = (page - 1) * per_page

    allowed_sort = {
        'nome': 'e.nome',
        'local_nome': 'l.nome',
        'tag': 'e.tag',
        'ano': 'CAST(COALESCE(e.ano_instalacao,0) AS INTEGER)',
        'quantidade': 'e.quantidade',
        'fabricante': 'e.fabricante',
        'modelo': 'e.modelo',
        'criticidade': 'e.criticidade'
    }
    sort_sql = allowed_sort.get(sort, 'l.nome, e.nome')
    order_sql = 'DESC' if order == 'desc' else 'ASC'

    where_clauses = ["COALESCE(e.deleted_at,'')=''"]
    params = []

    if q:
        _apply_advanced_query(q, where_clauses, params)

    if local_id and local_id.isdigit():
        local_ids_scope = get_descendant_local_ids(int(local_id), include_self=True)
        if not local_ids_scope:
            local_ids_scope = [int(local_id)]
        placeholders_local = ','.join('?' for _ in local_ids_scope)
        where_clauses.append(f"e.local_id IN ({placeholders_local})")
        params.extend(local_ids_scope)

    if categoria:
        where_clauses.append("COALESCE(e.categoria,'') LIKE ?")
        params.append(f"%{categoria}%")
    if fabricante:
        where_clauses.append("COALESCE(e.fabricante,'') LIKE ?")
        params.append(f"%{fabricante}%")
    if modelo:
        where_clauses.append("COALESCE(e.modelo,'') LIKE ?")
        params.append(f"%{modelo}%")
    if criticidade:
        where_clauses.append("COALESCE(e.criticidade,'') = ?")
        params.append(criticidade)
    if ano_min and ano_min.isdigit():
        where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?")
        params.append(int(ano_min))
    if ano_max and ano_max.isdigit():
        where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?")
        params.append(int(ano_max))

    if not incluir_inativos:
        where_clauses.append("COALESCE(e.ativo,1)=1")

    where_sql = "WHERE " + " AND ".join(where_clauses)

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()

    c.execute(f'''
        SELECT COUNT(*),
               SUM(CASE WHEN COALESCE(e.ativo,1)=1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN COALESCE(e.criticidade,'')='Alta' THEN 1 ELSE 0 END),
               SUM(CASE WHEN COALESCE(e.garantia_fim,'')<>'' AND date(e.garantia_fim)>=date('now') THEN 1 ELSE 0 END)
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
    ''', params)
    stat_row = c.fetchone() or (0,0,0,0)
    total = stat_row[0] or 0
    ativos = stat_row[1] or 0
    inativos = max((total or 0) - (ativos or 0), 0)
    criticos = stat_row[2] or 0
    em_garantia = stat_row[3] or 0

    c.execute(f'''
        SELECT e.id, e.nome, COALESCE(l.nome,''), COALESCE(e.tag,''), COALESCE(e.especificacao,''),
               e.ano_instalacao, COALESCE(e.quantidade,0), COALESCE(e.ativo,1),
               COALESCE(e.categoria,''), COALESCE(e.fabricante,''), COALESCE(e.modelo,''), COALESCE(e.numero_serie,''),
               e.custo_aquisicao, COALESCE(e.vida_util_anos,''), COALESCE(e.criticidade,''),
               e.potencia_kw, e.tensao_v, e.corrente_a, e.garantia_fim,
               COALESCE(cp.thumb_filename,(SELECT thumb_filename FROM equipamentos_photos WHERE equipamento_id = e.id ORDER BY uploaded_at DESC LIMIT 1),''),
               COALESCE(e.fornecedor,''), COALESCE(e.contrato_num,'')
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        LEFT JOIN equipamentos_photos cp ON cp.id = e.cover_photo_id
        {where_sql}
        ORDER BY {sort_sql} {order_sql}, e.id DESC
        LIMIT ? OFFSET ?
    ''', params + [per_page, offset])
    rows = c.fetchall()

    equipamentos = []
    for r in rows:
        ano_txt = _equip_clean_text(r[5], fallback='')
        equipamentos.append({
            'id': r[0],
            'nome': _equip_clean_text(r[1]),
            'local': _equip_clean_text(r[2]),
            'tag': _equip_clean_text(r[3]),
            'especificacao': _equip_clean_text(r[4], fallback=''),
            'ano': ano_txt if ano_txt != '—' else '',
            'quantidade': int(r[6] or 0),
            'ativo': _equip_bool(r[7]),
            'categoria': _equip_clean_text(r[8]),
            'fabricante': _equip_clean_text(r[9]),
            'modelo': _equip_clean_text(r[10]),
            'numero_serie': _equip_clean_text(r[11]),
            'custo': _equip_clean_number(r[12], decimals=2, fallback='0.00', zero_as_value=True),
            'vida_util': _equip_clean_text(r[13]),
            'criticidade': _equip_clean_text(r[14]),
            'potencia_kw': _equip_clean_number(r[15], decimals=2),
            'tensao_v': _equip_clean_number(r[16], decimals=0),
            'corrente_a': _equip_clean_number(r[17], decimals=2),
            'garantia_fim': _equip_clean_text(r[18]),
            'cover_thumb': r[19] or '',
            'fornecedor': _equip_clean_text(r[20]),
            'contrato_num': _equip_clean_text(r[21]),
        })

    c.execute('SELECT id, nome FROM locais ORDER BY nome')
    locais = c.fetchall()
    local_nome = ''
    if local_id and str(local_id).isdigit():
        for _lid, _lnome in locais:
            if str(_lid) == str(local_id):
                local_nome = _lnome or ''
                break
    conn.close()

    total_pages = max(1, math.ceil(total / per_page))

    return render_template('equipamentos.html',
                           equipamentos=equipamentos,
                           locais=locais,
                           q=q, sort=sort, order=order,
                           page=page, per_page=per_page,
                           total=total, total_pages=total_pages,
                           local_id=local_id, incluir_inativos=incluir_inativos,
                           categoria=categoria, fabricante=fabricante, modelo=modelo,
                           criticidade=criticidade, ano_min=ano_min, ano_max=ano_max,
                           ativos=ativos, inativos=inativos, criticos=criticos,
                           em_garantia=em_garantia, local_nome=local_nome)
@app.route('/equipamentos/adicionar', methods=['GET', 'POST'])
def adicionar_equipamento():
    locais = get_locais()
    prefill_local_id = (request.args.get('local_id') or '').strip()
    if request.method == 'POST':
        try:
            payload = _equip_form_payload()
        except ValueError:
            flash('Há valores numéricos inválidos no formulário.', 'danger')
            return render_template('adicionar_equipamento.html', locais=locais, form=request.form)

        errors = _equip_validate_payload(payload)
        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('adicionar_equipamento.html', locais=locais, form=request.form)

        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('''
            INSERT INTO equipamentos (
                nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo,
                created_at, updated_at, categoria, fabricante, modelo, numero_serie,
                custo_aquisicao, vida_util_anos, criticidade, potencia_kw, tensao_v,
                corrente_a, fornecedor, contrato_num, garantia_fim
            ) VALUES (
                ?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'), datetime('now','localtime'),
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
        ''', (
            payload['nome'], payload['local_id'], payload['tag'], payload['especificacao'], payload['ano_instalacao'],
            payload['quantidade'], payload['ativo'], payload['categoria'], payload['fabricante'], payload['modelo'],
            payload['numero_serie'], payload['custo_aquisicao'], payload['vida_util_anos'], payload['criticidade'],
            payload['potencia_kw'], payload['tensao_v'], payload['corrente_a'], payload['fornecedor'],
            payload['contrato_num'], payload['garantia_fim']
        ))
        equipamento_id = c.lastrowid
        conn.commit(); conn.close()
        log_equip_audit(equipamento_id, 'criar', f"nome={payload['nome']}")
        flash('Equipamento adicionado com sucesso.', 'success')
        return redirect(url_for('listar_equipamentos'))
    return render_template('adicionar_equipamento.html', locais=locais, form={'local_id': prefill_local_id} if prefill_local_id else None)
@app.route('/equipamentos/editar/<int:equipamento_id>', methods=['GET', 'POST'])
def editar_equipamento(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT * FROM equipamentos WHERE id=?', (equipamento_id,))
    equipamento = c.fetchone()
    locais = get_locais()
    if not equipamento:
        conn.close()
        flash('Equipamento não encontrado.', 'warning')
        return redirect(url_for('listar_equipamentos'))
    if request.method == 'POST':
        try:
            payload = _equip_form_payload()
        except ValueError:
            flash('Há valores numéricos inválidos no formulário.', 'danger')
            conn.close()
            return render_template('editar_equipamento.html', equipamento=equipamento, locais=locais, form=request.form, defaults=_equip_form_defaults(equipamento))

        errors = _equip_validate_payload(payload)
        if errors:
            for e in errors:
                flash(e, 'danger')
            conn.close()
            return render_template('editar_equipamento.html', equipamento=equipamento, locais=locais, form=request.form, defaults=_equip_form_defaults(equipamento))

        c.execute('''
            UPDATE equipamentos
            SET nome=?, local_id=?, tag=?, especificacao=?, ano_instalacao=?, quantidade=?, ativo=?,
                categoria=?, fabricante=?, modelo=?, numero_serie=?, custo_aquisicao=?, vida_util_anos=?,
                criticidade=?, potencia_kw=?, tensao_v=?, corrente_a=?, fornecedor=?, contrato_num=?,
                garantia_fim=?, updated_at=datetime('now','localtime')
            WHERE id=?
        ''', (
            payload['nome'], payload['local_id'], payload['tag'], payload['especificacao'], payload['ano_instalacao'],
            payload['quantidade'], payload['ativo'], payload['categoria'], payload['fabricante'], payload['modelo'],
            payload['numero_serie'], payload['custo_aquisicao'], payload['vida_util_anos'], payload['criticidade'],
            payload['potencia_kw'], payload['tensao_v'], payload['corrente_a'], payload['fornecedor'],
            payload['contrato_num'], payload['garantia_fim'], equipamento_id
        ))
        conn.commit(); conn.close()
        log_equip_audit(equipamento_id, 'editar', f"nome={payload['nome']}")
        flash('Equipamento atualizado.', 'success')
        return redirect(url_for('listar_equipamentos'))

    conn.close()
    return render_template('editar_equipamento.html', equipamento=equipamento, locais=locais, defaults=_equip_form_defaults(equipamento))
# === CONFIG DE EQUIPAMENTO ===
@app.route('/equipamentos/config/<int:equipamento_id>', methods=['GET', 'POST'])
def equipamento_config(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id, nome FROM equipamentos WHERE id=?', (equipamento_id,))
    equipamento = c.fetchone()
    if not equipamento:
        conn.close()
        return redirect(url_for('listar_equipamentos'))

    if request.method == 'POST':
        tensao = request.form.get('tensao_nominal') or None
        corrente = request.form.get('corrente_nominal') or None
        pnom = request.form.get('potencia_nominal_kw') or None
        fpnom = request.form.get('fp_nominal') or None
        efic = request.form.get('eficiencia_nominal') or None
        lim_i = request.form.get('limite_corrente') or None
        lim_fp = request.form.get('limite_fp') or None

        c.execute('INSERT OR IGNORE INTO equipamentos_cfg (equipamento_id) VALUES (?)', (equipamento_id,))
        c.execute('''
            UPDATE equipamentos_cfg
            SET tensao_nominal=?, corrente_nominal=?, potencia_nominal_kw=?, fp_nominal=?,
                eficiencia_nominal=?, limite_corrente=?, limite_fp=?
            WHERE equipamento_id=?
        ''', (tensao, corrente, pnom, fpnom, efic, lim_i, lim_fp, equipamento_id))
        conn.commit(); conn.close()
        return redirect(url_for('listar_equipamentos'))

    c.execute('SELECT tensao_nominal, corrente_nominal, potencia_nominal_kw, fp_nominal, eficiencia_nominal, limite_corrente, limite_fp FROM equipamentos_cfg WHERE equipamento_id=?',
              (equipamento_id,))
    cfg = c.fetchone()
    conn.close()
    return render_template('equipamento_config.html', equipamento=equipamento, cfg=cfg)

# === LEITURAS POR LOCAL ===

@app.route('/leituras/local/<int:local_id>')
def leituras_por_local(local_id):
    local = get_local_by_id(local_id)
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''
        SELECT * FROM leituras WHERE local=?
        ORDER BY datahora
    ''', (local[1],))
    leituras = c.fetchall()
    conn.close()
    return render_template('leituras_local.html', local=local, leituras=leituras)

# === LEITURA DIÁRIA ===

@app.route('/add', methods=['GET', 'POST'])
@app.route('/monitoria/nova', methods=['GET', 'POST'])
def add():
    """Nova leitura operacional diária.
    Este registo é técnico/pontual e não substitui a planilha mensal de faturação.
    """
    def _calc_aparente(a, b):
        try:
            a = float(a or 0); b = float(b or 0)
            return math.sqrt(a*a + b*b) if (a or b) else 0.0
        except Exception:
            return 0.0

    def _calc_fp(ativa, reativa, pot_ativa, pot_reativa, aparente, pot_aparente):
        try:
            if aparente and float(aparente) > 0 and ativa:
                return max(0.0, min(1.0, float(ativa) / float(aparente)))
            if pot_aparente and float(pot_aparente) > 0 and pot_ativa:
                return max(0.0, min(1.0, float(pot_ativa) / float(pot_aparente)))
            base = _calc_aparente(ativa, reativa)
            if base > 0 and ativa:
                return max(0.0, min(1.0, float(ativa or 0) / base))
            basep = _calc_aparente(pot_ativa, pot_reativa)
            if basep > 0 and pot_ativa:
                return max(0.0, min(1.0, float(pot_ativa or 0) / basep))
        except Exception:
            pass
        return 0.0

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    try:
        c.execute('''SELECT l.id, l.nome, COALESCE(cfg.fator_mult,1.0), COALESCE(cfg.pot_contratada,0.0), COALESCE(cfg.pot_instalada,0.0)
                     FROM locais l LEFT JOIN locais_cfg cfg ON cfg.local_id=l.id
                     WHERE COALESCE(l.ativo,1)=1
                     ORDER BY l.nome COLLATE NOCASE''')
        locais = [{'id': r[0], 'nome': r[1], 'fator_mult': float(r[2] or 1), 'pot_contratada': float(r[3] or 0), 'pot_instalada': float(r[4] or 0)} for r in c.fetchall()]
        c.execute('''SELECT e.id, e.nome, COALESCE(e.tag,''), COALESCE(e.local_id,0), COALESCE(l.nome,'')
                     FROM equipamentos e LEFT JOIN locais l ON l.id=e.local_id
                     ORDER BY e.nome COLLATE NOCASE''')
        equipamentos = [{'id': r[0], 'nome': r[1], 'tag': r[2], 'local_id': int(r[3] or 0), 'local_nome': r[4]} for r in c.fetchall()]
    finally:
        conn.close()

    if request.method == 'POST':
        datahora = request.form.get('datahora') or datetime.now().strftime('%Y-%m-%dT%H:%M')
        local = (request.form.get('local') or '').strip()
        equipamento = (request.form.get('equipamento') or '').strip()
        energia_ativa = _to_float(request.form.get('energia_ativa'))
        energia_reativa = _to_float(request.form.get('energia_reativa'))
        energia_aparente = _to_float(request.form.get('energia_aparente'))
        pot_ativa = _to_float(request.form.get('pot_ativa'))
        pot_reativa = _to_float(request.form.get('pot_reativa'))
        pot_aparente = _to_float(request.form.get('pot_aparente'))
        fp = _to_float(request.form.get('fp'))
        ponta = _to_float(request.form.get('ponta'))
        caudal_elevada = _to_float(request.form.get('caudal_elevada'))
        corrente = _to_float(request.form.get('corrente'))
        tensao = _to_float(request.form.get('tensao'))
        observacoes = (request.form.get('observacoes') or '').strip()

        if energia_aparente <= 0 and (energia_ativa > 0 or energia_reativa > 0):
            energia_aparente = _calc_aparente(energia_ativa, energia_reativa)
        if pot_aparente <= 0 and (pot_ativa > 0 or pot_reativa > 0):
            pot_aparente = _calc_aparente(pot_ativa, pot_reativa)
        if fp <= 0:
            fp = _calc_fp(energia_ativa, energia_reativa, pot_ativa, pot_reativa, energia_aparente, pot_aparente)

        avisos = []
        if fp and fp < 0.85:
            avisos.append(f'FP baixo: {fp:.3f}')
        if tensao and (tensao < 360 or tensao > 440):
            avisos.append(f'Tensão fora da faixa 360–440 V: {tensao:.1f} V')
        if corrente <= 0 and pot_ativa <= 0 and energia_ativa <= 0:
            avisos.append('Registo sem corrente, potência ou energia ativa informada')
        if avisos:
            observacoes = (observacoes + ' | ' if observacoes else '') + 'ALERTA OPERACIONAL: ' + '; '.join(avisos)

        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('''
            INSERT INTO leituras (datahora, local, equipamento, energia_ativa, energia_reativa, energia_aparente,
                                  pot_ativa, pot_reativa, pot_aparente, fp, ponta, caudal_elevada,
                                  corrente, tensao, observacoes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (datahora, local, equipamento, energia_ativa, energia_reativa, energia_aparente,
              pot_ativa, pot_reativa, pot_aparente, fp, ponta, caudal_elevada,
              corrente, tensao, observacoes))
        conn.commit(); conn.close()

        if request.form.get('continuar') == '1':
            return redirect(url_for('add', local=local, equipamento=equipamento))
        return redirect(url_for('leituras_list', local=local, inicio=datahora[:10], fim=datahora[:10]))

    now = datetime.now().strftime('%Y-%m-%dT%H:%M')
    selected_local = request.args.get('local','')
    selected_equipamento = request.args.get('equipamento','')
    return render_template('add.html', locais=locais, equipamentos=equipamentos, now=now, selected_local=selected_local, selected_equipamento=selected_equipamento)



# === LEITURAS DIÁRIAS — LISTAGEM/FILTER/CRUD/EXPORT/IMPORT/GRÁFICO ===



@app.route('/leituras', methods=['GET'])
@app.route('/monitoria', methods=['GET'])
def leituras_list():
    """Monitoria operacional diária: leituras pontuais por local/equipamento, KPIs, alertas e gráficos."""
    end = request.args.get('fim') or datetime.now().strftime('%Y-%m-%d')
    start = request.args.get('inicio') or (datetime.now() - timedelta(days=6)).strftime('%Y-%m-%d')
    local = request.args.get('local','').strip()
    equipamento = request.args.get('equipamento','').strip()
    q = request.args.get('q','').strip()

    try:
        page = int(request.args.get('page', 1))
        if page < 1: page = 1
    except Exception:
        page = 1
    try:
        per = int(request.args.get('per', 50))
        if per < 10: per = 10
        if per > 200: per = 200
    except Exception:
        per = 50
    offset = (page - 1) * per

    base_sql = " FROM leituras WHERE date(datahora) BETWEEN ? AND ?"
    params = [start, end]
    if local:
        base_sql += " AND local = ?"
        params.append(local)
    if equipamento:
        base_sql += " AND equipamento = ?"
        params.append(equipamento)
    if q:
        base_sql += " AND (equipamento LIKE ? OR observacoes LIKE ? OR local LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()

    total_rows = c.execute("SELECT COUNT(*)" + base_sql, params).fetchone()[0]
    tot = c.execute("""SELECT COALESCE(SUM(energia_ativa),0), COALESCE(AVG(pot_ativa),0),
                            COALESCE(MAX(ponta),0), COALESCE(AVG(fp),0), COALESCE(SUM(caudal_elevada),0),
                            COALESCE(AVG(tensao),0), COALESCE(MAX(corrente),0)
                     """ + base_sql, params).fetchone()
    total_ativa = float(tot[0] or 0)
    media_pot_ativa = float(tot[1] or 0)
    max_ponta = float(tot[2] or 0)
    fp_medio = float(tot[3] or 0)
    agua_total = float(tot[4] or 0)
    tensao_media = float(tot[5] or 0)
    corrente_max = float(tot[6] or 0)
    consumo_especifico = (total_ativa / agua_total) if agua_total > 0 else 0.0

    resumo = c.execute("""SELECT local, COALESCE(SUM(energia_ativa),0) kwh, COALESCE(AVG(pot_ativa),0) avgkw,
                               COALESCE(MAX(ponta),0) maxp, COALESCE(AVG(fp),0) avgfp, COALESCE(SUM(caudal_elevada),0) agua
                        """ + base_sql + " GROUP BY local ORDER BY kwh DESC", params).fetchall()

    resumo_equip = c.execute("""SELECT equipamento, local, COUNT(*) n, COALESCE(SUM(energia_ativa),0) kwh,
                                      COALESCE(AVG(pot_ativa),0) avgkw, COALESCE(AVG(fp),0) avgfp,
                                      COALESCE(MAX(corrente),0) imax
                               """ + base_sql + " GROUP BY equipamento, local ORDER BY kwh DESC LIMIT 15", params).fetchall()

    rows = c.execute("SELECT *" + base_sql + " ORDER BY datahora DESC LIMIT ? OFFSET ?", params + [per, offset]).fetchall()
    chart_rows = c.execute("""SELECT datahora, COALESCE(energia_ativa,0), COALESCE(pot_ativa,0), COALESCE(fp,0),
                                   COALESCE(ponta,0), COALESCE(caudal_elevada,0), COALESCE(corrente,0), COALESCE(tensao,0)
                            """ + base_sql + " ORDER BY datahora ASC LIMIT 500", params).fetchall()
    locais_opts = [r[0] for r in c.execute("SELECT DISTINCT nome FROM locais WHERE nome IS NOT NULL AND TRIM(nome)<>'' ORDER BY nome").fetchall()]
    # inclui locais que existam apenas nas leituras antigas
    for r in c.execute("SELECT DISTINCT local FROM leituras WHERE local IS NOT NULL AND TRIM(local)<>'' ORDER BY local").fetchall():
        if r[0] not in locais_opts:
            locais_opts.append(r[0])
    equipamentos_opts = [r[0] for r in c.execute("SELECT DISTINCT equipamento FROM leituras WHERE equipamento IS NOT NULL AND TRIM(equipamento)<>'' ORDER BY equipamento").fetchall()]
    local_config = None
    if local:
        hierarchy_lookup = {r['full_name']: r['nome'] for r in get_locais_hierarchy(include_inactive=True)}
        local_db_name = hierarchy_lookup.get(local, local)
        local_config = c.execute("""SELECT nome, COALESCE(potencia_contratada_kva,potencia_contratada,0),
                                         COALESCE(fator_multiplicativo,1), COALESCE(potencia_instalada_kw,0),
                                         COALESCE(estado_tecnico,''), COALESCE(prioridade,'')
                                  FROM locais WHERE nome=? LIMIT 1""", (local_db_name,)).fetchone()
    conn.close()

    # Anomalias simples e operacionais
    vals = [float(r[4] or 0) for r in rows]
    deltas = [abs(vals[i]-vals[i+1]) for i in range(len(vals)-1)] if len(vals)>1 else []
    med = sorted(deltas)[len(deltas)//2] if deltas else 0.0
    thr = max(3*(med or 0), 0)
    anomalias = set()
    for i in range(len(rows)-1):
        if deltas and deltas[i] > thr and thr>0:
            anomalias.add(rows[i][0]); anomalias.add(rows[i+1][0])
    low_fp_count = sum(1 for r in rows if (r[10] is not None and float(r[10] or 0) < 0.85 and float(r[10] or 0) > 0))
    tensao_alert_count = sum(1 for r in rows if (r[14] is not None and (float(r[14] or 0) < 360 or float(r[14] or 0) > 440)))
    corrente_alert_count = sum(1 for r in rows if (r[13] is not None and float(r[13] or 0) > 0 and float(r[13] or 0) == corrente_max and corrente_max > 0))

    grafico = {
        'horas': [r[0] for r in chart_rows],
        'ativa': [float(r[1] or 0) for r in chart_rows],
        'pot_ativa': [float(r[2] or 0) for r in chart_rows],
        'fp': [float(r[3] or 0) for r in chart_rows],
        'ponta': [float(r[4] or 0) for r in chart_rows],
        'agua': [float(r[5] or 0) for r in chart_rows],
        'corrente': [float(r[6] or 0) for r in chart_rows],
        'tensao': [float(r[7] or 0) for r in chart_rows],
    }

    total_pages = max(1, (total_rows + per - 1) // per)

    return render_template('leituras_list.html',
                           leituras=rows,
                           inicio=start, fim=end, local=local, equipamento=equipamento, q=q,
                           grafico=grafico, grafico_horas=grafico['horas'], grafico_ativa=grafico['ativa'],
                           page=page, total_pages=total_pages, per=per,
                           total_ativa=total_ativa, media_pot_ativa=media_pot_ativa, max_ponta=max_ponta,
                           fp_medio=fp_medio, agua_total=agua_total, consumo_especifico=consumo_especifico,
                           tensao_media=tensao_media, corrente_max=corrente_max,
                           low_fp_count=low_fp_count, tensao_alert_count=tensao_alert_count,
                           corrente_alert_count=corrente_alert_count,
                           anomalias=list(anomalias), resumo=resumo, resumo_equip=resumo_equip,
                           locais_opts=locais_opts, equipamentos_opts=equipamentos_opts,
                           local_config=local_config)



@app.route('/monitoria/controlo', methods=['GET'])
def monitoria_controlo():
    """Painel de controlo da monitoria operacional: criticidade, anomalias priorizadas e plano de ação."""
    end = request.args.get('fim') or datetime.now().strftime('%Y-%m-%d')
    start = request.args.get('inicio') or (datetime.now() - timedelta(days=6)).strftime('%Y-%m-%d')
    local = request.args.get('local','').strip()
    equipamento = request.args.get('equipamento','').strip()
    q = request.args.get('q','').strip()

    base_sql = " FROM leituras WHERE date(datahora) BETWEEN ? AND ?"
    params = [start, end]
    if local:
        base_sql += " AND local = ?"
        params.append(local)
    if equipamento:
        base_sql += " AND equipamento = ?"
        params.append(equipamento)
    if q:
        base_sql += " AND (equipamento LIKE ? OR observacoes LIKE ? OR local LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute("""SELECT id, datahora, local, equipamento, COALESCE(energia_ativa,0), COALESCE(energia_reativa,0),
                             COALESCE(energia_aparente,0), COALESCE(pot_ativa,0), COALESCE(pot_reativa,0),
                             COALESCE(pot_aparente,0), COALESCE(fp,0), COALESCE(ponta,0), COALESCE(caudal_elevada,0),
                             COALESCE(corrente,0), COALESCE(tensao,0), COALESCE(observacoes,'')
                      """ + base_sql + " ORDER BY datahora DESC LIMIT 2000", params).fetchall()
    locais_opts = [r[0] for r in c.execute("SELECT DISTINCT nome FROM locais WHERE nome IS NOT NULL AND TRIM(nome)<>'' ORDER BY nome").fetchall()]
    for r in c.execute("SELECT DISTINCT local FROM leituras WHERE local IS NOT NULL AND TRIM(local)<>'' ORDER BY local").fetchall():
        if r[0] not in locais_opts:
            locais_opts.append(r[0])
    equipamentos_opts = [r[0] for r in c.execute("SELECT DISTINCT equipamento FROM leituras WHERE equipamento IS NOT NULL AND TRIM(equipamento)<>'' ORDER BY equipamento").fetchall()]
    conn.close()

    total_registos = len(rows)
    total_kwh = sum(float(r[4] or 0) for r in rows)
    agua_total = sum(float(r[12] or 0) for r in rows)
    fp_vals = [float(r[10] or 0) for r in rows if float(r[10] or 0) > 0]
    tensao_vals = [float(r[14] or 0) for r in rows if float(r[14] or 0) > 0]
    corrente_vals = [float(r[13] or 0) for r in rows if float(r[13] or 0) > 0]
    fp_medio = sum(fp_vals)/len(fp_vals) if fp_vals else 0.0
    tensao_media = sum(tensao_vals)/len(tensao_vals) if tensao_vals else 0.0
    corrente_media = sum(corrente_vals)/len(corrente_vals) if corrente_vals else 0.0
    corrente_max = max(corrente_vals) if corrente_vals else 0.0
    max_ponta = max([float(r[11] or 0) for r in rows] or [0])
    consumo_especifico = (total_kwh / agua_total) if agua_total > 0 else 0.0

    # médias por equipamento para detectar desvios relativos
    eq_stats = {}
    for r in rows:
        key = (r[3] or 'Sem equipamento', r[2] or 'Sem local')
        st = eq_stats.setdefault(key, {'n':0,'kwh':0.0,'kw':0.0,'fp_sum':0.0,'fp_n':0,'corr_sum':0.0,'corr_n':0,'corr_max':0.0,'ponta':0.0,'agua':0.0,'alertas':0,'criticos':0})
        st['n'] += 1
        st['kwh'] += float(r[4] or 0)
        st['kw'] += float(r[7] or 0)
        fp = float(r[10] or 0); corr = float(r[13] or 0)
        if fp > 0: st['fp_sum'] += fp; st['fp_n'] += 1
        if corr > 0: st['corr_sum'] += corr; st['corr_n'] += 1; st['corr_max'] = max(st['corr_max'], corr)
        st['ponta'] = max(st['ponta'], float(r[11] or 0))
        st['agua'] += float(r[12] or 0)

    alertas = []
    def add_alerta(nivel, tipo, r, valor, impacto, acao):
        peso = 3 if nivel == 'Crítico' else 2 if nivel == 'Atenção' else 1
        alertas.append({'nivel':nivel, 'tipo':tipo, 'datahora':r[1], 'local':r[2] or '—', 'equipamento':r[3] or '—', 'valor':valor, 'impacto':impacto, 'acao':acao, 'peso':peso})
        key = (r[3] or 'Sem equipamento', r[2] or 'Sem local')
        if key in eq_stats:
            eq_stats[key]['alertas'] += 1
            if nivel == 'Crítico': eq_stats[key]['criticos'] += 1

    # Consumo específico por linha e referência do período
    ce_vals = []
    for r in rows:
        agua = float(r[12] or 0); kwh = float(r[4] or 0)
        if agua > 0 and kwh > 0:
            ce_vals.append(kwh/agua)
    ce_ref = (sum(ce_vals)/len(ce_vals)) if ce_vals else 0.0
    ponta_vals = [float(r[11] or 0) for r in rows if float(r[11] or 0)>0]
    ponta_ref = (sum(ponta_vals)/len(ponta_vals)) if ponta_vals else 0.0

    for r in rows:
        fp = float(r[10] or 0); tensao = float(r[14] or 0); corrente = float(r[13] or 0)
        ponta = float(r[11] or 0); kwh = float(r[4] or 0); agua = float(r[12] or 0)
        if 0 < fp < 0.75:
            add_alerta('Crítico','Fator de potência muito baixo',r,f'{fp:.3f}','Maior probabilidade de energia reativa excedente e perdas.','Inspecionar cargas, banco de capacitores, regime de operação e compensação reativa.')
        elif 0.75 <= fp < 0.85:
            add_alerta('Atenção','Fator de potência baixo',r,f'{fp:.3f}','Risco operacional e potencial penalização por reativa.','Acompanhar repetição do evento e avaliar necessidade de correção do FP.')
        if tensao > 0 and (tensao < 360 or tensao > 440):
            nivel = 'Crítico' if tensao < 340 or tensao > 460 else 'Atenção'
            add_alerta(nivel,'Tensão fora da faixa',r,f'{tensao:.1f} V','Pode causar aquecimento, disparos, baixo rendimento ou falha de equipamento.','Confirmar medição por fase, quedas de tensão, ligações e estado do transformador/quadro.')
        if corrente_media > 0 and corrente > 1.35 * corrente_media:
            add_alerta('Atenção','Corrente acima do padrão do período',r,f'{corrente:.1f} A','Possível sobrecarga, desequilíbrio ou alteração de regime.','Comparar com corrente nominal do equipamento e verificar carga mecânica/elétrica.')
        if ponta_ref > 0 and ponta > 1.25 * ponta_ref:
            add_alerta('Atenção','Ponta acima do padrão',r,f'{ponta:.2f} kW','Pode aumentar a procura máxima operacional e afetar custo mensal.','Verificar arranque simultâneo de cargas e possibilidade de escalonamento operacional.')
        if agua > 0 and ce_ref > 0 and (kwh/agua) > 1.30 * ce_ref:
            add_alerta('Atenção','Consumo específico acima do padrão',r,f'{(kwh/agua):.3f} kWh/m³','Indica possível perda de eficiência, queda de caudal ou operação fora do ponto ótimo.','Comparar caudal, pressão, válvulas, filtros e rendimento do conjunto motor-bomba.')
        if not (r[2] or '').strip() or not (r[3] or '').strip():
            add_alerta('Informativo','Registo incompleto',r,'Local/equipamento em falta','Reduz a qualidade da análise histórica.','Completar local e equipamento para melhorar rastreabilidade.')

    alertas.sort(key=lambda a: (-a['peso'], a['datahora'] or ''))
    criticos = sum(1 for a in alertas if a['nivel'] == 'Crítico')
    atencao = sum(1 for a in alertas if a['nivel'] == 'Atenção')
    informativos = sum(1 for a in alertas if a['nivel'] == 'Informativo')
    if criticos:
        estado = 'Crítico'; estado_classe = 'danger'; resumo_estado = 'Existem ocorrências críticas que exigem validação técnica antes de concluir o período.'
    elif atencao:
        estado = 'Atenção'; estado_classe = 'warn'; resumo_estado = 'Existem desvios operacionais que devem ser acompanhados e corrigidos.'
    elif total_registos:
        estado = 'Normal'; estado_classe = 'ok'; resumo_estado = 'Sem anomalias relevantes detectadas no intervalo filtrado.'
    else:
        estado = 'Sem dados'; estado_classe = 'info'; resumo_estado = 'Não existem leituras para o intervalo selecionado.'

    ranking = []
    for (eq, loc), st in eq_stats.items():
        avg_fp = st['fp_sum']/st['fp_n'] if st['fp_n'] else 0.0
        avg_kw = st['kw']/st['n'] if st['n'] else 0.0
        ce = st['kwh']/st['agua'] if st['agua'] > 0 else 0.0
        score = st['criticos']*3 + (st['alertas']-st['criticos'])*1
        ranking.append({'equipamento':eq, 'local':loc, 'n':st['n'], 'kwh':st['kwh'], 'avg_kw':avg_kw, 'avg_fp':avg_fp, 'imax':st['corr_max'], 'ponta':st['ponta'], 'agua':st['agua'], 'ce':ce, 'alertas':st['alertas'], 'criticos':st['criticos'], 'score':score})
    ranking.sort(key=lambda x: (-x['score'], -x['kwh']))

    plano = []
    if criticos:
        plano.append('Validar imediatamente as leituras críticas antes de usar estes dados em relatórios de desempenho.')
    if any(a['tipo'].startswith('Fator') for a in alertas):
        plano.append('Criar rotina de verificação do fator de potência por equipamento/local e cruzar com a fatura mensal de reativa excedente.')
    if any(a['tipo'].startswith('Tensão') for a in alertas):
        plano.append('Confirmar tensões por fase e avaliar quedas de tensão, transformador, cabos, barramentos e ligações.')
    if any(a['tipo'].startswith('Corrente') for a in alertas):
        plano.append('Comparar correntes medidas com corrente nominal e histórico do equipamento para identificar sobrecarga ou desequilíbrio.')
    if any('Consumo específico' in a['tipo'] for a in alertas):
        plano.append('Investigar causas hidráulicas/operacionais para consumo específico elevado, incluindo caudal, pressão, válvulas e ponto de operação da bomba.')
    if not plano:
        plano.append('Manter a monitoria diária e comparar tendências por local/equipamento semanalmente.')

    return render_template('monitoria_controle.html',
                           inicio=start, fim=end, local=local, equipamento=equipamento, q=q,
                           locais_opts=locais_opts, equipamentos_opts=equipamentos_opts,
                           total_registos=total_registos, total_kwh=total_kwh, agua_total=agua_total,
                           fp_medio=fp_medio, tensao_media=tensao_media, corrente_media=corrente_media,
                           corrente_max=corrente_max, max_ponta=max_ponta, consumo_especifico=consumo_especifico,
                           estado=estado, estado_classe=estado_classe, resumo_estado=resumo_estado,
                           criticos=criticos, atencao=atencao, informativos=informativos,
                           alertas=alertas[:80], alertas_total=len(alertas), ranking=ranking[:25], plano=plano)

@app.route('/monitoria/relatorio', methods=['GET'])
def monitoria_relatorio():
    """Relatório técnico-operacional da monitoria diária para impressão/análise."""
    end = request.args.get('fim') or datetime.now().strftime('%Y-%m-%d')
    start = request.args.get('inicio') or (datetime.now() - timedelta(days=6)).strftime('%Y-%m-%d')
    local = request.args.get('local','').strip()
    equipamento = request.args.get('equipamento','').strip()
    q = request.args.get('q','').strip()

    base_sql = " FROM leituras WHERE date(datahora) BETWEEN ? AND ?"
    params = [start, end]
    if local:
        base_sql += " AND local = ?"
        params.append(local)
    if equipamento:
        base_sql += " AND equipamento = ?"
        params.append(equipamento)
    if q:
        base_sql += " AND (equipamento LIKE ? OR observacoes LIKE ? OR local LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute("""SELECT id, datahora, local, equipamento, COALESCE(energia_ativa,0), COALESCE(energia_reativa,0),
                             COALESCE(pot_ativa,0), COALESCE(pot_reativa,0), COALESCE(fp,0), COALESCE(ponta,0),
                             COALESCE(caudal_elevada,0), COALESCE(corrente,0), COALESCE(tensao,0), COALESCE(observacoes,'')
                      """ + base_sql + " ORDER BY datahora ASC", params).fetchall()
    resumo_equip = c.execute("""SELECT COALESCE(equipamento,'Sem equipamento') equipamento, COALESCE(local,'Sem local') local, COUNT(*) n,
                                      COALESCE(SUM(energia_ativa),0) kwh, COALESCE(AVG(pot_ativa),0) avgkw,
                                      COALESCE(MAX(ponta),0) maxponta, COALESCE(AVG(fp),0) avgfp,
                                      COALESCE(MAX(corrente),0) imax, COALESCE(AVG(tensao),0) vmed,
                                      COALESCE(SUM(caudal_elevada),0) agua
                               """ + base_sql + " GROUP BY equipamento, local ORDER BY kwh DESC LIMIT 20", params).fetchall()
    resumo_local = c.execute("""SELECT COALESCE(local,'Sem local') local, COUNT(*) n, COALESCE(SUM(energia_ativa),0) kwh,
                                     COALESCE(AVG(pot_ativa),0) avgkw, COALESCE(MAX(ponta),0) maxponta,
                                     COALESCE(AVG(fp),0) avgfp, COALESCE(SUM(caudal_elevada),0) agua
                              """ + base_sql + " GROUP BY local ORDER BY kwh DESC", params).fetchall()
    local_config = None
    if local:
        local_config = c.execute("""SELECT nome, COALESCE(potencia_contratada_kva,potencia_contratada,0),
                                         COALESCE(fator_multiplicativo,1), COALESCE(potencia_instalada_kw,0)
                                  FROM locais WHERE nome=? LIMIT 1""", (local,)).fetchone()
    conn.close()

    total_registos = len(rows)
    total_kwh = sum(float(r[4] or 0) for r in rows)
    total_kvarh = sum(float(r[5] or 0) for r in rows)
    agua_total = sum(float(r[10] or 0) for r in rows)
    avg_kw = (sum(float(r[6] or 0) for r in rows) / total_registos) if total_registos else 0.0
    avg_fp = (sum(float(r[8] or 0) for r in rows if float(r[8] or 0) > 0) / max(1, sum(1 for r in rows if float(r[8] or 0) > 0))) if rows else 0.0
    max_ponta = max([float(r[9] or 0) for r in rows] or [0])
    tensao_media = (sum(float(r[12] or 0) for r in rows if float(r[12] or 0) > 0) / max(1, sum(1 for r in rows if float(r[12] or 0) > 0))) if rows else 0.0
    consumo_especifico = (total_kwh / agua_total) if agua_total > 0 else 0.0

    alertas = []
    for r in rows:
        fp = float(r[8] or 0); tensao = float(r[12] or 0); corrente = float(r[11] or 0)
        if 0 < fp < 0.85:
            alertas.append({'nivel':'Crítico', 'tipo':'FP baixo', 'datahora':r[1], 'local':r[2], 'equipamento':r[3], 'valor':f'{fp:.3f}', 'acao':'Verificar compensação reativa, carga parcial ou banco de capacitores.'})
        if tensao > 0 and (tensao < 360 or tensao > 440):
            alertas.append({'nivel':'Atenção', 'tipo':'Tensão fora da faixa', 'datahora':r[1], 'local':r[2], 'equipamento':r[3], 'valor':f'{tensao:.1f} V', 'acao':'Confirmar tensão por fase, estado das ligações e quedas de tensão.'})
        if corrente > 0 and avg_kw > 0 and corrente > 1.35 * max(1, sum(float(x[11] or 0) for x in rows)/max(1,total_registos)):
            alertas.append({'nivel':'Atenção', 'tipo':'Corrente elevada', 'datahora':r[1], 'local':r[2], 'equipamento':r[3], 'valor':f'{corrente:.1f} A', 'acao':'Comparar com corrente nominal e avaliar sobrecarga ou desequilíbrio.'})

    recomenda = []
    if avg_fp and avg_fp < 0.85:
        recomenda.append('Priorizar análise de fator de potência e compensação reativa nos equipamentos/períodos com FP baixo.')
    if tensao_media and (tensao_media < 380 or tensao_media > 420):
        recomenda.append('Validar tensão média operacional e investigar queda/elevação de tensão na alimentação.')
    if consumo_especifico > 0:
        recomenda.append('Acompanhar consumo específico kWh/m³ por equipamento e comparar contra o melhor dia/período operacional.')
    if not recomenda:
        recomenda.append('Manter a monitoria diária e comparar tendências por equipamento para detectar desvios antecipadamente.')

    return render_template('monitoria_relatorio.html',
                           inicio=start, fim=end, local=local, equipamento=equipamento, q=q,
                           rows=rows, resumo_equip=resumo_equip, resumo_local=resumo_local, local_config=local_config,
                           total_registos=total_registos, total_kwh=total_kwh, total_kvarh=total_kvarh,
                           agua_total=agua_total, avg_kw=avg_kw, avg_fp=avg_fp, max_ponta=max_ponta,
                           tensao_media=tensao_media, consumo_especifico=consumo_especifico,
                           alertas=alertas[:50], alertas_total=len(alertas), recomenda=recomenda)

@app.route('/leituras/<int:lid>/duplicate', methods=['POST','GET'])
def leituras_duplicate(lid):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    row = c.execute("SELECT datahora, local, equipamento, energia_ativa, energia_reativa, energia_aparente, pot_ativa, pot_reativa, pot_aparente, fp, ponta, caudal_elevada, corrente, tensao, observacoes FROM leituras WHERE id=?", (lid,)).fetchone()
    if not row:
        conn.close()
        flash("Registo não encontrado.", "warning")
        return redirect(url_for('leituras_list'))
    c.execute("""
        INSERT INTO leituras
        (datahora, local, equipamento, energia_ativa, energia_reativa, energia_aparente,
         pot_ativa, pot_reativa, pot_aparente, fp, ponta, caudal_elevada, corrente, tensao, observacoes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, row)
    conn.commit(); conn.close()
    flash("Leitura duplicada.", "success")
    return redirect(url_for('leituras_list'))

@app.route('/leituras/delete_batch', methods=['POST'])
def leituras_delete_batch():
    ids = request.form.getlist('ids')
    if not ids:
        flash("Nenhuma linha selecionada.", "warning")
        return redirect(url_for('leituras_list'))
    # filtra apenas dígitos
    ids_clean = [i for i in ids if str(i).isdigit()]
    if not ids_clean:
        flash("Seleção inválida.", "warning")
        return redirect(url_for('leituras_list'))
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    qmarks = ",".join(["?"]*len(ids_clean))
    c.execute(f"DELETE FROM leituras WHERE id IN ({qmarks})", ids_clean)
    conn.commit(); conn.close()
    flash(f"{len(ids_clean)} leitura(s) eliminada(s).", "success")
    return redirect(url_for('leituras_list'))
@app.route('/leituras/export', methods=['GET'])
def leituras_export_csv():
    """Exporta CSV de leituras filtradas."""
    end = request.args.get('fim') or datetime.now().strftime('%Y-%m-%d')
    start = request.args.get('inicio') or (datetime.now() - timedelta(days=6)).strftime('%Y-%m-%d')
    local = request.args.get('local','')
    q = request.args.get('q','').strip()

    sql = "SELECT * FROM leituras WHERE date(datahora) BETWEEN ? AND ?"
    params = [start, end]
    if local:
        sql += " AND local = ?"
        params.append(local)
    if q:
        sql += " AND (equipamento LIKE ? OR observacoes LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%"])
    sql += " ORDER BY datahora"

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute(sql, params).fetchall()
    conn.close()

    from io import StringIO
    si = StringIO()
    w = csv.writer(si, delimiter=';')
    header = [
        "id","datahora","local","equipamento","energia_ativa","energia_reativa","energia_aparente",
        "pot_ativa","pot_reativa","pot_aparente","fp","ponta","caudal_elevada","corrente","tensao","observacoes"
    ]
    w.writerow(header)
    for r in rows:
        w.writerow(r)
    return Response(si.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename=leituras_{start}_a_{end}.csv'})

@app.route('/leituras/import', methods=['GET','POST'])
def leituras_import():
    locais = [l[1] for l in get_locais()]
    if request.method == 'POST':
        f = request.files.get('arquivo')
        if not f or f.filename == '':
            flash('Selecione um ficheiro CSV.','warning')
            return redirect(url_for('leituras_import'))
        # parse CSV (cabecalhos como no export)
        import csv
        from io import TextIOWrapper
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        inseridos = 0
        with TextIOWrapper(f.stream, encoding='utf-8', errors='ignore') as fh:
            reader = csv.DictReader(fh, delimiter=';')
            for row in reader:
                try:
                    c.execute('''INSERT INTO leituras
                         (datahora, local, equipamento, energia_ativa, energia_reativa, energia_aparente,
                          pot_ativa, pot_reativa, pot_aparente, fp, ponta, caudal_elevada, corrente, tensao, observacoes)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                         (row.get('datahora'), row.get('local'), row.get('equipamento'),
                          _to_float(row.get('energia_ativa')), _to_float(row.get('energia_reativa')), _to_float(row.get('energia_aparente')),
                          _to_float(row.get('pot_ativa')), _to_float(row.get('pot_reativa')), _to_float(row.get('pot_aparente')),
                          _to_float(row.get('fp')), _to_float(row.get('ponta')), _to_float(row.get('caudal_elevada')),
                          _to_float(row.get('corrente')), _to_float(row.get('tensao')), row.get('observacoes')))
                    inseridos += 1
                except Exception:
                    pass
        conn.commit(); conn.close()
        flash(f'Importação concluída: {inseridos} registos inseridos.','success')
        return redirect(url_for('leituras_list'))
    return render_template('leituras_import.html', locais=locais)

@app.route('/leituras/<int:lid>/edit', methods=['GET','POST'])
def leituras_edit(lid):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    if request.method == 'POST':
        datahora = request.form.get('datahora')
        local = request.form.get('local')
        equipamento = request.form.get('equipamento')
        energia_ativa = _to_float(request.form.get('energia_ativa'))
        energia_reativa = _to_float(request.form.get('energia_reativa'))
        energia_aparente = _to_float(request.form.get('energia_aparente'))
        pot_ativa = _to_float(request.form.get('pot_ativa'))
        pot_reativa = _to_float(request.form.get('pot_reativa'))
        pot_aparente = _to_float(request.form.get('pot_aparente'))
        fp = _to_float(request.form.get('fp'))
        ponta = _to_float(request.form.get('ponta'))
        caudal_elevada = _to_float(request.form.get('caudal_elevada'))
        corrente = _to_float(request.form.get('corrente'))
        tensao = _to_float(request.form.get('tensao'))
        obs = request.form.get('observacoes','')
        c.execute('''UPDATE leituras SET datahora=?, local=?, equipamento=?, energia_ativa=?, energia_reativa=?, energia_aparente=?,
                     pot_ativa=?, pot_reativa=?, pot_aparente=?, fp=?, ponta=?, caudal_elevada=?, corrente=?, tensao=?, observacoes=?
                     WHERE id=?''',
                  (datahora, local, equipamento, energia_ativa, energia_reativa, energia_aparente,
                   pot_ativa, pot_reativa, pot_aparente, fp, ponta, caudal_elevada, corrente, tensao, obs, lid))
        conn.commit(); conn.close()
        flash('Leitura atualizada.','success')
        return redirect(url_for('leituras_list'))
    row = c.execute('SELECT * FROM leituras WHERE id=?', (lid,)).fetchone()
    conn.close()
    if not row:
        flash('Registo não encontrado.','warning')
        return redirect(url_for('leituras_list'))
    locais = [l[1] for l in get_locais()]
    return render_template('leituras_edit.html', row=row, locais=locais)

@app.route('/leituras/<int:lid>/delete', methods=['POST'])
def leituras_delete(lid):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('DELETE FROM leituras WHERE id=?', (lid,))
    conn.commit(); conn.close()
    flash('Leitura eliminada.','success')
    return redirect(url_for('leituras_list'))

@app.route('/leituras/visualizar', methods=['GET'])
def visualizar_diario_view():
    """Visualiza um dia (ou intervalo curto) com gráfico."""
    locais = [l[1] for l in get_locais()]
    data = request.args.get('data') or datetime.now().strftime('%Y-%m-%d')
    local = request.args.get('local','')
    sql = "SELECT * FROM leituras WHERE date(datahora)=?"; params=[data]
    if local:
        sql += " AND local=?"; params.append(local)
    sql += " ORDER BY datahora"
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    leit = c.execute(sql, params).fetchall()
    conn.close()
    horas = [r[1] for r in leit]
    ativa = [float(r[4] or 0) for r in leit]
    return render_template('visualizar_diario.html', data=data, locais=locais, local=local,
                           leituras=leit, grafico_horas=horas, grafico_ativa=ativa)

# === Export de Leituras Mensais (referenciado no template) ===
@app.route('/leituras_mensal/export')
def leituras_mensal_export():
    local = request.args.get('local','')
    mes = request.args.get('mes') or datetime.now().strftime('%m')
    ano = int(request.args.get('ano') or datetime.now().year)
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute('''SELECT local,data,hora,ativa,reativa,ponta,fp,potc,anterior,atual,diferenca,agua,esp,acum,valor
                        FROM leituras_mensais WHERE local=? AND mes=? AND ano=? ORDER BY data''',
                     (local, mes, ano)).fetchall()
    conn.close()
    si = io.StringIO(); w = csv.writer(si, delimiter=';')
    w.writerow(['local','data','hora','ativa','reativa','ponta','fp','potc','anterior','atual','diferenca','agua','esp','acum','valor'])
    for r in rows:
        w.writerow(r)
    return Response(si.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename=leituras_mensais_{local}_{mes}-{ano}.csv'})

# === LEITURA MENSAL ===


# === LEITURAS MENSAIS · FASE 2 OPERAÇÃO REAL ===

def ensure_leituras_mensais_phase2_schema():
    """Garante pequenos índices de performance sem alterar dados existentes.
    Mantém compatibilidade total com o sge.db atual e evita leituras lentas
    em meses com muitos registos.
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_mensais_local_mes_ano_data ON leituras_mensais(local, mes, ano, data)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_leituras_mensais_local_data ON leituras_mensais(local, data)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_locais_cfg_local_id ON locais_cfg(local_id)")
        conn.commit()
        conn.close()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
    return True


def _local_id_nome_from_request(raw_local, locais_db):
    raw = (str(raw_local or '').strip())
    if not raw and locais_db:
        raw = str(locais_db[0][0])
    selected_id = None
    local_nome = raw
    if raw.isdigit():
        selected_id = int(raw)
        nm = get_local_by_id(selected_id)
        local_nome = nm[1] if nm else raw
    else:
        for lid, lname in locais_db:
            if lname == raw:
                selected_id = lid
                local_nome = lname
                break
    return selected_id, local_nome


def _cfg_map_locais(locais_db):
    """Mapa de configurações dos locais em consulta única.
    Evita abrir uma ligação SQLite para cada local, o que deixava a página
    de leituras mensais lenta quando havia muitos locais cadastrados.
    """
    mapa = {}
    if not locais_db:
        return mapa
    nomes_por_id = {int(lid): lname for lid, lname in locais_db}
    ids = list(nomes_por_id.keys())
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        qmarks = ','.join(['?'] * len(ids))
        c.execute(f'''
            SELECT local_id, fator_mult, pot_contratada, tarifa_ativa, tarifa_reativa, tarifa_ponta,
                   tarifa_perdas, taxa_fixa, taxa_radio, taxa_lixo, iva, COALESCE(pot_instalada, 0.0)
            FROM locais_cfg
            WHERE local_id IN ({qmarks})
        ''', ids)
        rows = c.fetchall()
        conn.close()
        for lid, fator, potc, ta, tr, tp, perdas, fixa, radio, lixo, iva, poti in rows:
            lname = nomes_por_id.get(int(lid))
            if not lname:
                continue
            mapa[lname] = {
                'fator_mult': fator, 'pot_contratada': potc, 'tarifa_ativa': ta,
                'tarifa_reativa': tr, 'tarifa_ponta': tp, 'tarifa_perdas': perdas,
                'taxa_fixa': fixa, 'taxa_radio': radio, 'taxa_lixo': lixo,
                'iva': iva, 'pot_instalada': poti
            }
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
    # Garante defaults sem chamadas adicionais ao banco
    for lid, lname in locais_db:
        mapa.setdefault(lname, {
            'fator_mult': 1.0, 'pot_contratada': 0.0, 'tarifa_ativa': 0.0,
            'tarifa_reativa': 0.0, 'tarifa_ponta': 0.0, 'tarifa_perdas': 0.0,
            'taxa_fixa': 0.0, 'taxa_radio': 0.0, 'taxa_lixo': 0.0,
            'iva': 16.0, 'pot_instalada': 0.0
        })
    return mapa


def _safe_float(v, default=None):
    if v is None or v == '':
        return default
    try:
        return float(str(v).replace(',', '.'))
    except Exception:
        return default




def _fmt_mil(v, casas=2):
    """Formata números com separador de milhares no padrão PT: 1.234.567,89."""
    try:
        n = float(v or 0)
    except Exception:
        n = 0.0
    txt = f"{n:,.{int(casas)}f}"
    return txt.replace(',', 'X').replace('.', ',').replace('X', '.')


def _mzn_extenso(valor):
    """Valor monetário por extenso em português: meticais e centavos."""
    try:
        total_cent = int(round(float(valor or 0) * 100))
    except Exception:
        total_cent = 0
    meticais = total_cent // 100
    centavos = total_cent % 100

    unidades = ['', 'um', 'dois', 'três', 'quatro', 'cinco', 'seis', 'sete', 'oito', 'nove']
    especiais = {
        10:'dez', 11:'onze', 12:'doze', 13:'treze', 14:'catorze', 15:'quinze', 16:'dezasseis',
        17:'dezassete', 18:'dezoito', 19:'dezanove'
    }
    dezenas = ['', '', 'vinte', 'trinta', 'quarenta', 'cinquenta', 'sessenta', 'setenta', 'oitenta', 'noventa']
    centenas = ['', 'cento', 'duzentos', 'trezentos', 'quatrocentos', 'quinhentos', 'seiscentos', 'setecentos', 'oitocentos', 'novecentos']

    def ate_999(n):
        n = int(n)
        if n == 0:
            return ''
        if n == 100:
            return 'cem'
        parts = []
        c, r = divmod(n, 100)
        if c:
            parts.append(centenas[c])
        if r:
            if r < 10:
                parts.append(unidades[r])
            elif r < 20:
                parts.append(especiais[r])
            else:
                d, u = divmod(r, 10)
                if u:
                    parts.append(dezenas[d] + ' e ' + unidades[u])
                else:
                    parts.append(dezenas[d])
        return ' e '.join([x for x in parts if x])

    def inteiro_ext(n):
        n = int(n)
        if n == 0:
            return 'zero'
        grupos = []
        escala = [('', ''), ('mil', 'mil'), ('milhão', 'milhões'), ('mil milhões', 'mil milhões'), ('bilião', 'biliões')]
        i = 0
        while n > 0:
            grupos.append(n % 1000)
            n //= 1000
            i += 1
        partes = []
        for idx in range(len(grupos)-1, -1, -1):
            g = grupos[idx]
            if not g:
                continue
            if idx == 1 and g == 1:
                partes.append('mil')
            elif idx > 0:
                sing, plur = escala[idx] if idx < len(escala) else ('', '')
                partes.append(ate_999(g) + ' ' + (sing if g == 1 else plur))
            else:
                partes.append(ate_999(g))
        return ' e '.join(partes)

    texto = inteiro_ext(meticais) + (' metical' if meticais == 1 else ' meticais')
    if centavos:
        texto += ' e ' + inteiro_ext(centavos) + (' centavo' if centavos == 1 else ' centavos')
    return texto.capitalize()


try:
    app.jinja_env.filters['fmt_mil'] = _fmt_mil
    app.jinja_env.filters['mzn_extenso'] = _mzn_extenso
except Exception:
    pass


def _ponta_faturavel_edm(pot_contratada, ponta_lida_corrigida):
    """
    Regra operacional da ponta faturável EDM MT usada no SGE:
    20% da potência contratada + 80% da ponta lida corrigida pelo fator multiplicativo.
    A potência contratada vem diretamente da configuração do Local e NÃO leva fator multiplicativo.
    A ponta lida já deve chegar aqui corrigida pelo fator multiplicativo.
    """
    pc = _safe_float(pot_contratada, 0.0) or 0.0
    pl = _safe_float(ponta_lida_corrigida, 0.0) or 0.0
    return (0.20 * pc) + (0.80 * pl)



def _quantidades_fatura_mensal(local_nome: str, mes_str: str, ano_int: int):
    """
    Calcula as quantidades físicas de faturação a partir da planilha mensal.

    Regras aplicadas:
    - Leituras de ativa e reativa gravadas em leituras_mensais já estão corrigidas pelo fator multiplicativo.
    - Energia ativa do mês = última leitura ativa válida - leitura base inicial.
    - Energia reativa do mês = última leitura reativa válida - leitura base inicial.
    - Se existir leitura do mês anterior, ela é a base inicial.
    - Se não existir leitura do mês anterior, a primeira leitura válida do mês é apenas referência/base.
    - Ponta considerada = maior ponta corrigida registada no mês.
    - Diferenças negativas não são faturadas automaticamente; ficam sinalizadas para auditoria.
    """
    ensure_leituras_mensais_phase2_schema()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute("""
        SELECT data, ativa, reativa, ponta, agua, anterior
        FROM leituras_mensais
        WHERE local=? AND mes=? AND ano=?
          AND (ativa IS NOT NULL OR reativa IS NOT NULL OR ponta IS NOT NULL OR agua IS NOT NULL)
        ORDER BY data
    """, (local_nome, str(mes_str).zfill(2), int(ano_int))).fetchall()
    conn.close()

    prev_ativa, prev_reativa = get_prev_month_last_readings(local_nome, str(mes_str).zfill(2), int(ano_int))

    ativas = []
    reativas = []
    ponta_max = 0.0
    agua_total = 0.0
    avisos = []

    for r in rows:
        data = r['data']
        ativa = _safe_float(r['ativa'], None)
        reativa = _safe_float(r['reativa'], None)
        ponta = _safe_float(r['ponta'], None)
        agua = _safe_float(r['agua'], 0.0) or 0.0
        if ativa is not None and ativa > 0:
            ativas.append((data, ativa))
        if reativa is not None and reativa > 0:
            reativas.append((data, reativa))
        if ponta is not None and ponta > 0:
            ponta_max = max(ponta_max, ponta)
        agua_total += agua

    leitura_base_ativa = float(prev_ativa or 0.0) if prev_ativa and prev_ativa > 0 else (ativas[0][1] if ativas else 0.0)
    leitura_final_ativa = ativas[-1][1] if ativas else leitura_base_ativa
    kwh_ativa = leitura_final_ativa - leitura_base_ativa
    if kwh_ativa < 0:
        avisos.append('Leitura ativa final inferior à leitura base; consumo ativo faturável foi tratado como zero. Verificar leitura anterior, troca/reinício de contador ou fator multiplicativo.')
        kwh_ativa = 0.0

    leitura_base_reativa = float(prev_reativa or 0.0) if prev_reativa and prev_reativa > 0 else (reativas[0][1] if reativas else 0.0)
    leitura_final_reativa = reativas[-1][1] if reativas else leitura_base_reativa
    kvarh_reativa = leitura_final_reativa - leitura_base_reativa
    if kvarh_reativa < 0:
        avisos.append('Leitura reativa final inferior à leitura base; consumo reativo faturável foi tratado como zero. Verificar leitura anterior, troca/reinício de contador ou fator multiplicativo.')
        kvarh_reativa = 0.0

    limite_reativa = 0.75 * kwh_ativa
    kvarh_excedente = max(kvarh_reativa - limite_reativa, 0.0)
    consumo_especifico = (kwh_ativa / agua_total) if agua_total > 0 else None

    return {
        'rows': rows,
        'kwh_ativa': kwh_ativa,
        'kvarh_reativa': kvarh_reativa,
        'limite_reativa': limite_reativa,
        'kvarh_excedente': kvarh_excedente,
        'kw_ponta_lida': ponta_max,
        'agua_total': agua_total,
        'consumo_especifico': consumo_especifico,
        'leitura_base_ativa': leitura_base_ativa,
        'leitura_final_ativa': leitura_final_ativa,
        'leitura_base_reativa': leitura_base_reativa,
        'leitura_final_reativa': leitura_final_reativa,
        'tem_base_mes_anterior_ativa': bool(prev_ativa and prev_ativa > 0),
        'tem_base_mes_anterior_reativa': bool(prev_reativa and prev_reativa > 0),
        'avisos': avisos,
    }



# === Arquivo interno de faturas mensais e PDF em 1 página ===
def ensure_faturas_mensais_archive_schema():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS faturas_mensais_arquivo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            local TEXT NOT NULL,
            mes TEXT NOT NULL,
            ano INTEGER NOT NULL,
            periodo TEXT,
            total REAL DEFAULT 0,
            subtotal REAL DEFAULT 0,
            kwh_ativa REAL DEFAULT 0,
            kvarh_excedente REAL DEFAULT 0,
            demanda_ponta_kw REAL DEFAULT 0,
            agua_total REAL DEFAULT 0,
            consumo_especifico REAL,
            snapshot_json TEXT,
            criado_em TEXT DEFAULT (datetime('now','localtime')),
            atualizado_em TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(local, mes, ano)
        )
    ''')
    conn.commit(); conn.close()


def ensure_leituras_mensais_status_schema():
    """Tabela leve para controlo operacional do mês: aberto/fechado.
    Não altera leituras existentes. Serve para evitar que uma fatura já conferida
    seja modificada por engano sem reabertura consciente do período.
    """
    ensure_leituras_mensais_phase2_schema()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS leituras_mensais_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            local TEXT NOT NULL,
            mes TEXT NOT NULL,
            ano INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'aberto',
            fechado_em TEXT,
            fechado_por TEXT,
            reaberto_em TEXT,
            reaberto_por TEXT,
            observacao TEXT,
            atualizado_em TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(local, mes, ano)
        )
    ''')
    conn.commit(); conn.close()


def _get_periodo_status(local, mes, ano):
    ensure_leituras_mensais_status_schema()
    mes = str(mes).zfill(2); ano = int(ano)
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    row = c.execute('SELECT * FROM leituras_mensais_status WHERE local=? AND mes=? AND ano=?', (local, mes, ano)).fetchone()
    conn.close()
    if not row:
        return {'status':'aberto', 'fechado_em':'', 'fechado_por':'', 'reaberto_em':'', 'reaberto_por':'', 'observacao':'', 'fechado':False}
    d = dict(row); d['fechado'] = str(d.get('status') or '').lower() == 'fechado'
    return d


def _set_periodo_status(local, mes, ano, status='aberto', actor='operador', observacao=''):
    ensure_leituras_mensais_status_schema()
    mes = str(mes).zfill(2); ano = int(ano); status = (status or 'aberto').lower()
    now_sql = "datetime('now','localtime')"
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    if status == 'fechado':
        c.execute(f'''
            INSERT INTO leituras_mensais_status(local, mes, ano, status, fechado_em, fechado_por, observacao, atualizado_em)
            VALUES(?, ?, ?, 'fechado', {now_sql}, ?, ?, {now_sql})
            ON CONFLICT(local, mes, ano) DO UPDATE SET
                status='fechado', fechado_em={now_sql}, fechado_por=excluded.fechado_por,
                observacao=excluded.observacao, atualizado_em={now_sql}
        ''', (local, mes, ano, actor, observacao))
    else:
        c.execute(f'''
            INSERT INTO leituras_mensais_status(local, mes, ano, status, reaberto_em, reaberto_por, observacao, atualizado_em)
            VALUES(?, ?, ?, 'aberto', {now_sql}, ?, ?, {now_sql})
            ON CONFLICT(local, mes, ano) DO UPDATE SET
                status='aberto', reaberto_em={now_sql}, reaberto_por=excluded.reaberto_por,
                observacao=excluded.observacao, atualizado_em={now_sql}
        ''', (local, mes, ano, actor, observacao))
    conn.commit(); conn.close()


def _validar_periodo_mensal_operacional(local, mes, ano, pot_contratada=0, fator_mult=1):
    """Valida rapidamente o mês para orientar operador antes de gerar/fechar fatura."""
    mes = str(mes).zfill(2); ano = int(ano)
    qfat = _quantidades_fatura_mensal(local, mes, ano) if local else {}
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute('''SELECT data, ativa, reativa, ponta, fp, diferenca, agua
                        FROM leituras_mensais WHERE local=? AND mes=? AND ano=? ORDER BY data''', (local, mes, ano)).fetchall()
    conn.close()
    total_dias = calendar.monthrange(ano, int(mes))[1]
    dias_preenchidos = 0
    dias_fp_baixo = 0
    quedas_ativa = 0
    quedas_reativa = 0
    dias_sem_agua = 0
    ultima_ativa = None; ultima_reativa = None
    for r in rows:
        ativa = _safe_float(r['ativa'], None); reativa = _safe_float(r['reativa'], None)
        ponta = _safe_float(r['ponta'], None); agua = _safe_float(r['agua'], 0) or 0
        fp = _safe_float(r['fp'], None)
        tem_linha = any(v is not None and v != 0 for v in [ativa, reativa, ponta, agua])
        if tem_linha:
            dias_preenchidos += 1
            if agua <= 0:
                dias_sem_agua += 1
        if fp is not None and fp > 0 and fp < 0.85:
            dias_fp_baixo += 1
        if ativa is not None and ativa > 0:
            if ultima_ativa is not None and ativa < ultima_ativa:
                quedas_ativa += 1
            ultima_ativa = max(ultima_ativa or ativa, ativa)
        if reativa is not None and reativa > 0:
            if ultima_reativa is not None and reativa < ultima_reativa:
                quedas_reativa += 1
            ultima_reativa = max(ultima_reativa or reativa, reativa)

    avisos = list(qfat.get('avisos', []))
    criticos = []
    if dias_preenchidos == 0:
        criticos.append('Ainda não existem leituras preenchidas para este período.')
    if qfat.get('kwh_ativa', 0) <= 0 and dias_preenchidos > 1:
        criticos.append('Consumo ativo faturável igual a zero. Verificar leitura base, leitura final ou fator multiplicativo.')
    if not qfat.get('tem_base_mes_anterior_ativa'):
        avisos.append('Não foi encontrada base ativa do mês anterior. A primeira leitura válida do mês atual será tratada como base inicial.')
    if not qfat.get('tem_base_mes_anterior_reativa'):
        avisos.append('Não foi encontrada base reativa do mês anterior. A primeira leitura válida do mês atual será tratada como base inicial.')
    if qfat.get('kw_ponta_lida', 0) <= 0:
        avisos.append('Ponta máxima do mês ainda não foi registada.')
    if qfat.get('agua_total', 0) <= 0:
        avisos.append('Água elevada total está zerada; o consumo específico mensal não será representativo.')
    if quedas_ativa:
        criticos.append(f'Foram encontradas {quedas_ativa} ocorrência(s) de leitura ativa inferior à leitura válida anterior.')
    if quedas_reativa:
        criticos.append(f'Foram encontradas {quedas_reativa} ocorrência(s) de leitura reativa inferior à leitura válida anterior.')
    if dias_fp_baixo:
        avisos.append(f'{dias_fp_baixo} dia(s) com fator de potência abaixo de 0,85.')
    if dias_sem_agua and qfat.get('agua_total', 0) > 0:
        avisos.append(f'{dias_sem_agua} dia(s) preenchido(s) sem volume de água registado.')

    return {
        'ok_para_faturar': len(criticos) == 0,
        'criticos': criticos,
        'avisos': avisos,
        'dias_preenchidos': dias_preenchidos,
        'total_dias': total_dias,
        'progresso_pct': round((dias_preenchidos / total_dias) * 100, 1) if total_dias else 0,
        'qfat': qfat,
        'dias_fp_baixo': dias_fp_baixo,
        'quedas_ativa': quedas_ativa,
        'quedas_reativa': quedas_reativa,
    }


def _arquivar_fatura_mensal_snapshot(ctx):
    try:
        ensure_faturas_mensais_archive_schema()
        local = str(ctx.get('local') or '').strip()
        periodo = str(ctx.get('periodo') or '')
        mes = periodo.split('/')[0].zfill(2) if '/' in periodo else str(ctx.get('mes') or '').zfill(2)
        ano = int(periodo.split('/')[1]) if '/' in periodo else int(ctx.get('ano') or 0)
        if not local or not mes or not ano:
            return None
        serializable = {}
        for k, v in ctx.items():
            if k in ('request', 'qfat'):
                continue
            try:
                json.dumps(v)
                serializable[k] = v
            except Exception:
                serializable[k] = str(v)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            INSERT INTO faturas_mensais_arquivo
            (local, mes, ano, periodo, total, subtotal, kwh_ativa, kvarh_excedente, demanda_ponta_kw, agua_total, consumo_especifico, snapshot_json)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(local, mes, ano) DO UPDATE SET
                periodo=excluded.periodo,
                total=excluded.total,
                subtotal=excluded.subtotal,
                kwh_ativa=excluded.kwh_ativa,
                kvarh_excedente=excluded.kvarh_excedente,
                demanda_ponta_kw=excluded.demanda_ponta_kw,
                agua_total=excluded.agua_total,
                consumo_especifico=excluded.consumo_especifico,
                snapshot_json=excluded.snapshot_json,
                atualizado_em=datetime('now','localtime')
        ''', (
            local, mes, ano, periodo,
            float(ctx.get('total') or 0), float(ctx.get('subtotal') or 0),
            float(ctx.get('kwh_ativa') or 0), float(ctx.get('kvarh_excedente') or 0),
            float(ctx.get('demanda_ponta_kw') or 0), float(ctx.get('agua_total') or 0),
            ctx.get('consumo_especifico_medio'),
            json.dumps(serializable, ensure_ascii=False)
        ))
        conn.commit()
        rid = c.execute('SELECT id FROM faturas_mensais_arquivo WHERE local=? AND mes=? AND ano=?', (local, mes, ano)).fetchone()[0]
        conn.close()
        return rid
    except Exception as e:
        print('Falha ao arquivar fatura mensal:', e)
        return None


def _montar_contexto_fatura_mensal(local, mes, ano):
    mes_str = str(mes).zfill(2)
    ano_int = int(ano)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT id, nome FROM locais WHERE nome = ?', (local,))
    row_loc = c.fetchone()
    local_id = row_loc['id'] if row_loc else None
    fator_mult     = 1.0
    pot_contratada = 0.0
    tarifa_ativa   = 4.780
    tarifa_reativa = 1.430
    tarifa_ponta   = 497.00
    taxa_fixa      = 207.28
    taxa_radio     = 297.00
    taxa_lixo      = 150.00
    iva_cfg        = 16.0
    if local_id is not None:
        try:
            c.execute('''
                SELECT fator_mult, pot_contratada, tarifa_ativa, tarifa_reativa, tarifa_ponta,
                       tarifa_perdas, taxa_fixa, taxa_radio, taxa_lixo, iva
                FROM locais_cfg WHERE local_id = ?
            ''', (local_id,))
            cfg = c.fetchone()
            if cfg:
                (fator_mult, pot_contratada, tarifa_ativa, tarifa_reativa, tarifa_ponta,
                 _tarifa_perdas_ignorar, taxa_fixa, taxa_radio, taxa_lixo, iva_cfg) = cfg
        except Exception:
            pass
    conn.close()
    if not tarifa_reativa and tarifa_ativa:
        tarifa_reativa = tarifa_ativa * 0.30
    qfat = _quantidades_fatura_mensal(local, mes_str, ano_int)
    kwh_ativa = qfat['kwh_ativa']
    kvarh_reativa = qfat['kvarh_reativa']
    limite_reativa = qfat['limite_reativa']
    kvarh_excedente = qfat['kvarh_excedente']
    kw_ponta_lida = qfat['kw_ponta_lida']
    agua_total = qfat['agua_total']
    demanda_ponta_kw = _ponta_faturavel_edm(pot_contratada, kw_ponta_lida)
    valor_ativa = kwh_ativa * float(tarifa_ativa or 0)
    valor_reativa = kvarh_excedente * float(tarifa_reativa or 0)
    valor_ponta = demanda_ponta_kw * float(tarifa_ponta or 0)
    valor_perdas = 0.0
    subtotal_energia = valor_ativa + valor_reativa + valor_ponta
    subtotal_taxas = float(taxa_fixa or 0) + float(taxa_radio or 0) + float(taxa_lixo or 0)
    subtotal = subtotal_energia + subtotal_taxas
    IVA_ALIQUOTA = 0.16
    BASE_IVA_PERC = 0.62
    base_iva = subtotal * BASE_IVA_PERC
    valor_iva = base_iva * IVA_ALIQUOTA
    total = subtotal + valor_iva
    consumo_especifico_medio = (kwh_ativa / agua_total) if agua_total and kwh_ativa > 0 else None
    return dict(
        local=local, mes=mes_str, ano=ano_int, periodo=f"{mes_str}/{ano_int}",
        fator_mult=fator_mult, pot_contratada=pot_contratada,
        tarifa_ativa=tarifa_ativa, tarifa_reativa=tarifa_reativa, tarifa_ponta=tarifa_ponta,
        taxa_fixa=taxa_fixa, taxa_radio=taxa_radio, taxa_lixo=taxa_lixo,
        kwh_ativa=kwh_ativa, kvarh_reativa=kvarh_reativa, limite_reativa=limite_reativa,
        kvarh_excedente=kvarh_excedente, kw_ponta_lida=kw_ponta_lida,
        demanda_ponta_kw=demanda_ponta_kw,
        valor_ativa=valor_ativa, valor_reativa=valor_reativa, valor_ponta=valor_ponta,
        valor_perdas=valor_perdas, subtotal_energia=subtotal_energia,
        subtotal_taxas=subtotal_taxas, subtotal=subtotal, base_iva=base_iva,
        valor_iva=valor_iva, total=total, total_extenso=_mzn_extenso(total),
        consumo_especifico_medio=consumo_especifico_medio, agua_total=agua_total,
        iva_percent=IVA_ALIQUOTA * 100, base_iva_percent=BASE_IVA_PERC * 100,
        leitura_base_ativa=qfat.get('leitura_base_ativa', 0),
        leitura_final_ativa=qfat.get('leitura_final_ativa', 0),
        leitura_base_reativa=qfat.get('leitura_base_reativa', 0),
        leitura_final_reativa=qfat.get('leitura_final_reativa', 0),
        avisos_fatura=qfat.get('avisos', []), qfat=qfat
    )

@app.route('/leituras_mensal', methods=['GET', 'POST'])
def leituras_mensal():
    ensure_leituras_mensais_phase2_schema()
    hoje = datetime.now()
    locais_db = get_locais()
    meses = [(str(i).zfill(2), calendar.month_name[i]) for i in range(1, 13)]

    if not locais_db:
        flash("Nenhum local configurado. Primeiro cadastre pelo menos um Local.", "warning")
        return render_template('leituras_mensal.html', locais_db=[], selected_local_id=None, local='', meses=meses,
                               mes=hoje.strftime('%m'), ano=hoje.year, dias=[], leituras={}, fator_mult=1.0,
                               pot_contratada=0.0, pot_instalada=0.0, fp_medio=0.0, pot_max_ponta=0.0,
                               cfg_selected={}, cfg_map={}, first_prev_ativa=0.0, first_prev_reativa=0.0,
                               resumo={})

    # Leitura dos filtros: em POST, prioriza o formulário; em GET, usa a URL.
    # Isto evita o problema de a página manter o mês/ano antigo quando a URL
    # ainda trazia parâmetros anteriores.
    if request.method == 'POST':
        raw_local = request.form.get('local', str(locais_db[0][0]))
        mes_req = request.form.get('mes')
        ano_req = request.form.get('ano')
    else:
        raw_local = request.args.get('local', str(locais_db[0][0]))
        mes_req = request.args.get('mes')
        ano_req = request.args.get('ano')

    selected_local_id, local_nome = _local_id_nome_from_request(raw_local, locais_db)
    mes = (mes_req or hoje.strftime('%m')).zfill(2)
    ano = int(ano_req or hoje.year)

    cfg_selected = get_local_cfg_full(selected_local_id) if selected_local_id else {}
    fator_mult = float(cfg_selected.get('fator_mult') or 1.0)
    pot_contratada = float(cfg_selected.get('pot_contratada') or 0.0)
    pot_instalada = float(cfg_selected.get('pot_instalada') or 0.0)

    num_dias = calendar.monthrange(ano, int(mes))[1]
    dias = [f"{ano}-{mes}-{str(dia).zfill(2)}" for dia in range(1, num_dias + 1)]

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    rows_db = c.execute("""
        SELECT data,hora,ativa,reativa,ponta,fp,potc,anterior,atual,
               diferenca,agua,esp,acum,valor
        FROM leituras_mensais
        WHERE local=? AND mes=? AND ano=?
        ORDER BY data
    """, (local_nome, mes, ano)).fetchall()
    conn.close()

    leituras_map = {}
    for r in rows_db:
        fator = fator_mult if fator_mult else 1.0
        ativa_lida = ((r[2] or 0) / fator if r[2] not in (None, '') else '')
        reativa_lida = ((r[3] or 0) / fator if r[3] not in (None, '') else '')
        ponta_lida = ((r[4] or 0) / fator if r[4] not in (None, '') else '')
        leituras_map[r[0]] = {
            'hora': r[1] or '',
            'ativa': r[2] if r[2] is not None else '',
            'reativa': r[3] if r[3] is not None else '',
            'ponta': r[4] if r[4] is not None else '',
            'fp': r[5] if r[5] is not None else '',
            'potc': r[6] if r[6] is not None else pot_contratada,
            'anterior': r[7] if r[7] is not None else '',
            'atual': r[8] if r[8] is not None else '',
            'diferenca': r[9] if r[9] is not None else '',
            'agua': r[10] if r[10] is not None else '',
            'esp': r[11] if r[11] is not None else '',
            'acum': r[12] if r[12] is not None else '',
            'valor': r[13] if r[13] is not None else '',
            'ativa_lida': ativa_lida,
            'reativa_lida': reativa_lida,
            'ponta_lida': ponta_lida,
            'reativa_excedente': 0,
            'valor_ativa': 0,
            'valor_reativa': 0,
            'valor_ponta': 0,
            'valor_total_dia': r[13] or 0,
        }

    fp_vals = []
    pot_max_ponta = 0.0
    resumo = {'kwh_total':0.0, 'kvarh_total':0.0, 'ponta_max':0.0, 'agua_total':0.0, 'consumo_especifico':0.0,
              'valor_total':0.0, 'reativa_excedente':0.0, 'dias_preenchidos':0}
    for d in dias:
        row = leituras_map.get(d)
        if not row:
            continue
        resumo['dias_preenchidos'] += 1
        dif = _safe_float(row.get('diferenca'), 0.0) or 0.0
        agua = _safe_float(row.get('agua'), 0.0) or 0.0
        val = _safe_float(row.get('valor_total_dia'), _safe_float(row.get('valor'), 0.0)) or 0.0
        rea_exc = _safe_float(row.get('reativa_excedente'), 0.0) or 0.0
        ponta = _safe_float(row.get('ponta'), 0.0) or 0.0
        resumo['kwh_total'] += dif
        resumo['agua_total'] += agua
        resumo['valor_total'] += val
        resumo['reativa_excedente'] += rea_exc
        resumo['ponta_max'] = max(resumo['ponta_max'], ponta)
        try:
            if row['fp'] not in ('', None):
                fp_vals.append(float(row['fp']))
        except Exception:
            pass
    resumo['consumo_especifico'] = (resumo['kwh_total'] / resumo['agua_total']) if resumo['agua_total'] else 0.0
    fp_medio = round(sum(fp_vals) / len(fp_vals), 3) if fp_vals else 0.0
    pot_max_ponta = resumo['ponta_max']
    first_prev_ativa, first_prev_reativa = get_prev_month_last_readings(local_nome, mes, ano)
    periodo_status = _get_periodo_status(local_nome, mes, ano)
    validacao_periodo = _validar_periodo_mensal_operacional(local_nome, mes, ano, pot_contratada, fator_mult)

    return render_template('leituras_mensal.html', locais_db=locais_db, selected_local_id=selected_local_id,
                           local=local_nome, meses=meses, mes=mes, ano=ano, dias=dias, leituras=leituras_map,
                           fator_mult=fator_mult, pot_contratada=pot_contratada, pot_instalada=pot_instalada,
                           fp_medio=fp_medio, pot_max_ponta=pot_max_ponta, cfg_selected=cfg_selected,
                           cfg_map={}, first_prev_ativa=first_prev_ativa,
                           first_prev_reativa=first_prev_reativa, resumo=resumo,
                           periodo_status=periodo_status, validacao_periodo=validacao_periodo)


def get_prev_month_last_readings(local_nome: str, mes: str, ano: int):
    """Última leitura faturada de ativa/reativa do mês anterior."""
    ensure_leituras_mensais_phase2_schema()
    mes_int = int(mes); ano_int = int(ano)
    prev_mes = mes_int - 1; prev_ano = ano_int
    if prev_mes == 0:
        prev_mes = 12; prev_ano = ano_int - 1
    prev_mes_str = str(prev_mes).zfill(2)
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    row = c.execute("""
        SELECT ativa, reativa
        FROM leituras_mensais
        WHERE local=? AND mes=? AND ano=? AND ativa IS NOT NULL
        ORDER BY data DESC
        LIMIT 1
    """, (local_nome, prev_mes_str, prev_ano)).fetchone()
    conn.close()
    if row:
        return float(row[0] or 0), float(row[1] or 0)
    return 0.0, 0.0


@app.route('/leituras_mensal_salvar', methods=['POST'])
def leituras_mensal_salvar():
    ensure_leituras_mensais_phase2_schema()
    hoje = datetime.now()
    locais_db = get_locais()
    raw_local = (request.form.get('local') or '').strip()
    selected_local_id, local_nome = _local_id_nome_from_request(raw_local, locais_db)
    mes = (request.form.get('mes') or hoje.strftime('%m')).zfill(2)
    ano = int(request.form.get('ano') or hoje.year)
    cfg = get_local_cfg_full(selected_local_id) if selected_local_id else {}
    fator_mult = _safe_float(request.form.get('fator_mult'), cfg.get('fator_mult', 1.0)) or 1.0
    pot_contratada = _safe_float(cfg.get('pot_contratada'), 0.0) or 0.0
    t_ativa = _safe_float(cfg.get('tarifa_ativa'), 0.0) or 0.0
    t_reativa = _safe_float(cfg.get('tarifa_reativa'), 0.0) or 0.0
    t_ponta = _safe_float(cfg.get('tarifa_ponta'), 0.0) or 0.0
    acao = (request.form.get('acao') or '').strip()
    if _get_periodo_status(local_nome, mes, ano).get('fechado'):
        flash('Este mês está FECHADO. Para alterar leituras, reabra o período de forma controlada.', 'warning')
        return redirect(url_for('leituras_mensal', local=local_nome, mes=mes, ano=ano))
    num_dias = calendar.monthrange(ano, int(mes))[1]

    prev_ativa, prev_reativa = get_prev_month_last_readings(local_nome, mes, ano)
    # Se não existir leitura do mês anterior, a primeira leitura preenchida do mês
    # passa a ser a linha de base. Isto evita faturar indevidamente o valor
    # acumulado histórico do contador como se fosse consumo do mês.
    has_prev_ativa = bool(prev_ativa and prev_ativa > 0)
    has_prev_reativa = bool(prev_reativa and prev_reativa > 0)
    prev_ponta_corrigida = 0.0
    acum_mes = 0.0
    linhas_processadas = 0

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    for i in range(num_dias):
        data_str = request.form.get(f"data_{i}")
        if not data_str:
            continue
        hora = request.form.get(f"hora_{i}") or ""
        ativa_lida = _safe_float(request.form.get(f"ativa_lida_{i}"), None)
        reativa_lida = _safe_float(request.form.get(f"reativa_lida_{i}"), None)
        ponta_lida = _safe_float(request.form.get(f"ponta_lida_{i}"), None)
        agua_val = _safe_float(request.form.get(f"agua_{i}"), None)
        # compatibilidade com template antigo
        if ativa_lida is None:
            ativa_lida = _safe_float(request.form.get(f"ativa_{i}"), None)
        if reativa_lida is None:
            reativa_lida = _safe_float(request.form.get(f"reativa_{i}"), None)
        if ponta_lida is None:
            ponta_lida = _safe_float(request.form.get(f"ponta_{i}"), None)

        if ativa_lida is None and reativa_lida is None and ponta_lida is None and agua_val is None:
            continue

        ativa_fat = (ativa_lida * fator_mult) if ativa_lida is not None else None
        reativa_fat = (reativa_lida * fator_mult) if reativa_lida is not None else None
        ponta_fat = (ponta_lida * fator_mult) if ponta_lida is not None else None
        if ponta_fat is not None:
            # A ponta lida é tratada como registo máximo mensal: pode manter-se ou aumentar, nunca reduzir.
            if ponta_fat < prev_ponta_corrigida:
                ponta_fat = prev_ponta_corrigida
            else:
                prev_ponta_corrigida = ponta_fat
        # Energia ativa: contador acumulativo. Consumo do dia = leitura atual - leitura anterior.
        # Se não houver leitura anterior do mês anterior, a primeira leitura do mês é apenas referência inicial.
        anterior_val = prev_ativa if has_prev_ativa else (ativa_fat if ativa_fat is not None else prev_ativa)
        atual_val = ativa_fat if ativa_fat is not None else prev_ativa
        if ativa_fat is not None and has_prev_ativa:
            dif_val = atual_val - anterior_val
        else:
            dif_val = 0.0
        if dif_val < 0:
            # Leitura menor que a anterior: não fatura consumo negativo e NÃO atualiza
            # a referência anterior. Assim, o próximo dia válido continua a comparar
            # contra a última leitura correta do mês anterior ou do dia anterior válido.
            dif_operacional = 0.0
        else:
            dif_operacional = dif_val
        if ativa_fat is not None:
            if not has_prev_ativa:
                prev_ativa = atual_val
                has_prev_ativa = True
            elif dif_val >= 0:
                prev_ativa = atual_val

        # Energia reativa: mesma lógica da ativa. Nunca se deve usar a leitura acumulada total
        # como reativa excedente; usa-se apenas a diferença mensal/diária. Se a leitura recuar,
        # mantém-se a última referência válida para não contaminar os dias seguintes.
        if reativa_fat is not None:
            if has_prev_reativa:
                delta_reativa_real = reativa_fat - prev_reativa
            else:
                delta_reativa_real = 0.0
            if delta_reativa_real < 0:
                delta_reativa = 0.0
            else:
                delta_reativa = delta_reativa_real
            if not has_prev_reativa:
                prev_reativa = reativa_fat
                has_prev_reativa = True
            elif delta_reativa_real >= 0:
                prev_reativa = reativa_fat
        else:
            delta_reativa = 0.0

        if dif_operacional > 0 or delta_reativa > 0:
            fp_val = dif_operacional / math.sqrt((dif_operacional ** 2) + (delta_reativa ** 2)) if (dif_operacional or delta_reativa) else None
        else:
            fp_val = None
        reativa_excedente = max(delta_reativa - (0.75 * dif_operacional), 0.0)
        esp_val = (dif_operacional / agua_val) if agua_val else None
        acum_mes += dif_operacional
        valor_ativa = dif_operacional * t_ativa
        valor_reativa = reativa_excedente * t_reativa
        valor_ponta = 0.0  # A ponta é uma cobrança mensal: 20% PC + 80% ponta máxima corrigida. Não é somada por dia.
        valor_total_dia = valor_ativa + valor_reativa

        c.execute("""
            INSERT INTO leituras_mensais
            (local, data, hora, ativa, reativa, ponta, fp, potc, anterior, atual, diferenca,
             agua, esp, acum, valor, mes, ano)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(local, data) DO UPDATE SET
                hora=excluded.hora,
                ativa=excluded.ativa,
                reativa=excluded.reativa,
                ponta=excluded.ponta,
                fp=excluded.fp,
                potc=excluded.potc,
                anterior=excluded.anterior,
                atual=excluded.atual,
                diferenca=excluded.diferenca,
                agua=excluded.agua,
                esp=excluded.esp,
                acum=excluded.acum,
                valor=excluded.valor,
                mes=excluded.mes,
                ano=excluded.ano
        """, (local_nome, data_str, hora, ativa_fat, reativa_fat, ponta_fat, fp_val, pot_contratada,
              anterior_val, atual_val, dif_operacional, agua_val, esp_val, acum_mes, valor_total_dia,
              mes, ano))
        linhas_processadas += 1

    conn.commit(); conn.close()
    flash(f"Leituras mensais salvas e calculadas automaticamente ({linhas_processadas} linhas processadas).", "success")
    if acao == 'fatura_edm':
        return redirect(url_for('leituras_mensal_fatura_edm', local=local_nome, mes=mes, ano=str(ano)))
    return redirect(url_for('leituras_mensal', local=local_nome, mes=mes, ano=ano))


@app.route('/leituras_mensal/status', methods=['POST'])
def leituras_mensal_status_periodo():
    ensure_leituras_mensais_status_schema()
    local = (request.form.get('local') or request.args.get('local') or '').strip()
    mes = (request.form.get('mes') or request.args.get('mes') or datetime.now().strftime('%m')).zfill(2)
    ano = int(request.form.get('ano') or request.args.get('ano') or datetime.now().year)
    operacao = (request.form.get('operacao') or '').strip().lower()
    obs = (request.form.get('observacao') or '').strip()
    actor = session.get('username') if 'session' in globals() else None
    actor = actor or 'operador'
    if not local:
        flash('Selecione um local antes de alterar o estado do período.', 'warning')
        return redirect(url_for('leituras_mensal'))
    if operacao == 'fechar':
        validacao = _validar_periodo_mensal_operacional(local, mes, ano)
        if not validacao.get('ok_para_faturar'):
            flash('O mês não foi fechado porque existem inconsistências críticas. Abra a Auditoria Inteligente para corrigir.', 'danger')
            return redirect(url_for('leituras_mensal_audit', local=local, mes=mes, ano=ano))
        ctx = _montar_contexto_fatura_mensal(local, mes, ano)
        _arquivar_fatura_mensal_snapshot(ctx)
        _set_periodo_status(local, mes, ano, 'fechado', actor=actor, observacao=obs)
        flash('Período fechado com sucesso. A fatura foi arquivada e as leituras ficaram protegidas contra alterações acidentais.', 'success')
    elif operacao == 'reabrir':
        _set_periodo_status(local, mes, ano, 'aberto', actor=actor, observacao=obs)
        flash('Período reaberto. Pode alterar e salvar as leituras novamente.', 'info')
    else:
        flash('Operação de estado inválida.', 'warning')
    return redirect(url_for('leituras_mensal', local=local, mes=mes, ano=ano))


@app.route('/leituras_mensal/visualizar')
def visualizar_mensal():
    locais = [l[1] for l in get_locais()]
    mes = request.args.get('mes') or datetime.now().strftime('%Y-%m')
    # aceitar formatos YYYY-MM ou MM
    if len(mes)==2:
        mes_val = mes
        ano_val = int(request.args.get('ano') or datetime.now().year)
        mes_str = f"{ano_val}-{mes_val}"
    else:
        ano_val, mes_val = mes.split('-')
        ano_val = int(ano_val)
    local = request.args.get('local', locais[0] if locais else '')
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute('''SELECT data, diferenca, ativa, reativa, ponta, fp FROM leituras_mensais
                        WHERE local=? AND mes=? AND ano=? ORDER BY data''',
                     (local, mes_val, ano_val)).fetchall()
    conn.close()
    # converter para listas de plot
    dias = [r[0] for r in rows]
    difs = [r[1] or 0 for r in rows]
    ativa = [r[2] or 0 for r in rows]
    reativa = [r[3] or 0 for r in rows]
    ponta = [r[4] or 0 for r in rows]
    fp = [r[5] or 0 for r in rows]
    return render_template('visualizar_mensal.html', locais=locais, local=local, mes=f"{ano_val}-{mes_val}", dias=dias,
                           difs=difs, ativas=ativa, reativas=reativa, pontas=ponta, fps=fp)

# === Importar Leituras Mensais (CSV ; separado por ponto e vírgula) ===
@app.route('/leituras_mensal/import', methods=['GET', 'POST'])
def leituras_mensal_import():
    locais = [l[1] for l in get_locais()]
    if request.method == 'GET':
        return render_template('leituras_mensal_import.html', locais=locais)
    # POST
    local = request.form.get('local','')
    mes = request.form.get('mes') or datetime.now().strftime('%m')
    ano = int(request.form.get('ano') or datetime.now().year)
    file = request.files.get('arquivo')
    if not (file and file.filename):
        flash('Selecione um ficheiro CSV/Excel.', 'warning')
        return redirect(url_for('leituras_mensal_import'))
    filename = secure_filename(file.filename)
    data_bytes = file.read()
    # Detect CSV vs Excel
    try:
        import pandas as pd
        from io import BytesIO
        if filename.lower().endswith(('.xls','.xlsx')):
            df = pd.read_excel(BytesIO(data_bytes))
        else:
            df = pd.read_csv(io.BytesIO(data_bytes), sep=';', encoding='utf-8')
    except Exception as e:
        flash(f'Erro ao ler o ficheiro: {e}', 'danger')
        return redirect(url_for('leituras_mensal_import'))

    expected_cols = {'data','hora','ativa','reativa','ponta','fp','potc','anterior','atual','diferenca','agua','esp','acum','valor'}
    lower_map = {c: c.lower().strip() for c in df.columns}
    df.columns = [lower_map[c] for c in df.columns]
    missing = expected_cols - set(df.columns)
    if missing:
        flash('Colunas em falta: ' + ', '.join(sorted(missing)), 'warning')
        return redirect(url_for('leituras_mensal_import'))

    # Inserir linhas
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    inseridos = 0
    for _, r in df.iterrows():
        data = str(r.get('data'))
        hora = str(r.get('hora')) if not pd.isna(r.get('hora')) else ''
        vals = (
            local, data, hora,
            float(r.get('ativa',0) or 0), float(r.get('reativa',0) or 0), float(r.get('ponta',0) or 0),
            fp_val if 'fp_val' in locals() else float(r.get('fp',0) or 0), float(r.get('potc',0) or 0),
            float(r.get('anterior',0) or 0), float(r.get('atual',0) or 0), dif_val if 'dif_val' in locals() else float(r.get('diferenca',0) or 0),
            float(r.get('agua',0) or 0), float(r.get('esp',0) or 0), float(r.get('acum',0) or 0),
            float(r.get('valor',0) or 0), mes, ano
        )
        c.execute('DELETE FROM leituras_mensais WHERE local=? AND data=? AND mes=? AND ano=?',
                  (local, data, mes, ano))
        c.execute('''INSERT INTO leituras_mensais
            (local,data,hora,ativa,reativa,ponta,fp,potc,anterior,atual,diferenca,agua,esp,acum,valor,mes,ano)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', vals)
        inseridos += 1
    conn.commit(); conn.close()
    flash(f'Importação concluída: {inseridos} linhas.', 'success')
    return redirect(url_for('leituras_mensal', local=local, mes=mes, ano=ano))


# === FATURA (manual) ===

@app.route('/calcular_fatura', methods=['GET', 'POST'])
def calcular_fatura():
    locais = get_locais()
    if request.method == 'POST':
        local_nome = request.form['local']
        periodo = request.form['periodo']
        fator_mult = float(request.form['fator_mult'])
        demanda_max = float(request.form['demanda_max'])
        pot_contratada = float(request.form['pot_contratada'])

        ativa_ant = float(request.form['ativa_ant'])
        ativa_atu = float(request.form['ativa_atu'])
        reativa_ant = float(request.form['reativa_ant'])
        reativa_atu = float(request.form['reativa_atu'])
        perdas_ant = float(request.form['perdas_ant'])
        perdas_atu = float(request.form['perdas_atu'])
        ponta_ant = float(request.form['ponta_ant'])
        ponta_atu = float(request.form['ponta_atu'])

        tarifa_ativa = float(request.form['tarifa_ativa'])
        tarifa_reativa = float(request.form['tarifa_reativa'])
        tarifa_ponta = float(request.form['tarifa_ponta'])
        tarifa_perdas = float(request.form['tarifa_perdas'])
        taxa_fixa = float(request.form['taxa_fixa'])
        taxa_radio = float(request.form['taxa_radio'])
        taxa_lixo = float(request.form['taxa_lixo'])
        iva = float(request.form['iva'])
        saldo_ant = float(request.form['saldo_ant'])

        ativa = (ativa_atu - ativa_ant) * fator_mult
        reativa = (reativa_atu - reativa_ant) * fator_mult
        perdas = (perdas_atu - perdas_ant) * fator_mult
        ponta_lida_corrigida = (ponta_atu - ponta_ant) * fator_mult
        ponta = _ponta_faturavel_edm(pot_contratada, ponta_lida_corrigida)

        valor_ativa = ativa * tarifa_ativa
        valor_reativa = reativa_faturavel * tarifa_reativa
        valor_perdas = perdas * tarifa_perdas
        valor_ponta = ponta * tarifa_ponta

        subtotal = (valor_ativa + valor_reativa + valor_perdas + valor_ponta +
                    taxa_fixa + taxa_radio + taxa_lixo)
        valor_iva = subtotal * iva / 100
        total = subtotal + valor_iva + saldo_ant

        return render_template('fatura_resultado.html',
                               local=local_nome, periodo=periodo,
                               ativa=ativa, reativa=reativa,
                               reativa_faturavel=reativa_faturavel,
                               perdas=perdas, ponta=ponta,
                               valor_ativa=valor_ativa,
                               valor_reativa=valor_reativa,
                               valor_perdas=valor_perdas,
                               valor_ponta=valor_ponta,
                               subtotal=subtotal,
                               valor_iva=valor_iva,
                               total=total)
    # GET -> cfg_map para auto-preencher
    cfg_map = {}
    for lid, lname in locais:
        cfg = get_local_cfg(lid)
        cfg_map[str(lid)] = {
            "fator_mult": cfg[0],
            "pot_contratada": cfg[1],
            "tarifa_ativa": cfg[2],
            "tarifa_reativa": cfg[3],
            "tarifa_ponta": cfg[4],
            "tarifa_perdas": cfg[5],
            "taxa_fixa": cfg[6],
            "taxa_radio": cfg[7],
            "taxa_lixo": cfg[8],
            "iva": cfg[9],
        }
    return render_template('fatura.html', locais=locais, cfg_map=cfg_map)

# === FATURA (a partir do mês) ===
@app.route('/fatura/mes', methods=['GET', 'POST'])
def fatura_mes():
    locais = get_locais()
    hoje = datetime.now()
    if request.method == 'POST':
        local_id = int(request.form['local_id'])
        mes = request.form['mes']  # "YYYY-MM"
        ano_int = int(mes.split('-')[0]); mes_int = int(mes.split('-')[1])
        local_nome = [l[1] for l in locais if l[0]==local_id][0]

        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('SELECT fator_mult, pot_contratada, tarifa_ativa, tarifa_reativa, tarifa_ponta, tarifa_perdas, taxa_fixa, taxa_radio, taxa_lixo, iva FROM locais_cfg WHERE local_id=?',
                  (local_id,))
        cfg = c.fetchone()
        if not cfg:
            cfg = (1.0, 0.0, 4.780, 1.430, 4.970, 4.780, 207.28, 297.00, 150.00, 16.0)
        fator_mult, pot_contratada, tarifa_ativa, tarifa_reativa, tarifa_ponta, tarifa_perdas, taxa_fixa, taxa_radio, taxa_lixo, iva = cfg

        qfat = _quantidades_fatura_mensal(local_nome, str(mes_int).zfill(2), ano_int)
        ativa = qfat['kwh_ativa']
        reativa = qfat['kvarh_reativa']
        reativa_faturavel = qfat['kvarh_excedente']
        ponta_lida_corrigida = qfat['kw_ponta_lida']
        ponta = _ponta_faturavel_edm(pot_contratada, ponta_lida_corrigida)
        perdas = 0.0

        valor_ativa = ativa * tarifa_ativa
        reativa_faturavel = max(reativa - 0.75 * ativa, 0)
        valor_reativa = reativa_faturavel * tarifa_reativa
        valor_perdas = perdas * tarifa_perdas
        valor_ponta = ponta * tarifa_ponta

        subtotal = (valor_ativa + valor_reativa + valor_perdas + valor_ponta +
                    taxa_fixa + taxa_radio + taxa_lixo)
        valor_iva = subtotal * iva / 100
        total = subtotal + valor_iva

        periodo_leg = f"{mes}-01 a {mes}-{calendar.monthrange(ano_int, mes_int)[1]}"

        return render_template('fatura_resultado.html',
                               local=local_nome, periodo=periodo_leg,
                               ativa=ativa, reativa=reativa, reativa_faturavel=reativa_faturavel,
                               perdas=perdas, ponta=ponta,
                               valor_ativa=valor_ativa, valor_reativa=valor_reativa,
                               valor_perdas=valor_perdas, valor_ponta=valor_ponta,
                               subtotal=subtotal, valor_iva=valor_iva, total=total)

    return render_template('fatura_mes.html', locais=locais, hoje=hoje.strftime('%Y-%m'))

# =========================
# === MÓDULO "MOTORES" ===
# =========================
# Fase 4A — o módulo de Motores deixa de ser uma página isolada e passa a ser
# uma análise técnica especializada, integrada com Equipamentos, Monitoria e Alertas.

def _motor_intervalo_padrao():
    hoje = datetime.now().date()
    return (hoje - timedelta(days=30)).strftime('%Y-%m-%d'), hoje.strftime('%Y-%m-%d')


def _motor_float(v, default=0.0):
    try:
        if v is None or v == '':
            return default
        return float(str(v).replace(',', '.'))
    except Exception:
        return default


def _motor_dict_factory(cursor, row):
    return {col[0]: row[idx] for idx, col in enumerate(cursor.description)}


def _motor_load_equipamentos(local_id=None, equip_id=None):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = _motor_dict_factory
    c = conn.cursor()
    try:
        c.execute("PRAGMA table_info(equipamentos)")
        cols = {r['name'] for r in c.fetchall()}
    except Exception:
        cols = set()

    optional = []
    for col in ['tag', 'categoria', 'criticidade', 'fabricante', 'modelo', 'especificacao', 'potencia_kw', 'ativo']:
        if col in cols:
            optional.append(f"e.{col} AS {col}")
        else:
            optional.append(f"'' AS {col}")

    sql = f'''
        SELECT e.id, e.nome, e.local_id, l.nome AS local_nome,
               {', '.join(optional)},
               cfg.tensao_nominal, cfg.corrente_nominal, cfg.potencia_nominal_kw,
               cfg.fp_nominal, cfg.eficiencia_nominal, cfg.limite_corrente, cfg.limite_fp
        FROM equipamentos e
        LEFT JOIN locais l ON l.id=e.local_id
        LEFT JOIN equipamentos_cfg cfg ON cfg.equipamento_id=e.id
        WHERE 1=1
    '''
    params = []
    if local_id:
        local_ids_scope = get_descendant_local_ids(local_id, include_self=True)
        if not local_ids_scope:
            local_ids_scope = [int(local_id)]
        placeholders_local = ','.join('?' for _ in local_ids_scope)
        sql += f' AND e.local_id IN ({placeholders_local})'; params.extend(local_ids_scope)
    if equip_id:
        sql += ' AND e.id=?'; params.append(equip_id)
    sql += ' ORDER BY COALESCE(l.nome,\'\'), e.nome'
    c.execute(sql, params)
    rows = c.fetchall()
    conn.close()
    return rows


def _motor_collect_stats(local_id=None, equip_id=None, ini=None, fim=None):
    if not ini or not fim:
        ini, fim = _motor_intervalo_padrao()
    dt_ini = ini + ' 00:00:00'
    dt_fim = fim + ' 23:59:59'
    equipamentos = _motor_load_equipamentos(local_id, equip_id)
    by_id = {int(e['id']): e for e in equipamentos}
    by_name = {(e['nome'] or '').strip().lower(): int(e['id']) for e in equipamentos}

    stats = {}
    for e in equipamentos:
        stats[int(e['id'])] = {
            'equipamento': e,
            'n': 0, 'kwh': 0.0, 'fp_sum': 0.0, 'fp_n': 0,
            'corr_sum': 0.0, 'corr_n': 0, 'corr_max': 0.0,
            'kw_sum': 0.0, 'kw_n': 0, 'kw_max': 0.0,
            'ponta_max': 0.0, 'tensao_min': None, 'tensao_max': None,
            'agua': 0.0, 'horas': 0.0, 'arranques': 0, 'last_ts': '',
            'alertas': [], 'fontes': set()
        }

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = _motor_dict_factory
    c = conn.cursor()

    # Medições específicas do módulo Motores
    params = [dt_ini, dt_fim]
    sql = '''
        SELECT m.*, e.nome AS equipamento_nome, e.local_id
        FROM motor_medicoes m
        JOIN equipamentos e ON e.id=m.equipamento_id
        WHERE datetime(m.datahora) BETWEEN datetime(?) AND datetime(?)
    '''
    if local_id:
        local_ids_scope = get_descendant_local_ids(local_id, include_self=True)
        if not local_ids_scope:
            local_ids_scope = [int(local_id)]
        placeholders_local = ','.join('?' for _ in local_ids_scope)
        sql += f' AND e.local_id IN ({placeholders_local})'; params.extend(local_ids_scope)
    if equip_id:
        sql += ' AND e.id=?'; params.append(equip_id)
    sql += ' ORDER BY datetime(m.datahora)'
    try:
        c.execute(sql, params)
        rows = c.fetchall()
    except Exception:
        rows = []

    for r in rows:
        eid = int(r['equipamento_id'])
        if eid not in stats:
            continue
        st = stats[eid]
        st['fontes'].add('Motores')
        st['n'] += 1
        st['last_ts'] = max(st['last_ts'], r.get('datahora') or '')
        kwh = _motor_float(r.get('energia_kwh'))
        st['kwh'] += max(kwh, 0)
        fp = r.get('fator_potencia')
        if fp is not None:
            st['fp_sum'] += _motor_float(fp); st['fp_n'] += 1
        corr = r.get('corrente_a')
        if corr is not None:
            corr = _motor_float(corr); st['corr_sum'] += corr; st['corr_n'] += 1; st['corr_max'] = max(st['corr_max'], corr)
        kw = r.get('pot_ativa_kw')
        if kw is not None:
            kw = _motor_float(kw); st['kw_sum'] += kw; st['kw_n'] += 1; st['kw_max'] = max(st['kw_max'], kw)
        tensao = r.get('tensao_v')
        if tensao is not None:
            tensao = _motor_float(tensao)
            st['tensao_min'] = tensao if st['tensao_min'] is None else min(st['tensao_min'], tensao)
            st['tensao_max'] = tensao if st['tensao_max'] is None else max(st['tensao_max'], tensao)

    # Dados da Monitoria Operacional: tabela leituras, associada pelo nome do equipamento.
    params = [dt_ini, dt_fim]
    sql = '''
        SELECT datahora, local, equipamento, energia_ativa, pot_ativa, fp, ponta, caudal_elevada, corrente, tensao
        FROM leituras
        WHERE datetime(datahora) BETWEEN datetime(?) AND datetime(?)
          AND COALESCE(equipamento,'') <> ''
    '''
    if equip_id and equip_id in by_id:
        sql += ' AND lower(equipamento)=lower(?)'; params.append(by_id[equip_id]['nome'])
    sql += ' ORDER BY datetime(datahora)'
    try:
        c.execute(sql, params)
        leituras = c.fetchall()
    except Exception:
        leituras = []

    for r in leituras:
        name = (r.get('equipamento') or '').strip().lower()
        eid = by_name.get(name)
        if not eid or eid not in stats:
            continue
        # se filtrou local, confirma pelo equipamento cadastrado
        if local_id and int(stats[eid]['equipamento'].get('local_id') or 0) != int(local_id):
            continue
        st = stats[eid]
        st['fontes'].add('Monitoria')
        st['n'] += 1
        st['last_ts'] = max(st['last_ts'], r.get('datahora') or '')
        st['kwh'] += max(_motor_float(r.get('energia_ativa')), 0)
        fp = r.get('fp')
        if fp is not None:
            st['fp_sum'] += _motor_float(fp); st['fp_n'] += 1
        corr = r.get('corrente')
        if corr is not None:
            corr = _motor_float(corr); st['corr_sum'] += corr; st['corr_n'] += 1; st['corr_max'] = max(st['corr_max'], corr)
        kw = r.get('pot_ativa')
        if kw is not None:
            kw = _motor_float(kw); st['kw_sum'] += kw; st['kw_n'] += 1; st['kw_max'] = max(st['kw_max'], kw)
        st['ponta_max'] = max(st['ponta_max'], _motor_float(r.get('ponta')))
        st['agua'] += max(_motor_float(r.get('caudal_elevada')), 0)
        tensao = r.get('tensao')
        if tensao is not None:
            tensao = _motor_float(tensao)
            st['tensao_min'] = tensao if st['tensao_min'] is None else min(st['tensao_min'], tensao)
            st['tensao_max'] = tensao if st['tensao_max'] is None else max(st['tensao_max'], tensao)

    # Horas de funcionamento / arranques
    params = [dt_ini, dt_fim, dt_ini, dt_fim, dt_fim]
    sql = '''
        SELECT r.equipamento_id, r.start_time, r.stop_time, r.duracao_min
        FROM motor_runs r
        JOIN equipamentos e ON e.id=r.equipamento_id
        WHERE ((datetime(r.start_time) BETWEEN datetime(?) AND datetime(?))
            OR (r.stop_time IS NOT NULL AND datetime(r.stop_time) BETWEEN datetime(?) AND datetime(?))
            OR (r.stop_time IS NULL AND datetime(r.start_time) <= datetime(?)))
    '''
    if local_id:
        sql += ' AND e.local_id=?'; params.append(local_id)
    if equip_id:
        sql += ' AND e.id=?'; params.append(equip_id)
    try:
        c.execute(sql, params)
        runs = c.fetchall()
    except Exception:
        runs = []
    fim_dt = datetime.fromisoformat(dt_fim)
    for r in runs:
        eid = int(r['equipamento_id'])
        if eid not in stats:
            continue
        stats[eid]['arranques'] += 1
        stats[eid]['fontes'].add('Horas')
        dur = r.get('duracao_min')
        if dur is None:
            try:
                st_dt = datetime.fromisoformat((r.get('start_time') or '').replace(' ', 'T'))
                dur = max((fim_dt - st_dt).total_seconds()/60.0, 0)
            except Exception:
                dur = 0
        stats[eid]['horas'] += _motor_float(dur)/60.0

    conn.close()

    analyzed = []
    total_alertas = 0
    criticos = 0
    for eid, st in stats.items():
        e = st['equipamento']
        avg_fp = st['fp_sum']/st['fp_n'] if st['fp_n'] else None
        avg_corr = st['corr_sum']/st['corr_n'] if st['corr_n'] else None
        avg_kw = st['kw_sum']/st['kw_n'] if st['kw_n'] else None
        pot_nom = _motor_float(e.get('potencia_nominal_kw')) or _motor_float(e.get('potencia_kw'))
        corr_nom = _motor_float(e.get('corrente_nominal'))
        lim_corr = _motor_float(e.get('limite_corrente')) or (corr_nom * 1.10 if corr_nom else 0)
        lim_fp = _motor_float(e.get('limite_fp'), 0.80) or 0.80
        carga_pct = (avg_kw / pot_nom * 100.0) if avg_kw is not None and pot_nom else None
        ce = (st['kwh']/st['agua']) if st['agua'] else None
        alertas = []
        if st['n'] == 0:
            alertas.append(('Informativo', 'Sem medições no período', 'Lançar medições operacionais ou associar leituras da Monitoria.'))
        if avg_fp is not None and avg_fp < lim_fp:
            nivel = 'Crítico' if avg_fp < 0.75 else 'Atenção'
            alertas.append((nivel, f'FP médio baixo: {avg_fp:.3f}', 'Avaliar compensação reativa, regime de carga e banco de capacitores.'))
        if lim_corr and st['corr_max'] > lim_corr:
            alertas.append(('Crítico', f'Corrente máxima acima do limite: {st["corr_max"]:.1f} A > {lim_corr:.1f} A', 'Verificar sobrecarga, desalinhamento, rolamentos, bomba travada ou desequilíbrio.'))
        if st['tensao_min'] is not None and (st['tensao_min'] < 360 or st['tensao_max'] > 440):
            alertas.append(('Atenção', f'Tensão fora da faixa: {st["tensao_min"]:.1f}–{st["tensao_max"]:.1f} V', 'Confirmar tensão por fase, queda de tensão e estado do quadro/transformador.'))
        if carga_pct is not None and carga_pct < 35:
            alertas.append(('Atenção', f'Baixo carregamento estimado: {carga_pct:.1f}%', 'Motor pode estar sobredimensionado ou a operar fora do ponto ótimo.'))
        if carga_pct is not None and carga_pct > 105:
            alertas.append(('Crítico', f'Sobrecarga estimada: {carga_pct:.1f}%', 'Reduzir carga e verificar dimensionamento/proteções.'))
        if st['horas'] > 0 and st['arranques'] / max(st['horas'], 1) > 6:
            alertas.append(('Atenção', 'Frequência elevada de arranques', 'Avaliar lógica de comando, pressostatos/níveis e impacto na vida útil.'))
        status = 'Normal'
        if any(a[0] == 'Crítico' for a in alertas): status = 'Crítico'
        elif any(a[0] == 'Atenção' for a in alertas): status = 'Atenção'
        elif st['n'] == 0: status = 'Sem dados'
        total_alertas += len(alertas)
        if status == 'Crítico': criticos += 1
        analyzed.append({
            'id': eid, 'equipamento': e, 'status': status, 'alertas': alertas,
            'n': st['n'], 'kwh': st['kwh'], 'avg_fp': avg_fp, 'avg_corr': avg_corr,
            'corr_max': st['corr_max'], 'avg_kw': avg_kw, 'kw_max': st['kw_max'],
            'ponta_max': st['ponta_max'], 'horas': st['horas'], 'arranques': st['arranques'],
            'tensao_min': st['tensao_min'], 'tensao_max': st['tensao_max'],
            'agua': st['agua'], 'ce': ce, 'carga_pct': carga_pct, 'last_ts': st['last_ts'],
            'fontes': ', '.join(sorted(st['fontes'])) or '—',
        })
    order = {'Crítico':0, 'Atenção':1, 'Normal':2, 'Sem dados':3}
    analyzed.sort(key=lambda x: (order.get(x['status'], 9), -(x['kwh'] or 0), x['equipamento']['nome'] or ''))
    resumo = {
        'equipamentos': len(analyzed),
        'com_dados': sum(1 for x in analyzed if x['n'] > 0),
        'criticos': criticos,
        'alertas': total_alertas,
        'kwh': sum(x['kwh'] for x in analyzed),
        'horas': sum(x['horas'] for x in analyzed),
        'fp_medio': None,
        'corrente_max': max([x['corr_max'] for x in analyzed] or [0]),
    }
    fp_vals = [x['avg_fp'] for x in analyzed if x['avg_fp'] is not None]
    if fp_vals:
        resumo['fp_medio'] = sum(fp_vals)/len(fp_vals)
    return analyzed, resumo, ini, fim



def _motor_kvar_compensacao(avg_kw, fp_atual, fp_alvo=0.92):
    """Estimativa de kVAr para corrigir FP de um motor/carga.
    Retorna None quando os dados não permitem cálculo coerente.
    """
    try:
        p_kw = float(avg_kw or 0)
        fp1 = float(fp_atual or 0)
        fp2 = float(fp_alvo or 0.92)
        if p_kw <= 0 or fp1 <= 0 or fp1 >= fp2 or fp2 >= 1:
            return None
        import math
        phi1 = math.acos(max(min(fp1, 0.999999), 0.000001))
        phi2 = math.acos(max(min(fp2, 0.999999), 0.000001))
        kvar = p_kw * (math.tan(phi1) - math.tan(phi2))
        return max(kvar, 0)
    except Exception:
        return None


def _motor_recomendacoes_detalhadas(analise):
    recs = []
    avg_fp = analise.get('avg_fp')
    carga = analise.get('carga_pct')
    corr_max = analise.get('corr_max') or 0
    horas = analise.get('horas') or 0
    arr = analise.get('arranques') or 0
    if avg_fp is not None and avg_fp < 0.80:
        recs.append(('Compensação reativa', 'Prioritário', 'Verificar banco de capacitores, correção individual/coletiva e operação em baixo carregamento.'))
    if carga is not None and carga < 35:
        recs.append(('Baixo carregamento', 'Atenção', 'Avaliar se o motor está sobredimensionado, se existe estrangulamento hidráulico ou se o ponto de operação da bomba está fora do ideal.'))
    if carga is not None and carga > 100:
        recs.append(('Sobrecarga', 'Crítico', 'Confirmar corrente por fase, vibração, rolamentos, alinhamento, estado da bomba e proteções térmicas.'))
    if arr and horas and arr / max(horas, 1) > 6:
        recs.append(('Arranques frequentes', 'Atenção', 'Rever lógica de comando, níveis, pressostatos, VFD/soft-starter e proteção contra partidas excessivas.'))
    if corr_max:
        recs.append(('Inspeção eléctrica', 'Rotina', 'Comparar corrente medida com corrente nominal do motor e verificar desequilíbrio entre fases quando houver medição trifásica disponível.'))
    if not recs:
        recs.append(('Operação normal', 'Rotina', 'Manter monitoria periódica, limpeza, reaperto de terminais, verificação de ventilação e atualização do histórico operacional.'))
    return recs

@app.route('/motores')
def motores_menu():
    local_id = request.args.get('local_id', type=int)
    equip_id = request.args.get('equipamento_id', type=int)
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id, nome FROM locais ORDER BY nome')
    locais = c.fetchall()
    c.execute('SELECT id, nome FROM equipamentos ORDER BY nome')
    equipamentos_select = c.fetchall()
    conn.close()
    analises, resumo, data_ini, data_fim = _motor_collect_stats(local_id, equip_id, data_ini, data_fim)
    return render_template('motores.html', locais=locais, equipamentos_select=equipamentos_select,
                           analises=analises, resumo=resumo, local_id=local_id or '', equip_id=equip_id or '',
                           data_ini=data_ini, data_fim=data_fim)


@app.route('/motores/medir', methods=['GET', 'POST'])
@app.route('/motores/nova', methods=['GET', 'POST'])
def motor_medir():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''
        SELECT e.id, e.nome, COALESCE(l.nome,'')
        FROM equipamentos e LEFT JOIN locais l ON l.id=e.local_id
        ORDER BY COALESCE(l.nome,''), e.nome
    ''')
    equipamentos = c.fetchall()
    conn.close()

    if request.method == 'POST':
        equipamento_id = int(request.form['equipamento_id'])
        datahora = request.form.get('datahora') or datetime.now().strftime('%Y-%m-%dT%H:%M')
        tensao_v = _motor_float(request.form.get('tensao_v'))
        corrente_a = _motor_float(request.form.get('corrente_a'))
        fp = _motor_float(request.form.get('fator_potencia'))
        freq = _motor_float(request.form.get('frequencia_hz'), 50)
        fases = int(_motor_float(request.form.get('fases'), 3) or 3)
        pot_ativa_kw = request.form.get('pot_ativa_kw')
        pot_reativa_kvar = request.form.get('pot_reativa_kvar')
        pot_aparente_kva = request.form.get('pot_aparente_kva')
        eficiencia = request.form.get('eficiencia')
        observacoes = request.form.get('observacoes', '').strip()

        pot_ativa_kw = _motor_float(pot_ativa_kw, None) if pot_ativa_kw not in (None, '',) else None
        pot_reativa_kvar = _motor_float(pot_reativa_kvar, None) if pot_reativa_kvar not in (None, '',) else None
        pot_aparente_kva = _motor_float(pot_aparente_kva, None) if pot_aparente_kva not in (None, '',) else None
        eficiencia = _motor_float(eficiencia, None) if eficiencia not in (None, '',) else None

        if not pot_aparente_kva and fases == 3 and tensao_v > 0 and corrente_a > 0:
            pot_aparente_kva = (math.sqrt(3) * tensao_v * corrente_a) / 1000.0
        elif not pot_aparente_kva and fases == 1 and tensao_v > 0 and corrente_a > 0:
            pot_aparente_kva = (tensao_v * corrente_a) / 1000.0
        if not pot_ativa_kw and pot_aparente_kva and fp > 0:
            pot_ativa_kw = pot_aparente_kva * fp
        if not pot_reativa_kvar and pot_aparente_kva and (pot_ativa_kw is not None):
            pot_reativa_kvar = math.sqrt(max(pot_aparente_kva**2 - pot_ativa_kw**2, 0))

        energia_kwh = None
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('SELECT datahora, pot_ativa_kw FROM motor_medicoes WHERE equipamento_id=? ORDER BY datetime(datahora) DESC LIMIT 1', (equipamento_id,))
        last = c.fetchone()
        if last and pot_ativa_kw is not None:
            try:
                dt_last = datetime.fromisoformat(last[0].replace(' ', 'T'))
                dt_now = datetime.fromisoformat(datahora)
                dh = max((dt_now - dt_last).total_seconds() / 3600.0, 0)
                pot_media = ((_motor_float(last[1])) + pot_ativa_kw) / 2.0
                energia_kwh = pot_media * dh
            except Exception:
                energia_kwh = None

        avisos = []
        if fp and fp < 0.80: avisos.append(f'FP baixo: {fp:.3f}')
        if tensao_v and (tensao_v < 360 or tensao_v > 440): avisos.append(f'Tensão fora da faixa: {tensao_v:.1f} V')
        if avisos:
            observacoes = (observacoes + ' | ' if observacoes else '') + 'ALERTA MOTOR: ' + '; '.join(avisos)

        c.execute('''
            INSERT INTO motor_medicoes
            (equipamento_id, datahora, tensao_v, corrente_a, fator_potencia, frequencia_hz, fases,
             pot_ativa_kw, pot_reativa_kvar, pot_aparente_kva, eficiencia, energia_kwh, observacoes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (equipamento_id, datahora.replace('T',' '), tensao_v, corrente_a, fp, freq, fases,
              pot_ativa_kw, pot_reativa_kvar, pot_aparente_kva, eficiencia, energia_kwh, observacoes))
        conn.commit(); conn.close()
        if request.form.get('continuar') == '1':
            return redirect(url_for('motor_medir'))
        return redirect(url_for('motores_menu'))

    now = datetime.now().strftime('%Y-%m-%dT%H:%M')
    return render_template('motor_medicao_form.html', equipamentos=equipamentos, now=now)


@app.route('/motores/runs', methods=['GET'])
def motor_runs_page():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT e.id, e.nome, l.nome FROM equipamentos e LEFT JOIN locais l ON e.local_id=l.id ORDER BY l.nome, e.nome')
    equipamentos = c.fetchall()
    c.execute('''
        SELECT r.id, e.nome, COALESCE(l.nome,''), r.start_time, r.stop_time, r.duracao_min
        FROM motor_runs r
        LEFT JOIN equipamentos e ON e.id=r.equipamento_id
        LEFT JOIN locais l ON l.id=e.local_id
        ORDER BY r.id DESC LIMIT 50
    ''')
    runs = c.fetchall()
    conn.close()
    return render_template('motor_runs.html', equipamentos=equipamentos, runs=runs)


@app.route('/motores/run/start', methods=['POST'])
def motor_run_start():
    equipamento_id = int(request.form['equipamento_id'])
    agora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id FROM motor_runs WHERE equipamento_id=? AND stop_time IS NULL', (equipamento_id,))
    aberto = c.fetchone()
    if not aberto:
        c.execute('INSERT INTO motor_runs (equipamento_id, start_time) VALUES (?, ?)', (equipamento_id, agora))
        conn.commit()
    conn.close()
    return redirect(url_for('motor_runs_page'))


@app.route('/motores/run/stop', methods=['POST'])
def motor_run_stop():
    equipamento_id = int(request.form['equipamento_id'])
    agora_dt = datetime.now()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id, start_time FROM motor_runs WHERE equipamento_id=? AND stop_time IS NULL ORDER BY id DESC LIMIT 1', (equipamento_id,))
    row = c.fetchone()
    if row:
        run_id, start_str = row
        try:
            start_dt = datetime.fromisoformat(start_str)
            dur_min = max((agora_dt - start_dt).total_seconds() / 60.0, 0)
        except Exception:
            dur_min = 0
        c.execute('UPDATE motor_runs SET stop_time=?, duracao_min=? WHERE id=?',
                  (agora_dt.strftime('%Y-%m-%d %H:%M:%S'), dur_min, run_id))
        conn.commit()
    conn.close()
    return redirect(url_for('motor_runs_page'))


@app.route('/motores/graficos', methods=['GET'])
def motor_graficos():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT e.id, e.nome, l.nome FROM equipamentos e LEFT JOIN locais l ON e.local_id=l.id ORDER BY l.nome, e.nome')
    equipamentos = c.fetchall()
    conn.close()
    equip_id = request.args.get('equipamento_id', type=int)
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    series = []
    if equip_id:
        rows, _, _, _ = _motor_collect_stats(None, equip_id, data_ini, data_fim)
        # gráfico detalhado usa medições reais do motor
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('''
            SELECT datahora, tensao_v, corrente_a, fator_potencia, frequencia_hz,
                   pot_ativa_kw, pot_reativa_kvar, pot_aparente_kva
            FROM motor_medicoes
            WHERE equipamento_id=? AND datetime(datahora) BETWEEN datetime(?) AND datetime(?)
            ORDER BY datetime(datahora)
        ''', (equip_id, data_ini+' 00:00:00', data_fim+' 23:59:59'))
        series = c.fetchall(); conn.close()
    datas = [s[0] for s in series]
    return render_template('motor_graficos.html', equipamentos=equipamentos, equip_id=equip_id or '', data_ini=data_ini, data_fim=data_fim,
                           datas=datas, tensao=[s[1] for s in series], corrente=[s[2] for s in series], fp=[s[3] for s in series],
                           freq=[s[4] for s in series], p_kw=[s[5] for s in series], q_kvar=[s[6] for s in series], s_kva=[s[7] for s in series])


@app.route('/motores/relatorio', methods=['GET'])
def motor_relatorio():
    local_id = request.args.get('local_id', type=int)
    equip_id = request.args.get('equipamento_id', type=int)
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id, nome FROM locais ORDER BY nome')
    locais = c.fetchall()
    c.execute('SELECT id, nome FROM equipamentos ORDER BY nome')
    equipamentos_select = c.fetchall()
    conn.close()
    analises, resumo, data_ini, data_fim = _motor_collect_stats(local_id, equip_id, data_ini, data_fim)
    return render_template('motor_relatorio.html', locais=locais, equipamentos_select=equipamentos_select,
                           analises=analises, resumo=resumo, local_id=local_id or '', equip_id=equip_id or '',
                           data_ini=data_ini, data_fim=data_fim)



@app.route('/motores/detalhe/<int:equipamento_id>', methods=['GET'])
def motor_detalhe(equipamento_id):
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()

    analises, resumo, data_ini, data_fim = _motor_collect_stats(None, equipamento_id, data_ini, data_fim)
    analise = analises[0] if analises else None

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = _motor_dict_factory
    c = conn.cursor()
    c.execute("""
        SELECT e.*, l.nome AS local_nome,
               cfg.tensao_nominal, cfg.corrente_nominal, cfg.potencia_nominal_kw,
               cfg.fp_nominal, cfg.eficiencia_nominal, cfg.limite_corrente, cfg.limite_fp
        FROM equipamentos e
        LEFT JOIN locais l ON l.id=e.local_id
        LEFT JOIN equipamentos_cfg cfg ON cfg.equipamento_id=e.id
        WHERE e.id=?
    """, (equipamento_id,))
    equipamento = c.fetchone()
    if not equipamento:
        conn.close()
        flash('Equipamento não encontrado.', 'warning')
        return redirect(url_for('motores_menu'))

    c.execute("""
        SELECT datahora, tensao_v, corrente_a, fator_potencia, frequencia_hz,
               pot_ativa_kw, pot_reativa_kvar, pot_aparente_kva, energia_kwh, observacoes
        FROM motor_medicoes
        WHERE equipamento_id=? AND datetime(datahora) BETWEEN datetime(?) AND datetime(?)
        ORDER BY datetime(datahora) DESC LIMIT 25
    """, (equipamento_id, data_ini+' 00:00:00', data_fim+' 23:59:59'))
    medicoes = c.fetchall()

    c.execute("""
        SELECT id, start_time, stop_time, duracao_min
        FROM motor_runs
        WHERE equipamento_id=? AND datetime(start_time) BETWEEN datetime(?) AND datetime(?)
        ORDER BY datetime(start_time) DESC LIMIT 20
    """, (equipamento_id, data_ini+' 00:00:00', data_fim+' 23:59:59'))
    runs = c.fetchall()

    c.execute("""
        SELECT datahora, local, equipamento, energia_ativa, pot_ativa, fp, ponta, corrente, tensao, observacoes
        FROM leituras
        WHERE lower(COALESCE(equipamento,''))=lower(?)
          AND datetime(datahora) BETWEEN datetime(?) AND datetime(?)
        ORDER BY datetime(datahora) DESC LIMIT 25
    """, ((equipamento.get('nome') or ''), data_ini+' 00:00:00', data_fim+' 23:59:59'))
    leituras_monitoria = c.fetchall()
    conn.close()

    if not analise:
        analise = {
            'id': equipamento_id, 'equipamento': equipamento, 'status': 'Sem dados', 'alertas': [],
            'n': 0, 'kwh': 0, 'avg_fp': None, 'avg_corr': None, 'corr_max': 0, 'avg_kw': None,
            'kw_max': 0, 'ponta_max': 0, 'horas': 0, 'arranques': 0, 'tensao_min': None,
            'tensao_max': None, 'agua': 0, 'ce': None, 'carga_pct': None, 'last_ts': '', 'fontes': '—'
        }

    kvar_sugerido = _motor_kvar_compensacao(analise.get('avg_kw'), analise.get('avg_fp'), 0.92)
    recomendacoes = _motor_recomendacoes_detalhadas(analise)
    return render_template('motor_detalhe.html', equipamento=equipamento, analise=analise,
                           medicoes=medicoes, runs=runs, leituras_monitoria=leituras_monitoria,
                           data_ini=data_ini, data_fim=data_fim, kvar_sugerido=kvar_sugerido,
                           recomendacoes=recomendacoes)

@app.route('/motores/export/medicoes_csv')
def export_medicoes_csv():
    equip_id = request.args.get('equipamento_id', type=int)
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not equip_id:
        return Response('equipamento_id é obrigatório', status=400)
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''SELECT datahora,tensao_v,corrente_a,fator_potencia,frequencia_hz,pot_ativa_kw,pot_reativa_kvar,pot_aparente_kva,energia_kwh,observacoes
                 FROM motor_medicoes WHERE equipamento_id=? AND datetime(datahora) BETWEEN datetime(?) AND datetime(?) ORDER BY datetime(datahora)''',
              (equip_id, data_ini+' 00:00:00', data_fim+' 23:59:59'))
    rows = c.fetchall(); conn.close()
    si = StringIO(); w = csv.writer(si, delimiter=';')
    w.writerow(['datahora','tensao_v','corrente_a','fator_potencia','frequencia_hz','pot_ativa_kw','pot_reativa_kvar','pot_aparente_kva','energia_kwh','observacoes'])
    for r in rows: w.writerow(r)
    return Response(si.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment;filename=medicoes_motor_{equip_id}_{data_ini}_a_{data_fim}.csv'})


@app.route('/motores/export/runs_csv')
def export_runs_csv():
    equip_id = request.args.get('equipamento_id', type=int)
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not equip_id:
        return Response('equipamento_id é obrigatório', status=400)
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''SELECT id,start_time,stop_time,duracao_min FROM motor_runs WHERE equipamento_id=? AND datetime(start_time) BETWEEN datetime(?) AND datetime(?) ORDER BY datetime(start_time)''',
              (equip_id, data_ini+' 00:00:00', data_fim+' 23:59:59'))
    rows = c.fetchall(); conn.close()
    si = StringIO(); w = csv.writer(si, delimiter=';')
    w.writerow(['id','start_time','stop_time','duracao_min'])
    for r in rows: w.writerow(r)
    return Response(si.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment;filename=horas_motor_{equip_id}_{data_ini}_a_{data_fim}.csv'})




# === MOTORES - FECHO DO MÓDULO: MANUTENÇÃO, EXPORTAÇÃO E IMPRESSÃO ===

def _motor_contexto_detalhe(equipamento_id, data_ini=None, data_fim=None):
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    analises, resumo, data_ini, data_fim = _motor_collect_stats(None, equipamento_id, data_ini, data_fim)
    analise = analises[0] if analises else None

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = _motor_dict_factory
    c = conn.cursor()
    c.execute("""
        SELECT e.*, l.nome AS local_nome,
               cfg.tensao_nominal, cfg.corrente_nominal, cfg.potencia_nominal_kw,
               cfg.fp_nominal, cfg.eficiencia_nominal, cfg.limite_corrente, cfg.limite_fp
        FROM equipamentos e
        LEFT JOIN locais l ON l.id=e.local_id
        LEFT JOIN equipamentos_cfg cfg ON cfg.equipamento_id=e.id
        WHERE e.id=?
    """, (equipamento_id,))
    equipamento = c.fetchone()
    if not equipamento:
        conn.close()
        return None

    c.execute("""
        SELECT datahora, tensao_v, corrente_a, fator_potencia, frequencia_hz,
               pot_ativa_kw, pot_reativa_kvar, pot_aparente_kva, energia_kwh, observacoes
        FROM motor_medicoes
        WHERE equipamento_id=? AND datetime(datahora) BETWEEN datetime(?) AND datetime(?)
        ORDER BY datetime(datahora) DESC LIMIT 25
    """, (equipamento_id, data_ini+' 00:00:00', data_fim+' 23:59:59'))
    medicoes = c.fetchall()

    c.execute("""
        SELECT id, start_time, stop_time, duracao_min
        FROM motor_runs
        WHERE equipamento_id=? AND datetime(start_time) BETWEEN datetime(?) AND datetime(?)
        ORDER BY datetime(start_time) DESC LIMIT 20
    """, (equipamento_id, data_ini+' 00:00:00', data_fim+' 23:59:59'))
    runs = c.fetchall()

    c.execute("""
        SELECT datahora, local, equipamento, energia_ativa, pot_ativa, fp, ponta, corrente, tensao, observacoes
        FROM leituras
        WHERE lower(COALESCE(equipamento,''))=lower(?)
          AND datetime(datahora) BETWEEN datetime(?) AND datetime(?)
        ORDER BY datetime(datahora) DESC LIMIT 25
    """, ((equipamento.get('nome') or ''), data_ini+' 00:00:00', data_fim+' 23:59:59'))
    leituras_monitoria = c.fetchall()
    conn.close()

    if not analise:
        analise = {
            'id': equipamento_id, 'equipamento': equipamento, 'status': 'Sem dados', 'alertas': [],
            'n': 0, 'kwh': 0, 'avg_fp': None, 'avg_corr': None, 'corr_max': 0, 'avg_kw': None,
            'kw_max': 0, 'ponta_max': 0, 'horas': 0, 'arranques': 0, 'tensao_min': None,
            'tensao_max': None, 'agua': 0, 'ce': None, 'carga_pct': None, 'last_ts': '', 'fontes': '—'
        }
    kvar_sugerido = _motor_kvar_compensacao(analise.get('avg_kw'), analise.get('avg_fp'), 0.92)
    recomendacoes = _motor_recomendacoes_detalhadas(analise)
    return dict(equipamento=equipamento, analise=analise, medicoes=medicoes, runs=runs,
                leituras_monitoria=leituras_monitoria, data_ini=data_ini, data_fim=data_fim,
                kvar_sugerido=kvar_sugerido, recomendacoes=recomendacoes,
                gerado_em=datetime.now().strftime('%d/%m/%Y %H:%M'))


@app.route('/motores/manutencao', methods=['GET'])
def motor_manutencao():
    local_id = request.args.get('local_id', type=int)
    equip_id = request.args.get('equipamento_id', type=int)
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id, nome FROM locais ORDER BY nome')
    locais = c.fetchall()
    c.execute('SELECT id, nome FROM equipamentos ORDER BY nome')
    equipamentos_select = c.fetchall()
    conn.close()

    analises, resumo, data_ini, data_fim = _motor_collect_stats(local_id, equip_id, data_ini, data_fim)

    planos = []
    for a in analises:
        score = 0
        if a.get('status') == 'Crítico': score += 70
        elif a.get('status') == 'Atenção': score += 40
        elif a.get('status') == 'Sem dados': score += 15
        score += min(len(a.get('alertas') or []) * 8, 30)
        if a.get('avg_fp') is not None and a.get('avg_fp') < 0.80: score += 12
        if a.get('corr_max') and a.get('avg_corr') and a.get('corr_max') > max(a.get('avg_corr')*1.35, 1): score += 8
        if a.get('carga_pct') is not None and (a.get('carga_pct') < 35 or a.get('carga_pct') > 105): score += 10
        score = min(score, 100)
        if score >= 75:
            prioridade = 'Alta'
            prazo = '0–7 dias'
        elif score >= 45:
            prioridade = 'Média'
            prazo = '7–30 dias'
        else:
            prioridade = 'Baixa'
            prazo = 'Próxima ronda'
        planos.append(dict(a=a, score=score, prioridade=prioridade, prazo=prazo,
                           recomendacoes=_motor_recomendacoes_detalhadas(a)))
    planos.sort(key=lambda x: x['score'], reverse=True)
    return render_template('motor_manutencao.html', locais=locais, equipamentos_select=equipamentos_select,
                           analises=analises, planos=planos, resumo=resumo, local_id=local_id or '',
                           equip_id=equip_id or '', data_ini=data_ini, data_fim=data_fim,
                           gerado_em=datetime.now().strftime('%d/%m/%Y %H:%M'))


@app.route('/motores/detalhe/<int:equipamento_id>/imprimir', methods=['GET'])
def motor_detalhe_imprimir(equipamento_id):
    ctx = _motor_contexto_detalhe(equipamento_id, request.args.get('ini'), request.args.get('fim'))
    if not ctx:
        flash('Equipamento não encontrado.', 'warning')
        return redirect(url_for('motores_menu'))
    return render_template('motor_detalhe_print.html', **ctx)


@app.route('/motores/export/diagnostico_csv', methods=['GET'])
def export_diagnostico_motores_csv():
    local_id = request.args.get('local_id', type=int)
    equip_id = request.args.get('equipamento_id', type=int)
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    analises, resumo, data_ini, data_fim = _motor_collect_stats(local_id, equip_id, data_ini, data_fim)
    si = StringIO(); w = csv.writer(si, delimiter=';')
    w.writerow(['periodo_inicial','periodo_final','estado','equipamento','local','fontes','medicoes','energia_kwh','pot_media_kw','carga_pct','corrente_media_a','corrente_max_a','fp_medio','horas','arranques','tensao_min_v','tensao_max_v','alertas'])
    for a in analises:
        alertas = ' | '.join([f'{al[0]}: {al[1]}' for al in (a.get('alertas') or [])])
        eq = a.get('equipamento') or {}
        w.writerow([data_ini, data_fim, a.get('status'), eq.get('nome'), eq.get('local_nome'), a.get('fontes'), a.get('n'),
                    f"{a.get('kwh') or 0:.3f}", f"{a.get('avg_kw') or 0:.3f}" if a.get('avg_kw') is not None else '',
                    f"{a.get('carga_pct') or 0:.2f}" if a.get('carga_pct') is not None else '',
                    f"{a.get('avg_corr') or 0:.3f}" if a.get('avg_corr') is not None else '',
                    f"{a.get('corr_max') or 0:.3f}", f"{a.get('avg_fp') or 0:.4f}" if a.get('avg_fp') is not None else '',
                    f"{a.get('horas') or 0:.2f}", a.get('arranques') or 0,
                    f"{a.get('tensao_min') or 0:.2f}" if a.get('tensao_min') is not None else '',
                    f"{a.get('tensao_max') or 0:.2f}" if a.get('tensao_max') is not None else '', alertas])
    return Response(si.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment;filename=diagnostico_motores_{data_ini}_a_{data_fim}.csv'})


# === ALERTAS ===


def _ensure_alertas_perf_indexes():
    """Índices leves para acelerar o Centro de Alertas. Idempotente."""
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('CREATE INDEX IF NOT EXISTS idx_leituras_datahora_local ON leituras(datahora, local)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_leituras_local_equip ON leituras(local, equipamento)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_lm_data_local ON leituras_mensais(data, local)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_motor_medicoes_data_equip ON motor_medicoes(datahora, equipamento_id)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_motor_runs_equip_start ON motor_runs(equipamento_id, start_time)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_equipamentos_local ON equipamentos(local_id)')
        conn.commit(); conn.close()
    except Exception:
        try:
            if conn: conn.close()
        except Exception:
            pass

def _alertas_hash(*parts):
    import hashlib
    raw = '|'.join(str(p or '') for p in parts)
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()[:16]


def _ensure_alertas_acoes_schema():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS alertas_acoes (
            alerta_id TEXT PRIMARY KEY,
            estado TEXT DEFAULT 'Novo',
            responsavel TEXT,
            observacao TEXT,
            atualizado_em TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    existentes = {r[1] for r in c.execute('PRAGMA table_info(alertas_acoes)').fetchall()}
    extras = {
        'prazo': 'TEXT',
        'acao_tomada': 'TEXT',
        'fechado_em': 'TEXT',
        'prioridade_manual': 'TEXT',
        'evidencia': 'TEXT',
        'custo_estimado': 'REAL DEFAULT 0',
        'snapshot_nivel': 'TEXT',
        'snapshot_origem': 'TEXT',
        'snapshot_categoria': 'TEXT',
        'snapshot_local': 'TEXT',
        'snapshot_equipamento': 'TEXT',
        'snapshot_tipo': 'TEXT',
        'snapshot_causa': 'TEXT',
        'snapshot_impacto': 'TEXT',
        'snapshot_acao': 'TEXT',
        'snapshot_ultima': 'TEXT',
        'snapshot_link': 'TEXT',
        'manual': 'INTEGER DEFAULT 0'
    }
    for col, typ in extras.items():
        if col not in existentes:
            try:
                c.execute(f'ALTER TABLE alertas_acoes ADD COLUMN {col} {typ}')
            except Exception:
                pass
    try:
        c.execute('CREATE INDEX IF NOT EXISTS idx_alertas_acoes_estado ON alertas_acoes(estado)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_alertas_acoes_manual ON alertas_acoes(manual)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_alertas_acoes_atualizado ON alertas_acoes(atualizado_em)')
    except Exception:
        pass
    conn.commit(); conn.close()


def _load_alertas_acoes():
    _ensure_alertas_acoes_schema()
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    rows = conn.execute('SELECT * FROM alertas_acoes').fetchall()
    conn.close()
    return {r['alerta_id']: dict(r) for r in rows}




def _snapshot_to_event(row):
    """Converte alertas manuais/arquivados em eventos exibíveis, mesmo que a origem dinâmica já não gere o alerta."""
    return {
        'id': row.get('alerta_id') or '',
        'nivel': row.get('snapshot_nivel') or 'Informativo',
        'origem': row.get('snapshot_origem') or ('Manual / Operador' if row.get('manual') else 'Arquivo'),
        'local': row.get('snapshot_local') or '—',
        'equipamento': row.get('snapshot_equipamento') or '—',
        'tipo': row.get('snapshot_tipo') or 'Alerta registado',
        'causa': row.get('snapshot_causa') or 'Registo manual ou histórico preservado.',
        'impacto': row.get('snapshot_impacto') or 'Acompanhar impacto técnico/operacional.',
        'acao': row.get('snapshot_acao') or 'Definir e executar acção correctiva.',
        'ultima': row.get('snapshot_ultima') or row.get('atualizado_em') or '—',
        'impacto_mt': float(row.get('custo_estimado') or 0),
        'link': row.get('snapshot_link') or '',
        'estado': row.get('estado') or 'Novo',
        'responsavel': row.get('responsavel') or '',
        'observacao': row.get('observacao') or '',
        'acao_tomada': row.get('acao_tomada') or '',
        'prazo': row.get('prazo') or '',
        'fechado_em': row.get('fechado_em') or '',
        'atualizado_em': row.get('atualizado_em') or '',
        'evidencia': row.get('evidencia') or '',
        'categoria': row.get('snapshot_categoria') or _categoria_alerta(row.get('snapshot_tipo'), row.get('snapshot_origem')),
        'score': 0,
        'sla': 'Sem prazo',
        'prazo_sugerido': '',
        'manual': int(row.get('manual') or 0),
    }


def _alertas_saved_snapshot_rows(only_manual=False):
    _ensure_alertas_acoes_schema()
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    sql = 'SELECT * FROM alertas_acoes'
    params = []
    if only_manual:
        sql += ' WHERE COALESCE(manual,0)=1'
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def _fmt_mt(v):
    try:
        return f"{float(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') + ' MT'
    except Exception:
        return '0,00 MT'


def _nivel_peso(nivel):
    return {'Crítico': 0, 'Atenção': 1, 'Informativo': 2}.get(nivel, 9)


def _categoria_alerta(tipo, origem=''):
    t = (tipo or '').lower(); o = (origem or '').lower()
    if 'fp' in t or 'fator' in t or 'reativa' in t:
        return 'Energia reativa / FP'
    if 'tensão' in t or 'tensao' in t:
        return 'Qualidade de energia'
    if 'corrente' in t or 'sobrecarga' in t or 'arranque' in t or 'motor' in o:
        return 'Motores e cargas'
    if 'ponta' in t or 'demanda' in t:
        return 'Ponta / demanda'
    if 'específico' in t or 'agua' in t or 'água' in t:
        return 'Eficiência hidráulica'
    if 'fatura' in t or 'factura' in t:
        return 'Faturação'
    if 'leitura' in t:
        return 'Dados / leituras'
    return 'Operacional'


def _prazo_sugerido(nivel, tipo):
    hoje = datetime.now().date()
    t = (tipo or '').lower()
    if nivel == 'Crítico':
        dias = 2
    elif nivel == 'Atenção':
        dias = 7
    else:
        dias = 15
    if 'tensão' in t or 'corrente' in t or 'sobrecarga' in t:
        dias = min(dias, 3 if nivel == 'Crítico' else 5)
    return (hoje + timedelta(days=dias)).isoformat()


def _score_alerta(e):
    base = {'Crítico': 90, 'Atenção': 60, 'Informativo': 25}.get(e.get('nivel'), 10)
    tipo = (e.get('tipo') or '').lower()
    if any(x in tipo for x in ['reativa', 'ponta', 'fatura', 'factura']): base += 8
    if any(x in tipo for x in ['sobrecarga', 'tensão', 'corrente']): base += 10
    if e.get('estado') == 'Resolvido': base -= 60
    if e.get('estado') == 'Ignorado': base -= 50
    if e.get('estado') == 'Em análise': base -= 10
    return max(0, min(100, base))


def _classificar_sla(e):
    estado = e.get('estado') or 'Novo'
    if estado in ('Resolvido', 'Ignorado'):
        return 'Fechado'
    prazo = e.get('prazo') or e.get('prazo_sugerido') or ''
    try:
        d = datetime.strptime(prazo[:10], '%Y-%m-%d').date()
        hoje = datetime.now().date()
        if d < hoje:
            return 'Vencido'
        if (d - hoje).days <= 2:
            return 'A vencer'
        return 'No prazo'
    except Exception:
        return 'Sem prazo'


def _add_alerta_evento(eventos, nivel, origem, local, equipamento, tipo, causa, impacto, acao, ultima='—', impacto_mt=0, link=None, chave_extra=''):
    alerta_id = _alertas_hash(nivel, origem, local, equipamento, tipo, ultima, chave_extra)
    eventos.append({
        'id': alerta_id,
        'nivel': nivel,
        'origem': origem,
        'local': local or '—',
        'equipamento': equipamento or '—',
        'tipo': tipo,
        'causa': causa,
        'impacto': impacto,
        'acao': acao,
        'ultima': ultima or '—',
        'impacto_mt': float(impacto_mt or 0),
        'link': link or '',
        'estado': 'Novo',
        'responsavel': '',
        'observacao': '',
        'acao_tomada': '',
        'prazo': '',
        'fechado_em': '',
        'atualizado_em': '',
        'categoria': _categoria_alerta(tipo, origem),
        'score': 0,
        'sla': 'Sem prazo',
        'prazo_sugerido': '',
    })


def _collect_alertas_monitoria(local_nome=None, data_ini=None, data_fim=None):
    eventos = []
    try:
        conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
        c = conn.cursor()
        sql = """SELECT id, datahora, local, equipamento, energia_ativa, energia_reativa, pot_ativa, fp, ponta, caudal_elevada, corrente, tensao
                 FROM leituras WHERE date(substr(datahora,1,10)) BETWEEN date(?) AND date(?)"""
        params = [data_ini, data_fim]
        if local_nome:
            if isinstance(local_nome, (list, tuple, set)):
                nomes = [x for x in local_nome if str(x).strip()]
                if nomes:
                    sql += ' AND local IN (' + ','.join('?' for _ in nomes) + ')'; params.extend(nomes)
            else:
                sql += ' AND local=?'; params.append(local_nome)
        sql += ' ORDER BY datetime(datahora) DESC LIMIT 1500'
        rows = c.execute(sql, params).fetchall()
        conn.close()
        correntes = [float(r['corrente'] or 0) for r in rows if float(r['corrente'] or 0) > 0]
        pontas = [float(r['ponta'] or 0) for r in rows if float(r['ponta'] or 0) > 0]
        corrente_ref = (sum(correntes) / len(correntes) * 1.35) if correntes else 0
        ponta_ref = (sum(pontas) / len(pontas) * 1.40) if pontas else 0
        for r in rows:
            fp = float(r['fp'] or 0)
            tensao = float(r['tensao'] or 0)
            corrente = float(r['corrente'] or 0)
            ponta = float(r['ponta'] or 0)
            kwh = float(r['energia_ativa'] or 0)
            agua = float(r['caudal_elevada'] or 0)
            ult = r['datahora'] or '—'
            local = r['local'] or '—'; eq = r['equipamento'] or '—'
            if fp and fp < 0.80:
                _add_alerta_evento(eventos, 'Crítico', 'Monitoria Operacional', local, eq, 'Fator de potência muito baixo', 'FP medido abaixo de 0,80.', 'Aumenta perdas, aquecimento e risco de reativa excedente na instalação.', 'Verificar banco de capacitores, cargas em vazio e regime de operação.', ult, 0, '/monitoria', r['id'])
            elif fp and fp < 0.85:
                _add_alerta_evento(eventos, 'Atenção', 'Monitoria Operacional', local, eq, 'Fator de potência baixo', 'FP medido abaixo do limite operacional recomendado.', 'Pode contribuir para penalizações e baixa eficiência.', 'Acompanhar recorrência e avaliar necessidade de correção do FP.', ult, 0, '/monitoria', r['id'])
            if tensao and (tensao < 360 or tensao > 440):
                nivel = 'Crítico' if tensao < 340 or tensao > 460 else 'Atenção'
                _add_alerta_evento(eventos, nivel, 'Monitoria Operacional', local, eq, 'Tensão fora da faixa', f'Tensão registada: {tensao:.1f} V.', 'Pode provocar falhas, aquecimento ou disparos de proteção.', 'Confirmar medição por fase, queda de tensão, ligações e estado do transformador/quadro.', ult, 0, '/monitoria', r['id'])
            if corrente_ref and corrente > corrente_ref:
                _add_alerta_evento(eventos, 'Atenção', 'Monitoria Operacional', local, eq, 'Corrente acima do padrão', f'Corrente registada: {corrente:.1f} A.', 'Possível sobrecarga, desequilíbrio ou alteração mecânica na carga.', 'Comparar com corrente nominal do equipamento e inspecionar a carga.', ult, 0, '/monitoria', r['id'])
            if ponta_ref and ponta > ponta_ref:
                _add_alerta_evento(eventos, 'Atenção', 'Monitoria Operacional', local, eq, 'Ponta operacional elevada', f'Ponta registada: {ponta:.2f} kW.', 'Pode elevar a procura máxima mensal.', 'Avaliar arranques simultâneos e escalonamento de cargas.', ult, 0, '/monitoria', r['id'])
            if kwh > 0 and agua > 0 and (kwh/agua) > 5:
                _add_alerta_evento(eventos, 'Atenção', 'Monitoria Operacional', local, eq, 'Consumo específico elevado', f'{(kwh/agua):.3f} kWh/m³.', 'Indica possível queda de eficiência no sistema de bombagem.', 'Verificar caudal, pressão, filtros, válvulas e ponto de operação da bomba.', ult, 0, '/monitoria', r['id'])
    except Exception:
        pass
    return eventos


def _collect_alertas_mensais(local_nome=None, data_ini=None, data_fim=None):
    eventos = []
    try:
        ensure_faturas_mensais_archive_schema()
        conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
        c = conn.cursor()
        sql = """SELECT local, data, mes, ano, ativa, reativa, ponta, fp, diferenca, agua, esp, valor
                 FROM leituras_mensais WHERE date(data) BETWEEN date(?) AND date(?)"""
        params = [data_ini, data_fim]
        if local_nome:
            if isinstance(local_nome, (list, tuple, set)):
                nomes = [x for x in local_nome if str(x).strip()]
                if nomes:
                    sql += ' AND local IN (' + ','.join('?' for _ in nomes) + ')'; params.extend(nomes)
            else:
                sql += ' AND local=?'; params.append(local_nome)
        sql += ' ORDER BY date(data) DESC LIMIT 1500'
        rows = c.execute(sql, params).fetchall()
        fatur_sql = """SELECT local, mes, ano, total, kvarh_excedente, demanda_ponta_kw, consumo_especifico, atualizado_em
                             FROM faturas_mensais_arquivo"""
        fatur_params = []
        if local_nome:
            if isinstance(local_nome, (list, tuple, set)):
                nomes = [x for x in local_nome if str(x).strip()]
                if nomes:
                    fatur_sql += ' WHERE local IN (' + ','.join('?' for _ in nomes) + ')'; fatur_params.extend(nomes)
            else:
                fatur_sql += ' WHERE local=?'; fatur_params.append(local_nome)
        fatur_sql += ' ORDER BY atualizado_em DESC LIMIT 300'
        fatur = c.execute(fatur_sql, fatur_params).fetchall()
        conn.close()
        for r in rows:
            ult = r['data'] or '—'; local = r['local'] or '—'
            fp = float(r['fp'] or 0); dif = float(r['diferenca'] or 0); esp = float(r['esp'] or 0); agua = float(r['agua'] or 0)
            if dif < 0:
                _add_alerta_evento(eventos, 'Crítico', 'Leituras Mensais', local, 'Instalação', 'Leitura ativa decrescente', 'A leitura atual ficou inferior à leitura anterior.', 'Pode gerar fatura incorreta e distorcer energia ativa.', 'Conferir leitura do contador, fator multiplicativo e leitura do mês anterior.', ult, 0, '/leituras_mensal', f"{local}-{ult}-dif")
            if fp and fp < 0.80:
                _add_alerta_evento(eventos, 'Crítico', 'Leituras Mensais', local, 'Instalação', 'FP mensal diário muito baixo', f'FP = {fp:.3f}.', 'Risco direto de reativa excedente e perdas internas.', 'Avaliar compensação reativa e operação das cargas indutivas.', ult, 0, '/leituras_mensal', f"{local}-{ult}-fp")
            elif fp and fp < 0.85:
                _add_alerta_evento(eventos, 'Atenção', 'Leituras Mensais', local, 'Instalação', 'FP mensal diário baixo', f'FP = {fp:.3f}.', 'Pode aumentar reativa excedente no fecho do mês.', 'Monitorar e acionar plano de correção caso se repita.', ult, 0, '/leituras_mensal', f"{local}-{ult}-fp")
            if dif > 0 and agua <= 0:
                _add_alerta_evento(eventos, 'Informativo', 'Leituras Mensais', local, 'Instalação', 'Água não registada', 'Existe consumo de energia sem volume de água informado.', 'Impede cálculo confiável do consumo específico.', 'Preencher água elevada/produzida para análise energética.', ult, 0, '/leituras_mensal', f"{local}-{ult}-agua")
            if esp and esp > 5:
                _add_alerta_evento(eventos, 'Atenção', 'Leituras Mensais', local, 'Instalação', 'Consumo específico mensal elevado', f'{esp:.3f} kWh/m³.', 'Pode indicar operação fora do ponto eficiente ou perda hidráulica.', 'Comparar com histórico, verificar caudal, pressão, válvulas e bombas.', ult, 0, '/leituras_mensal', f"{local}-{ult}-esp")
        for f in fatur:
            if local_nome and f['local'] != local_nome:
                continue
            total = float(f['total'] or 0); kvar = float(f['kvarh_excedente'] or 0); ce = float(f['consumo_especifico'] or 0)
            ult = f"{str(f['mes']).zfill(2)}/{f['ano']}"
            if kvar > 0:
                _add_alerta_evento(eventos, 'Atenção', 'Fatura EDM', f['local'], 'Instalação', 'Reativa excedente faturada', f'Reativa excedente: {kvar:,.2f} kVArh.'.replace(',', 'X').replace('.', ',').replace('X','.'), 'Aumenta o valor da fatura e indica baixo fator de potência.', 'Verificar banco de capacitores e perfis de carga do período.', ult, 0, '/leituras_mensal/faturas', f"{f['local']}-{ult}-kvar")
            if total > 0:
                _add_alerta_evento(eventos, 'Informativo', 'Fatura EDM', f['local'], 'Instalação', 'Fatura arquivada', f'Total: {_fmt_mt(total)}.', 'Fatura disponível no arquivo para consulta e descarga.', 'Conferir valores e manter o mês fechado após validação.', ult, total, '/leituras_mensal/faturas', f"{f['local']}-{ult}-fat")
            if ce and ce > 5:
                _add_alerta_evento(eventos, 'Atenção', 'Fatura EDM', f['local'], 'Instalação', 'Consumo específico mensal elevado', f'{ce:.3f} kWh/m³.', 'Pode representar custo excessivo de bombagem.', 'Priorizar auditoria hidráulica e elétrica da instalação.', ult, 0, '/leituras_mensal/faturas', f"{f['local']}-{ult}-ce")
    except Exception:
        pass
    return eventos


def _collect_alertas_motores(local_id=None, data_ini=None, data_fim=None):
    eventos = []
    try:
        analises, resumo, _, _ = _motor_collect_stats(local_id, None, data_ini, data_fim)
        max_motor_alertas = 220
        for a in analises:
            if len(eventos) >= max_motor_alertas:
                break
            for nivel, titulo, acao in a.get('alertas') or []:
                titulo_norm = (titulo or '').lower()
                # Evita gerar centenas de alertas informativos "Sem medições", que deixavam a página lenta.
                if 'sem medi' in titulo_norm or 'sem dados' in titulo_norm:
                    continue
                eq = a.get('equipamento') or {}
                _add_alerta_evento(
                    eventos, nivel, 'Motores', eq.get('local_nome') or '—', eq.get('nome') or '—', titulo,
                    'Diagnóstico automático do desempenho electromecânico.',
                    'Pode afetar rendimento, disponibilidade, consumo energético ou manutenção.',
                    acao, a.get('last_ts') or '—', 0, f"/motores/detalhe/{a.get('id')}", a.get('id')
                )
                if len(eventos) >= max_motor_alertas:
                    break
        return eventos, resumo
    except Exception:
        return eventos, {'equipamentos': 0}


def _preparar_eventos_alertas(local_id=None, data_ini=None, data_fim=None, origem_filtro='', nivel_filtro='', estado_filtro='', categoria_filtro='', sla_filtro=''):
    _ensure_alertas_perf_indexes()
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    locais_rows = conn.execute('SELECT id, nome FROM locais ORDER BY nome').fetchall()
    conn.close()
    local_nome = ''
    local_ids_scope = []
    local_names_scope = []
    if local_id:
        local_ids_scope = get_descendant_local_ids(local_id, include_self=True)
        local_names_scope = get_local_names_for_ids(local_ids_scope)
        for l in locais_rows:
            if int(l['id']) == int(local_id):
                local_nome = l['nome']; break

    eventos = []
    motor_resumo = {'equipamentos': 0}
    # Coleta selectiva por origem: evita processar todos os módulos quando o utilizador filtra uma origem.
    if not origem_filtro or origem_filtro == 'Motores':
        motor_eventos, motor_resumo = _collect_alertas_motores(local_id, data_ini, data_fim)
        eventos.extend(motor_eventos)
    if not origem_filtro or origem_filtro == 'Monitoria Operacional':
        eventos.extend(_collect_alertas_monitoria((local_names_scope or None) if local_id else None, data_ini, data_fim))
    if not origem_filtro or origem_filtro in ('Leituras Mensais', 'Fatura EDM'):
        eventos.extend(_collect_alertas_mensais((local_names_scope or None) if local_id else None, data_ini, data_fim))

    acoes = _load_alertas_acoes()
    dyn_ids = {e.get('id') for e in eventos}
    # Acrescenta alertas manuais e alertas arquivados que já não aparecem nas origens dinâmicas,
    # preservando histórico e rastreabilidade operacional.
    for aid, row in acoes.items():
        if aid not in dyn_ids and (row.get('manual') or row.get('snapshot_tipo')):
            eventos.append(_snapshot_to_event(row))

    for e in eventos:
        st = acoes.get(e['id'], {})
        if st:
            e['estado'] = st.get('estado') or e.get('estado') or 'Novo'
            e['responsavel'] = st.get('responsavel') or ''
            e['observacao'] = st.get('observacao') or ''
            e['acao_tomada'] = st.get('acao_tomada') or ''
            e['prazo'] = st.get('prazo') or e.get('prazo') or ''
            e['fechado_em'] = st.get('fechado_em') or ''
            e['atualizado_em'] = st.get('atualizado_em') or ''
            e['evidencia'] = st.get('evidencia') or ''
            if st.get('custo_estimado') not in (None, ''):
                try: e['impacto_mt'] = float(st.get('custo_estimado') or e.get('impacto_mt') or 0)
                except Exception: pass
        e['categoria'] = e.get('categoria') or _categoria_alerta(e.get('tipo'), e.get('origem'))
        e['prazo_sugerido'] = _prazo_sugerido(e.get('nivel'), e.get('tipo'))
        if not e.get('prazo'):
            e['prazo'] = e['prazo_sugerido']
        e['sla'] = _classificar_sla(e)
        e['score'] = _score_alerta(e)

    if origem_filtro:
        eventos = [e for e in eventos if e['origem'] == origem_filtro]
    if nivel_filtro:
        eventos = [e for e in eventos if e['nivel'] == nivel_filtro]
    if estado_filtro:
        eventos = [e for e in eventos if e['estado'] == estado_filtro]
    if categoria_filtro:
        eventos = [e for e in eventos if e['categoria'] == categoria_filtro]
    if sla_filtro:
        eventos = [e for e in eventos if e['sla'] == sla_filtro]

    eventos.sort(key=lambda x: (_nivel_peso(x['nivel']), {'Vencido':0,'A vencer':1,'No prazo':2,'Sem prazo':3,'Fechado':4}.get(x.get('sla','Sem prazo'),9), -x.get('score',0), {'Novo':0,'Em análise':1,'Resolvido':2,'Ignorado':3}.get(x.get('estado','Novo'),9)))
    # Protecção de desempenho: a tela principal mostra alertas prioritários em vez de renderizar milhares de linhas.
    if len(eventos) > 650:
        eventos = eventos[:650]
    resumo = {
        'total': len(eventos),
        'criticos': sum(1 for e in eventos if e['nivel'] == 'Crítico'),
        'atencao': sum(1 for e in eventos if e['nivel'] == 'Atenção'),
        'informativos': sum(1 for e in eventos if e['nivel'] == 'Informativo'),
        'novos': sum(1 for e in eventos if e['estado'] == 'Novo'),
        'analise': sum(1 for e in eventos if e['estado'] == 'Em análise'),
        'resolvidos': sum(1 for e in eventos if e['estado'] == 'Resolvido'),
        'ignorados': sum(1 for e in eventos if e['estado'] == 'Ignorado'),
        'pendentes': sum(1 for e in eventos if e['estado'] not in ('Resolvido','Ignorado')),
        'vencidos': sum(1 for e in eventos if e['sla'] == 'Vencido'),
        'a_vencer': sum(1 for e in eventos if e['sla'] == 'A vencer'),
        'impacto_mt': sum(float(e.get('impacto_mt') or 0) for e in eventos),
        'equipamentos': motor_resumo.get('equipamentos', 0) if isinstance(motor_resumo, dict) else 0,
    }
    resumo['taxa_resolucao'] = round((resumo['resolvidos'] / resumo['total'] * 100), 1) if resumo['total'] else 0
    if resumo['criticos'] > 0 or resumo['vencidos'] > 0:
        estado_geral = 'Crítico'
    elif resumo['atencao'] > 0 or resumo['a_vencer'] > 0:
        estado_geral = 'Atenção'
    elif resumo['total'] > 0:
        estado_geral = 'Informativo'
    else:
        estado_geral = 'Normal'

    origem_counts, categoria_counts, sla_counts, local_counts = {}, {}, {}, {}
    for e in eventos:
        origem_counts[e['origem']] = origem_counts.get(e['origem'], 0) + 1
        categoria_counts[e['categoria']] = categoria_counts.get(e['categoria'], 0) + 1
        sla_counts[e['sla']] = sla_counts.get(e['sla'], 0) + 1
        local_counts[e['local']] = local_counts.get(e['local'], 0) + 1
    ranking_locais = sorted(local_counts.items(), key=lambda x: x[1], reverse=True)[:8]
    return eventos, resumo, locais_rows, estado_geral, origem_counts, categoria_counts, sla_counts, ranking_locais




def _alertas_request_filters():
    local_id = request.args.get('local_id', type=int)
    origem_filtro = request.args.get('origem', '').strip()
    nivel_filtro = request.args.get('nivel', '').strip()
    estado_filtro = request.args.get('estado', '').strip()
    categoria_filtro = request.args.get('categoria', '').strip()
    sla_filtro = request.args.get('sla', '').strip()
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    return local_id, origem_filtro, nivel_filtro, estado_filtro, categoria_filtro, sla_filtro, data_ini, data_fim


@app.route('/alertas/kanban')
def alertas_kanban():
    local_id, origem_filtro, nivel_filtro, estado_filtro, categoria_filtro, sla_filtro, data_ini, data_fim = _alertas_request_filters()
    eventos, resumo, locais_rows, estado_geral, origem_counts, categoria_counts, sla_counts, ranking_locais = _preparar_eventos_alertas(local_id, data_ini, data_fim, origem_filtro, nivel_filtro, '', categoria_filtro, sla_filtro)
    colunas = {k: [] for k in ['Novo','Em análise','Resolvido','Ignorado']}
    for e in eventos:
        colunas.setdefault(e.get('estado') or 'Novo', []).append(e)
    return render_template('alertas_kanban.html', colunas=colunas, eventos=eventos, resumo=resumo,
                           data_ini=data_ini, data_fim=data_fim, estado_geral=estado_geral,
                           local_id=local_id or '', origem_filtro=origem_filtro, nivel_filtro=nivel_filtro,
                           categoria_filtro=categoria_filtro, sla_filtro=sla_filtro)


@app.route('/alertas/historico')
def alertas_historico():
    _ensure_alertas_acoes_schema()
    estado = request.args.get('estado','').strip()
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    sql = 'SELECT * FROM alertas_acoes'
    params = []
    if estado:
        sql += ' WHERE estado=?'; params.append(estado)
    sql += ' ORDER BY COALESCE(atualizado_em, fechado_em, prazo) DESC'
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    eventos = [_snapshot_to_event(r) for r in rows]
    for e in eventos:
        e['categoria'] = e.get('categoria') or _categoria_alerta(e.get('tipo'), e.get('origem'))
        e['prazo_sugerido'] = _prazo_sugerido(e.get('nivel'), e.get('tipo'))
        if not e.get('prazo'): e['prazo'] = e['prazo_sugerido']
        e['sla'] = _classificar_sla(e)
        e['score'] = _score_alerta(e)
    return render_template('alertas_historico.html', eventos=eventos, estado=estado)


@app.route('/alertas/manual', methods=['POST'])
def alertas_manual():
    _ensure_alertas_acoes_schema()
    nivel = request.form.get('nivel','Atenção').strip() or 'Atenção'
    local = request.form.get('local','').strip() or '—'
    equipamento = request.form.get('equipamento','').strip() or '—'
    tipo = request.form.get('tipo','Alerta manual').strip() or 'Alerta manual'
    causa = request.form.get('causa','Registo manual do operador.').strip() or 'Registo manual do operador.'
    impacto = request.form.get('impacto','Impacto operacional a acompanhar.').strip() or 'Impacto operacional a acompanhar.'
    acao = request.form.get('acao','Avaliar e executar acção correctiva.').strip() or 'Avaliar e executar acção correctiva.'
    responsavel = request.form.get('responsavel','').strip()
    prazo = request.form.get('prazo','').strip() or _prazo_sugerido(nivel, tipo)
    custo = request.form.get('custo_estimado','0').replace(',','.')
    try: custo = float(custo or 0)
    except Exception: custo = 0.0
    aid = 'manual_' + _alertas_hash(datetime.now().isoformat(), nivel, local, equipamento, tipo)
    categoria = _categoria_alerta(tipo, 'Manual / Operador')
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""INSERT INTO alertas_acoes(alerta_id, estado, responsavel, observacao, acao_tomada, prazo, atualizado_em,
                 evidencia, custo_estimado, snapshot_nivel, snapshot_origem, snapshot_categoria, snapshot_local,
                 snapshot_equipamento, snapshot_tipo, snapshot_causa, snapshot_impacto, snapshot_acao, snapshot_ultima,
                 snapshot_link, manual)
                 VALUES(?,?,?,?,?,?,datetime('now','localtime'),?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",
              (aid, 'Novo', responsavel, '', '', prazo, '', custo, nivel, 'Manual / Operador', categoria, local,
               equipamento, tipo, causa, impacto, acao, datetime.now().strftime('%Y-%m-%d %H:%M'), '',))
    conn.commit(); conn.close()
    return redirect(request.form.get('next') or url_for('alertas'))

@app.route('/alertas', methods=['GET'])
def alertas():
    local_id = request.args.get('local_id', type=int)
    origem_filtro = request.args.get('origem', '').strip()
    nivel_filtro = request.args.get('nivel', '').strip()
    estado_filtro = request.args.get('estado', '').strip()
    categoria_filtro = request.args.get('categoria', '').strip()
    sla_filtro = request.args.get('sla', '').strip()
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    eventos, resumo, locais_rows, estado_geral, origem_counts, categoria_counts, sla_counts, ranking_locais = _preparar_eventos_alertas(local_id, data_ini, data_fim, origem_filtro, nivel_filtro, estado_filtro, categoria_filtro, sla_filtro)
    return render_template('alertas.html', eventos=eventos, resumo=resumo, locais=locais_rows,
                           local_id=local_id or '', data_ini=data_ini, data_fim=data_fim,
                           estado_geral=estado_geral, origem_counts=origem_counts,
                           categoria_counts=categoria_counts, sla_counts=sla_counts,
                           ranking_locais=ranking_locais,
                           origem_filtro=origem_filtro, nivel_filtro=nivel_filtro,
                           estado_filtro=estado_filtro, categoria_filtro=categoria_filtro, sla_filtro=sla_filtro)


@app.route('/alertas/acao', methods=['POST'])
def alertas_acao():
    _ensure_alertas_acoes_schema()
    alerta_id = request.form.get('alerta_id', '').strip()
    estado = request.form.get('estado', 'Em análise').strip()
    responsavel = request.form.get('responsavel', '').strip()
    observacao = request.form.get('observacao', '').strip()
    acao_tomada = request.form.get('acao_tomada', '').strip()
    evidencia = request.form.get('evidencia', '').strip()
    prazo = request.form.get('prazo', '').strip()
    custo_raw = request.form.get('custo_estimado', '').strip().replace(',','.')
    try: custo_estimado = float(custo_raw) if custo_raw != '' else 0.0
    except Exception: custo_estimado = 0.0
    snap = {k: request.form.get(k, '').strip() for k in ['snapshot_nivel','snapshot_origem','snapshot_categoria','snapshot_local','snapshot_equipamento','snapshot_tipo','snapshot_causa','snapshot_impacto','snapshot_acao','snapshot_ultima','snapshot_link']}
    next_url = request.form.get('next') or url_for('alertas')
    if not alerta_id:
        return redirect(next_url)
    if estado not in ('Novo', 'Em análise', 'Resolvido', 'Ignorado'):
        estado = 'Em análise'
    fechado_em = "datetime('now','localtime')" if estado in ('Resolvido','Ignorado') else 'NULL'
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f"""INSERT INTO alertas_acoes(alerta_id, estado, responsavel, observacao, acao_tomada, prazo, fechado_em, atualizado_em,
                    evidencia, custo_estimado, snapshot_nivel, snapshot_origem, snapshot_categoria, snapshot_local, snapshot_equipamento,
                    snapshot_tipo, snapshot_causa, snapshot_impacto, snapshot_acao, snapshot_ultima, snapshot_link)
                 VALUES(?,?,?,?,?,?,{fechado_em},datetime('now','localtime'),?,?,?,?,?,?,?,?,?,?,?,?,?)
                 ON CONFLICT(alerta_id) DO UPDATE SET
                    estado=excluded.estado,
                    responsavel=excluded.responsavel,
                    observacao=excluded.observacao,
                    acao_tomada=excluded.acao_tomada,
                    prazo=excluded.prazo,
                    fechado_em={fechado_em},
                    evidencia=excluded.evidencia,
                    custo_estimado=excluded.custo_estimado,
                    snapshot_nivel=CASE WHEN excluded.snapshot_nivel!='' THEN excluded.snapshot_nivel ELSE alertas_acoes.snapshot_nivel END,
                    snapshot_origem=CASE WHEN excluded.snapshot_origem!='' THEN excluded.snapshot_origem ELSE alertas_acoes.snapshot_origem END,
                    snapshot_categoria=CASE WHEN excluded.snapshot_categoria!='' THEN excluded.snapshot_categoria ELSE alertas_acoes.snapshot_categoria END,
                    snapshot_local=CASE WHEN excluded.snapshot_local!='' THEN excluded.snapshot_local ELSE alertas_acoes.snapshot_local END,
                    snapshot_equipamento=CASE WHEN excluded.snapshot_equipamento!='' THEN excluded.snapshot_equipamento ELSE alertas_acoes.snapshot_equipamento END,
                    snapshot_tipo=CASE WHEN excluded.snapshot_tipo!='' THEN excluded.snapshot_tipo ELSE alertas_acoes.snapshot_tipo END,
                    snapshot_causa=CASE WHEN excluded.snapshot_causa!='' THEN excluded.snapshot_causa ELSE alertas_acoes.snapshot_causa END,
                    snapshot_impacto=CASE WHEN excluded.snapshot_impacto!='' THEN excluded.snapshot_impacto ELSE alertas_acoes.snapshot_impacto END,
                    snapshot_acao=CASE WHEN excluded.snapshot_acao!='' THEN excluded.snapshot_acao ELSE alertas_acoes.snapshot_acao END,
                    snapshot_ultima=CASE WHEN excluded.snapshot_ultima!='' THEN excluded.snapshot_ultima ELSE alertas_acoes.snapshot_ultima END,
                    snapshot_link=CASE WHEN excluded.snapshot_link!='' THEN excluded.snapshot_link ELSE alertas_acoes.snapshot_link END,
                    atualizado_em=datetime('now','localtime')""",
              (alerta_id, estado, responsavel, observacao, acao_tomada, prazo, evidencia, custo_estimado,
               snap['snapshot_nivel'], snap['snapshot_origem'], snap['snapshot_categoria'], snap['snapshot_local'], snap['snapshot_equipamento'],
               snap['snapshot_tipo'], snap['snapshot_causa'], snap['snapshot_impacto'], snap['snapshot_acao'], snap['snapshot_ultima'], snap['snapshot_link']))
    conn.commit(); conn.close()
    return redirect(next_url)


@app.route('/alertas/acao_lote', methods=['POST'])
def alertas_acao_lote():
    _ensure_alertas_acoes_schema()
    ids = request.form.getlist('alerta_ids')
    estado = request.form.get('estado_lote', 'Em análise')
    responsavel = request.form.get('responsavel_lote', '').strip()
    observacao = request.form.get('observacao_lote', '').strip()
    next_url = request.form.get('next') or url_for('alertas')
    if estado not in ('Novo','Em análise','Resolvido','Ignorado'):
        estado = 'Em análise'
    if ids:
        fechado_expr = "datetime('now','localtime')" if estado in ('Resolvido','Ignorado') else 'NULL'
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        for alerta_id in ids:
            c.execute(f"""INSERT INTO alertas_acoes(alerta_id, estado, responsavel, observacao, atualizado_em, fechado_em)
                         VALUES(?,?,?,?,datetime('now','localtime'),{fechado_expr})
                         ON CONFLICT(alerta_id) DO UPDATE SET
                            estado=excluded.estado,
                            responsavel=COALESCE(NULLIF(excluded.responsavel,''), alertas_acoes.responsavel),
                            observacao=CASE WHEN excluded.observacao!='' THEN excluded.observacao ELSE alertas_acoes.observacao END,
                            atualizado_em=datetime('now','localtime'),
                            fechado_em={fechado_expr}""",
                      (alerta_id, estado, responsavel, observacao))
        conn.commit(); conn.close()
    return redirect(next_url)


@app.route('/alertas/relatorio')
def alertas_relatorio():
    local_id = request.args.get('local_id', type=int)
    origem_filtro = request.args.get('origem', '').strip()
    nivel_filtro = request.args.get('nivel', '').strip()
    estado_filtro = request.args.get('estado', '').strip()
    categoria_filtro = request.args.get('categoria', '').strip()
    sla_filtro = request.args.get('sla', '').strip()
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    eventos, resumo, locais_rows, estado_geral, origem_counts, categoria_counts, sla_counts, ranking_locais = _preparar_eventos_alertas(local_id, data_ini, data_fim, origem_filtro, nivel_filtro, estado_filtro, categoria_filtro, sla_filtro)
    return render_template('alertas_relatorio.html', eventos=eventos[:80], resumo=resumo, locais=locais_rows,
                           data_ini=data_ini, data_fim=data_fim, estado_geral=estado_geral,
                           origem_counts=origem_counts, categoria_counts=categoria_counts,
                           sla_counts=sla_counts, ranking_locais=ranking_locais, gerado_em=datetime.now().strftime('%d/%m/%Y %H:%M'))


@app.route('/alertas/export/csv')
def alertas_export_csv():
    local_id = request.args.get('local_id', type=int)
    origem_filtro = request.args.get('origem', '').strip()
    nivel_filtro = request.args.get('nivel', '').strip()
    estado_filtro = request.args.get('estado', '').strip()
    categoria_filtro = request.args.get('categoria', '').strip()
    sla_filtro = request.args.get('sla', '').strip()
    data_ini = request.args.get('ini')
    data_fim = request.args.get('fim')
    if not data_ini or not data_fim:
        data_ini, data_fim = _motor_intervalo_padrao()
    eventos, resumo, *_ = _preparar_eventos_alertas(local_id, data_ini, data_fim, origem_filtro, nivel_filtro, estado_filtro, categoria_filtro, sla_filtro)
    si = StringIO(); w = csv.writer(si, delimiter=';')
    w.writerow(['id','nivel','score','sla','prazo','estado','origem','categoria','local','equipamento','tipo','causa','impacto','acao_recomendada','acao_tomada','evidencia','ultima','responsavel','observacao','impacto_mt'])
    for e in eventos:
        w.writerow([e['id'], e['nivel'], e.get('score'), e.get('sla'), e.get('prazo'), e['estado'], e['origem'], e.get('categoria',''), e['local'], e['equipamento'], e['tipo'], e['causa'], e['impacto'], e['acao'], e.get('acao_tomada',''), e.get('evidencia',''), e['ultima'], e.get('responsavel',''), e.get('observacao',''), f"{float(e.get('impacto_mt') or 0):.2f}"])
    return Response(si.getvalue(), mimetype='text/csv; charset=utf-8', headers={'Content-Disposition': f'attachment;filename=alertas_sge_{data_ini}_a_{data_fim}.csv'})


# ==============================
# === DIMENSIONAMENTO SOLAR ===
# ==============================
# (coloca este bloco só uma vez no ficheiro)

@app.route('/solar', methods=['GET'])
def solar_home():
    """Centro unificado de Energia Solar.
    Mantém o dimensionamento FV e lâmpadas solares como submódulos, sem apagar as rotas antigas.
    """
    stats = {
        'projetos_fv': 0,
        'projetos_lampadas': 0,
        'itens_catalogo': 0,
        'kwp_total': 0.0,
        'economia_mensal_total': 0.0,
        'co2_total': 0.0,
        'ultimos_fv': [],
        'ultimos_lampadas': []
    }
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        try:
            row = c.execute("SELECT COUNT(*) AS n FROM solar_projetos").fetchone()
            stats['projetos_fv'] = int(row['n'] or 0) if row else 0
            rows = c.execute("SELECT id, created_at, nome_projeto, resultado_json FROM solar_projetos ORDER BY id DESC LIMIT 5").fetchall()
            for r in rows:
                try:
                    data = json.loads(r['resultado_json'] or '{}')
                except Exception:
                    data = {}
                stats['ultimos_fv'].append({
                    'id': r['id'],
                    'created_at': r['created_at'],
                    'nome': r['nome_projeto'] or data.get('local_nome') or 'Projeto fotovoltaico',
                    'kwp': float(data.get('kwp_real') or 0),
                    'payback': data.get('payback_anos'),
                    'economia': float(data.get('economia_mensal') or 0)
                })
                stats['kwp_total'] += float(data.get('kwp_real') or 0)
                stats['economia_mensal_total'] += float(data.get('economia_mensal') or 0)
                stats['co2_total'] += float(data.get('co2_t_ano') or 0)
        except Exception:
            pass
        try:
            row = c.execute("SELECT COUNT(*) AS n FROM solar_lampadas").fetchone()
            stats['projetos_lampadas'] = int(row['n'] or 0) if row else 0
            rows = c.execute("SELECT id, created_at, nome, resultado_json FROM solar_lampadas ORDER BY id DESC LIMIT 5").fetchall()
            for r in rows:
                try:
                    data = json.loads(r['resultado_json'] or '{}')
                except Exception:
                    data = {}
                stats['ultimos_lampadas'].append({
                    'id': r['id'],
                    'created_at': r['created_at'],
                    'nome': r['nome'] or 'Projeto de iluminação solar',
                    'painel_wp': data.get('painel_wp'),
                    'bateria_wh': data.get('bateria_wh_bruto'),
                    'capex': data.get('capex_estimado')
                })
        except Exception:
            pass
        try:
            row = c.execute("SELECT COUNT(*) AS n FROM solar_lampadas_catalogo").fetchone()
            stats['itens_catalogo'] = int(row['n'] or 0) if row else 0
        except Exception:
            pass
        conn.close()
    except Exception:
        pass
    return render_template('solar_home.html', stats=stats)



# === Energia Solar Expert: consumo robusto a partir das Leituras Mensais ===
def solar_consumo_mensal_robusto(local_nome, mes, ano, fator_mult=1.0):
    """Calcula consumo mensal para dimensionamento solar sem duplicar o factor multiplicativo."""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT data, ativa, anterior, atual, diferenca
              FROM leituras_mensais
             WHERE local=? AND mes=? AND ano=?
             ORDER BY data ASC
        """, (local_nome, str(mes).zfill(2), int(ano))).fetchall()
        conn.close()
    except Exception:
        return 0.0, 0, 'sem dados'
    if not rows:
        return 0.0, 0, 'sem dados'
    def fv(x, default=None):
        try:
            if x is None or x == '': return default
            return float(str(x).replace(',', '.'))
        except Exception:
            return default
    difs_valid = []
    for r in rows:
        d = fv(r['diferenca'], None)
        if d is not None and d >= 0:
            difs_valid.append(d)
    soma_dif = sum(difs_valid)
    if soma_dif > 0:
        return float(soma_dif), len(difs_valid), 'soma das diferenças diárias faturáveis'
    bases = []
    atuais = []
    for r in rows:
        b = fv(r['anterior'], None)
        a = fv(r['atual'], None)
        if b is not None: bases.append(b)
        if a is not None: atuais.append(a)
    if bases and atuais and max(atuais) >= bases[0]:
        return float(max(atuais) - bases[0]), len(atuais), 'diferença leitura final - leitura base'
    ativas = []
    for r in rows:
        a = fv(r['ativa'], None)
        if a is not None: ativas.append(a)
    if ativas:
        return float(sum(ativas) * float(fator_mult or 1.0)), len(ativas), 'fallback: soma ativa lida x factor multiplicativo'
    return 0.0, 0, 'sem dados úteis'


def solar_irr(fluxos, low=-0.9, high=1.5, iterations=80):
    """IRR simples por bissecção; retorna None quando não há raiz coerente."""
    def npv_rate(rate):
        total = 0.0
        for i, cf in enumerate(fluxos):
            total += cf / ((1 + rate) ** i)
        return total
    try:
        f_low, f_high = npv_rate(low), npv_rate(high)
        if f_low * f_high > 0:
            return None
        for _ in range(iterations):
            mid = (low + high) / 2
            f_mid = npv_rate(mid)
            if abs(f_mid) < 1e-6:
                return mid
            if f_low * f_mid <= 0:
                high = mid; f_high = f_mid
            else:
                low = mid; f_low = f_mid
        return (low + high) / 2
    except Exception:
        return None
@app.route('/solar/dimensionamento', methods=['GET'])
@app.route('/solar/fotovoltaico', methods=['GET'])
def solar_form():
    # precisa existir a lista de locais
    locais = get_locais()
    hoje = datetime.now()
    default_periodo = hoje.strftime('%Y-%m')

    defaults = {
        "psh": 5.0,
        "derate": 0.77,
        "panel_wp": 550,
        "panel_area": 2.2,
        "tarifa_kwh": 4.78,
        "inv_dcac": 1.2,
        "autonomy_days": 1.0,
        "battery_dod": 0.8,
        "battery_eff": 0.9,
        "system_voltage": 48,
        "battery_module_kwh": 5.12,
        "capex_kwp": 90000.0,
        "opex_pct": 1.0,
        "tarifa_esc": 0.0,
        "desconto": 10.0,
        "anos_analise": 20,
        "co2_factor": 0.6,
        "cobertura_pct": 80.0,
        "autoconsumo_pct": 90.0,
        "crescimento_carga_pct": 0.0,
        "pico_kw": 0.0,
        "reserva_inversor_pct": 15.0,
        "area_disponivel": 0.0,
        "sombreamento_pct": 0.0,
        "perdas_cabos_pct": 2.0,
        "perdas_sujidade_pct": 3.0,
        "perdas_temp_pct": 8.0,
    }

    perfil_sazonal = [1.00, 1.02, 1.05, 1.08, 1.10, 1.05,
                      0.98, 0.95, 0.97, 0.99, 1.01, 1.03]

    # ATENÇÃO: o template solar.html deve usar {{ url_for('solar_projetos') }}
    # e como esta rota está logo abaixo, agora o Flask já conhece o endpoint
    return render_template(
        'solar.html',
        locais=locais,
        periodo=default_periodo,
        defaults=defaults,
        perfil_sazonal=perfil_sazonal
    )


@app.route('/solar/calcular', methods=['POST'])
def solar_calcular():
    def f(nome, default=0.0):
        v = request.form.get(nome, "")
        if v is None:
            return float(default)
        v = str(v).strip().replace(' ', '').replace(',', '.')
        if v == "":
            return float(default)
        try:
            return float(v)
        except ValueError:
            return float(default)

    def clamp(v, lo, hi):
        return max(lo, min(hi, v))

    modo = request.form.get('modo', 'manual')
    local_id = request.form.get('local_id')
    periodo = request.form.get('periodo')
    tipo = request.form.get('tipo_sistema', 'ongrid')

    psh = max(f('psh', 5.0), 0.1)
    derate = clamp(f('derate', 0.77), 0.30, 0.98)
    panel_wp = max(f('panel_wp', 550), 1.0)
    panel_area = max(f('panel_area', 2.2), 0.1)
    inv_dcac = max(f('inv_dcac', 1.2), 0.1)
    fator_mult = max(f('fator_mult', 1.0), 0.0001)
    cobertura_pct = clamp(f('cobertura_pct', 80.0), 1.0, 150.0)
    autoconsumo_pct = clamp(f('autoconsumo_pct', 90.0), 0.0, 100.0)
    crescimento_carga_pct = f('crescimento_carga_pct', 0.0)
    pico_kw = max(f('pico_kw', 0.0), 0.0)
    reserva_inversor_pct = clamp(f('reserva_inversor_pct', 15.0), 0.0, 100.0)
    area_disponivel = max(f('area_disponivel', 0.0), 0.0)
    sombreamento_pct = clamp(f('sombreamento_pct', 0.0), 0.0, 60.0)
    perdas_cabos_pct = clamp(f('perdas_cabos_pct', 2.0), 0.0, 15.0)
    perdas_sujidade_pct = clamp(f('perdas_sujidade_pct', 3.0), 0.0, 20.0)
    perdas_temp_pct = clamp(f('perdas_temp_pct', 8.0), 0.0, 30.0)
    derate_expert = derate * (1 - sombreamento_pct/100.0) * (1 - perdas_cabos_pct/100.0) * (1 - perdas_sujidade_pct/100.0) * (1 - perdas_temp_pct/100.0)
    derate_expert = clamp(derate_expert, 0.20, 0.98)

    tarifa_kwh = f('tarifa_kwh', 4.78)
    capex_kwp = f('capex_kwp', 90000)
    opex_pct = f('opex_pct', 1.0)
    tarifa_esc = f('tarifa_esc', 0.0)
    desconto = f('desconto', 10.0)
    anos_analise = int(max(1, f('anos_analise', 20)))
    co2_factor = f('co2_factor', 0.6)

    autonomy_days = f('autonomy_days', 1.0)
    battery_dod = clamp(f('battery_dod', 0.8), 0.05, 1.0)
    battery_eff = clamp(f('battery_eff', 0.9), 0.05, 1.0)
    system_voltage = f('system_voltage', 48)
    battery_module_kwh = max(f('battery_module_kwh', 5.12), 0.1)

    perfil_sazonal_raw = request.form.get('perfil_sazonal_json')
    try:
        perfil_sazonal = json.loads(perfil_sazonal_raw) if perfil_sazonal_raw else None
    except Exception:
        perfil_sazonal = None
    if not perfil_sazonal or len(perfil_sazonal) != 12:
        perfil_sazonal = [1.00,1.02,1.05,1.08,1.10,1.05,0.98,0.95,0.97,0.99,1.01,1.03]

    consumo_metodo = 'manual'
    local_nome = request.form.get('local_nome_manual', 'Sem Local') or 'Sem Local'
    mes = ano = None
    dias_utilizados = None
    if modo == 'manual':
        daily_kwh_base = max(f('daily_kwh', 0), 0.0)
        total_mes_kwh = daily_kwh_base * 30.0
    else:
        total_mes_kwh = 0.0
        daily_kwh_base = 0.0
        dias_utilizados = 0
        if local_id and periodo:
            local_row = get_local_by_id(int(local_id))
            local_nome = local_row[1] if local_row else 'Sem Local'
            try:
                cfg = get_local_cfg_full(int(local_id))
                fator_mult = float(cfg.get('fator_mult') or fator_mult)
                tarifa_kwh = float(cfg.get('tarifa_ativa') or tarifa_kwh)
                if not pico_kw:
                    pico_kw = float(cfg.get('pot_contratada') or 0)
            except Exception:
                pass
            ano = int(periodo.split('-')[0]); mes = periodo.split('-')[1]
            total_mes_kwh, dias_utilizados, consumo_metodo = solar_consumo_mensal_robusto(local_nome, mes, ano, fator_mult)
            nd = calendar.monthrange(ano, int(mes))[1]
            daily_kwh_base = (total_mes_kwh / dias_utilizados) if dias_utilizados else (total_mes_kwh / nd if nd else 0.0)
        else:
            consumo_metodo = 'local não selecionado'

    daily_kwh_corrigido = daily_kwh_base * (1 + crescimento_carga_pct/100.0)
    total_mes_corrigido = total_mes_kwh * (1 + crescimento_carga_pct/100.0)
    daily_kwh_solar = daily_kwh_corrigido * (cobertura_pct/100.0)
    consumo_anual_estimado = daily_kwh_corrigido * 365.0

    kwp_necessario = daily_kwh_solar / (psh * derate_expert) if psh > 0 and derate_expert > 0 else 0.0
    n_paineis = math.ceil((kwp_necessario * 1000.0) / panel_wp) if panel_wp > 0 else 0
    kwp_real = (n_paineis * panel_wp) / 1000.0
    area_total = n_paineis * panel_area
    inversor_por_dcac = kwp_real / inv_dcac if inv_dcac > 0 else kwp_real
    inversor_por_pico = pico_kw * (1 + reserva_inversor_pct/100.0) if pico_kw > 0 else 0.0
    inversor_kw = max(inversor_por_dcac, inversor_por_pico)
    cabivel_area = True if area_disponivel <= 0 else area_total <= area_disponivel
    area_excedente = max(area_total - area_disponivel, 0.0) if area_disponivel > 0 else 0.0

    bateria_kwh_util = daily_kwh_corrigido * autonomy_days if tipo in ['offgrid','hibrido','hybrid'] else 0.0
    bateria_kwh_bruta = bateria_kwh_util / (battery_dod * battery_eff) if bateria_kwh_util > 0 else 0.0
    n_modulos_bateria = math.ceil(bateria_kwh_bruta / battery_module_kwh) if bateria_kwh_bruta > 0 else 0

    prod_mensal = []
    prod_anual = 0.0
    ano_ref = ano or datetime.now().year
    for m in range(1, 13):
        dias_m = calendar.monthrange(ano_ref, m)[1]
        psh_m = psh * float(perfil_sazonal[m-1])
        e_m = kwp_real * derate_expert * psh_m * dias_m
        prod_mensal.append(round(e_m, 2))
        prod_anual += e_m
    prod_anual = round(prod_anual, 2)
    limite_cobertura = consumo_anual_estimado * (cobertura_pct/100.0 if cobertura_pct <= 100 else 1.0)
    energia_util_anual = min(prod_anual * (autoconsumo_pct/100.0), limite_cobertura) if consumo_anual_estimado > 0 else 0.0
    cobertura_real_pct = (energia_util_anual / consumo_anual_estimado * 100.0) if consumo_anual_estimado > 0 else 0.0
    economia_anual = energia_util_anual * tarifa_kwh
    economia_mensal = economia_anual / 12.0

    capex_baterias = n_modulos_bateria * battery_module_kwh * capex_kwp * 0.15 if n_modulos_bateria else 0.0
    capex_total = kwp_real * capex_kwp + capex_baterias
    opex_anual = capex_total * (opex_pct / 100.0)

    r_desc = desconto / 100.0 if desconto > 0 else 0.0
    g = tarifa_esc / 100.0
    npv = -capex_total
    cumul = -capex_total
    payback_anos = None
    fluxos = [-capex_total]
    tarifa_t = tarifa_kwh
    energia_desc_total = 0.0
    custo_desc_total = capex_total
    for t in range(1, anos_analise + 1):
        receita_t = energia_util_anual * tarifa_t
        cf_t = receita_t - opex_anual
        fluxos.append(cf_t)
        fator_desc = ((1 + r_desc) ** t) if r_desc > 0 else 1.0
        npv += cf_t / fator_desc
        energia_desc_total += energia_util_anual / fator_desc
        custo_desc_total += opex_anual / fator_desc
        cumul += cf_t
        if payback_anos is None and cumul >= 0:
            prev_cumul = cumul - cf_t
            frac = 0 if cf_t == 0 else (0 - prev_cumul) / cf_t
            payback_anos = (t - 1) + max(0, min(1, frac))
        tarifa_t *= (1 + g)
    irr = solar_irr(fluxos)
    lcoe = (custo_desc_total / energia_desc_total) if energia_desc_total > 0 else None
    co2_t_ano = (energia_util_anual / 1000.0) * co2_factor

    alertas = []
    if daily_kwh_corrigido <= 0:
        alertas.append(('danger', 'Consumo não informado', 'Informe consumo diário ou selecione um local com leituras mensais gravadas.'))
    if not cabivel_area:
        alertas.append(('warning', 'Área disponível insuficiente', f'Área necessária {area_total:.1f} m²; área disponível {area_disponivel:.1f} m².'))
    if derate_expert < 0.65:
        alertas.append(('warning', 'Perdas elevadas', 'As perdas combinadas reduzem bastante a produção. Rever sombreamento, sujidade, cabos e temperatura.'))
    if tipo in ['offgrid','hibrido','hybrid'] and autonomy_days > 0 and n_modulos_bateria == 0:
        alertas.append(('info', 'Baterias não dimensionadas', 'Verifique o valor do módulo de bateria e a autonomia pretendida.'))
    if payback_anos is None and capex_total > 0:
        alertas.append(('warning', 'Payback não atingido no período', 'Rever CAPEX, tarifa, cobertura pretendida ou autoconsumo.'))

    r_dict = {
        'modo': modo, 'local_nome': local_nome, 'periodo': periodo, 'mes': mes, 'ano': ano,
        'dias_utilizados': dias_utilizados, 'consumo_metodo': consumo_metodo,
        'daily_kwh': daily_kwh_corrigido, 'daily_kwh_base': daily_kwh_base,
        'total_mes_kwh': total_mes_corrigido, 'total_mes_base_kwh': total_mes_kwh,
        'consumo_anual_estimado': consumo_anual_estimado,
        'psh': psh, 'derate': derate, 'derate_expert': derate_expert,
        'panel_wp': panel_wp, 'panel_area': panel_area, 'n_paineis': n_paineis,
        'kwp_necessario': kwp_necessario, 'kwp_real': kwp_real, 'area_total': area_total,
        'area_disponivel': area_disponivel, 'cabivel_area': cabivel_area, 'area_excedente': area_excedente,
        'inv_dcac': inv_dcac, 'inversor_kw': inversor_kw, 'inversor_por_dcac': inversor_por_dcac,
        'inversor_por_pico': inversor_por_pico, 'pico_kw': pico_kw, 'reserva_inversor_pct': reserva_inversor_pct,
        'tipo_sistema': tipo, 'tarifa_kwh': tarifa_kwh, 'economia_mensal': economia_mensal,
        'economia_anual': economia_anual, 'energia_util_anual': energia_util_anual,
        'cobertura_pct': cobertura_pct, 'cobertura_real_pct': cobertura_real_pct,
        'autoconsumo_pct': autoconsumo_pct, 'crescimento_carga_pct': crescimento_carga_pct,
        'sombreamento_pct': sombreamento_pct, 'perdas_cabos_pct': perdas_cabos_pct,
        'perdas_sujidade_pct': perdas_sujidade_pct, 'perdas_temp_pct': perdas_temp_pct,
        'autonomy_days': autonomy_days, 'battery_dod': battery_dod, 'battery_eff': battery_eff,
        'system_voltage': system_voltage, 'battery_module_kwh': battery_module_kwh,
        'bateria_kwh_util': bateria_kwh_util, 'bateria_kwh_bruta': bateria_kwh_bruta,
        'n_modulos_bateria': n_modulos_bateria, 'fator_mult': fator_mult,
        'producao_mensal': prod_mensal, 'producao_anual': prod_anual,
        'capex_kwp': capex_kwp, 'capex_baterias': capex_baterias, 'capex_total': capex_total,
        'opex_pct': opex_pct, 'opex_anual': opex_anual, 'tarifa_esc': tarifa_esc, 'desconto': desconto,
        'anos_analise': anos_analise, 'payback_anos': payback_anos,
        'npv': npv, 'irr': irr, 'lcoe': lcoe, 'co2_factor': co2_factor, 'co2_t_ano': co2_t_ano,
        'perfil_sazonal': perfil_sazonal, 'alertas': alertas
    }
    params = dict(r_dict)
    params['local_id'] = int(local_id) if local_id else None
    return render_template('solar_resultado.html', r=r_dict, params=params)


# === SALVAR PROJETO SOLAR ===

@app.route('/solar/salvar', methods=['POST'])
def solar_salvar():
    import sqlite3, json, time
    from datetime import datetime

    # dados que vieram escondidos do formulário
    r_json = request.form.get('r_json', '{}')
    params_json = request.form.get('params_json', '{}')

    # NOVOS CAMPOS visíveis no formulário
    nome_projeto = request.form.get('nome_projeto') or None
    obs = request.form.get('obs') or None

    # parse seguro
    try:
        r = json.loads(r_json) if r_json else {}
    except Exception:
        r = {}
    try:
        params = json.loads(params_json) if params_json else {}
    except Exception:
        params = {}

    agora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    local_id = params.get('local_id')
    local_nome = r.get('local_nome') or params.get('local_nome') or None
    periodo = r.get('periodo')
    modo = r.get('modo')
    tipo = r.get('tipo_sistema')

    # (resto da função continua aqui…)


    # parte comum a qualquer versão da tabela
    comuns = [
        local_id, local_nome, periodo, modo, tipo,
        r.get('daily_kwh'), r.get('total_mes_kwh'), r.get('psh'), r.get('derate'),
        r.get('panel_wp'), r.get('panel_area'), r.get('n_paineis'),
        r.get('kwp_necessario'), r.get('kwp_real'), r.get('area_total'),
        r.get('inv_dcac'), r.get('inversor_kw'), r.get('tarifa_kwh'),
        r.get('economia_mensal'),
        r.get('autonomy_days'), r.get('battery_dod'), r.get('battery_eff'),
        r.get('system_voltage'), r.get('battery_module_kwh'),
        r.get('bateria_kwh_util'), r.get('bateria_kwh_bruta'),
        r.get('n_modulos_bateria'), r.get('mes'), r.get('ano'),
        r.get('dias_utilizados'), r.get('fator_mult'),
        json.dumps(r, ensure_ascii=False),
        json.dumps(params, ensure_ascii=False),
        r.get('capex_kwp'), r.get('capex_total'), r.get('opex_pct'), r.get('opex_anual'),
        r.get('tarifa_esc'), r.get('desconto'), r.get('anos_analise'),
        r.get('payback_anos'), r.get('npv'), r.get('co2_factor'), r.get('co2_t_ano'),
        r.get('producao_anual'),
        json.dumps(r.get('producao_mensal'), ensure_ascii=False),
        json.dumps(r.get('perfil_sazonal'), ensure_ascii=False),
        nome_projeto,
        obs,
    ]

    # tua tabela antiga tinha created_at
    colunas_old = [
        'created_at','local_id','local_nome','periodo','modo','tipo_sistema',
        'daily_kwh','total_mes_kwh','psh','derate','panel_wp','panel_area',
        'n_paineis','kwp_necessario','kwp_real','area_total','inv_dcac','inversor_kw',
        'tarifa_kwh','economia_mensal','autonomy_days','battery_dod','battery_eff',
        'system_voltage','battery_module_kwh','bateria_kwh_util','bateria_kwh_bruta',
        'n_modulos_bateria','mes','ano','dias_utilizados','fator_mult',
        'resultado_json','params_json',
        'capex_kwp','capex_total','opex_pct','opex_anual','tarifa_esc','desconto','anos_analise',
        'payback_anos','npv','co2_factor','co2_t_ano','producao_anual_kwh',
        'producao_mensal_json','perfil_sazonal_json',
        'nome_projeto','obs'
    ]
    valores_old = [agora] + comuns

    # nossa tabela nova usa criado_em
    colunas_new = colunas_old.copy()
    colunas_new[0] = 'criado_em'
    valores_new = [agora] + comuns

    def tentar_criar_colunas_extra(conn):
        """cria nome_projeto e obs se não existirem (ignora erro)"""
        try:
            conn.execute("ALTER TABLE solar_projetos ADD COLUMN nome_projeto TEXT;")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE solar_projetos ADD COLUMN obs TEXT;")
        except Exception:
            pass

    def do_insert(colunas, valores):
        conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
        try:
            conn.execute("PRAGMA busy_timeout=10000;")
            # garante que a tabela existe (sem as colunas novas ainda)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS solar_projetos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT
                )
            """)
            # tenta criar as colunas novas (se já existem, ignora)
            tentar_criar_colunas_extra(conn)

            placeholders = ",".join(["?"] * len(valores))
            sql = f"INSERT INTO solar_projetos ({','.join(colunas)}) VALUES ({placeholders})"
            cur = conn.cursor()
            cur.execute(sql, valores)
            conn.commit()
        finally:
            conn.close()

    # vamos tentar até 5 vezes por causa de "database is locked"
    ultimo_erro = None
    for _ in range(5):
        try:
            try:
                # tenta com o nome que o teu banco já tinha
                do_insert(colunas_old, valores_old)
            except sqlite3.OperationalError as e:
                # se for erro de coluna, tenta com o outro nome
                if "no such column" in str(e).lower():
                    do_insert(colunas_new, valores_new)
                else:
                    raise
            ultimo_erro = None
            break
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower():
                ultimo_erro = e
                time.sleep(0.5)
                continue
            else:
                ultimo_erro = e
                break
        except sqlite3.IntegrityError as e:
            # se foi NOT NULL em created_at, tenta com criado_em
            if "created_at" in str(e):
                do_insert(colunas_new, valores_new)
                ultimo_erro = None
                break
            else:
                ultimo_erro = e
                break

    if ultimo_erro is not None:
        raise ultimo_erro

    flash("Projeto solar salvo.", "success")
    return redirect(url_for('solar_projetos'))

# === LISTAR PROJETOS SALVOS ===
@app.route('/solar/projetos', methods=['GET'])
def solar_projetos():
    import sqlite3
    conn = sqlite3.connect(DB_PATH, timeout=15, check_same_thread=False)
    conn.execute("PRAGMA busy_timeout=5000;")
    c = conn.cursor()
    # tenta trazer também nome_projeto e obs; se não existir, fica NULL
    c.execute("""
        SELECT
            id,
            criado_em,
            created_at,
            local_nome,
            periodo,
            tipo_sistema,
            kwp_real,
            n_paineis,
            inversor_kw,
            economia_mensal,
            payback_anos,
            co2_t_ano,
            nome_projeto,
            obs
        FROM solar_projetos
        ORDER BY id DESC
    """)
    rows = c.fetchall()
    conn.close()
    return render_template('solar_projetos.html', projetos=rows)


# === DETALHAR PROJETO SALVO ===
@app.route('/solar/projeto/<int:pid>', methods=['GET'])
def solar_projeto_detalhe(pid):
    import sqlite3, json

    conn = sqlite3.connect(DB_PATH, timeout=15, check_same_thread=False)
    conn.execute("PRAGMA busy_timeout=5000;")
    c = conn.cursor()
    # vamos tentar pegar tudo que pode existir
    c.execute("""
        SELECT
            resultado_json,
            params_json,
            nome_projeto,
            obs,
            local_nome,
            periodo,
            tipo_sistema,
            kwp_real,
            economia_mensal,
            payback_anos,
            co2_t_ano
        FROM solar_projetos
        WHERE id=?
    """, (pid,))
    row = c.fetchone()
    conn.close()

    if not row:
        return "Projeto não encontrado.", 404

    resultado_json = row[0]
    params_json = row[1]
    nome_projeto = row[2]
    obs = row[3]
    local_nome = row[4]
    periodo = row[5]
    tipo_sistema = row[6]
    kwp_real = row[7]
    economia_mensal = row[8]
    payback_anos = row[9]
    co2_t_ano = row[10]

    # parse do resultado salvo
    try:
        r = json.loads(resultado_json) if resultado_json else {}
    except Exception:
        r = {}
    try:
        params = json.loads(params_json) if params_json else {}
    except Exception:
        params = {}

    # caso o nome/local não estejam no JSON, usa os da tabela
    if nome_projeto and not r.get("nome_projeto"):
        r["nome_projeto"] = nome_projeto
    if local_nome and not r.get("local_nome"):
        r["local_nome"] = local_nome
    if periodo and not r.get("periodo"):
        r["periodo"] = periodo
    if obs:
        r["obs"] = obs

    # alguns campos podem não existir no JSON antigo
    if kwp_real and not r.get("kwp_real"):
        r["kwp_real"] = kwp_real
    if economia_mensal and not r.get("economia_mensal"):
        r["economia_mensal"] = economia_mensal
    if payback_anos is not None and not r.get("payback_anos"):
        r["payback_anos"] = payback_anos
    if co2_t_ano and not r.get("co2_t_ano"):
        r["co2_t_ano"] = co2_t_ano
    if tipo_sistema and not r.get("tipo_sistema"):
        r["tipo_sistema"] = tipo_sistema

    return render_template(
        'solar_projeto_detalhe.html',
        pid=pid,
        r=r,
        params=params
    )



# === EXPORTAR PRODUÇÃO MENSAL DE UM PROJETO (CSV) ===
@app.route('/solar/export/<int:pid>.csv')
def solar_export_csv(pid):
    conn = sqlite3.connect(DB_PATH, timeout=15, check_same_thread=False)
    conn.execute("PRAGMA busy_timeout=5000;")
    c = conn.cursor()
    c.execute("SELECT producao_mensal_json FROM solar_projetos WHERE id=?", (pid,))
    row = c.fetchone()
    conn.close()

    if not row or not row[0]:
        return Response("Projeto/produção não encontrada.", status=404)

    try:
        serie = json.loads(row[0])
    except Exception:
        return Response("Formato inválido.", status=400)

    si = StringIO()
    w = csv.writer(si, delimiter=';')
    w.writerow(["Mes", "Producao_kWh"])
    meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
             "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    for i, val in enumerate(serie):
        mes = meses[i] if i < len(meses) else f"M{i+1}"
        w.writerow([mes, val])

    output = si.getvalue()
    filename = f"producao_mensal_projeto_{pid}.csv"
    return Response(
        output,
        mimetype='text/csv',
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )
# ==============================
# === DIMENSIONAMENTO LÂMPADAS SOLARES
# ==============================

import math, json, sqlite3, csv, io
from io import StringIO
from datetime import datetime
from flask import (
    render_template, render_template_string, request, redirect,
    url_for, Response, flash, abort
)

# ---------- Conexão robusta ----------
def _db_conn():
    conn = sqlite3.connect(DB_PATH, timeout=15, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=5000;")
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn

# --- Catálogo padrão (mantido como no teu código) ---
def _catalogo_lampadas_padrao():
    return [
        {"modelo": "SL-30W", "potencia_w": 30, "fluxo_lm": 4200, "bateria_wh": 192, "autonomia_h": 12, "altura_poste_m": 4},
        {"modelo": "SL-50W", "potencia_w": 50, "fluxo_lm": 7000, "bateria_wh": 384, "autonomia_h": 12, "altura_poste_m": 6},
        {"modelo": "SL-80W", "potencia_w": 80, "fluxo_lm": 11000, "bateria_wh": 480, "autonomia_h": 12, "altura_poste_m": 8},
    ]

# ---------- Tabela de catálogo (garantia) ----------
def _ensure_lamp_catalog_table():
    with _db_conn() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS solar_lampadas_catalogo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            modelo TEXT UNIQUE,
            potencia_w REAL,
            fluxo_lm REAL,
            bateria_wh REAL,
            autonomia_h REAL,
            altura_poste_m REAL,
            fabricante TEXT,
            preco_mt REAL,
            nota TEXT
        )
        """)

# ---------- Catálogo via BD com fallback ----------
def _catalogo_lampadas_from_db():
    """
    Lê catálogo da BD; caso esteja vazio, devolve o catálogo padrão.
    """
    _ensure_lamp_catalog_table()
    with _db_conn() as conn:
        rows = conn.execute("""
            SELECT modelo, potencia_w, fluxo_lm, bateria_wh, autonomia_h, altura_poste_m
            FROM solar_lampadas_catalogo
            ORDER BY fluxo_lm ASC
        """).fetchall()
    if rows:
        return [
            {
                "modelo": r["modelo"],
                "potencia_w": r["potencia_w"] or 0.0,
                "fluxo_lm": r["fluxo_lm"] or 0.0,
                "bateria_wh": r["bateria_wh"] or 0.0,
                "autonomia_h": r["autonomia_h"] or 0.0,
                "altura_poste_m": r["altura_poste_m"] or 0.0,
            }
            for r in rows
        ]
    return _catalogo_lampadas_padrao()

# ---------- FORM (GET) ----------
@app.route('/solar/lampadas', methods=['GET'], endpoint='solar_lampadas_form')
def solar_lampadas_form():
    """Centro expert de dimensionamento de lâmpadas solares/autónomas."""
    catalogo = _catalogo_lampadas_from_db()
    altura_qs = request.args.get('altura', type=float)
    return render_template('solar_lampadas.html', resultado=None, catalogo=catalogo, sugestoes=[], altura_qs=altura_qs)

@app.route('/solar/lampadas/usar/<int:cat_id>', methods=['GET'])
def solar_lampadas_usar(cat_id):
    _ensure_lamp_catalog_table()
    with _db_conn() as conn:
        row = conn.execute("SELECT altura_poste_m FROM solar_lampadas_catalogo WHERE id=?",(cat_id,)).fetchone()
    altura = float(row["altura_poste_m"]) if row and row["altura_poste_m"] is not None else 5.0
    return redirect(url_for('solar_lampadas_form', altura=altura))

def _solar_float(form, nome, default=0.0):
    v = (form.get(nome) or '').strip()
    if v == '':
        return float(default)
    try:
        return float(v.replace(' ', '').replace(',', '.'))
    except Exception:
        return float(default)

def _solar_int(form, nome, default=0):
    try:
        return int(round(_solar_float(form, nome, default)))
    except Exception:
        return int(default)

def _lampadas_classe_presets():
    return {
        'vias_pedonais': {'nome':'Vias pedonais / passeios', 'lux': 7.5, 'altura': 4.0, 'shr': 4.0},
        'parques': {'nome':'Parques / jardins', 'lux': 15.0, 'altura': 5.0, 'shr': 4.0},
        'rua_local': {'nome':'Rua local / residencial', 'lux': 12.0, 'altura': 6.0, 'shr': 4.8},
        'rua_coletora': {'nome':'Rua coletora', 'lux': 20.0, 'altura': 8.0, 'shr': 5.5},
        'estacionamento': {'nome':'Estacionamento / pátio', 'lux': 12.0, 'altura': 6.0, 'shr': 4.5},
        'armazem_ext': {'nome':'Recinto industrial exterior', 'lux': 20.0, 'altura': 8.0, 'shr': 5.0},
        'seguranca': {'nome':'Segurança perimetral', 'lux': 10.0, 'altura': 5.0, 'shr': 4.5},
    }

def _lampadas_numero_postes(modo, area_m2, comprimento_via, largura_via, altura_poste, classe, dupla_fileira, qtd_manual):
    presets = _lampadas_classe_presets()
    shr = presets.get(classe, {}).get('shr', 5.0)
    espacamento_m = max(altura_poste * shr, 1.0)
    if modo == 'manual' and qtd_manual > 0:
        qtd = qtd_manual
    elif modo == 'via':
        fileiras = 2 if dupla_fileira else 1
        qtd = max(int(math.ceil(max(comprimento_via, 0) / espacamento_m)) + 1, 1) * fileiras
        area_m2 = max(area_m2, comprimento_via * max(largura_via, 1.0))
    else:
        # Estimativa por área: área de influência aproximada por poste.
        area_por_poste = max((espacamento_m * espacamento_m * 0.55), 1.0)
        qtd = max(int(math.ceil(max(area_m2, 1.0) / area_por_poste)), 1)
    return int(qtd), float(espacamento_m), float(area_m2)

@app.route('/solar/lampadas/calcular', methods=['POST'])
def solar_lampadas_calcular():
    """Dimensionamento expert de iluminação solar autónoma."""
    form = request.form
    classe = (form.get('classe') or '').strip()
    modo_implantacao = (form.get('modo_implantacao') or 'area').strip()
    nome_projeto = (form.get('nome_projeto') or '').strip()

    presets = _lampadas_classe_presets()
    preset = presets.get(classe, {})

    area_m2 = _solar_float(form, 'area_m2', 200)
    comprimento_via_m = _solar_float(form, 'comprimento_via_m', 100)
    largura_via_m = _solar_float(form, 'largura_via_m', 6)
    iluminancia_lux = _solar_float(form, 'iluminancia_lux', preset.get('lux', 10))
    altura_poste = _solar_float(form, 'altura_poste', preset.get('altura', 5))
    if not (form.get('iluminancia_lux') or '').strip() and preset:
        iluminancia_lux = preset['lux']
    if not (form.get('altura_poste') or '').strip() and preset:
        altura_poste = preset['altura']

    qtd_manual = _solar_int(form, 'qtd_manual', 0)
    dupla_fileira = (form.get('dupla_fileira') == '1')
    qtd_postes, espacamento_m, area_corrigida = _lampadas_numero_postes(
        modo_implantacao, area_m2, comprimento_via_m, largura_via_m, altura_poste,
        classe, dupla_fileira, qtd_manual
    )
    area_por_luminaria = area_corrigida / max(qtd_postes, 1)

    autonomia_h = _solar_float(form, 'autonomia_h', 12)
    autonomia_dias = _solar_float(form, 'autonomia_dias', 2)
    psh = _solar_float(form, 'horas_carga', 5)
    fator_util = _solar_float(form, 'fator_util', 0.80)
    fator_man = _solar_float(form, 'fator_man', 0.90)
    fator_seg = _solar_float(form, 'fator_seg', 1.20)
    lm_por_w = _solar_float(form, 'lm_por_w', 150)

    dim_0_6 = _solar_float(form, 'dim_0_6', 100)
    dim_6_12 = _solar_float(form, 'dim_6_12', 50)
    dim_extra = _solar_float(form, 'dim_extra', 30)

    dod = _solar_float(form, 'dod', 0.80)
    eficiencia_bateria = _solar_float(form, 'eficiencia_bateria', 0.90)
    eficiencia_controlador = _solar_float(form, 'eficiencia_controlador', 0.92)
    bateria_v = _solar_float(form, 'bateria_v', 12)
    margem_painel = _solar_float(form, 'margem_painel', 1.25)

    custo_luminaria = _solar_float(form, 'custo_luminaria', 15000)
    custo_painel_w = _solar_float(form, 'custo_painel_w', 30)
    custo_bateria_wh = _solar_float(form, 'custo_bateria_wh', 12)
    custo_poste = _solar_float(form, 'custo_poste', 18000)
    custo_instalacao = _solar_float(form, 'custo_instalacao', 8000)
    opex_pct = _solar_float(form, 'opex_pct', 2.0)
    tarifa_kwh = _solar_float(form, 'tarifa_kwh', 8.0)

    # Fluxo e potência por luminária
    fluxo_min_lm = (iluminancia_lux * area_por_luminaria) / max(fator_util * fator_man, 1e-6)
    fluxo_projeto_lm = fluxo_min_lm * max(fator_seg, 1.0)
    potencia_led_w_nom = fluxo_projeto_lm / max(lm_por_w, 1e-6)

    # Perfil noturno com dimerização em três blocos
    h1 = min(autonomia_h, 6.0)
    h2 = min(max(autonomia_h - h1, 0.0), 6.0)
    h3 = max(autonomia_h - h1 - h2, 0.0)
    energia_noite_wh = potencia_led_w_nom * ((dim_0_6/100.0)*h1 + (dim_6_12/100.0)*h2 + (dim_extra/100.0)*h3)
    potencia_media_w = energia_noite_wh / max(autonomia_h, 1e-6)

    # Painel, bateria e controlador por poste
    painel_wp = energia_noite_wh / max(psh * eficiencia_controlador, 1e-6) * margem_painel
    bateria_wh_util = energia_noite_wh * max(autonomia_dias, 1.0)
    bateria_wh_nominal = bateria_wh_util / max(dod * eficiencia_bateria, 1e-6)
    bateria_ah = bateria_wh_nominal / max(bateria_v, 1e-6)
    controlador_a = max((painel_wp / max(bateria_v, 1e-6)) * 1.25, 5.0)

    energia_noite_total_kwh = (energia_noite_wh * qtd_postes) / 1000.0
    energia_anual_kwh = energia_noite_total_kwh * 365.0
    economia_anual_mt = energia_anual_kwh * tarifa_kwh

    custo_unit = (custo_luminaria + painel_wp*custo_painel_w + bateria_wh_nominal*custo_bateria_wh + custo_poste + custo_instalacao)
    capex_total = custo_unit * qtd_postes
    opex_anual = capex_total * (opex_pct/100.0)
    economia_liquida_anual = max(economia_anual_mt - opex_anual, 0)
    payback = capex_total / economia_liquida_anual if economia_liquida_anual > 0 else 0
    co2_t_ano = energia_anual_kwh * 0.0007

    # Catálogo e seleção inteligente
    catalogo = _catalogo_lampadas_from_db()
    sugestoes = []
    for item in catalogo:
        fluxo = float(item.get('fluxo_lm') or 0)
        pot = float(item.get('potencia_w') or 0)
        bat = float(item.get('bateria_wh') or 0)
        aut = float(item.get('autonomia_h') or 0)
        alt = float(item.get('altura_poste_m') or 0)
        atende_fluxo = fluxo >= fluxo_projeto_lm
        atende_pot = pot >= potencia_led_w_nom * 0.85
        atende_bateria = bat >= bateria_wh_util
        atende_autonomia = aut >= autonomia_h
        altura_ok = abs(alt - altura_poste) <= 2.0 if alt else True
        score = 0
        score += 35 if atende_fluxo else max(0, 20 * fluxo / max(fluxo_projeto_lm, 1))
        score += 25 if atende_bateria else max(0, 15 * bat / max(bateria_wh_util, 1))
        score += 15 if atende_autonomia else 0
        score += 15 if altura_ok else 5
        score += 10 if atende_pot else 3
        sugestoes.append({
            'item': item, 'score': round(min(score, 100), 1),
            'atende_fluxo': atende_fluxo, 'atende_bateria': atende_bateria,
            'atende_autonomia': atende_autonomia, 'altura_ok': altura_ok,
            'margem_fluxo': round(fluxo - fluxo_projeto_lm, 1)
        })
    sugestoes.sort(key=lambda x: (-x['score'], abs((x['item'].get('fluxo_lm') or 0)-fluxo_projeto_lm)))

    alertas = []
    if psh < 4.0: alertas.append('PSH baixo: aumentar painel ou autonomia para garantir carga em dias críticos.')
    if bateria_ah > 250: alertas.append('Banco de baterias elevado por poste: considerar tensão maior, bateria modular ou reduzir potência/dimerização.')
    if painel_wp > 350: alertas.append('Painel por poste elevado: verificar sombreamento, PSH e potência da luminária.')
    if fluxo_projeto_lm < 1500: alertas.append('Fluxo baixo: confirme se a aplicação permite este nível de iluminação.')
    if payback and payback > 10: alertas.append('Payback elevado: validar custos unitários, tarifa evitada e necessidade operacional do projecto.')

    resultado = {
        'nome_projeto': nome_projeto,
        'classe': classe,
        'classe_nome': preset.get('nome', classe or 'Personalizado'),
        'modo_implantacao': modo_implantacao,
        'area_m2': round(area_corrigida, 2),
        'comprimento_via_m': round(comprimento_via_m, 2),
        'largura_via_m': round(largura_via_m, 2),
        'qtd_postes': qtd_postes,
        'espacamento_m': round(espacamento_m, 2),
        'area_por_luminaria': round(area_por_luminaria, 2),
        'dupla_fileira': dupla_fileira,
        'iluminancia_lux': round(iluminancia_lux, 2),
        'altura_poste': round(altura_poste, 2),
        'autonomia_h': round(autonomia_h, 2),
        'autonomia_dias': round(autonomia_dias, 2),
        'horas_carga': round(psh, 2),
        'fator_util': round(fator_util, 3),
        'fator_man': round(fator_man, 3),
        'fator_seg': round(fator_seg, 2),
        'lm_por_w': round(lm_por_w, 1),
        'dim_0_6': round(dim_0_6, 1), 'dim_6_12': round(dim_6_12, 1), 'dim_extra': round(dim_extra, 1),
        'fluxo_min_lm': round(fluxo_min_lm, 1),
        'fluxo_projeto_lm': round(fluxo_projeto_lm, 1),
        'potencia_led_w': round(potencia_led_w_nom, 1),
        'potencia_media_w': round(potencia_media_w, 1),
        'energia_noite_wh': round(energia_noite_wh, 1),
        'painel_wp': round(painel_wp, 1),
        'bateria_wh_util': round(bateria_wh_util, 1),
        'bateria_wh_bruto': round(bateria_wh_nominal, 1),
        'bateria_ah': round(bateria_ah, 1),
        'bateria_v': round(bateria_v, 1),
        'controlador_a': round(controlador_a, 1),
        'energia_noite_total_kwh': round(energia_noite_total_kwh, 2),
        'energia_anual_kwh': round(energia_anual_kwh, 2),
        'tarifa_kwh': round(tarifa_kwh, 2),
        'economia_anual_mt': round(economia_anual_mt, 2),
        'capex_unitario': round(custo_unit, 2),
        'capex_estimado': round(capex_total, 2),
        'opex_anual': round(opex_anual, 2),
        'payback_anos': round(payback, 2),
        'co2_t_ano': round(co2_t_ano, 2),
        'custo_luminaria': round(custo_luminaria, 2),
        'custo_painel_w': round(custo_painel_w, 2),
        'custo_bateria_wh': round(custo_bateria_wh, 2),
        'custo_poste': round(custo_poste, 2),
        'custo_instalacao': round(custo_instalacao, 2),
        'alertas': alertas,
    }
    return render_template('solar_lampadas.html', resultado=resultado, catalogo=[s['item'] for s in sugestoes], sugestoes=sugestoes, altura_qs=None)

@app.route('/solar/lampadas/espacamento', methods=['POST'])
def solar_lampadas_espacamento():
    classe = (request.form.get('classe') or '').strip()
    try:
        altura = float((request.form.get('altura_poste') or '0').replace(',','.'))
        comp_via = float((request.form.get('comprimento_via_m') or '0').replace(',','.'))
    except Exception:
        return {"ok": False, "erro": "altura/comprimento inválidos"}, 400
    if altura <= 0 or comp_via <= 0:
        return {"ok": False, "erro": "altura/comprimento inválidos"}, 400
    shr = _lampadas_classe_presets().get(classe, {}).get('shr', 5.0)
    espac_m = max(altura * shr, 1.0)
    fileiras = 2 if (request.form.get('dupla_fileira') == '1') else 1
    qtd = max(int(math.ceil(comp_via / espac_m)) + 1, 1) * fileiras
    return {"ok": True, "espac_m": round(espac_m, 2), "qtd_postes": int(qtd)}

@app.route('/solar/lampadas/orcamento', methods=['POST'])
def solar_lampadas_orcamento():
    _ensure_lamp_catalog_table()
    modelo = (request.form.get('modelo') or '').strip()
    try:
        qtd = int((request.form.get('qtd') or '0').strip())
    except Exception:
        qtd = 0
    if not modelo or qtd <= 0:
        return {"ok": False, "erro": "modelo/quantidade inválidos"}, 400
    with _db_conn() as conn:
        row = conn.execute("SELECT preco_mt FROM solar_lampadas_catalogo WHERE modelo=?", (modelo,)).fetchone()
    if not row or row["preco_mt"] is None:
        return {"ok": False, "erro": "modelo sem preço"}, 400
    preco = float(row["preco_mt"])
    total = preco * qtd
    return {"ok": True, "preco_unit": round(preco,2), "qtd": qtd, "total": round(total,2)}

@app.route('/solar/lampadas/salvar', methods=['POST'])
def solar_lampadas_salvar():
    r_json = request.form.get('r_json', '{}')
    nome_projeto = request.form.get('nome_projeto') or None
    obs = request.form.get('obs') or None
    try:
        r = json.loads(r_json) if r_json else {}
    except Exception:
        r = {}
    with _db_conn() as conn:
        conn.execute("""
          CREATE TABLE IF NOT EXISTS solar_lampadas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            nome TEXT,
            obs TEXT,
            resultado_json TEXT NOT NULL
          )
        """)
        conn.execute(
            "INSERT INTO solar_lampadas (created_at, nome, obs, resultado_json) VALUES (?, ?, ?, ?)",
            (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), nome_projeto, obs, json.dumps(r, ensure_ascii=False))
        )
    flash("Projeto de iluminação solar salvo.", "success")
    return redirect(url_for('solar_lampadas_projetos'))

# ==============================
# === CATÁLOGO: LÂMPADAS SOLARES (CRUD) ===
# ==============================

@app.route('/solar/lampadas/catalogo', methods=['GET'])
def solar_lampadas_catalogo_list():
    _ensure_lamp_catalog_table()
    with _db_conn() as conn:
        rows = conn.execute("""
            SELECT id, modelo, fabricante, potencia_w, fluxo_lm, bateria_wh,
                   autonomia_h, altura_poste_m, preco_mt, nota
            FROM solar_lampadas_catalogo
            ORDER BY fluxo_lm DESC, potencia_w DESC
        """).fetchall()
    return render_template('solar_lampadas_catalogo_list.html', items=rows)

@app.route('/api/solar/lampadas/catalogo.json')
def api_solar_lampadas_catalogo():
    _ensure_lamp_catalog_table()
    with _db_conn() as conn:
        rows = conn.execute("""
            SELECT id, modelo, fabricante, fluxo_lm, potencia_w, preco_mt
            FROM solar_lampadas_catalogo
            ORDER BY modelo ASC
        """).fetchall()
    data = []
    for r in rows:
        data.append({"id": r["id"], "modelo": r["modelo"], "fabricante": r["fabricante"],
                     "fluxo_lm": float(r["fluxo_lm"] or 0), "potencia_w": float(r["potencia_w"] or 0),
                     "preco_mt": float(r["preco_mt"] or 0)})
    return {"ok": True, "itens": data}

@app.route('/solar/lampadas/catalogo/novo', methods=['GET','POST'])
def solar_lampadas_catalogo_novo():
    _ensure_lamp_catalog_table()
    if request.method == 'POST':
        modelo = (request.form.get('modelo') or '').strip()
        fabricante = (request.form.get('fabricante') or '').strip()
        potencia_w = _solar_float(request.form, 'potencia_w', 0)
        fluxo_lm = _solar_float(request.form, 'fluxo_lm', 0)
        bateria_wh = _solar_float(request.form, 'bateria_wh', 0)
        autonomia_h = _solar_float(request.form, 'autonomia_h', 0)
        altura_poste_m = _solar_float(request.form, 'altura_poste_m', 0)
        preco_mt = _solar_float(request.form, 'preco_mt', 0)
        nota = (request.form.get('nota') or '').strip()
        if not modelo:
            flash("Informe o modelo.", "warning")
            return redirect(url_for('solar_lampadas_catalogo_novo'))
        try:
            with _db_conn() as conn:
                conn.execute("""INSERT INTO solar_lampadas_catalogo
                                (modelo, fabricante, potencia_w, fluxo_lm, bateria_wh, autonomia_h, altura_poste_m, preco_mt, nota)
                                VALUES (?,?,?,?,?,?,?,?,?)""",
                             (modelo, fabricante, potencia_w, fluxo_lm, bateria_wh, autonomia_h, altura_poste_m, preco_mt, nota))
            flash("Modelo adicionado.", "success")
        except sqlite3.IntegrityError:
            flash("Modelo já existente.", "danger")
        return redirect(url_for('solar_lampadas_catalogo_list'))
    return render_template('solar_lampadas_catalogo_form.html', item=None)

@app.route('/solar/lampadas/catalogo/<int:cat_id>/editar', methods=['GET','POST'])
def solar_lampadas_catalogo_editar(cat_id):
    _ensure_lamp_catalog_table()
    with _db_conn() as conn:
        if request.method == 'POST':
            modelo = (request.form.get('modelo') or '').strip()
            fabricante = (request.form.get('fabricante') or '').strip()
            potencia_w = _solar_float(request.form, 'potencia_w', 0)
            fluxo_lm = _solar_float(request.form, 'fluxo_lm', 0)
            bateria_wh = _solar_float(request.form, 'bateria_wh', 0)
            autonomia_h = _solar_float(request.form, 'autonomia_h', 0)
            altura_poste_m = _solar_float(request.form, 'altura_poste_m', 0)
            preco_mt = _solar_float(request.form, 'preco_mt', 0)
            nota = (request.form.get('nota') or '').strip()
            try:
                conn.execute("""UPDATE solar_lampadas_catalogo
                                SET modelo=?, fabricante=?, potencia_w=?, fluxo_lm=?, bateria_wh=?,
                                    autonomia_h=?, altura_poste_m=?, preco_mt=?, nota=?
                                WHERE id=?""",
                             (modelo, fabricante, potencia_w, fluxo_lm, bateria_wh, autonomia_h, altura_poste_m, preco_mt, nota, cat_id))
                flash("Modelo atualizado.", "success")
            except sqlite3.IntegrityError:
                flash("Modelo duplicado.", "danger")
            return redirect(url_for('solar_lampadas_catalogo_list'))
        row = conn.execute("""SELECT id, modelo, fabricante, potencia_w, fluxo_lm, bateria_wh,
                                     autonomia_h, altura_poste_m, preco_mt, nota
                              FROM solar_lampadas_catalogo WHERE id=?""", (cat_id,)).fetchone()
    if not row:
        flash("Modelo não encontrado.", "warning")
        return redirect(url_for('solar_lampadas_catalogo_list'))
    return render_template('solar_lampadas_catalogo_form.html', item=row)

@app.route('/solar/lampadas/catalogo/<int:cat_id>/apagar', methods=['POST'])
def solar_lampadas_catalogo_apagar(cat_id):
    _ensure_lamp_catalog_table()
    with _db_conn() as conn:
        conn.execute("DELETE FROM solar_lampadas_catalogo WHERE id=?", (cat_id,))
    flash("Modelo removido.", "info")
    return redirect(url_for('solar_lampadas_catalogo_list'))

@app.route('/solar/lampadas/catalogo/export.csv', methods=['GET'])
def solar_lampadas_catalogo_export():
    _ensure_lamp_catalog_table()
    with _db_conn() as conn:
        rows = conn.execute("""SELECT modelo, fabricante, potencia_w, fluxo_lm, bateria_wh,
                                      autonomia_h, altura_poste_m, preco_mt, nota
                               FROM solar_lampadas_catalogo
                               ORDER BY modelo ASC""").fetchall()
    si = io.StringIO(); w = csv.writer(si, delimiter=';')
    w.writerow(["modelo","fabricante","potencia_w","fluxo_lm","bateria_wh","autonomia_h","altura_poste_m","preco_mt","nota"])
    for r in rows:
        w.writerow([r["modelo"], r["fabricante"], r["potencia_w"], r["fluxo_lm"], r["bateria_wh"], r["autonomia_h"], r["altura_poste_m"], r["preco_mt"], r["nota"]])
    return Response(si.getvalue(), mimetype='text/csv', headers={"Content-Disposition": "attachment;filename=catalogo_lampadas.csv"})

@app.route('/solar/lampadas/catalogo/import', methods=['GET','POST'])
def solar_lampadas_catalogo_import():
    _ensure_lamp_catalog_table()
    if request.method == 'POST':
        data = request.form.get('csv_text', '').strip()
        delim = request.form.get('delim', ';')
        if not data:
            flash("Cole o conteúdo CSV.", "warning")
            return redirect(url_for('solar_lampadas_catalogo_import'))
        reader = csv.DictReader(io.StringIO(data), delimiter=delim)
        inseridos = atualizados = 0
        with _db_conn() as conn:
            for row in reader:
                modelo = (row.get('modelo') or '').strip()
                if not modelo: continue
                def _flt(x):
                    try: return float((x or '0').replace(',','.'))
                    except: return 0.0
                vals = ((row.get('fabricante') or '').strip(), _flt(row.get('potencia_w')), _flt(row.get('fluxo_lm')), _flt(row.get('bateria_wh')), _flt(row.get('autonomia_h')), _flt(row.get('altura_poste_m')), _flt(row.get('preco_mt')), (row.get('nota') or '').strip())
                ex = conn.execute("SELECT id FROM solar_lampadas_catalogo WHERE modelo=?", (modelo,)).fetchone()
                if ex:
                    conn.execute("""UPDATE solar_lampadas_catalogo SET fabricante=?, potencia_w=?, fluxo_lm=?, bateria_wh=?, autonomia_h=?, altura_poste_m=?, preco_mt=?, nota=? WHERE modelo=?""", vals + (modelo,))
                    atualizados += 1
                else:
                    conn.execute("""INSERT INTO solar_lampadas_catalogo (modelo, fabricante, potencia_w, fluxo_lm, bateria_wh, autonomia_h, altura_poste_m, preco_mt, nota) VALUES (?,?,?,?,?,?,?,?,?)""", (modelo,) + vals)
                    inseridos += 1
        flash(f"Importação concluída. Inseridos: {inseridos}, Atualizados: {atualizados}", "success")
        return redirect(url_for('solar_lampadas_catalogo_list'))
    return render_template('solar_lampadas_catalogo_import.html')

@app.route('/solar/lampadas/projetos', methods=['GET'])
def solar_lampadas_projetos():
    with _db_conn() as conn:
        rows = conn.execute("SELECT id, created_at, nome, obs, resultado_json FROM solar_lampadas ORDER BY id DESC").fetchall()
    projetos=[]
    for r in rows:
        try: data=json.loads(r['resultado_json'] or '{}')
        except Exception: data={}
        projetos.append({'id':r['id'], 'created_at':r['created_at'], 'nome':r['nome'], 'obs':r['obs'], 'r':data})
    return render_template('solar_lampadas_projetos.html', projetos=projetos)

@app.route('/solar/lampadas/projeto/<int:pid>', methods=['GET'])
def solar_lampadas_projeto(pid):
    with _db_conn() as conn:
        row = conn.execute("SELECT id, created_at, nome, obs, resultado_json FROM solar_lampadas WHERE id=?", (pid,)).fetchone()
    if not row:
        return Response("Projeto não encontrado", status=404)
    try: r = json.loads(row["resultado_json"] or '{}')
    except Exception: r = {}
    return render_template('solar_lampadas_projeto_detalhe.html', pid=row['id'], criado=row['created_at'], nome=row['nome'], obs=row['obs'], r=r)

@app.route('/solar/lampadas/export/<int:pid>.csv')
def solar_lampadas_export(pid):
    with _db_conn() as conn:
        row = conn.execute("SELECT resultado_json FROM solar_lampadas WHERE id=?", (pid,)).fetchone()
    if not row: return Response("Projeto não encontrado.", status=404)
    try: r = json.loads(row["resultado_json"])
    except Exception: return Response("Formato inválido.", status=400)
    si = StringIO(); w = csv.writer(si, delimiter=';')
    w.writerow(["campo","valor"])
    for k in sorted(r.keys()):
        val = r.get(k)
        if isinstance(val, (dict, list)): val = json.dumps(val, ensure_ascii=False)
        w.writerow([k, val])
    return Response(si.getvalue(), mimetype='text/csv', headers={"Content-Disposition": f"attachment;filename=lampadas_projeto_{pid}.csv"})



# ==============================
# === SOLAR: PORTFÓLIO, COMPARAÇÃO E RELATÓRIO CONSOLIDADO ===
# ==============================

def _solar_num(v, default=0.0):
    try:
        if v is None or v == '':
            return float(default)
        return float(str(v).replace(',', '.'))
    except Exception:
        return float(default)

def _solar_json_load(txt):
    try:
        return json.loads(txt or '{}')
    except Exception:
        return {}

def _solar_fmt_mt(v):
    try:
        return ("{:,.2f}".format(float(v or 0))).replace(',', 'X').replace('.', ',').replace('X', '.') + ' MT'
    except Exception:
        return '0,00 MT'

def _solar_get_fv_projects(limit=None):
    with _db_conn() as conn:
        try:
            sql = "SELECT * FROM solar_projetos ORDER BY id DESC"
            if limit:
                sql += " LIMIT " + str(int(limit))
            rows = conn.execute(sql).fetchall()
        except Exception:
            rows = []
    projetos = []
    for row in rows:
        keys = row.keys()
        r = _solar_json_load(row['resultado_json'] if 'resultado_json' in keys else '{}')
        nome = (row['nome_projeto'] if 'nome_projeto' in keys else None) or r.get('nome_projeto') or f"Projeto FV #{row['id']}"
        criado = (row['criado_em'] if 'criado_em' in keys else None) or (row['created_at'] if 'created_at' in keys else '')
        def col(name, fallback=0):
            try:
                return row[name] if name in keys and row[name] not in [None, ''] else r.get(name, fallback)
            except Exception:
                return r.get(name, fallback)
        projetos.append({
            'tipo': 'FV', 'uid': f"fv:{row['id']}", 'id': row['id'], 'nome': nome, 'criado': criado,
            'local': col('local_nome', r.get('local_nome','')), 'sistema': col('tipo_sistema', r.get('tipo_sistema','')),
            'kwp': _solar_num(col('kwp_real', r.get('kwp_real',0))), 'paineis': int(_solar_num(col('n_paineis', r.get('n_paineis',0)),0)),
            'inversor_kw': _solar_num(col('inversor_kw', r.get('inversor_kw',0))), 'capex': _solar_num(col('capex_total', r.get('capex_total',0))),
            'economia_mensal': _solar_num(col('economia_mensal', r.get('economia_mensal',0))),
            'economia_anual': _solar_num(col('economia_mensal', r.get('economia_mensal',0))) * 12,
            'payback': _solar_num(col('payback_anos', r.get('payback_anos',0))), 'co2': _solar_num(col('co2_t_ano', r.get('co2_t_ano',0))),
            'energia_anual': _solar_num(col('producao_anual_kwh', r.get('producao_anual',0))),
            'url': url_for('solar_projeto_detalhe', pid=row['id']),
            'csv_url': url_for('solar_export_csv', pid=row['id'])
        })
    return projetos

def _solar_get_lamp_projects(limit=None):
    with _db_conn() as conn:
        try:
            conn.execute("""CREATE TABLE IF NOT EXISTS solar_lampadas (
                id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT NOT NULL, nome TEXT, obs TEXT, resultado_json TEXT NOT NULL
            )""")
            sql = "SELECT id, created_at, nome, obs, resultado_json FROM solar_lampadas ORDER BY id DESC"
            if limit:
                sql += " LIMIT " + str(int(limit))
            rows = conn.execute(sql).fetchall()
        except Exception:
            rows = []
    projetos=[]
    for row in rows:
        r = _solar_json_load(row['resultado_json'])
        projetos.append({
            'tipo': 'Iluminação', 'uid': f"lamp:{row['id']}", 'id': row['id'], 'nome': row['nome'] or r.get('nome_projeto') or f"Projeto Iluminação #{row['id']}",
            'criado': row['created_at'], 'local': r.get('classe_nome') or r.get('modo_implantacao') or '', 'sistema': 'Lâmpadas solares autónomas',
            'kwp': 0.0, 'paineis': int(_solar_num(r.get('qtd_postes'),0)), 'postes': int(_solar_num(r.get('qtd_postes'),0)),
            'painel_wp': _solar_num(r.get('painel_wp')), 'bateria_wh': _solar_num(r.get('bateria_wh_bruto')), 'capex': _solar_num(r.get('capex_estimado')),
            'economia_mensal': _solar_num(r.get('economia_anual_mt')) / 12 if _solar_num(r.get('economia_anual_mt')) else 0,
            'economia_anual': _solar_num(r.get('economia_anual_mt')), 'payback': _solar_num(r.get('payback_anos')), 'co2': _solar_num(r.get('co2_t_ano')),
            'energia_anual': _solar_num(r.get('energia_anual_kwh')),
            'url': url_for('solar_lampadas_projeto', pid=row['id']),
            'csv_url': url_for('solar_lampadas_export', pid=row['id'])
        })
    return projetos

def _solar_portfolio_data():
    fv = _solar_get_fv_projects()
    lamps = _solar_get_lamp_projects()
    todos = fv + lamps
    stats = {
        'total': len(todos), 'fv_count': len(fv), 'lamp_count': len(lamps),
        'kwp_total': sum(p.get('kwp',0) for p in fv),
        'postes_total': sum(p.get('postes',0) for p in lamps),
        'capex_total': sum(p.get('capex',0) for p in todos),
        'economia_anual_total': sum(p.get('economia_anual',0) for p in todos),
        'co2_total': sum(p.get('co2',0) for p in todos),
        'energia_anual_total': sum(p.get('energia_anual',0) for p in todos),
    }
    stats['payback_medio'] = (stats['capex_total'] / stats['economia_anual_total']) if stats['economia_anual_total'] > 0 else 0
    ranking = sorted(todos, key=lambda x: x.get('economia_anual',0), reverse=True)[:10]
    alertas=[]
    if stats['total'] == 0:
        alertas.append(('info','Ainda não existem simulações solares guardadas.'))
    if any((p.get('capex',0) <= 0) for p in todos):
        alertas.append(('warning','Existem projectos sem CAPEX informado; isso reduz a precisão financeira do portfólio.'))
    if any((p.get('payback',0) > 10) for p in todos if p.get('payback',0)):
        alertas.append(('warning','Há projectos com payback superior a 10 anos; convém rever custos, tarifa evitada e cobertura solar.'))
    if stats['economia_anual_total'] > 0 and stats['payback_medio'] <= 5:
        alertas.append(('success','O portfólio apresenta retorno global atractivo pela simulação guardada.'))
    return {'fv': fv, 'lamps': lamps, 'todos': todos, 'stats': stats, 'ranking': ranking, 'alertas': alertas}

@app.route('/solar/portfolio')
def solar_portfolio():
    data = _solar_portfolio_data()
    return render_template('solar_portfolio.html', **data)

@app.route('/solar/relatorio')
def solar_relatorio_consolidado():
    data = _solar_portfolio_data()
    return render_template('solar_relatorio_consolidado.html', **data, gerado_em=datetime.now().strftime('%d/%m/%Y %H:%M'))

@app.route('/solar/portfolio/export.csv')
def solar_portfolio_export_csv():
    data = _solar_portfolio_data()
    si = StringIO(); w = csv.writer(si, delimiter=';')
    w.writerow(['tipo','id','nome','local/classe','sistema','potencia_kWp','postes_ou_paineis','energia_anual_kWh','economia_anual_MT','capex_MT','payback_anos','co2_t_ano'])
    for p in data['todos']:
        w.writerow([p.get('tipo'), p.get('id'), p.get('nome'), p.get('local'), p.get('sistema'), p.get('kwp',0), p.get('postes') or p.get('paineis'), p.get('energia_anual',0), p.get('economia_anual',0), p.get('capex',0), p.get('payback',0), p.get('co2',0)])
    return Response(si.getvalue(), mimetype='text/csv', headers={'Content-Disposition':'attachment;filename=portfolio_solar.csv'})

@app.route('/solar/comparador')
def solar_comparador():
    data = _solar_portfolio_data()
    opcoes = data['todos']
    a_uid = request.args.get('a') or (opcoes[0]['uid'] if opcoes else '')
    b_uid = request.args.get('b') or (opcoes[1]['uid'] if len(opcoes) > 1 else '')
    mapa = {p['uid']: p for p in opcoes}
    a = mapa.get(a_uid)
    b = mapa.get(b_uid)
    metricas = [
        ('Energia anual', 'energia_anual', 'kWh/ano'),
        ('Economia anual', 'economia_anual', 'MT/ano'),
        ('CAPEX', 'capex', 'MT'),
        ('Payback', 'payback', 'anos'),
        ('CO₂ evitado', 'co2', 't/ano'),
        ('Potência FV', 'kwp', 'kWp'),
        ('Postes', 'postes', 'un'),
    ]
    return render_template('solar_comparador.html', opcoes=opcoes, a=a, b=b, a_uid=a_uid, b_uid=b_uid, metricas=metricas)



# === Energia Solar: decisão executiva, premissas e proposta técnica ===
def _solar_to_float(v, default=0.0):
    try:
        if v is None or v == '':
            return default
        return float(str(v).replace(' ', '').replace(',', '.'))
    except Exception:
        return default

def _solar_decision_data():
    data = _solar_portfolio_data()
    projetos = []
    max_economia = max([_solar_to_float(p.get('economia_anual')) for p in data.get('todos', [])] + [1.0])
    max_co2 = max([_solar_to_float(p.get('co2')) for p in data.get('todos', [])] + [1.0])
    for p in data.get('todos', []):
        payback = _solar_to_float(p.get('payback'))
        economia = _solar_to_float(p.get('economia_anual'))
        co2 = _solar_to_float(p.get('co2'))
        capex = _solar_to_float(p.get('capex'))
        energia = _solar_to_float(p.get('energia_anual'))
        score = 0
        # retorno financeiro
        if payback > 0:
            if payback <= 3: score += 35
            elif payback <= 5: score += 28
            elif payback <= 7: score += 20
            elif payback <= 10: score += 12
            else: score += 5
        # impacto económico relativo
        score += min(25, 25 * (economia / max_economia if max_economia else 0))
        # impacto ambiental relativo
        score += min(15, 15 * (co2 / max_co2 if max_co2 else 0))
        # qualidade dos dados
        if capex > 0: score += 10
        if energia > 0: score += 10
        if p.get('tipo') == 'Fotovoltaico' and _solar_to_float(p.get('kwp')) > 0: score += 5
        if p.get('tipo') != 'Fotovoltaico' and _solar_to_float(p.get('postes')) > 0: score += 5
        score = round(min(100, score), 1)
        if score >= 75:
            classe = 'Prioridade alta'
            decisao = 'Avançar para estudo executivo / proposta comercial'
        elif score >= 55:
            classe = 'Prioridade média'
            decisao = 'Rever premissas e validar custos antes de avançar'
        elif score >= 35:
            classe = 'Prioridade baixa'
            decisao = 'Manter em carteira; optimizar CAPEX, autonomia ou cobertura'
        else:
            classe = 'Incompleto'
            decisao = 'Completar dados técnicos e financeiros'
        riscos=[]
        if capex <= 0: riscos.append('CAPEX não informado')
        if payback <= 0: riscos.append('Payback não calculado')
        if payback > 10: riscos.append('Payback elevado')
        if energia <= 0: riscos.append('Energia anual não estimada')
        if p.get('tipo') == 'Fotovoltaico' and _solar_to_float(p.get('kwp')) <= 0: riscos.append('Potência FV não calculada')
        if p.get('tipo') != 'Fotovoltaico' and _solar_to_float(p.get('postes')) <= 0: riscos.append('Quantidade de postes não definida')
        pp = dict(p)
        pp.update({'score': score, 'classe': classe, 'decisao': decisao, 'riscos': riscos})
        projetos.append(pp)
    projetos.sort(key=lambda x: x.get('score',0), reverse=True)
    data['projetos_decisao'] = projetos
    data['score_medio'] = round(sum([p.get('score',0) for p in projetos]) / len(projetos), 1) if projetos else 0
    data['alta_prioridade'] = len([p for p in projetos if p.get('score',0) >= 75])
    data['incompletos'] = len([p for p in projetos if p.get('score',0) < 35])
    data['gerado_em'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    return data

@app.route('/solar/decisao')
def solar_decisao():
    data = _solar_decision_data()
    return render_template('solar_decisao.html', **data)

@app.route('/solar/premissas')
def solar_premissas():
    return render_template('solar_premissas.html', gerado_em=datetime.now().strftime('%d/%m/%Y %H:%M'))

@app.route('/solar/proposta')
def solar_proposta():
    data = _solar_decision_data()
    return render_template('solar_proposta_executiva.html', **data)

@app.route('/solar/portfolio/export.json')
def solar_portfolio_export_json():
    data = _solar_portfolio_data()
    payload = {
        'gerado_em': datetime.now().isoformat(timespec='seconds'),
        'stats': data.get('stats', {}),
        'projetos': data.get('todos', [])
    }
    return Response(json.dumps(payload, ensure_ascii=False, indent=2), mimetype='application/json', headers={'Content-Disposition':'attachment;filename=portfolio_solar_sge.json'})

# === API: config por local (JSON) (mantida) ===
@app.route('/api/local_cfg/<int:local_id>')
def api_local_cfg(local_id):
    local = get_local_by_id(local_id)
    if not local:
        return {"ok": False, "error": "Local não encontrado"}, 404
    cfg = get_local_cfg_full(local_id)
    return {"ok": True, "local_id": local_id, "local_nome": local[1], "cfg": cfg}, 200

# === NOVO: API Locais JSON (mantida) ===
@app.route('/api/locais.json')
def api_locais_json():
    incluir_inativos = (request.args.get('inativos') == '1')
    q = (request.args.get('q') or '').strip()
    data = get_locais_with_cfg(q if q else None, incluir_inativos=incluir_inativos)
    return {"ok": True, "locais": data}

# ==============================

# ==============================

@app.route('/equipamentos/remover/<int:equipamento_id>')
def remover_equipamento(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET deleted_at=datetime('now','localtime'), updated_at=datetime('now','localtime') WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "arquivar", "")
    flash("Equipamento removido da lista ativa.", "warning")
    return redirect(url_for('listar_equipamentos'))



@app.route('/equipamentos/export/csv')
def exportar_equipamentos_csv():
    q = request.args.get('q','').strip()
    local_id = request.args.get('local_id','').strip()
    where = []; params=[]
    if q:
        like = f"%{q}%"
        where.append("(e.nome LIKE ? OR e.tag LIKE ? OR e.especificacao LIKE ?)")
        params += [like, like, like]
    if local_id and local_id.isdigit():
        where.append("e.local_id=?"); params.append(int(local_id))
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.id, e.nome, COALESCE(l.nome,''), e.tag, e.especificacao,
               COALESCE(e.ano_instalacao,''), COALESCE(e.quantidade,0)
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    rows = c.fetchall(); conn.close()

    output = io.StringIO()
    writer = csv.writer(output, lineterminator='\n')
    writer.writerow(["ID","Nome","Local","TAG","Especificação","Ano instalação","Quantidade"])
    for r in rows:
        writer.writerow(r)
    csv_data = output.getvalue().encode('utf-8')
    return Response(csv_data, mimetype='text/csv',
                    headers={"Content-Disposition":"attachment; filename=equipamentos.csv"})



@app.route('/equipamentos/export/pdf')
def exportar_equipamentos_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm

    q = request.args.get('q','').strip()
    local_id = request.args.get('local_id','').strip()
    where = []; params=[]
    if q:
        like = f"%{q}%"
        where.append("(e.nome LIKE ? OR e.tag LIKE ? OR e.especificacao LIKE ?)")
        params += [like, like, like]
    if local_id and local_id.isdigit():
        where.append("e.local_id=?"); params.append(int(local_id))
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.id, e.nome, COALESCE(l.nome,''), e.tag, e.especificacao,
               COALESCE(e.ano_instalacao,''), COALESCE(e.quantidade,0)
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    rows = c.fetchall(); conn.close()

    buffer = io.BytesIO()
    page_size = landscape(A4)
    cpdf = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    title = f"Equipamentos - Relatório"
    cpdf.setFont("Helvetica-Bold", 16)
    cpdf.drawString(2*cm, height-1.5*cm, title)
    cpdf.setFont("Helvetica", 9)
    cpdf.drawString(2*cm, height-2.2*cm, f"Filtro: q='{q}'  local_id='{local_id}'")

    # Table header
    headers = ["ID","Nome","Local","TAG","Especificação","Ano","Qtd"]
    col_x = [1.0*cm, 2.5*cm, 8.5*cm, 12.5*cm, 15.0*cm, 24.0*cm, 26.0*cm]
    y = height - 3.0*cm
    cpdf.setFont("Helvetica-Bold", 8)
    for i, h in enumerate(headers):
        cpdf.drawString(col_x[i], y, h)
    cpdf.line(1*cm, y-0.2*cm, width-1*cm, y-0.2*cm)
    y -= 0.5*cm
    cpdf.setFont("Helvetica", 8)

    for r in rows:
        if y < 1.5*cm:
            cpdf.showPage()
            cpdf.setFont("Helvetica-Bold", 8)
            y = height - 1.5*cm
            for i, h in enumerate(headers):
                cpdf.drawString(col_x[i], y, h)
            cpdf.line(1*cm, y-0.2*cm, width-1*cm, y-0.2*cm)
            y -= 0.5*cm
            cpdf.setFont("Helvetica", 8)

        values = [str(r[0]), r[1] or "", r[2] or "", r[3] or "", (r[4] or "")[:70], str(r[5] or ""), str(r[6] or "")]
        for i, val in enumerate(values):
            cpdf.drawString(col_x[i], y, val)
        y -= 0.45*cm

    cpdf.showPage()
    cpdf.save()
    pdf_data = buffer.getvalue()
    buffer.close()
    return Response(pdf_data, mimetype='application/pdf',
                    headers={"Content-Disposition":"attachment; filename=equipamentos.pdf"})



def log_equip_audit(equipamento_id, acao, detalhes=""):
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("INSERT INTO mecanismos_dummy (x) VALUES (1)") if False else None  # no-op to keep pattern
        c.execute("INSERT INTO equipamentos_audit (equipamento_id, acao, detalhes) VALUES (?, ?, ?)",
                  (equipamento_id, acao, detalhes))
        conn.commit(); conn.close()
    except Exception:
        pass


@app.route('/equipamentos/desativar/<int:equipamento_id>')
def desativar_equipamento(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET ativo=0, updated_at=datetime('now','localtime') WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "desativar", "")
    flash("Equipamento desativado.", "warning")
    return redirect(url_for('listar_equipamentos'))

@app.route('/equipamentos/ativar/<int:equipamento_id>')
def ativar_equipamento(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET ativo=1, updated_at=datetime('now','localtime') WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "ativar", "")
    flash("Equipamento reativado.", "success")
    return redirect(url_for('listar_equipamentos'))


@app.route('/equipamentos/<int:equipamento_id>', methods=['GET', 'POST'])
def equipamento_detalhe(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''
        SELECT e.id, e.nome, COALESCE(l.nome,''), COALESCE(e.tag,''), COALESCE(e.especificacao,''),
               COALESCE(e.ano_instalacao,''), COALESCE(e.quantidade,0), COALESCE(e.ativo,1),
               e.created_at, e.updated_at, COALESCE(e.categoria,''), COALESCE(e.fabricante,''),
               COALESCE(e.modelo,''), COALESCE(e.numero_serie,''), COALESCE(e.custo_aquisicao,0.0),
               COALESCE(e.vida_util_anos,''), COALESCE(e.criticidade,''), COALESCE(e.potencia_kw,''),
               COALESCE(e.tensao_v,''), COALESCE(e.corrente_a,''), COALESCE(e.fornecedor,''),
               COALESCE(e.contrato_num,''), COALESCE(e.garantia_fim,'')
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        WHERE e.id=?
    ''', (equipamento_id,))
    eq = c.fetchone()

    if not eq:
        conn.close()
        flash("Equipamento não encontrado.", "warning")
        return redirect(url_for('listar_equipamentos'))

    c.execute('SELECT id, original_name, filename, mime, size, uploaded_at FROM equipamentos_files WHERE equipamento_id=? ORDER BY uploaded_at DESC', (equipamento_id,))
    files = c.fetchall()
    c.execute('SELECT id, filename, thumb_filename, caption FROM equipamentos_photos WHERE equipamento_id=? ORDER BY uploaded_at DESC', (equipamento_id,))
    photos = c.fetchall()

    try:
        c.execute('SELECT tensao_nominal, corrente_nominal, potencia_nominal_kw, fp_nominal, eficiencia_nominal FROM equipamentos_cfg WHERE equipamento_id=?', (equipamento_id,))
        cfg = c.fetchone()
    except Exception:
        cfg = None

    conn.close()

    detalhe = {
        'id': eq[0],
        'nome': _equip_clean_text(eq[1]),
        'local': _equip_clean_text(eq[2]),
        'tag': _equip_clean_text(eq[3]),
        'especificacao': _equip_clean_text(eq[4]),
        'ano': _equip_clean_text(eq[5]),
        'quantidade': int(eq[6] or 0),
        'ativo': _equip_bool(eq[7]),
        'categoria': _equip_clean_text(eq[10]),
        'fabricante': _equip_clean_text(eq[11]),
        'modelo': _equip_clean_text(eq[12]),
        'numero_serie': _equip_clean_text(eq[13]),
        'custo': _equip_clean_number(eq[14], decimals=2, fallback='0.00', zero_as_value=True),
        'vida_util': _equip_clean_text(eq[15]),
        'criticidade': _equip_clean_text(eq[16]),
        'potencia_kw': _equip_clean_number(eq[17], decimals=2),
        'tensao_v': _equip_clean_number(eq[18], decimals=0),
        'corrente_a': _equip_clean_number(eq[19], decimals=2),
        'fornecedor': _equip_clean_text(eq[20]),
        'contrato_num': _equip_clean_text(eq[21]),
        'garantia_fim': _equip_clean_text(eq[22]),
        'created_at': _equip_clean_text(eq[8]),
        'updated_at': _equip_clean_text(eq[9]),
    }
    return render_template('equipamento_detalhe.html', eq=eq, detalhe=detalhe, files=files, photos=photos, cfg=cfg)

# Upload endpoint
@app.route('/equipamentos/<int:equipamento_id>/upload', methods=['POST'])
def equipamento_upload(equipamento_id):
    file = request.files.get('file')
    if not file or file.filename == '':
        flash("Nenhum ficheiro selecionado.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    filename = secure_filename(file.filename)
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    file.save(save_path)
    size = os.path.getsize(save_path)
    mime = file.mimetype or ""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('INSERT INTO equipamentos_files (equipamento_id, filename, original_name, mime, size) VALUES (?,?,?,?,?)',
              (equipamento_id, filename, file.filename, mime, size))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "upload", file.filename)
    flash("Ficheiro carregado.", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/files/<path:filename>')
def equipamento_download(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)


@app.route('/equipamentos/import', methods=['GET', 'POST'])
def importar_equipamentos():
    if request.method == 'POST':
        file = request.files.get('csv')
        if not file or file.filename == '':
            flash("Selecione um ficheiro CSV.", "danger")
            return redirect(url_for('importar_equipamentos'))
        import csv, io
        content = file.read().decode('utf-8', errors='ignore')
        reader = csv.DictReader(io.StringIO(content))
        inserted = 0; skipped = 0
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        for row in reader:
            nome = (row.get('nome') or row.get('Nome') or '').strip()
            if not nome:
                skipped += 1; continue
            local_nome = (row.get('local') or row.get('Local') or '').strip()
            # resolve local_id by name if provided
            local_id = None
            if local_nome:
                c.execute("SELECT id FROM locais WHERE nome=?", (local_nome,))
                r = c.fetchone()
                if r: local_id = r[0]
            tag = (row.get('tag') or row.get('TAG') or '').strip()
            especificacao = (row.get('especificacao') or row.get('Especificacao') or row.get('Especificação') or '').strip()
            ano = (row.get('ano_instalacao') or row.get('ano') or '').strip()
            qtd = (row.get('quantidade') or row.get('qtd') or '1').strip()
            try:
                qtd_i = int(qtd)
            except Exception:
                qtd_i = 1

            c.execute('''
                INSERT INTO equipamentos (nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'), datetime('now','localtime'))
            ''', (nome, local_id, tag, especificacao, ano or None, qtd_i))
            inserted += 1
        conn.commit(); conn.close()
        flash(f"Importação concluída: {inserted} inseridos, {skipped} ignorados.", "success")
        return redirect(url_for('listar_equipamentos'))
    return render_template('importar_equipamentos.html')





@app.route('/equipamentos/<int:equipamento_id>/upload_photo', methods=['POST'])
def equipamento_upload_photo(equipamento_id):
    file = request.files.get('photo')
    if not file or file.filename == '':
        flash("Selecione uma imagem.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    if not (file.mimetype or "").lower().startswith("image/"):
        flash("O ficheiro deve ser uma imagem.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    filename, thumb_name, w, h = _save_image_and_thumb(file, dest_dir=app.config['UPLOAD_FOLDER']+'/photos', thumb_dir=app.config['UPLOAD_FOLDER']+'/thumbs')
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('INSERT INTO equipamentos_photos (equipamento_id, filename, thumb_filename, width, height) VALUES (?,?,?,?,?)',
              (equipamento_id, 'photos/'+filename, ('thumbs/'+thumb_name if thumb_name else None), w, h))
    photo_id = c.lastrowid
    # set as cover if not set
    c.execute('SELECT cover_photo_id FROM equipamentos WHERE id=?', (equipamento_id,))
    r = c.fetchone()
    if not r or not r[0]:
        c.execute('UPDATE equipamentos SET cover_photo_id=? WHERE id=?', (photo_id, equipamento_id))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "upload_foto", filename)
    flash("Foto adicionada.", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/<int:equipamento_id>/upload_photos', methods=['POST'])
def equipamento_upload_photos(equipamento_id):
    files = request.files.getlist('photos[]')
    if not files:
        flash("Selecione uma ou mais imagens.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    added = 0
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    for file in files:
        if not file or file.filename == '': continue
        if not (file.mimetype or "").lower().startswith("image/"): continue
        filename, thumb_name, w, h = _save_image_and_thumb(file, dest_dir=app.config['UPLOAD_FOLDER']+'/photos', thumb_dir=app.config['UPLOAD_FOLDER']+'/thumbs')
        c.execute('INSERT INTO equipamentos_photos (equipamento_id, filename, thumb_filename, width, height) VALUES (?,?,?,?,?)',
                  (equipamento_id, 'photos/'+filename, ('thumbs/'+thumb_name if thumb_name else None), w, h))
        if added == 0:
            # set cover if empty
            c.execute('UPDATE equipamentos SET cover_photo_id=COALESCE(cover_photo_id, last_insert_rowid()) WHERE id=?', (equipamento_id,))
        added += 1
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "upload_fotos_multi", f"{added} fotos")
    flash(f"{added} foto(s) adicionada(s).", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/<int:equipamento_id>/photo/<int:photo_id>/delete')
def equipamento_photo_delete(equipamento_id, photo_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT filename, thumb_filename FROM equipamentos_photos WHERE id=? AND equipamento_id=?', (photo_id, equipamento_id))
    r = c.fetchone()
    if r:
        fn, tfn = r[0], r[1]
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], fn))
        except Exception: pass
        try:
            if tfn: os.remove(os.path.join(app.config['UPLOAD_FOLDER'], tfn))
        except Exception: pass
    c.execute('DELETE FROM equipamentos_photos WHERE id=? AND equipamento_id=?', (photo_id, equipamento_id))
    # if it was cover, unset
    c.execute('UPDATE equipamentos SET cover_photo_id=NULL WHERE id=? AND cover_photo_id=?', (equipamento_id, photo_id))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "apagar_foto", str(photo_id))
    flash("Foto removida.", "warning")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/<int:equipamento_id>/photo/<int:photo_id>/cover')
def equipamento_photo_cover(equipamento_id, photo_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('UPDATE equipamentos SET cover_photo_id=? WHERE id=?', (photo_id, equipamento_id))
    conn.commit(); conn.close()
    log_equip_audit(equipamento_id, "definir_capa", str(photo_id))
    flash("Foto definida como capa.", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))



@app.route('/equipamentos/<int:equipamento_id>/label')
def equipamento_label(equipamento_id):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A7, landscape
    from reportlab.lib.units import mm
    from reportlab.graphics.barcode import qr
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id, nome, tag FROM equipamentos WHERE id=?', (equipamento_id,))
    r = c.fetchone(); conn.close()
    if not r:
        flash("Equipamento não encontrado.", "warning")
        return redirect(url_for('listar_equipamentos'))
    _id, nome, tag = r
    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=landscape(A7))
    w, h = landscape(A7)
    cpdf.setFont("Helvetica-Bold", 10)
    cpdf.drawString(10*mm, h-10*mm, f"EQ #{_id}")
    cpdf.setFont("Helvetica", 9)
    cpdf.drawString(10*mm, h-16*mm, (nome or "")[:40])
    if tag:
        cpdf.drawString(10*mm, h-22*mm, f"TAG: {tag[:30]}")
    # QR code content
    code_val = f"SGE:EQ:{_id}"
    qr_code = qr.QrCodeWidget(code_val)
    bounds = qr_code.getBounds()
    size = 30*mm
    width = bounds[2] - bounds[0]
    height = bounds[3] - bounds[1]
    d = size / max(width, height)
    from reportlab.graphics.shapes import Drawing
    drawing = Drawing(size, size, transform=[d,0,0,d,0,0])
    drawing.add(qr_code)
    from reportlab.graphics import renderPDF
    renderPDF.draw(drawing, cpdf, w-40*mm, h-40*mm)
    cpdf.showPage(); cpdf.save()
    pdf_data = buf.getvalue(); buf.close()
    return Response(pdf_data, mimetype='application/pdf',
                    headers={"Content-Disposition": f"attachment; filename=label_eq_{_id}.pdf"})

import openpyxl


@app.route('/equipamentos/bulk', methods=['POST'])
def equipamentos_bulk():
    action = request.form.get('action')
    ids = request.form.getlist('ids')
    ids = [int(x) for x in ids if x.isdigit()]
    if not ids:
        flash("Selecione pelo menos um item.", "warning")
        return redirect(url_for('listar_equipamentos'))
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()

    if action == 'ativar':
        c.executemany("UPDATE equipamentos SET ativo=1, updated_at=datetime('now','localtime') WHERE id=?", [(i,) for i in ids])
        conn.commit(); conn.close()
        flash(f"{len(ids)} ativado(s).", "success")
        return redirect(url_for('listar_equipamentos'))
    if action == 'desativar':
        c.executemany("UPDATE equipamentos SET ativo=0, updated_at=datetime('now','localtime') WHERE id=?", [(i,) for i in ids])
        conn.commit(); conn.close()
        flash(f"{len(ids)} desativado(s).", "warning")
        return redirect(url_for('listar_equipamentos'))
    if action == 'remover':
        c.executemany("UPDATE equipamentos SET deleted_at=datetime('now','localtime'), updated_at=datetime('now','localtime') WHERE id=?", [(i,) for i in ids])
        conn.commit(); conn.close()
        flash(f"{len(ids)} removido(s) da lista ativa.", "warning")
        return redirect(url_for('listar_equipamentos'))
    if action == 'labels':
        conn.close()
        ids_str = ",".join([str(i) for i in ids])
        return redirect(url_for('equipamentos_labels_pdf', ids=ids_str))
    if action == 'export_csv':
        conn.close()
        ids_str = ",".join([str(i) for i in ids])
        return redirect(url_for('exportar_equipamentos_csv') + f"?ids={ids_str}")
    if action == 'export_pdf':
        conn.close()
        ids_str = ",".join([str(i) for i in ids])
        return redirect(url_for('exportar_equipamentos_pdf') + f"?ids={ids_str}")

    conn.close()
    flash("Ação desconhecida.", "danger")
    return redirect(url_for('listar_equipamentos'))


@app.route('/equipamentos/duplicados')
def equipamentos_duplicados():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    # Duplicados por (nome, local)
    c.execute('''
        SELECT e.nome, e.local_id, COUNT(*)
        FROM equipamentos e
        GROUP BY e.nome, e.local_id
        HAVING COUNT(*) > 1
        ORDER BY COUNT(*) DESC, e.nome
    ''')
    by_nome_local = c.fetchall()

    # Duplicados por TAG
    c.execute('''
        SELECT e.tag, COUNT(*)
        FROM equipamentos e
        WHERE e.tag IS NOT NULL AND e.tag <> ''
        GROUP BY e.tag
        HAVING COUNT(*) > 1
        ORDER BY COUNT(*) DESC, e.tag
    ''')
    by_tag = c.fetchall()

    # Duplicados por número de série
    try:
        c.execute('''
            SELECT e.numero_serie, COUNT(*)
            FROM equipamentos e
            WHERE e.numero_serie IS NOT NULL AND e.numero_serie <> ''
            GROUP BY e.numero_serie
            HAVING COUNT(*) > 1
            ORDER BY COUNT(*) DESC, e.numero_serie
        ''')
        by_serial = c.fetchall()
    except Exception:
        by_serial = []

    conn.close()
    return render_template('equipamentos_duplicados.html',
                           by_nome_local=by_nome_local, by_tag=by_tag, by_serial=by_serial)


@app.route('/equipamentos/export/detalhe/<int:equipamento_id>.pdf')
def equipamento_detalhe_pdf(equipamento_id):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('''
        SELECT e.id, e.nome, COALESCE(l.nome,''), e.tag, e.especificacao, e.ano_instalacao,
               e.quantidade, e.categoria, e.fabricante, e.modelo, e.numero_serie, e.custo_aquisicao,
               e.vida_util_anos, e.criticidade, e.created_at, e.updated_at
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        WHERE e.id=?
    ''', (equipamento_id,))
    eq = c.fetchone()

    cover = None
    c.execute('SELECT filename FROM equipamentos_photos WHERE id=(SELECT cover_photo_id FROM equipamentos WHERE id=?)', (equipamento_id,))
    r = c.fetchone()
    if r: cover = r[0]
    conn.close()

    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    cpdf.setFont("Helvetica-Bold", 14)
    cpdf.drawString(2*cm, h-2*cm, f"Ficha do Equipamento #{equipamento_id}")
    cpdf.setFont("Helvetica", 10)
    y = h-3*cm

    labels = ["Nome","Local","TAG","Especificação","Ano","Qtd","Categoria","Fabricante","Modelo","Nº Série","Custo (MZN)","Vida Útil (anos)","Criticidade","Criado","Atualizado"]
    vals = [eq[1],eq[2],eq[3],eq[4],eq[5],eq[6],eq[7],eq[8],eq[9],eq[10],eq[11],eq[12],eq[13],eq[14],eq[15]]
    for L,V in zip(labels,vals):
        cpdf.drawString(2*cm, y, f"{L}: {V if V is not None else '-'}")
        y -= 0.6*cm
        if y < 4*cm:
            cpdf.showPage(); y = h-2*cm

    if cover:
        try:
            from reportlab.lib.utils import ImageReader
            img = ImageReader(os.path.join(app.config['UPLOAD_FOLDER'], cover))
            cpdf.drawImage(img, 2*cm, 2*cm, width=12*cm, height=8*cm, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    cpdf.showPage(); cpdf.save()
    pdf = buf.getvalue(); buf.close()
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=equip_{equipamento_id}.pdf"})


@app.route('/equipamentos/labels.pdf')
def equipamentos_labels_pdf():
    ids_str = request.args.get('ids','').strip()
    if not ids_str:
        flash("Nenhum ID selecionado.", "warning")
        return redirect(url_for('listar_equipamentos'))
    try:
        ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
    except Exception:
        ids = []
    if not ids:
        flash("IDs inválidos.", "danger")
        return redirect(url_for('listar_equipamentos'))

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF
    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    # grid 3x8 (aprox) de etiquetas
    cols, rows = 3, 8
    margin_x, margin_y = 10*mm, 10*mm
    cell_w = (w - 2*margin_x) / cols
    cell_h = (h - 2*margin_y) / rows

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f"SELECT id, nome, tag FROM equipamentos WHERE id IN ({','.join(['?']*len(ids))}) ORDER BY id", ids)
    rows_data = c.fetchall(); conn.close()

    i = 0
    for r in rows_data:
        x = i % cols
        y = i // cols
        if y >= rows:
            cpdf.showPage()
            y = 0; i = 0
        pos_x = margin_x + x*cell_w
        pos_y = h - margin_y - (y+1)*cell_h
        cpdf.rect(pos_x, pos_y, cell_w, cell_h)
        cpdf.setFont("Helvetica-Bold", 10)
        cpdf.drawString(pos_x+5*mm, pos_y+cell_h-10*mm, f"EQ #{r[0]}")
        cpdf.setFont("Helvetica", 9)
        cpdf.drawString(pos_x+5*mm, pos_y+cell_h-16*mm, (r[1] or "")[:32])
        if r[2]:
            cpdf.drawString(pos_x+5*mm, pos_y+cell_h-22*mm, f"TAG: {r[2][:24]}")
        qr_code = qr.QrCodeWidget(f"SGE:EQ:{r[0]}")
        bounds = qr_code.getBounds()
        size = min(cell_w, cell_h) * 0.5
        width = bounds[2] - bounds[0]; height = bounds[3] - bounds[1]
        scale = size / max(width, height)
        drawing = Drawing(size, size, transform=[scale,0,0,scale,0,0])
        drawing.add(qr_code)
        renderPDF.draw(drawing, cpdf, pos_x + cell_w - size - 5*mm, pos_y + 5*mm)
        i += 1

    cpdf.showPage(); cpdf.save()
    data = buf.getvalue(); buf.close()
    return Response(data, mimetype="application/pdf",
                    headers={"Content-Disposition":"attachment; filename=labels_equipamentos.pdf"})


@app.route('/equipamentos/import/xlsx', methods=['GET','POST'])
def importar_equipamentos_xlsx():
    if request.method == 'POST':
        file = request.files.get('xlsx')
        if not file or file.filename == '':
            flash("Selecione um ficheiro XLSX.", "danger")
            return redirect(url_for('importar_equipamentos_xlsx'))
        from openpyxl import load_workbook
        data = io.BytesIO(file.read())
        wb = load_workbook(filename=data, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h:i for i,h in enumerate(headers)}
        def get(row, key):
            i = idx.get(key)
            return (str(row[i].value).strip() if (i is not None and row[i].value is not None) else '')
        inserted=0; skipped=0
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        for row in ws.iter_rows(min_row=2):
            nome = get(row,'nome')
            if not nome: skipped+=1; continue
            local_nome = get(row,'local')
            local_id = None
            if local_nome:
                c.execute("SELECT id FROM locais WHERE nome=?", (local_nome,))
                r = c.fetchone()
                if r: local_id = r[0]
            tag = get(row,'tag')
            especificacao = get(row,'especificacao') or get(row,'especificação')
            ano = get(row,'ano') or get(row,'ano_instalacao')
            quantidade = get(row,'quantidade') or '1'
            categoria = get(row,'categoria')
            fabricante = get(row,'fabricante')
            modelo = get(row,'modelo')
            numero_serie = get(row,'numero_serie') or get(row,'nº série')
            custo_aquisicao = get(row,'custo_aquisicao')
            vida_util_anos = get(row,'vida_util_anos')
            criticidade = get(row,'criticidade')

            try: qtd_i = int(quantidade)
            except: qtd_i = 1
            try: custo_val = float(custo_aquisicao) if custo_aquisicao else None
            except: custo_val = None
            try: vida_val = int(vida_util_anos) if vida_util_anos else None
            except: vida_val = None

            c.execute('''
                INSERT INTO equipamentos (nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo, created_at, updated_at,
                                          categoria, fabricante, modelo, numero_serie, custo_aquisicao, vida_util_anos, criticidade)
                VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'), datetime('now','localtime'),
                        ?, ?, ?, ?, ?, ?, ?)
            ''', (nome, local_id, tag, especificacao, ano or None, qtd_i, categoria, fabricante, modelo, numero_serie, custo_val, vida_val, criticidade))
            inserted += 1
        conn.commit(); conn.close()
        flash(f"Importação XLSX concluída: {inserted} inseridos, {skipped} ignorados.", "success")
        return redirect(url_for('listar_equipamentos'))
    return render_template('importar_equipamentos_xlsx.html')


@app.route('/equipamentos/<int:equipamento_id>/upload_zip', methods=['POST'])
def equipamento_upload_zip(equipamento_id):
    file = request.files.get('zip')
    if not file or file.filename == '':
        flash("Selecione um ZIP.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    zdata = io.BytesIO(file.read())
    try:
        with zipfile.ZipFile(zdata) as z:
            added = 0
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            for name in z.namelist():
                if name.endswith('/'):
                    continue
                data = z.read(name)
                # create a FileStorage-like wrapper
                class _FS:
                    def __init__(self, filename, data):
                        self.filename = filename
                        self.data = data
                        self.mimetype = "image/jpeg"
                    def save(self, path):
                        with open(path, 'wb') as f: f.write(self.data)
                fs = _FS(os.path.basename(name), data)
                filename, thumb_name, w, h = _save_image_and_thumb(fs, dest_dir=app.config['UPLOAD_FOLDER']+'/photos', thumb_dir=app.config['UPLOAD_FOLDER']+'/thumbs')
                c.execute('INSERT INTO equipamentos_photos (equipamento_id, filename, thumb_filename, width, height) VALUES (?,?,?,?,?)',
                          (equipamento_id, 'photos/'+filename, ('thumbs/'+thumb_name if thumb_name else None), w, h))
                added += 1
            conn.commit(); conn.close()
            flash(f"{added} foto(s) importadas do ZIP.", "success")
    except Exception as e:
        flash(f"ZIP inválido: {e}", "danger")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))


# === SETTINGS HELPER ===
def get_setting(key, default='0'):
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("SELECT value FROM settings WHERE key=?", (key,))
        r = c.fetchone(); conn.close()
        return r[0] if r and r[0] is not None else default
    except Exception:
        return default


def _apply_advanced_query(q, where_clauses, params):
    """Operadores suportados:
    fabricante:, categoria:, local:, tag:, modelo:, serie:/nserie:,
    crit:/criticidade:, ano>=, ano<=, além de termos livres.
    """
    import shlex
    try:
        parts = shlex.split(q)
    except Exception:
        parts = q.split()
    rest = []
    for p in parts:
        pl = p.lower()
        if pl.startswith('fabricante:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('categoria:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('local:'):
            v = p.split(':',1)[1]; where_clauses.append("EXISTS (SELECT 1 FROM locais lx WHERE lx.id=e.local_id AND lx.nome LIKE ?)"); params.append(f"%{v}%"); continue
        if pl.startswith('tag:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.tag,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('modelo:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('serie:') or pl.startswith('nserie:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.numero_serie,'') LIKE ?"); params.append(f"%{v}%"); continue
        if pl.startswith('crit:') or pl.startswith('criticidade:'):
            v = p.split(':',1)[1]; where_clauses.append("COALESCE(e.criticidade,'') = ?"); params.append(v); continue
        if pl.startswith('ano>='):
            try:
                v = int(p.split('>=',1)[1]); where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(v); continue
            except: pass
        if pl.startswith('ano<='):
            try:
                v = int(p.split('<=',1)[1]); where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(v); continue
            except: pass
        rest.append(p)
    if rest:
        like_q = "%" + " ".join(rest) + "%"
        where_clauses.append("(e.nome LIKE ? OR e.tag LIKE ? OR e.especificacao LIKE ? OR e.modelo LIKE ? OR e.fabricante LIKE ?)")
        params.extend([like_q, like_q, like_q, like_q, like_q])


# === EQUIPAMENTOS: Export JSON ===
@app.route('/equipamentos/export/json')
def equipamentos_export_json():
    ids = request.args.get('ids','').strip()
    where = []; params = []
    if ids:
        try:
            arr = [int(x) for x in ids.split(',') if x.strip().isdigit()]
            if arr:
                where.append("e.id IN (" + ",".join(["?"]*len(arr)) + ")")
                params += arr
        except Exception:
            pass
    if not where:
        q = request.args.get('q','').strip()
        categoria = request.args.get('categoria','').strip()
        fabricante = request.args.get('fabricante','').strip()
        modelo = request.args.get('modelo','').strip()
        criticidade = request.args.get('criticidade','').strip()
        local_id = request.args.get('local_id','').strip()
        ano_min = request.args.get('ano_min','').strip()
        ano_max = request.args.get('ano_max','').strip()
        incluir_inativos = request.args.get('incluir_inativos','0')=='1'
        if q: _apply_advanced_query(q, where, params)
        if local_id and local_id.isdigit(): where.append("e.local_id=?"); params.append(int(local_id))
        if categoria: where.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
        if fabricante: where.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
        if modelo: where.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
        if criticidade: where.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
        if ano_min and ano_min.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
        if ano_max and ano_max.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
        if not incluir_inativos: where.append("COALESCE(e.ativo,1)=1")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.*, COALESCE(l.nome,'') AS local_nome
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    cols = [d[0] for d in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    conn.close()
    return Response(json.dumps(rows, ensure_ascii=False, indent=2), mimetype="application/json")


# === EQUIPAMENTOS: Import JSON ===
@app.route('/equipamentos/import/json', methods=['GET','POST'])
def equipamentos_import_json():
    if request.method == 'POST':
        file = request.files.get('json')
        if not file or file.filename == '':
            flash("Selecione um ficheiro JSON.", "danger")
            return redirect(url_for('equipamentos_import_json'))
        try:
            data = json.loads(file.read().decode('utf-8', errors='ignore'))
            if not isinstance(data, list):
                raise ValueError("JSON deve ser uma lista de objetos.")
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            inserted=0; skipped=0
            for obj in data:
                nome = (obj.get('nome') or obj.get('NOME') or '').strip()
                if not nome: skipped+=1; continue
                local_id = obj.get('local_id')
                tag = obj.get('tag') or ''
                especificacao = obj.get('especificacao') or obj.get('especificação') or ''
                ano = obj.get('ano_instalacao') or obj.get('ano') or None
                qtd = obj.get('quantidade') or 1
                categoria = obj.get('categoria') or ''
                fabricante = obj.get('fabricante') or ''
                modelo = obj.get('modelo') or ''
                numero_serie = obj.get('numero_serie') or ''
                custo = obj.get('custo_aquisicao')
                vida = obj.get('vida_util_anos')
                criticidade = obj.get('criticidade') or ''
                try: qtd_i = int(qtd)
                except: qtd_i = 1
                try: custo_val = float(custo) if custo is not None else None
                except: custo_val = None
                try: vida_val = int(vida) if vida is not None else None
                except: vida_val = None
                c.execute('''
                    INSERT INTO equipamentos (nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo, created_at, updated_at,
                                              categoria, fabricante, modelo, numero_serie, custo_aquisicao, vida_util_anos, criticidade)
                    VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'), datetime('now','localtime'),
                            ?, ?, ?, ?, ?, ?, ?)
                ''', (nome, local_id, tag, especificacao, ano, qtd_i, categoria, fabricante, modelo, numero_serie, custo_val, vida_val, criticidade))
                inserted+=1
            conn.commit(); conn.close()
            flash(f"Importação JSON concluída: {inserted} inseridos, {skipped} ignorados.", "success")
            return redirect(url_for('listar_equipamentos'))
        except Exception as e:
            flash(f"JSON inválido: {e}", "danger")
            return redirect(url_for('equipamentos_import_json'))
    return render_template('importar_equipamentos_json.html')


# === XLSX template & export filtrado ===
@app.route('/equipamentos/export/xlsx_template')
def equipamentos_xlsx_template():
    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out, {'in_memory': True})
    ws = wb.add_worksheet('Equipamentos')
    headers = ["nome","local","tag","especificacao","ano","quantidade","categoria","fabricante","modelo","numero_serie","custo_aquisicao","vida_util_anos","criticidade"]
    for i,hx in enumerate(headers): ws.write(0, i, hx)
    ws.data_validation(1, 12, 10000, 12, {'validate': 'list', 'source': ['Baixa','Média','Alta']})
    wb.close(); out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=template_equipamentos.xlsx"})


@app.route('/equipamentos/export/xlsx')
def equipamentos_export_xlsx():
    q = request.args.get('q','').strip()
    local_id = request.args.get('local_id','').strip()
    incluir_inativos = request.args.get('incluir_inativos','0')=='1'
    categoria = request.args.get('categoria','').strip()
    fabricante = request.args.get('fabricante','').strip()
    modelo = request.args.get('modelo','').strip()
    criticidade = request.args.get('criticidade','').strip()
    ano_min = request.args.get('ano_min','').strip()
    ano_max = request.args.get('ano_max','').strip()

    where=[]; params=[]
    if q: _apply_advanced_query(q, where, params)
    if local_id and local_id.isdigit(): where.append("e.local_id=?"); params.append(int(local_id))
    if categoria: where.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
    if fabricante: where.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
    if modelo: where.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
    if criticidade: where.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
    if ano_min and ano_min.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
    if ano_max and ano_max.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
    if not incluir_inativos: where.append("COALESCE(e.ativo,1)=1")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.nome, COALESCE(l.nome,''), e.tag, e.especificacao, e.ano_instalacao, e.quantidade,
               e.categoria, e.fabricante, e.modelo, e.numero_serie, e.custo_aquisicao, e.vida_util_anos, e.criticidade
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    rows = c.fetchall(); conn.close()

    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out, {'in_memory': True})
    ws = wb.add_worksheet('Equipamentos')
    headers = ["nome","local","tag","especificacao","ano","quantidade","categoria","fabricante","modelo","numero_serie","custo_aquisicao","vida_util_anos","criticidade"]
    for i,hx in enumerate(headers): ws.write(0, i, hx)
    for r_i, r in enumerate(rows, start=1):
        for c_i, v in enumerate(r): ws.write(r_i, c_i, v)
    ws.data_validation(1, 12, 10000, 12, {'validate': 'list', 'source': ['Baixa','Média','Alta']})
    wb.close(); out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=equipamentos_filtrados.xlsx"})


# === Relatório Consolidado (PDF) ===
@app.route('/equipamentos/export/relatorio.pdf')
def equipamentos_relatorio_pdf():
    q = request.args.get('q','').strip()
    local_id = request.args.get('local_id','').strip()
    incluir_inativos = request.args.get('incluir_inativos','0')=='1'
    categoria = request.args.get('categoria','').strip()
    fabricante = request.args.get('fabricante','').strip()
    modelo = request.args.get('modelo','').strip()
    criticidade = request.args.get('criticidade','').strip()
    ano_min = request.args.get('ano_min','').strip()
    ano_max = request.args.get('ano_max','').strip()

    where_clauses=[]; params=[]
    if q: _apply_advanced_query(q, where_clauses, params)
    if local_id and local_id.isdigit(): where_clauses.append("e.local_id=?"); params.append(int(local_id))
    if categoria: where_clauses.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
    if fabricante: where_clauses.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
    if modelo: where_clauses.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
    if criticidade: where_clauses.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
    if ano_min and ano_min.isdigit(): where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
    if ano_max and ano_max.isdigit(): where_clauses.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
    if not incluir_inativos: where_clauses.append("COALESCE(e.ativo,1)=1")
    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f'''
        SELECT e.id, e.nome, COALESCE(l.nome,''), e.tag, e.especificacao, e.ano_instalacao, e.quantidade,
               COALESCE(e.categoria,''), COALESCE(e.fabricante,''), COALESCE(e.modelo,''), COALESCE(e.criticidade,''),
               COALESCE(e.custo_aquisicao,0.0)
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id = l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    ''', params)
    rows = c.fetchall(); conn.close()

    from collections import Counter
    def summarize(values):
        c = Counter([v or '' for v in values])
        return sorted(c.items(), key=lambda x:(-x[1], x[0]))

    sum_local = summarize([r[2] for r in rows])
    sum_cat = summarize([r[7] for r in rows])
    sum_fab = summarize([r[8] for r in rows])

    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    from reportlab.lib.units import cm
    cpdf.setFont("Helvetica-Bold", 14)
    cpdf.drawString(2*cm, h-2*cm, "Relatório Consolidado - Equipamentos")
    cpdf.setFont("Helvetica", 9)
    cpdf.drawString(2*cm, h-2.7*cm, f"Filtros: q='{q}' local_id='{local_id}' cat='{categoria}' fab='{fabricante}' mod='{modelo}' crit='{criticidade}' ano[{ano_min},{ano_max}]")

    y = h-3.5*cm
    cpdf.setFont("Helvetica-Bold", 10); cpdf.drawString(2*cm, y, "Sumário por Local"); y -= 0.5*cm
    cpdf.setFont("Helvetica", 9)
    for k,v in sum_local[:22]:
        cpdf.drawString(2*cm, y, f"{k or '-'}: {v}"); y -= 0.4*cm
        if y < 3*cm: cpdf.showPage(); y=h-2*cm

    cpdf.setFont("Helvetica-Bold", 10); cpdf.drawString(10*cm, h-3.5*cm, "Sumário por Categoria")
    y2 = h-4.0*cm; cpdf.setFont("Helvetica", 9)
    for k,v in sum_cat[:22]:
        cpdf.drawString(10*cm, y2, f"{k or '-'}: {v}"); y2 -= 0.4*cm

    cpdf.showPage()
    cpdf.setFont("Helvetica-Bold", 10)
    cpdf.drawString(2*cm, h-2*cm, "Lista Detalhada")
    y = h-2.8*cm; cpdf.setFont("Helvetica", 8)
    headers = ["ID","Nome","Local","TAG","Ano","Qtd","Cat","Fab","Mod","Crit","Custo","Custo Total"]
    col = [1.0, 2.0, 7.0, 11.0, 15.0, 16.5, 18.0, 20.0, 23.0, 26.0, 28.0, 31.0]
    for i,hx in enumerate(headers): cpdf.drawString(col[i]*cm, y, hx)
    y -= 0.4*cm
    total_custo = 0.0
    for r in rows:
        custo = float(r[11] or 0.0); qtd = int(r[6] or 0)
        custo_total = custo * qtd; total_custo += custo_total
        vals = [r[0], r[1], r[2], r[3] or "", r[5] or "", qtd, r[7] or "", r[8] or "", r[9] or "", r[10] or "", f"{custo:.2f}", f"{custo_total:.2f}"]
        for i,v in enumerate(vals):
            cpdf.drawString(col[i]*cm, y, str(v)[:18])
        y -= 0.35*cm
        if y < 2.5*cm:
            cpdf.drawString(28.0*cm, 1.5*cm, f"Total custo: {total_custo:.2f}")
            cpdf.showPage(); cpdf.setFont("Helvetica", 8); y = h-2.5*cm
    cpdf.drawString(28.0*cm, 1.5*cm, f"Total custo: {total_custo:.2f}")
    cpdf.showPage(); cpdf.save()
    data = buf.getvalue(); buf.close()
    return Response(data, mimetype="application/pdf",
                    headers={"Content-Disposition":"attachment; filename=relatorio_equipamentos.pdf"})


# === Links: add/remove ===
@app.route('/equipamentos/<int:equipamento_id>/links/add', methods=['POST'])
def equipamento_add_link(equipamento_id):
    urlv = (request.form.get('url') or '').strip()
    title = (request.form.get('title') or '').strip()
    if not urlv:
        flash("URL é obrigatório.", "danger")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("INSERT INTO equipamentos_links (equipamento_id, url, title) VALUES (?,?,?)", (equipamento_id, urlv, title or urlv))
    conn.commit(); conn.close()
    flash("Link adicionado.", "success")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))

@app.route('/equipamentos/<int:equipamento_id>/links/<int:link_id>/delete')
def equipamento_del_link(equipamento_id, link_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("DELETE FROM equipamentos_links WHERE id=? AND equipamento_id=?", (link_id, equipamento_id))
    conn.commit(); conn.close()
    flash("Link removido.", "warning")
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))


# === API simples com cover_thumb ===


@app.route('/api/equipamentos')
def api_equipamentos():
    q = request.args.get('q','').strip()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    if q:
        like = f"%{q}%"
        c.execute('''
            SELECT e.id, e.nome, COALESCE(l.nome,''), COALESCE(e.tag,''),
                   COALESCE(cp.thumb_filename,(SELECT thumb_filename FROM equipamentos_photos WHERE equipamento_id=e.id ORDER BY uploaded_at DESC LIMIT 1),'') as cover_thumb
            FROM equipamentos e
            LEFT JOIN locais l ON e.local_id=l.id
            LEFT JOIN equipamentos_photos cp ON cp.id = e.cover_photo_id
            WHERE e.nome LIKE ? OR COALESCE(e.tag,'') LIKE ? OR COALESCE(l.nome,'') LIKE ?
            ORDER BY e.nome ASC
            LIMIT 200
        ''', (like, like, like))
    else:
        c.execute('''
            SELECT e.id, e.nome, COALESCE(l.nome,''), COALESCE(e.tag,''),
                   COALESCE(cp.thumb_filename,(SELECT thumb_filename FROM equipamentos_photos WHERE equipamento_id=e.id ORDER BY uploaded_at DESC LIMIT 1),'') as cover_thumb
            FROM equipamentos e
            LEFT JOIN locais l ON e.local_id=l.id
            LEFT JOIN equipamentos_photos cp ON cp.id = e.cover_photo_id
            ORDER BY e.nome ASC
            LIMIT 200
        ''')
    data = [{"id": r[0], "nome": r[1], "local": r[2], "tag": r[3], "cover_thumb": r[4]} for r in c.fetchall()]
    conn.close()
    return Response(json.dumps(data, ensure_ascii=False), mimetype='application/json')


@app.route('/equipamentos/<int:equipamento_id>/files/<int:file_id>/delete')
def delete_equip_file(equipamento_id, file_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT filename, original_name FROM equipamentos_files WHERE id=? AND equipamento_id=?", (file_id, equipamento_id))
    row = c.fetchone()
    if not row:
        conn.close(); flash("Documento não encontrado.", "warning")
        return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], row[0])
        if os.path.exists(filepath):
            os.remove(filepath)
        c.execute("DELETE FROM equipamentos_files WHERE id=?", (file_id,))
        conn.commit()
        flash("Documento removido.", "success")
        log_equip_audit(equipamento_id, "delete_file", row[1] or row[0])
    except Exception as ex:
        flash(f"Erro ao apagar documento: {ex}", "danger")
    finally:
        conn.close()
    return redirect(url_for('equipamento_detalhe', equipamento_id=equipamento_id))


@app.route('/api/equipamentos/<int:equipamento_id>')
def api_equipamento_detalhe(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT id, nome, local_id, tag, especificacao, ano_instalacao, quantidade, categoria, custo_aquisicao, criticidade, fabricante, modelo, numero_serie, vida_util_anos, ativo, potencia_kw, tensao_v, corrente_a, fornecedor, contrato_num, garantia_fim FROM equipamentos WHERE id=?", (equipamento_id,))
    e = c.fetchone()
    if not e:
        conn.close(); return Response(json.dumps({"error":"not found"}), status=404, mimetype="application/json")
    c.execute("SELECT id, filename, caption FROM equipamentos_photos WHERE equipamento_id=? ORDER BY uploaded_at DESC", (equipamento_id,))
    photos = c.fetchall()
    c.execute("SELECT id, filename, size, original_name FROM equipamentos_files WHERE equipamento_id=? ORDER BY uploaded_at DESC", (equipamento_id,))
    files = c.fetchall()
    c.execute("SELECT id, url, title FROM equipamentos_links WHERE equipamento_id=? ORDER BY added_at DESC", (equipamento_id,))
    links = c.fetchall()
    conn.close()
    return Response(json.dumps({
        "equipamento": e, "photos": photos, "files": files, "links": links
    }, ensure_ascii=False), mimetype="application/json")

def _equip_where_from_request(prefer_ids=True):
    ids = (request.args.get('ids') or '').strip()
    q = (request.args.get('q') or '').strip()
    local_id = (request.args.get('local_id') or '').strip()
    incluir_inativos = (request.args.get('incluir_inativos','0') == '1')
    categoria = (request.args.get('categoria') or '').strip()
    fabricante = (request.args.get('fabricante') or '').strip()
    modelo = (request.args.get('modelo') or '').strip()
    criticidade = (request.args.get('criticidade') or '').strip()
    ano_min = (request.args.get('ano_min') or '').strip()
    ano_max = (request.args.get('ano_max') or '').strip()
    where = []; params = []
    if prefer_ids and ids:
        try:
            arr = [int(x) for x in ids.split(',') if x.strip().isdigit()]
            if arr:
                where.append("e.id IN (" + ",".join(["?"]*len(arr)) + ")")
                params += arr
        except Exception:
            pass
    if not where:
        if q:
            _apply_advanced_query(q, where, params)
        if local_id and local_id.isdigit():
            where.append("e.local_id=?"); params.append(int(local_id))
        if categoria:
            where.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
        if fabricante:
            where.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
        if modelo:
            where.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
        if criticidade:
            where.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
        if ano_min and ano_min.isdigit():
            where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
        if ano_max and ano_max.isdigit():
            where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
        if not incluir_inativos:
            where.append("COALESCE(e.ativo,1)=1")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    return where_sql, params

# ==== SOFT DELETE ====
@app.route('/equipamentos/<int:equipamento_id>/arquivar', methods=['POST'])
def equipamentos_arquivar(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET deleted_at=datetime('now','localtime') WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    flash("Equipamento arquivado.", "warning")
    return redirect(url_for('listar_equipamentos'))

@app.route('/equipamentos/<int:equipamento_id>/restaurar', methods=['POST'])
def equipamentos_restaurar(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("UPDATE equipamentos SET deleted_at=NULL WHERE id=?", (equipamento_id,))
    conn.commit(); conn.close()
    flash("Equipamento restaurado.", "success")
    return redirect(url_for('listar_equipamentos'))
# ==== /SOFT DELETE ====


# ==== HISTÓRICO ====
@app.route('/equipamentos/<int:equipamento_id>/historico')
def equipamentos_historico(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    try:
        c.execute("SELECT acao, detalhes, datetime(ts,'localtime') FROM equipamentos_audit WHERE equipamento_id=? ORDER BY ts DESC", (equipamento_id,))
        rows = c.fetchall()
    except Exception:
        rows = []
    conn.close()
    return render_template('equipamentos_historico.html', equipamento_id=equipamento_id, rows=rows)
# ==== /HISTÓRICO ====


# ==== IMPORT PREVIEW ====
@app.route('/equipamentos/import/preview', methods=['GET','POST'])
def equipamentos_import_preview():
    if request.method == 'GET':
        return render_template('importar_equipamentos_preview.html')
    file = request.files.get('file')
    ftype = (request.form.get('tipo') or '').lower()
    if not file or file.filename == '' or ftype not in ('csv','xlsx','json'):
        flash("Selecione um ficheiro e o tipo correto (csv/xlsx/json).", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    import csv, io, json
    rows, errors = [], []
    try:
        if ftype=='csv':
            content = file.read().decode('utf-8', errors='ignore')
            rd = csv.DictReader(io.StringIO(content))
            for i,row in enumerate(rd, start=2):
                nome = (row.get('nome') or row.get('Nome') or '').strip()
                if not nome: errors.append((i,"Nome vazio")); continue
                rows.append(row)
        elif ftype=='xlsx':
            from openpyxl import load_workbook
            data = io.BytesIO(file.read()); wb = load_workbook(filename=data, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1,max_row=1))]
            for r_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                obj = {headers[i]: (str(cell.value).strip() if cell.value is not None else '') for i,cell in enumerate(row)}
                if not (obj.get('nome') or ''): errors.append((r_i,"Nome vazio")); continue
                rows.append(obj)
        else:
            data = json.loads(file.read().decode('utf-8', errors='ignore'))
            if not isinstance(data, list): raise ValueError("JSON deve ser uma lista")
            for i,obj in enumerate(data, start=1):
                nome = (obj.get('nome') or obj.get('Nome') or '').strip()
                if not nome: errors.append((i,"Nome vazio")); continue
                rows.append(obj)
    except Exception as e:
        flash(f"Erro a ler ficheiro: {e}", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    preview_key = f"equip_preview_{int(datetime.timestamp(datetime.now()))}"
    tmp_path = os.path.join(UPLOAD_DIR, preview_key + '.json')
    try:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump({'ftype': ftype, 'rows': rows}, f, ensure_ascii=False)
    except Exception:
        tmp_path = None
    return render_template('importar_equipamentos_preview.html', rows=rows[:200], errors=errors, preview_key=preview_key)

@app.route('/equipamentos/import/confirm', methods=['POST'])
def equipamentos_import_confirm():
    key = (request.form.get('preview_key') or '').strip()
    if not key:
        flash("Pré-visualização expirada.", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    path = os.path.join(UPLOAD_DIR, key + '.json')
    if not os.path.exists(path):
        flash("Pré-visualização não encontrada.", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        flash(f"Erro ao carregar pré-visualização: {e}", "danger")
        return redirect(url_for('equipamentos_import_preview'))
    rows = data.get('rows') or []
    inserted=0; skipped=0
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    for row in rows:
        nome = (row.get('nome') or row.get('Nome') or '').strip()
        if not nome: skipped+=1; continue
        local_nome = (row.get('local') or row.get('Local') or '').strip()
        local_id = None
        if local_nome:
            c.execute("SELECT id FROM locais WHERE nome=?", (local_nome,)); r = c.fetchone()
            if r: local_id = r[0]
        tag = (row.get('tag') or row.get('TAG') or '').strip()
        especificacao = (row.get('especificacao') or row.get('Especificacao') or row.get('Especificação') or '').strip()
        ano = (row.get('ano_instalacao') or row.get('ano') or '').strip() or None
        try: qtd_i = int((row.get('quantidade') or row.get('qtd') or '1').strip())
        except: qtd_i = 1
        c.execute('''
            INSERT INTO equipamentos (nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'), datetime('now','localtime'))
        ''', (nome, local_id, tag, especificacao, ano, qtd_i))
        inserted+=1
    conn.commit(); conn.close()
    try: os.remove(path)
    except Exception: pass
    flash(f"Importação concluída: {inserted} inseridos, {skipped} ignorados.", "success")
    return redirect(url_for('listar_equipamentos'))
# ==== /IMPORT PREVIEW ====


# ==== DASHBOARD ====
@app.route('/equipamentos/dashboard')
def equipamentos_dashboard():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM equipamentos WHERE COALESCE(deleted_at,'')=''"); total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM equipamentos WHERE COALESCE(ativo,1)=1 AND COALESCE(deleted_at,'')=''"); ativos = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM equipamentos WHERE COALESCE(ativo,1)=0 AND COALESCE(deleted_at,'')=''"); inativos = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM equipamentos WHERE garantia_fim IS NOT NULL AND garantia_fim<>'' AND date(garantia_fim)>=date('now') AND COALESCE(deleted_at,'')=''")
    em_garantia = c.fetchone()[0]
    c.execute("SELECT COALESCE(categoria,''), COUNT(*) FROM equipamentos WHERE COALESCE(deleted_at,'')='' GROUP BY COALESCE(categoria,'') ORDER BY COUNT(*) DESC")
    por_cat = c.fetchall()
    c.execute("SELECT COALESCE(l.nome,''), COUNT(*) FROM equipamentos e LEFT JOIN locais l ON e.local_id=l.id WHERE COALESCE(e.deleted_at,'')='' GROUP BY COALESCE(l.nome,'') ORDER BY COUNT(*) DESC")
    por_local = c.fetchall()
    conn.close()
    return render_template('equipamentos_dashboard.html', total=total, ativos=ativos, inativos=inativos, em_garantia=em_garantia, por_cat=por_cat, por_local=por_local)
# ==== /DASHBOARD ====


# ==== FILTROS SALVOS ====
@app.route('/equipamentos/filtros', methods=['GET'])
def equipamentos_filtros():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT id, user, nome, query_json, datetime(created_at,'localtime') FROM saved_filters WHERE modulo='equipamentos' ORDER BY created_at DESC")
    rows = c.fetchall(); conn.close()
    return render_template('equipamentos_filtros.html', rows=rows)

@app.route('/equipamentos/filtros/salvar', methods=['POST'])
def equipamentos_filtros_salvar():
    user = (request.form.get('user') or 'admin').strip()
    nome = (request.form.get('nome') or '').strip()
    query_json = (request.form.get('query_json') or '').strip()
    if not nome or not query_json:
        flash("Informe nome e filtros.", "danger")
        return redirect(url_for('equipamentos_filtros'))
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("INSERT INTO saved_filters (user, modulo, nome, query_json) VALUES (?, 'equipamentos', ?, ?)", (user, nome, query_json))
    conn.commit(); conn.close()
    flash("Filtro salvo.", "success")
    return redirect(url_for('equipamentos_filtros'))
# ==== /FILTROS SALVOS ====


# ==== ETIQUETAS CUSTOM ====
@app.route('/equipamentos/labels/options')
def equipamentos_labels_options():
    return render_template('equipamentos_labels_options.html')

@app.route('/equipamentos/labels_custom.pdf')
def equipamentos_labels_custom():
    ids_str = (request.args.get('ids') or '').strip()
    if not ids_str:
        flash("IDs obrigatórios", "warning")
        return redirect(url_for('listar_equipamentos'))
    ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
    show_modelo = (request.args.get('modelo','0')=='1')
    show_serie = (request.args.get('serie','0')=='1')
    layout = (request.args.get('layout') or 'A4_3x8')
    include_logo = (request.args.get('logo','0')=='1')

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF
    import math

    buf = io.BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    cols, rows = (3,8) if layout=='A4_3x8' else (2,6)
    margin_x, margin_y = 10, 10
    cell_w = (w - 2*margin_x) / cols
    cell_h = (h - 2*margin_y) / rows

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    qmarks = ",".join(["?"]*len(ids))
    c.execute(f"SELECT id, nome, COALESCE(tag,''), COALESCE(modelo,''), COALESCE(numero_serie,'') FROM equipamentos WHERE id IN ({qmarks}) ORDER BY id", ids)
    rows_data = c.fetchall(); conn.close()

    i = 0
    for r in rows_data:
        x = i % cols; y = i // cols
        if y >= rows:
            cpdf.showPage(); y = 0; i = 0
        pos_x = margin_x + x*cell_w
        pos_y = h - margin_y - (y+1)*cell_h
        cpdf.rect(pos_x, pos_y, cell_w, cell_h)
        cpdf.setFont("Helvetica-Bold", 10)
        cpdf.drawString(pos_x+6, pos_y+cell_h-14, f"EQ #{r[0]}")
        cpdf.setFont("Helvetica", 9)
        cpdf.drawString(pos_x+6, pos_y+cell_h-26, (r[1] or "")[:36])
        if r[2]:
            cpdf.drawString(pos_x+6, pos_y+cell_h-38, f"TAG: {r[2][:26]}")
        yoff = 50
        if show_modelo and r[3]:
            cpdf.drawString(pos_x+6, pos_y+cell_h-yoff, f"Modelo: {r[3][:26]}"); yoff += 12
        if show_serie and r[4]:
            cpdf.drawString(pos_x+6, pos_y+cell_h-yoff, f"Série: {r[4][:26]}"); yoff += 12
        qr_code = qr.QrCodeWidget(f"SGE:EQ:{r[0]}")
        bounds = qr_code.getBounds()
        size = min(cell_w, cell_h) * 0.40
        width = bounds[2] - bounds[0]; height = bounds[3] - bounds[1]
        scale = size / max(width, height)
        drawing = Drawing(size, size, transform=[scale,0,0,scale,0,0])
        drawing.add(qr_code)
        renderPDF.draw(drawing, cpdf, pos_x + cell_w - size - 8, pos_y + 8)
        if include_logo:
            cpdf.setFont("Helvetica-Bold", 8)
            cpdf.drawString(pos_x+6, pos_y+6, "LOGO")
        i += 1

    cpdf.showPage(); cpdf.save()
    data = buf.getvalue(); buf.close()
    return Response(data, mimetype="application/pdf",
                    headers={"Content-Disposition":"attachment; filename=labels_custom.pdf"})
# ==== /ETIQUETAS CUSTOM ====


# ==== API v2 (GET) ====
@app.route('/api/v2/equipamentos')
def api_v2_equip_list():
    where=[]; params=[]
    q = (request.args.get('q') or '').strip()
    if q:
        _apply_advanced_query(q, where, params)
    local_id = (request.args.get('local_id') or '').strip()
    categoria = (request.args.get('categoria') or '').strip()
    fabricante = (request.args.get('fabricante') or '').strip()
    modelo = (request.args.get('modelo') or '').strip()
    criticidade = (request.args.get('criticidade') or '').strip()
    ano_min = (request.args.get('ano_min') or '').strip()
    ano_max = (request.args.get('ano_max') or '').strip()
    incluir_inativos = (request.args.get('incluir_inativos','0')=='1')
    if local_id.isdigit(): where.append("e.local_id=?"); params.append(int(local_id))
    if categoria: where.append("COALESCE(e.categoria,'') LIKE ?"); params.append(f"%{categoria}%")
    if fabricante: where.append("COALESCE(e.fabricante,'') LIKE ?"); params.append(f"%{fabricante}%")
    if modelo: where.append("COALESCE(e.modelo,'') LIKE ?"); params.append(f"%{modelo}%")
    if criticidade: where.append("COALESCE(e.criticidade,'') = ?"); params.append(criticidade)
    if ano_min.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) >= ?"); params.append(int(ano_min))
    if ano_max.isdigit(): where.append("CAST(COALESCE(e.ano_instalacao,0) AS INTEGER) <= ?"); params.append(int(ano_max))
    if not incluir_inativos: where.append("COALESCE(e.ativo,1)=1")
    where.append("COALESCE(e.deleted_at,'')=''")
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f"""
        SELECT e.id, e.nome, e.local_id, COALESCE(l.nome,''), e.tag, e.modelo, e.fabricante,
               e.criticidade, e.ativo, e.deleted_at, e.garantia_fim
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id=l.id
        {where_sql}
        ORDER BY l.nome, e.nome
    """, params)
    rows = c.fetchall(); conn.close()
    data = [{
        "id":r[0],"nome":r[1],"local_id":r[2],"local":r[3],"tag":r[4],"modelo":r[5],"fabricante":r[6],
        "criticidade":r[7],"ativo":r[8],"arquivado": bool(r[9]), "garantia_fim": r[10]
    } for r in rows]
    return jsonify(data)

@app.route('/api/v2/equipamentos/<int:equipamento_id>')
def api_v2_equip_detail(equipamento_id):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""
        SELECT e.*, COALESCE(l.nome,'') as local_nome
        FROM equipamentos e
        LEFT JOIN locais l ON e.local_id=l.id
        WHERE e.id=?
    """, (equipamento_id,))
    e = c.fetchone()
    if not e:
        conn.close(); return jsonify({"error":"not found"}), 404
    c.execute("SELECT id, nome, COALESCE(fabricante,''), COALESCE(modelo,''), COALESCE(qtd,1) FROM equipamentos_componentes WHERE equipamento_id=? ORDER BY nome", (equipamento_id,))
    comps = c.fetchall()
    try:
        c.execute("SELECT id, filename, size FROM equipamentos_files WHERE equipamento_id=? ORDER BY uploaded_at DESC", (equipamento_id,)); files = c.fetchall()
    except Exception:
        files = []
    try:
        c.execute("SELECT id, filename, thumb_filename, width, height FROM equipamentos_photos WHERE equipamento_id=? ORDER BY uploaded_at DESC", (equipamento_id,)); photos = c.fetchall()
    except Exception:
        photos = []
    cols = [d[0] for d in c.description] if c.description else []
    conn.close()
    return jsonify({
        "equipamento": dict(zip(cols, e)) if cols else {},
        "componentes": [{"id":r[0],"nome":r[1],"fabricante":r[2],"modelo":r[3],"qtd":r[4]} for r in comps],
        "files": [{"id":r[0],"filename":r[1],"size":r[2]} for r in files],
        "photos": [{"id":r[0],"filename":r[1],"thumb":r[2],"w":r[3],"h":r[4]} for r in photos]
    })
# ==== /API v2 ====



@app.route('/equipamentos/filtros/delete/<int:fid>')
def equipamentos_filtros_delete(fid):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("DELETE FROM saved_filters WHERE id=?", (fid,))
    conn.commit(); conn.close()
    flash("Filtro removido.", "warning")
    return redirect(url_for('equipamentos_filtros'))

# === Util: configuração completa por Local ===
def _db_columns(conn, table):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    return {row[1] for row in cur.fetchall()}

def _get_local_id_by_any(conn, any_id_or_name):
    cur = conn.cursor()
    try:
        lid = int(str(any_id_or_name).strip())
        cur.execute("SELECT id, nome FROM locais WHERE id = ?", (lid,))
        row = cur.fetchone()
        if row: return row[0], row[1]
    except Exception:
        pass
    cur.execute("SELECT id, nome FROM locais WHERE nome = ?", (str(any_id_or_name).strip(),))
    row = cur.fetchone()
    if row: return row[0], row[1]
    return None, None

def _get_local_cfg_full(conn, local_id):
    cfg = {
        "fator_mult": 1.0,
        "pot_contratada": 0.0,
        "pot_instalada": 0.0,
        "tarifa_ativa": 0.0,
        "tarifa_reativa": 0.0,
        "tarifa_ponta": 0.0,
        "tarifa_perdas": 0.0,
        "taxa_fixa": 0.0,
        "taxa_radio": 0.0,
        "taxa_lixo": 0.0,
        "iva": 0.0,
    }
    cur = conn.cursor()
    cols_cfg = _db_columns(conn, "locais_cfg")
    select_cols = [c for c in ["fator_mult","pot_contratada","pot_instalada",
                               "tarifa_ativa","tarifa_reativa","tarifa_ponta","tarifa_perdas",
                               "taxa_fixa","taxa_radio","taxa_lixo","iva"] if c in cols_cfg]
    if select_cols:
        sql = "SELECT " + ",".join(select_cols) + " FROM locais_cfg WHERE local_id = ?"
        cur.execute(sql, (local_id,))
        row = cur.fetchone()
        if row:
            for idx, col in enumerate(select_cols):
                cfg[col] = _to_float(row[idx], cfg[col])

    cols_locais = _db_columns(conn, "locais")
    if "pot_instalada" in cols_locais:
        cur.execute("SELECT pot_instalada FROM locais WHERE id = ?", (local_id,))
        r2 = cur.fetchone()
        if r2 and r2[0] is not None:
            cfg["pot_instalada"] = _to_float(r2[0], cfg["pot_instalada"])

    return cfg

@app.get('/api/local_cfg', endpoint='api_local_cfg_v2')
def api_local_cfg_v2():
    try:
        local_any = request.args.get('local', '').strip()
    except Exception:
        local_any = ''
    conn = sqlite3.connect(DB_PATH if 'DB_PATH' in globals() else 'sge.db')
    try:
        local_id, _ = _get_local_id_by_any(conn, local_any)
        if not local_id:
            return jsonify({"error":"local não encontrado"}), 404
        cfg = _get_local_cfg_full(conn, local_id)
        return jsonify(cfg), 200
    finally:
        conn.close()


# --- Alias para compatibilidade com o frontend: /api/local_cfg/<id_ou_nome> ---
@app.get('/api/local_cfg/<path:local_any>', endpoint='api_local_cfg_alias')
def api_local_cfg_alias(local_any):
    try:
        any_val = (local_any or '').strip()
    except Exception:
        any_val = ''
    conn = sqlite3.connect(DB_PATH if 'DB_PATH' in globals() else 'sge.db')
    try:
        local_id, _ = _get_local_id_by_any(conn, any_val)
        if not local_id:
            return jsonify({"error": "local não encontrado"}), 404
        cfg = _get_local_cfg_full(conn, local_id)
        return jsonify(cfg), 200
    finally:
        conn.close()

@app.post('/api/leituras_mensal/calc_fatura', endpoint='api_calc_fatura_mensal_v2')
def api_calc_fatura_mensal_v2():
    data = request.get_json(silent=True) or {}

    fator_mult  = _to_float(data.get('fator_mult'), 1.0)
    kwh_ativa   = _to_float(data.get('kwh_ativa'))   * fator_mult
    kwh_reativa = _to_float(data.get('kwh_reativa')) * fator_mult
    kwh_ponta   = _to_float(data.get('kwh_ponta'))   * fator_mult
    kwh_perdas  = _to_float(data.get('kwh_perdas'))  * fator_mult  # mantido só para referência

    t_ativa   = _to_float(data.get('tarifa_ativa'))
    t_reativa = _to_float(data.get('tarifa_reativa'))
    t_ponta   = _to_float(data.get('tarifa_ponta'))
    t_perdas  = _to_float(data.get('tarifa_perdas'))  # já não será usado

    taxa_fixa  = _to_float(data.get('taxa_fixa'))
    taxa_radio = _to_float(data.get('taxa_radio'))
    taxa_lixo  = _to_float(data.get('taxa_lixo'))
    iva        = _to_float(data.get('iva'))

    # Reativa excedente faturável (apenas o que ultrapassa 0,75 × kWh ativa)
    limite_reativa   = 0.75 * kwh_ativa
    kwh_reativa_fat  = max(kwh_reativa - limite_reativa, 0.0)

    c_ativa   = kwh_ativa      * t_ativa
    c_reativa = kwh_reativa_fat * t_reativa
    c_ponta   = kwh_ponta      * t_ponta
    c_perdas  = 0.0  # retirado do cálculo porque o contador está do lado de MT/AT

    energia_subtotal = c_ativa + c_reativa + c_ponta
    taxas_subtotal   = taxa_fixa + taxa_radio + taxa_lixo
    subtotal         = energia_subtotal + taxas_subtotal
    valor_iva        = subtotal * iva
    total            = subtotal + valor_iva

    result = {
        "kwh": {
            "ativa":   round(kwh_ativa,       3),
            "reativa": round(kwh_reativa_fat, 3),  # só excedente faturável
            "ponta":   round(kwh_ponta,       3),
            "perdas":  0.0
        },
        "custos": {
            "ativa":   round(c_ativa,   2),
            "reativa": round(c_reativa, 2),
            "ponta":   round(c_ponta,   2),
            "perdas":  round(c_perdas,  2)
        },
        "energia_subtotal": round(energia_subtotal, 2),
        "taxas": {
            "fixa":  round(taxa_fixa,  2),
            "radio": round(taxa_radio, 2),
            "lixo":  round(taxa_lixo,  2)
        },
        "taxas_subtotal": round(taxas_subtotal, 2),
        "subtotal":       round(subtotal, 2),
        "iva":            round(valor_iva, 2),
        "total":          round(total, 2)
    }
    return jsonify(result), 200

@app.route('/leituras/save_filter', methods=['POST'])
def leituras_save_filter():
    """Guarda filtros (por utilizador, nome)."""
    user = (request.form.get('user') or 'default').strip()
    nome = (request.form.get('nome') or '').strip() or datetime.now().strftime('filtro_%Y%m%d_%H%M')
    query = {
        'inicio': request.form.get('inicio') or '',
        'fim': request.form.get('fim') or '',
        'local': request.form.get('local') or '',
        'equipamento': request.form.get('equipamento') or '',
        'q': request.form.get('q') or '',
        'per': request.form.get('per') or '50'
    }
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('INSERT INTO saved_filters (user, modulo, nome, query_json) VALUES (?, ?, ?, ?)',
              (user, 'leituras', nome, json.dumps(query)))
    conn.commit(); conn.close()
    flash("Filtro guardado.", "success")
    return redirect(url_for('leituras_list', **query))

@app.route('/leituras/list_filters')
def leituras_list_filters():
    user = (request.args.get('user') or 'default').strip()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute("SELECT id, nome, query_json, created_at FROM saved_filters WHERE modulo='leituras' AND user=? ORDER BY created_at DESC", (user,)).fetchall()
    conn.close()
    return jsonify([{'id': r[0], 'nome': r[1], 'query': json.loads(r[2] or '{}'), 'created_at': r[3]} for r in rows])

@app.route('/leituras/apply_filter/<int:fid>')
def leituras_apply_filter(fid):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    r = c.execute("SELECT query_json FROM saved_filters WHERE id=? AND modulo='leituras'", (fid,)).fetchone()
    conn.close()
    if not r:
        flash("Filtro não encontrado.", "warning")
        return redirect(url_for('leituras_list'))
    q = json.loads(r[0] or "{}")
    return redirect(url_for('leituras_list', **q))



@app.route('/leituras/<int:lid>/update_field', methods=['POST'])
def leituras_update_field(lid):
    """Atualiza um único campo (edição inline)."""
    data = request.get_json(silent=True) or {}
    field = (data.get('field') or '').strip()
    value = data.get('value')

    allowed = {
        'datahora','local','equipamento','energia_ativa','energia_reativa','energia_aparente',
        'pot_ativa','pot_reativa','pot_aparente','fp','ponta','caudal_elevada','corrente','tensao','observacoes'
    }
    if field not in allowed:
        return jsonify({'ok': False, 'error': 'Campo não permitido.'}), 400

    # normalização básica
    num_fields = {'energia_ativa','energia_reativa','energia_aparente','pot_ativa','pot_reativa','pot_aparente','fp','ponta','caudal_elevada','corrente','tensao'}
    if field in num_fields:
        try:
            s = ('' if value is None else str(value)).replace(',', '.').strip()
            value = None if s=='' else float(s)
        except Exception:
            return jsonify({'ok': False, 'error': 'Valor numérico inválido.'}), 400

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(f"UPDATE leituras SET {field}=? WHERE id=?", (value, lid))
    conn.commit(); conn.close()
    return jsonify({'ok': True})



@app.route('/leituras/export_xlsx')
def leituras_export_xlsx():
    """Exporta o filtro actual para XLSX."""
    end = request.args.get('fim') or datetime.now().strftime('%Y-%m-%d')
    start = request.args.get('inicio') or (datetime.now() - timedelta(days=6)).strftime('%Y-%m-%d')
    local = request.args.get('local','').strip()
    q = request.args.get('q','').strip()

    base_sql = " FROM leituras WHERE date(datahora) BETWEEN ? AND ?"
    params = [start, end]
    if local:
        base_sql += " AND local = ?"
        params.append(local)
    if q:
        base_sql += " AND (equipamento LIKE ? OR observacoes LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%"])

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute("SELECT *" + base_sql + " ORDER BY datahora DESC", params).fetchall()
    conn.close()

    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out, {'in_memory': True})
    ws = wb.add_worksheet('Leituras')
    header = ["id","datahora","local","equipamento","energia_ativa","energia_reativa","energia_aparente","pot_ativa","pot_reativa","pot_aparente","fp","ponta","caudal_elevada","corrente","tensao","observacoes"]
    for j,h in enumerate(header): ws.write(0,j,h)
    for i,row in enumerate(rows, start=1):
        for j,val in enumerate(row): ws.write(i,j,val)
    wb.close()
    out.seek(0)
    filename = f"leituras_{start}_a_{end}.xlsx"
    return Response(out.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})



@app.route('/leituras/import_preview', methods=['GET','POST'])
def leituras_import_preview():
    """Pré-visualização do CSV antes de importar."""
    if request.method == 'POST':
        f = request.files.get('arquivo')
        if not f or f.filename == '':
            flash("Selecione um CSV.", "warning")
            return redirect(url_for('leituras_import_preview'))
        content = f.read().decode('utf-8', errors='ignore')
        delimiter = ';' if content.count(';')>=content.count(',') else ','
        import csv as _csv
        reader = _csv.reader(content.splitlines(), delimiter=delimiter)
        rows = list(reader)
        head = rows[0] if rows else []
        preview = rows[1:101]  # primeiras 100
        # guarda em sessão mínima (fallback: reenvia o arquivo no próximo passo se necessário)
        return render_template('leituras_import_preview.html', header=head, preview=preview, delimiter=delimiter, raw=content)
    return render_template('leituras_import_preview.html', header=[], preview=[], delimiter=';', raw='')

@app.route('/leituras/import_commit', methods=['POST'])
def leituras_import_commit():
    """Confirma import do CSV recebido da pré-visualização."""
    content = request.form.get('raw','')
    delimiter = request.form.get('delimiter',';')
    if not content:
        flash("Conteúdo do CSV ausente.", "danger")
        return redirect(url_for('leituras_import_preview'))
    import csv as _csv
    reader = _csv.DictReader(content.splitlines(), delimiter=delimiter)
    required = {'datahora','local','equipamento'}
    if not required.issubset({(h or '').strip().lower() for h in (reader.fieldnames or [])}):
        flash("Cabeçalho obrigatório: datahora, local, equipamento.", "danger")
        return redirect(url_for('leituras_import_preview'))

    to_float = lambda v: (None if v is None or str(v).strip()=='' else float(str(v).replace(',','.')))
    ok, err = 0, 0
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    for row in reader:
        try:
            c.execute("""
                INSERT INTO leituras
                (datahora, local, equipamento, energia_ativa, energia_reativa, energia_aparente,
                 pot_ativa, pot_reativa, pot_aparente, fp, ponta, caudal_elevada, corrente, tensao, observacoes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                row.get('datahora'), row.get('local'), row.get('equipamento'),
                to_float(row.get('energia_ativa')), to_float(row.get('energia_reativa')), to_float(row.get('energia_aparente')),
                to_float(row.get('pot_ativa')), to_float(row.get('pot_reativa')), to_float(row.get('pot_aparente')),
                to_float(row.get('fp')), to_float(row.get('ponta')), to_float(row.get('caudal_elevada')),
                to_float(row.get('corrente')), to_float(row.get('tensao')), row.get('observacoes')
            ))
            ok += 1
        except Exception:
            err += 1
    conn.commit(); conn.close()
    flash(f"Importação concluída: {ok} ok, {err} erros.", "success" if err==0 else "warning")
    return redirect(url_for('leituras_list'))


def _audit_leitura(leitura_id, acao, field=None, old_value=None, new_value=None, actor=None):
    try:
        actor = actor or request.headers.get('X-User') or 'anon'
    except Exception:
        actor = 'anon'
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('''INSERT INTO leituras_audit (leitura_id, acao, field, old_value, new_value, actor)
                     VALUES (?,?,?,?,?,?)''', (leitura_id, acao, field, str(old_value) if old_value is not None else None,
                                               str(new_value) if new_value is not None else None, actor))
        conn.commit(); conn.close()
    except Exception:
        pass


@app.route('/leituras/bulk', methods=['GET'])
def leituras_bulk_form():
    return render_template('leituras_bulk_edit.html')

@app.route('/leituras/bulk_apply', methods=['POST'])
def leituras_bulk_apply():
    ids = request.form.get('ids','').strip()
    op = request.form.get('op','')
    actor = request.headers.get('X-User') or 'anon'
    if not ids:
        flash("Indique IDs separados por vírgula.", "warning")
        return redirect(url_for('leituras_bulk_form'))
    try:
        ids_list = [int(x) for x in ids.split(',') if x.strip().isdigit()]
    except Exception:
        flash("IDs inválidos.", "danger")
        return redirect(url_for('leituras_bulk_form'))

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    count = 0

    if op == 'shift_time':
        minutes = int(request.form.get('minutes','0') or 0)
        for lid in ids_list:
            old = c.execute("SELECT datahora FROM leituras WHERE id=?", (lid,)).fetchone()
            if not old or not old[0]: continue
            try:
                dt = datetime.strptime(old[0][:16], "%Y-%m-%d %H:%M")
            except Exception:
                try:
                    dt = datetime.fromisoformat(old[0].replace('Z',''))
                except Exception:
                    continue
            new_dt = dt + timedelta(minutes=minutes)
            nds = new_dt.strftime("%Y-%m-%d %H:%M")
            c.execute("UPDATE leituras SET datahora=? WHERE id=?", (nds, lid))
            _audit_leitura(lid, "bulk_shift_time", "datahora", old[0], nds, actor=actor)
            count += 1

    elif op == 'set_local':
        new_local = (request.form.get('new_local') or '').strip()
        if not new_local:
            conn.close()
            flash("Informe o novo local.", "warning")
            return redirect(url_for('leituras_bulk_form'))
        for lid in ids_list:
            old = c.execute("SELECT local FROM leituras WHERE id=?", (lid,)).fetchone()
            c.execute("UPDATE leituras SET local=? WHERE id=?", (new_local, lid))
            _audit_leitura(lid, "bulk_set_local", "local", old[0] if old else None, new_local, actor=actor)
            count += 1

    elif op == 'set_equip':
        new_eq = (request.form.get('new_equip') or '').strip()
        if not new_eq:
            conn.close()
            flash("Informe o novo equipamento.", "warning")
            return redirect(url_for('leituras_bulk_form'))
        for lid in ids_list:
            old = c.execute("SELECT equipamento FROM leituras WHERE id=?", (lid,)).fetchone()
            c.execute("UPDATE leituras SET equipamento=? WHERE id=?", (new_eq, lid))
            _audit_leitura(lid, "bulk_set_equip", "equipamento", old[0] if old else None, new_eq, actor=actor)
            count += 1

    conn.commit(); conn.close()
    flash(f"Edição em massa concluída: {count} linhas.", "success")
    return redirect(url_for('leituras_list'))


@app.route('/leituras/export_pdf')
def leituras_export_pdf():
    end = request.args.get('fim') or datetime.now().strftime('%Y-%m-%d')
    start = request.args.get('inicio') or (datetime.now() - timedelta(days=6)).strftime('%Y-%m-%d')
    local = request.args.get('local','').strip()
    q = request.args.get('q','').strip()

    base_sql = " FROM leituras WHERE date(datahora) BETWEEN ? AND ?"
    params = [start, end]
    if local:
        base_sql += " AND local = ?"; params.append(local)
    if q:
        base_sql += " AND (equipamento LIKE ? OR observacoes LIKE ?)"; params.extend([f"%{q}%", f"%{q}%"])

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute("SELECT id, datahora, local, equipamento, energia_ativa, ponta, fp " + base_sql + " ORDER BY datahora DESC LIMIT 1000", params).fetchall()
    conn.close()

    buffer = io.BytesIO()
    cpdf = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    y = h - 30
    cpdf.setFont("Helvetica-Bold", 12)
    cpdf.drawString(30, y, f"Leituras {start} a {end}  (Local: {local or 'todos'})"); y -= 20
    cpdf.setFont("Helvetica", 9)
    header = "ID   Data/Hora           Local            Equipamento         kWh     Ponta   FP"
    cpdf.drawString(30, y, header); y -= 12
    cpdf.line(30, y+5, w-30, y+5)

    for rid, dh, loc, eq, kwh, ponta, fp in rows:
        line = f"{str(rid).ljust(4)} {str(dh or '')[:16].ljust(18)} {str(loc or '')[:14].ljust(16)} {str(eq or '')[:18].ljust(20)} {kwh or 0:6.1f}  {ponta or 0:6.1f}  {fp or 0:4.2f}"
        if y < 50:
            cpdf.showPage(); y = h - 30
            cpdf.setFont("Helvetica", 9)
        cpdf.drawString(30, y, line); y -= 12

    cpdf.showPage(); cpdf.save()
    pdf = buffer.getvalue()
    buffer.close()
    return Response(pdf, mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename=leituras_{start}_{end}.pdf"})


# ===== API REST (token simples via ?token= ou Header Authorization: Bearer) =====
API_TOKEN = os.environ.get('SGE_API_TOKEN', 'sge-api-token')

def _api_check_token():
    tok = request.args.get('token') or ''
    if not tok:
        auth = request.headers.get('Authorization','')
        if auth.lower().startswith('bearer '):
            tok = auth.split(' ',1)[1].strip()
    return tok == API_TOKEN

@app.route('/api/leituras', methods=['GET'])
def api_leituras_list():
    if not _api_check_token():
        return jsonify({'error':'unauthorized'}), 401
    end = request.args.get('fim') or datetime.now().strftime('%Y-%m-%d')
    start = request.args.get('inicio') or (datetime.now() - timedelta(days=6)).strftime('%Y-%m-%d')
    local = request.args.get('local','').strip()
    q = request.args.get('q','').strip()

    base_sql = " FROM leituras WHERE date(datahora) BETWEEN ? AND ?"
    params = [start, end]
    if local: base_sql += " AND local = ?"; params.append(local)
    if q: base_sql += " AND (equipamento LIKE ? OR observacoes LIKE ?)"; params.extend([f"%{q}%", f"%{q}%"])

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute("SELECT *" + base_sql + " ORDER BY datahora DESC LIMIT 1000", params).fetchall()
    conn.close()
    out = []
    for r in rows:
        out.append({
            "id": r[0], "datahora": r[1], "local": r[2], "equipamento": r[3],
            "energia_ativa": r[4], "energia_reativa": r[5], "energia_aparente": r[6],
            "pot_ativa": r[7], "pot_reativa": r[8], "pot_aparente": r[9],
            "fp": r[10], "ponta": r[11], "caudal_elevada": r[12], "corrente": r[13], "tensao": r[14],
            "observacoes": r[15]
        })
    return jsonify(out)

@app.route('/api/leituras', methods=['POST'])
def api_leituras_create():
    if not _api_check_token():
        return jsonify({'error':'unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    fields = ['datahora','local','equipamento','energia_ativa','energia_reativa','energia_aparente','pot_ativa','pot_reativa','pot_aparente','fp','ponta','caudal_elevada','corrente','tensao','observacoes']
    vals = [data.get(f) for f in fields]
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""
        INSERT INTO leituras
        (datahora,local,equipamento,energia_ativa,energia_reativa,energia_aparente,pot_ativa,pot_reativa,pot_aparente,fp,ponta,caudal_elevada,corrente,tensao,observacoes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, tuple(vals))
    lid = c.lastrowid
    conn.commit(); conn.close()
    _audit_leitura(lid, "api_create", actor="api")
    return jsonify({'ok':True,'id':lid}), 201

@app.route('/api/leituras/<int:lid>', methods=['PATCH'])
def api_leituras_patch(lid):
    if not _api_check_token():
        return jsonify({'error':'unauthorized'}), 401
    data = request.get_json(silent=True) or {}
    allowed = {'datahora','local','equipamento','energia_ativa','energia_reativa','energia_aparente','pot_ativa','pot_reativa','pot_aparente','fp','ponta','caudal_elevada','corrente','tensao','observacoes'}
    sets, params = [], []
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    for k,v in data.items():
        if k in allowed:
            old = c.execute(f"SELECT {k} FROM leituras WHERE id=?", (lid,)).fetchone()
            sets.append(f"{k}=?"); params.append(v)
            _audit_leitura(lid, "api_patch", k, old[0] if old else None, v, actor="api")
    if not sets:
        conn.close()
        return jsonify({'ok':False, 'error':'sem campos válidos'}), 400
    params.append(lid)
    c.execute(f"UPDATE leituras SET {', '.join(sets)} WHERE id=?", params)
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/leituras/<int:lid>', methods=['DELETE'])
def api_leituras_delete(lid):
    if not _api_check_token():
        return jsonify({'error':'unauthorized'}), 401
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("DELETE FROM leituras WHERE id=?", (lid,))
    conn.commit(); conn.close()
    _audit_leitura(lid, "api_delete", actor="api")
    return jsonify({'ok':True})


@app.route('/admin/backup_db')
def admin_backup_db():
    # faz dump do DB para o cliente (download)
    if not os.path.exists(DB_PATH):
        return "DB inexistente", 404
    return send_from_directory(os.path.dirname(DB_PATH), os.path.basename(DB_PATH), as_attachment=True)

@app.route('/admin/health')
def admin_health():
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("SELECT 1"); conn.close()
        return jsonify({"ok": True, "db": "connected"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/leituras_mensal/export_xlsx')
def leituras_mensal_export_xlsx():
    import io
    import xlsxwriter
    local = request.args.get('local','')
    mes = request.args.get('mes') or datetime.now().strftime('%m')
    ano = int(request.args.get('ano') or datetime.now().year)
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute('''SELECT data,hora,ativa,reativa,ponta,fp,potc,anterior,atual,diferenca,agua,esp,acum,valor
                        FROM leituras_mensais WHERE local=? AND mes=? AND ano=? ORDER BY data''',
                     (local, mes, ano)).fetchall()
    conn.close()
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet('Mensal')
    headers = ['Data','Hora','Ativa','Reativa','Ponta','FP','PotC','Anterior','Atual','Diferença','Água','Esp','Acum','Valor']
    for j,h in enumerate(headers): ws.write(0,j,h)
    for i,row in enumerate(rows, start=1):
        for j,val in enumerate(row):
            ws.write(i, j, val)
    wb.close()
    output.seek(0)
    return Response(output.read(), headers={
        'Content-Disposition': f'attachment; filename=leituras_mensais_{local}_{ano}-{mes}.xlsx'
    }, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/leituras_mensal_stats')
def api_leituras_mensal_stats():
    local = request.args.get('local','')
    mes = request.args.get('mes') or datetime.now().strftime('%m')
    ano = int(request.args.get('ano') or datetime.now().year)
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    data = c.execute('''SELECT COUNT(*), SUM(ativa), SUM(diferenca), AVG(fp), MAX(ativa), MIN(ativa)
                        FROM leituras_mensais WHERE local=? AND mes=? AND ano=?''', (local, mes, ano)).fetchone()
    conn.close()
    total_dias, soma_kwh, soma_dif, fp_medio, pico, minimo = data or (0,0,0,0,0,0)
    return jsonify({
        'total_dias': total_dias or 0,
        'kwh_total': float(soma_kwh or 0),
        'consumo_total_diferenca': float(soma_dif or 0),
        'fp_medio': float(fp_medio or 0),
        'pico_ativa': float(pico or 0),
        'min_ativa': float(minimo or 0)
    })

@app.route('/leituras_mensal/clone_prev', methods=['POST'])
def leituras_mensal_clone_prev():
    # Clona o último mês preenchido para o mês atual (por local)
    local = request.form.get('local','')
    mes = request.form.get('mes') or datetime.now().strftime('%m')
    ano = int(request.form.get('ano') or datetime.now().year)
    # mês anterior
    prev_ano, prev_mes = ano, int(mes)-1
    if prev_mes == 0:
        prev_mes = 12; prev_ano = ano-1
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    prev_rows = c.execute('''SELECT data,hora,ativa,reativa,ponta,fp,potc,anterior,atual,diferenca,agua,esp,acum,valor
                             FROM leituras_mensais WHERE local=? AND mes=? AND ano=? ORDER BY data''',
                          (local, str(prev_mes).zfill(2), prev_ano)).fetchall()
    inseridos = 0
    num_dias = calendar.monthrange(ano, int(mes))[1]
    for i in range(num_dias):
        day = str(i+1).zfill(2)
        # find matching day from previous month if exists (same day index)
        if i < len(prev_rows):
            pr = prev_rows[i]
            data = f"{ano}-{mes}-{day}"
            # upsert
            c.execute('''INSERT INTO leituras_mensais(local,data,hora,ativa,reativa,ponta,fp,potc,anterior,atual,diferenca,agua,esp,acum,valor,mes,ano)
                         VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                         ON CONFLICT(local,data) DO UPDATE SET
                           hora=excluded.hora, ativa=excluded.ativa, reativa=excluded.reativa, ponta=excluded.ponta,
                           fp=excluded.fp, potc=excluded.potc, anterior=excluded.anterior, atual=excluded.atual,
                           diferenca=excluded.diferenca, agua=excluded.agua, esp=excluded.esp, acum=excluded.acum,
                           valor=excluded.valor, mes=excluded.mes, ano=excluded.ano''',
                      (local, data, pr[1], pr[2], pr[3], pr[4], pr[5], pr[6], pr[7], pr[8], pr[9], pr[10], pr[11], pr[12], pr[13], mes, ano))
            inseridos += 1
    conn.commit(); conn.close()
    flash(f'Clonado {inseridos} dias do mês anterior.', 'success')
    return redirect(url_for('leituras_mensal'))

@app.route('/api/leituras_mensal_series')
def api_leituras_mensal_series():
    local = request.args.get('local','')
    mes = request.args.get('mes') or datetime.now().strftime('%m')
    ano = int(request.args.get('ano') or datetime.now().year)
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute('''SELECT data, diferenca, ativa FROM leituras_mensais 
                        WHERE local=? AND mes=? AND ano=? ORDER BY data''',
                        (local, mes, ano)).fetchall()
    conn.close()
    dias = [r[0] for r in rows]
    difs = [float(r[1] or 0) for r in rows]
    atv = [float(r[2] or 0) for r in rows]
    return jsonify({'labels': dias, 'diferenca': difs, 'ativa': atv})


@app.route('/leituras_mensal/template_csv')
def leituras_mensal_template_csv():
    # Template CSV com cabeçalho padrão para importação
    import io, csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['data','hora','ativa','reativa','ponta','fp','potc','anterior','atual','diferenca','agua','esp','acum','valor'])
    # exemplo de 3 linhas
    writer.writerow(['2025-01-01','00:00', '', '', '', '', '', '', '', '', '', '', '', ''])
    writer.writerow(['2025-01-02','00:00', '', '', '', '', '', '', '', '', '', '', '', ''])
    writer.writerow(['2025-01-03','00:00', '', '', '', '', '', '', '', '', '', '', '', ''])
    data = output.getvalue().encode('utf-8')
    return Response(data, headers={'Content-Disposition':'attachment; filename=template_leituras_mensais.csv'},
                    mimetype='text/csv')


@app.route('/config_validacao', methods=['GET','POST'])
def config_validacao():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    # populate locais list
    locais = [r[0] for r in c.execute("SELECT DISTINCT nome FROM locais ORDER BY nome").fetchall()]
    msg = None
    if request.method == 'POST':
        local = request.form.get('local','').strip()
        fp_min = float(request.form.get('fp_min', 0.85) or 0.85)
        kwh_dia_max = request.form.get('kwh_dia_max','').strip()
        kwh_val = float(kwh_dia_max) if kwh_dia_max != '' else None
        permitir_reg = 1 if request.form.get('permitir_regressivo') == 'on' else 0
        set_validacao_local(local, fp_min, kwh_val, permitir_reg)
        msg = "Configuração gravada com sucesso."
    # load configs
    rows = []
    for l in locais:
        cfg = get_validacao_local(l)
        rows.append({'local': l, **cfg})
    conn.close()
    return render_template('config_validacao.html', locais=locais, rows=rows, msg=msg)

@app.route('/leituras_mensal/audit')
def leituras_mensal_audit():
    local = request.args.get('local','')
    mes = (request.args.get('mes') or datetime.now().strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or datetime.now().year)

    # Auditoria operacional: além do histórico de alterações, avalia a coerência
    # técnica das leituras já gravadas no mês. Isto torna o botão útil para
    # validação antes de emitir a fatura.
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        logs = c.execute('''SELECT a.ts, a.field, a.old_value, a.new_value, a.acao, a.actor, lm.data
                            FROM leituras_mensais_audit a
                            JOIN leituras_mensais lm ON lm.rowid = a.lm_id
                            WHERE lm.local=? AND lm.mes=? AND lm.ano=?
                            ORDER BY a.ts DESC''', (local, mes, ano)).fetchall()
    except Exception:
        logs = []

    rows = c.execute('''SELECT data, ativa, reativa, ponta, fp, diferenca, agua, esp, valor
                        FROM leituras_mensais
                        WHERE local=? AND mes=? AND ano=?
                        ORDER BY data''', (local, mes, ano)).fetchall()
    conn.close()

    qfat = _quantidades_fatura_mensal(local, mes, ano) if local else {}

    anomalies = []
    prev_ativa = qfat.get('leitura_base_ativa', 0) if isinstance(qfat, dict) else 0
    prev_reativa = qfat.get('leitura_base_reativa', 0) if isinstance(qfat, dict) else 0
    prev_ponta = 0.0
    dias_preenchidos = 0
    fp_validos = []
    for r in rows:
        data = r['data']
        ativa = _safe_float(r['ativa'], None)
        reativa = _safe_float(r['reativa'], None)
        ponta = _safe_float(r['ponta'], None)
        fpv = _safe_float(r['fp'], None)
        dif = _safe_float(r['diferenca'], 0.0) or 0.0
        agua = _safe_float(r['agua'], 0.0) or 0.0
        if any(v is not None and v != 0 for v in (ativa, reativa, ponta)) or agua > 0:
            dias_preenchidos += 1
        if ativa is not None and ativa > 0:
            if prev_ativa and ativa < prev_ativa:
                anomalies.append({'tipo':'Leitura ativa decrescente', 'nivel':'critico', 'data':data, 'detalhe':f'Ativa {ativa:,.2f} menor que a base/anterior {prev_ativa:,.2f}. Verificar digitação, contador ou fator multiplicativo.'})
            else:
                prev_ativa = ativa
        if reativa is not None and reativa > 0:
            if prev_reativa and reativa < prev_reativa:
                anomalies.append({'tipo':'Leitura reativa decrescente', 'nivel':'critico', 'data':data, 'detalhe':f'Reativa {reativa:,.2f} menor que a base/anterior {prev_reativa:,.2f}. Verificar leitura reativa.'})
            else:
                prev_reativa = reativa
        if ponta is not None and ponta > 0:
            if prev_ponta and ponta < prev_ponta:
                anomalies.append({'tipo':'Ponta inferior à máxima anterior', 'nivel':'alerta', 'data':data, 'detalhe':f'Ponta {ponta:,.2f} menor que a máxima anterior {prev_ponta:,.2f}. Para faturação será considerada a maior ponta do mês.'})
            prev_ponta = max(prev_ponta, ponta)
        if fpv is not None and fpv > 0:
            fp_validos.append(fpv)
            if fpv < 0.85:
                anomalies.append({'tipo':'Fator de potência baixo', 'nivel':'alerta', 'data':data, 'detalhe':f'FP = {fpv:.3f}. Avaliar compensação reativa/capacitores.'})
        if dif > 0 and agua <= 0:
            anomalies.append({'tipo':'Energia sem água registada', 'nivel':'aviso', 'data':data, 'detalhe':'Há consumo de energia no dia, mas a água elevada está zero. O consumo específico fica incompleto.'})

    resumo_audit = {
        'dias_preenchidos': dias_preenchidos,
        'total_linhas': len(rows),
        'anomalias': len(anomalies),
        'fp_medio': (sum(fp_validos)/len(fp_validos)) if fp_validos else 0,
        'kwh_ativa': qfat.get('kwh_ativa', 0) if isinstance(qfat, dict) else 0,
        'kvarh_excedente': qfat.get('kvarh_excedente', 0) if isinstance(qfat, dict) else 0,
        'ponta_max': qfat.get('kw_ponta_lida', 0) if isinstance(qfat, dict) else 0,
        'agua_total': qfat.get('agua_total', 0) if isinstance(qfat, dict) else 0,
        'consumo_especifico': qfat.get('consumo_especifico', None) if isinstance(qfat, dict) else None,
        'avisos_fatura': qfat.get('avisos', []) if isinstance(qfat, dict) else [],
    }

    return render_template('leituras_mensal_audit.html', local=local, mes=mes, ano=ano, logs=logs, rows=rows, anomalies=anomalies, resumo=resumo_audit, qfat=qfat)


# === Pack 4: Importação CSV e Resumo Financeiro ===
@app.route('/leituras_mensal/import_csv', methods=['GET','POST'])
def leituras_mensal_import_csv():
    import io, csv
    msg = None; report = None
    hoje = datetime.now()
    if request.method == 'POST':
        local = (request.form.get('local') or '').strip()
        mes = (request.form.get('mes') or hoje.strftime('%m')).zfill(2)
        ano = int(request.form.get('ano') or hoje.year)
        fator_mult = float(request.form.get('fator_mult') or 1)
        file = request.files.get('csv_file')
        if not file or file.filename == '':
            msg = ('warning', 'Selecione um ficheiro CSV.')
        else:
            content = file.stream.read().decode('utf-8', errors='ignore')
            reader = csv.DictReader(io.StringIO(content))
            expected = ['data','hora','ativa','reativa','ponta','fp','potc','anterior','atual','diferenca','agua','esp','acum','valor']
            if [h.strip() for h in reader.fieldnames or []] != expected:
                msg = ('danger', 'Cabeçalho inválido. Baixe o template CSV e use exatamente as colunas esperadas.')
            else:
                conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                ins = upd = err = 0
                for r in reader:
                    try:
                        data = r['data'].strip()
                        hora = r['hora'].strip()
                        ativa = float(r['ativa'] or 0) * fator_mult
                        reativa = float(r['reativa'] or 0) * fator_mult
                        ponta = float(r['ponta'] or 0) * fator_mult
                        fpv = float(r['fp'] or 0); fpv = max(0.0, min(1.0, fpv))
                        potc = float(r['potc'] or 0)
                        anterior = float(r['anterior'] or 0)
                        atual = float(r['atual'] or 0)
                        dif = float(r['diferenca'] or (max(0.0, atual - anterior)))
                        agua = float(r['agua'] or 0)
                        esp = float(r['esp'] or 0)
                        acum = float(r['acum'] or 0)
                        valor = float(r['valor'] or 0)
                        prev = c.execute("SELECT 1 FROM leituras_mensais WHERE local=? AND data=?", (local, data)).fetchone()
                        c.execute('''INSERT INTO leituras_mensais(local,data,hora,ativa,reativa,ponta,fp,potc,anterior,atual,diferenca,agua,esp,acum,valor,mes,ano)
                                     VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) 
                                     ON CONFLICT(local,data) DO UPDATE SET
                                       hora=excluded.hora, ativa=excluded.ativa, reativa=excluded.reativa, ponta=excluded.ponta,
                                       fp=excluded.fp, potc=excluded.potc, anterior=excluded.anterior, atual=excluded.atual,
                                       diferenca=excluded.diferenca, agua=excluded.agua, esp=excluded.esp, acum=excluded.acum,
                                       valor=excluded.valor, mes=excluded.mes, ano=excluded.ano''',
                                  (local, data, hora, ativa, reativa, ponta, fpv, potc, anterior, atual, dif, agua, esp, acum, valor, mes, ano))
                        if prev: upd += 1
                        else: ins += 1
                    except Exception:
                        err += 1
                conn.commit(); conn.close()
                report = {'ins': ins, 'upd': upd, 'err': err, 'local': local, 'mes': mes, 'ano': ano}
                msg = ('success', f'Importação concluída. Inseridos={ins}, Atualizados={upd}, Erros={err}.')
    # GET or after POST render
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    locais = [r[0] for r in c.execute("SELECT DISTINCT nome FROM locais ORDER BY nome").fetchall()]
    conn.close()
    return render_template('leituras_mensal_import_csv.html', locais=locais, msg=msg, report=report)
# === FATURA EDM A PARTIR DAS LEITURAS MENSAIS ===

@app.route('/leituras_mensal/fatura_edm')
def leituras_mensal_fatura_edm():
    """
    Gera um resumo de fatura em layout tipo EDM, a partir das leituras_mensais,
    para um determinado local/mês/ano.
    Abre uma página própria, optimizada para impressão / PDF.
    """
    local = request.args.get('local', '').strip()
    mes   = request.args.get('mes', '').strip()   # "01".."12"
    ano   = request.args.get('ano', '').strip()

    if not local or not mes or not ano:
        # volta para leituras_mensal se faltar parâmetro
        return redirect(url_for('leituras_mensal'))

    try:
        ano_int = int(ano)
    except ValueError:
        ano_int = datetime.now().year

    mes_str = str(mes).zfill(2)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # 1) Descobrir o id do local e a configuração (tarifas, potência, etc.)
    c.execute('SELECT id, nome FROM locais WHERE nome = ?', (local,))
    row_loc = c.fetchone()
    local_id = row_loc['id'] if row_loc else None

    # Valores padrão caso não haja cfg no locais_cfg
    fator_mult     = 1.0
    pot_contratada = 0.0
    tarifa_ativa   = 4.780
    tarifa_reativa = 1.430
    tarifa_ponta   = 497.00
    taxa_fixa      = 207.28
    taxa_radio     = 297.00
    taxa_lixo      = 150.00
    iva_cfg        = 16.0  # só para referência, mas a regra EDM nova está abaixo

    if local_id is not None:
        c.execute('''
            SELECT fator_mult,
                   pot_contratada,
                   tarifa_ativa,
                   tarifa_reativa,
                   tarifa_ponta,
                   tarifa_perdas,
                   taxa_fixa,
                   taxa_radio,
                   taxa_lixo,
                   iva
            FROM locais_cfg
            WHERE local_id = ?
        ''', (local_id,))
        cfg = c.fetchone()
        if cfg:
            (fator_mult,
             pot_contratada,
             tarifa_ativa,
             tarifa_reativa,
             tarifa_ponta,
             _tarifa_perdas_ignorar,
             taxa_fixa,
             taxa_radio,
             taxa_lixo,
             iva_cfg) = cfg

    # Se a tarifa reativa não estiver configurada, usa-se a regra-base do tarifário:
    # energia reativa excedente cobrada a 30% do preço da energia ativa.
    if not tarifa_reativa and tarifa_ativa:
        tarifa_reativa = tarifa_ativa * 0.30

    # 2) Quantidades do mês a partir das leituras mensais.
    #    A função central abaixo evita erros comuns: somar leituras acumuladas,
    #    faturar a primeira leitura histórica quando não existe base anterior,
    #    ou usar reativa acumulada como se fosse consumo do mês.
    qfat = _quantidades_fatura_mensal(local, mes_str, ano_int)
    kwh_ativa = qfat['kwh_ativa']
    kvarh_reativa = qfat['kvarh_reativa']
    limite_reativa = qfat['limite_reativa']
    kvarh_excedente = qfat['kvarh_excedente']
    kw_ponta_lida = qfat['kw_ponta_lida']
    agua_total = qfat['agua_total']
    kwh_delta_total = kwh_ativa
    # 4) Demanda de ponta faturável (kW)
    #    Fórmula que combinámos: 20% * P contratada + 80% * Ponta medida (máxima do mês)
    demanda_ponta_kw = _ponta_faturavel_edm(pot_contratada, kw_ponta_lida)

    # 5) Custos de energia (sem perdas)
    valor_ativa   = kwh_ativa      * tarifa_ativa
    valor_reativa = kvarh_excedente * tarifa_reativa
    valor_ponta   = demanda_ponta_kw * tarifa_ponta
    valor_perdas  = 0.0  # retirado do cálculo, contador do lado de MT

    subtotal_energia = valor_ativa + valor_reativa + valor_ponta

    # 6) Taxas fixas
    subtotal_taxas = taxa_fixa + taxa_radio + taxa_lixo

    # 7) Subtotal antes de IVA
    subtotal = subtotal_energia + subtotal_taxas

    # 8) IVA: 16% sobre 62% do subtotal
    IVA_ALIQUOTA = 0.16
    BASE_IVA_PERC = 0.62
    base_iva  = subtotal * BASE_IVA_PERC
    valor_iva = base_iva * IVA_ALIQUOTA

    # 9) Total final
    total = subtotal + valor_iva

    # 10) Alguns indicadores extra (consumo específico médio, etc.)
    consumo_especifico_medio = None
    if agua_total > 0 and kwh_ativa > 0:
        consumo_especifico_medio = kwh_ativa / agua_total

    periodo_str = f"{mes_str}/{ano_int}"

    invoice_ctx = dict(
        local=local, periodo=periodo_str,
        kwh_ativa=kwh_ativa, kvarh_reativa=kvarh_reativa, kvarh_excedente=kvarh_excedente,
        kw_ponta_lida=kw_ponta_lida, demanda_ponta_kw=demanda_ponta_kw,
        valor_ativa=valor_ativa, valor_reativa=valor_reativa, valor_ponta=valor_ponta, valor_perdas=valor_perdas,
        subtotal_energia=subtotal_energia, taxa_fixa=taxa_fixa, taxa_radio=taxa_radio, taxa_lixo=taxa_lixo,
        subtotal_taxas=subtotal_taxas, subtotal=subtotal, base_iva=base_iva, valor_iva=valor_iva, total=total,
        total_extenso=_mzn_extenso(total), consumo_especifico_medio=consumo_especifico_medio, agua_total=agua_total,
        pot_contratada=pot_contratada, tarifa_ativa=tarifa_ativa, tarifa_reativa=tarifa_reativa, tarifa_ponta=tarifa_ponta,
        iva_percent=IVA_ALIQUOTA * 100, base_iva_percent=BASE_IVA_PERC * 100, limite_reativa=limite_reativa,
        leitura_base_ativa=qfat.get('leitura_base_ativa', 0), leitura_final_ativa=qfat.get('leitura_final_ativa', 0),
        leitura_base_reativa=qfat.get('leitura_base_reativa', 0), leitura_final_reativa=qfat.get('leitura_final_reativa', 0),
        avisos_fatura=qfat.get('avisos', [])
    )
    invoice_id = _arquivar_fatura_mensal_snapshot(invoice_ctx)

    return render_template(
        'leituras_mensal_fatura_edm.html',
        local=local,
        periodo=periodo_str,
        invoice_id=invoice_id,
        kwh_ativa=kwh_ativa,
        kvarh_reativa=kvarh_reativa,
        kvarh_excedente=kvarh_excedente,
        kw_ponta_lida=kw_ponta_lida,
        demanda_ponta_kw=demanda_ponta_kw,
        valor_ativa=valor_ativa,
        valor_reativa=valor_reativa,
        valor_ponta=valor_ponta,
        valor_perdas=valor_perdas,
        subtotal_energia=subtotal_energia,
        taxa_fixa=taxa_fixa,
        taxa_radio=taxa_radio,
        taxa_lixo=taxa_lixo,
        subtotal_taxas=subtotal_taxas,
        subtotal=subtotal,
        base_iva=base_iva,
        valor_iva=valor_iva,
        total=total,
        total_extenso=_mzn_extenso(total),
        consumo_especifico_medio=consumo_especifico_medio,
        agua_total=agua_total,
        pot_contratada=pot_contratada,
        tarifa_ativa=tarifa_ativa,
        tarifa_reativa=tarifa_reativa,
        tarifa_ponta=tarifa_ponta,
        iva_percent=IVA_ALIQUOTA * 100,
        base_iva_percent=BASE_IVA_PERC * 100,
        limite_reativa=limite_reativa,
        leitura_base_ativa=qfat.get('leitura_base_ativa', 0),
        leitura_final_ativa=qfat.get('leitura_final_ativa', 0),
        leitura_base_reativa=qfat.get('leitura_base_reativa', 0),
        leitura_final_reativa=qfat.get('leitura_final_reativa', 0),
        avisos_fatura=qfat.get('avisos', []),
    )

@app.route('/leituras_mensal/financeiro')
def leituras_mensal_financeiro():
    local = request.args.get('local','').strip()
    mes = (request.args.get('mes') or datetime.now().strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or datetime.now().year)

    lid = None
    for (lid_, nome) in get_locais():
        if nome == local:
            lid = lid_
            break
    cfg = get_local_cfg_full(lid) if lid is not None else {}

    pot_contratada = float(cfg.get('pot_contratada', 0) or 0)
    tarifa_ativa   = float(cfg.get('tarifa_ativa', 0) or 0)
    tarifa_reativa = float(cfg.get('tarifa_reativa', 0) or 0)
    tarifa_ponta   = float(cfg.get('tarifa_ponta', 0) or 0)
    taxa_fixa      = float(cfg.get('taxa_fixa', 0) or 0)
    taxa_radio     = float(cfg.get('taxa_radio', 0) or 0)
    taxa_lixo      = float(cfg.get('taxa_lixo', 0) or 0)
    if not tarifa_reativa and tarifa_ativa:
        tarifa_reativa = tarifa_ativa * 0.30

    qfat = _quantidades_fatura_mensal(local, mes, ano) if local else {}
    kwh_ativa = qfat.get('kwh_ativa', 0)
    kvarh_reativa = qfat.get('kvarh_reativa', 0)
    limite_reativa = qfat.get('limite_reativa', 0)
    kvarh_excedente = qfat.get('kvarh_excedente', 0)
    kw_ponta_lida = qfat.get('kw_ponta_lida', 0)
    demanda_ponta_kw = _ponta_faturavel_edm(pot_contratada, kw_ponta_lida)

    valor_ativa = kwh_ativa * tarifa_ativa
    valor_reativa = kvarh_excedente * tarifa_reativa
    valor_ponta = demanda_ponta_kw * tarifa_ponta
    subtotal_energia = valor_ativa + valor_reativa + valor_ponta
    subtotal_taxas = taxa_fixa + taxa_radio + taxa_lixo
    subtotal = subtotal_energia + subtotal_taxas
    IVA_ALIQUOTA = 0.16
    BASE_IVA_PERC = 0.62
    base_iva = subtotal * BASE_IVA_PERC
    valor_iva = base_iva * IVA_ALIQUOTA
    total = subtotal + valor_iva

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    rows = c.execute('''SELECT data, ativa, reativa, ponta, fp, diferenca, agua, esp, valor
                        FROM leituras_mensais WHERE local=? AND mes=? AND ano=?
                        ORDER BY data''', (local, mes, ano)).fetchall()
    conn.close()
    fp_vals = [_safe_float(r['fp'], None) for r in rows if _safe_float(r['fp'], None) is not None and _safe_float(r['fp'], None) > 0]
    fp_medio = (sum(fp_vals)/len(fp_vals)) if fp_vals else 0

    resumo = {
        'valor_total': total,
        'total_extenso': _mzn_extenso(total),
        'subtotal': subtotal,
        'subtotal_energia': subtotal_energia,
        'subtotal_taxas': subtotal_taxas,
        'base_iva': base_iva,
        'valor_iva': valor_iva,
        'kwh_ativa': kwh_ativa,
        'kvarh_reativa': kvarh_reativa,
        'limite_reativa': limite_reativa,
        'kvarh_excedente': kvarh_excedente,
        'kw_ponta_lida': kw_ponta_lida,
        'demanda_ponta_kw': demanda_ponta_kw,
        'valor_ativa': valor_ativa,
        'valor_reativa': valor_reativa,
        'valor_ponta': valor_ponta,
        'agua_total': qfat.get('agua_total', 0),
        'consumo_especifico': qfat.get('consumo_especifico', None),
        'fp_medio': fp_medio,
        'leitura_base_ativa': qfat.get('leitura_base_ativa', 0),
        'leitura_final_ativa': qfat.get('leitura_final_ativa', 0),
        'leitura_base_reativa': qfat.get('leitura_base_reativa', 0),
        'leitura_final_reativa': qfat.get('leitura_final_reativa', 0),
        'avisos': qfat.get('avisos', []),
        'iva_percent': IVA_ALIQUOTA * 100,
        'base_iva_percent': BASE_IVA_PERC * 100,
    }

    return render_template('leituras_mensal_financeiro.html',
        local=local, mes=mes, ano=ano, cfg=cfg, resumo=resumo, rows=rows, qfat=qfat,
        tarifa_ativa=tarifa_ativa, tarifa_reativa=tarifa_reativa, tarifa_ponta=tarifa_ponta,
        taxa_fixa=taxa_fixa, taxa_radio=taxa_radio, taxa_lixo=taxa_lixo,
        pot_contratada=pot_contratada)


@app.route('/leituras_mensal/faturas')
def leituras_mensal_faturas_arquivo():
    ensure_faturas_mensais_archive_schema()
    local = request.args.get('local','').strip()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    if local:
        faturas = c.execute('''SELECT * FROM faturas_mensais_arquivo WHERE local=? ORDER BY ano DESC, mes DESC''', (local,)).fetchall()
    else:
        faturas = c.execute('''SELECT * FROM faturas_mensais_arquivo ORDER BY atualizado_em DESC LIMIT 200''').fetchall()
    locais = [r[1] for r in get_locais()] if 'get_locais' in globals() else []
    conn.close()
    return render_template('leituras_mensal_faturas_arquivo.html', faturas=faturas, local=local, locais=locais)


@app.route('/leituras_mensal/arquivo')
def leituras_mensal_arquivo():
    ensure_leituras_mensais_phase2_schema()
    local = request.args.get('local','').strip()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    if local:
        rows = c.execute('''
            SELECT local, mes, ano,
                   COUNT(*) AS total_linhas,
                   SUM(CASE WHEN (ativa IS NOT NULL AND ativa>0) OR (reativa IS NOT NULL AND reativa>0) OR (ponta IS NOT NULL AND ponta>0) OR (agua IS NOT NULL AND agua>0) THEN 1 ELSE 0 END) AS dias_preenchidos,
                   MAX(data) AS ultima_data,
                   SUM(COALESCE(agua,0)) AS agua_total,
                   SUM(COALESCE(diferenca,0)) AS soma_diferencas
            FROM leituras_mensais
            WHERE local=?
            GROUP BY local, mes, ano
            ORDER BY ano DESC, mes DESC
        ''', (local,)).fetchall()
    else:
        rows = c.execute('''
            SELECT local, mes, ano,
                   COUNT(*) AS total_linhas,
                   SUM(CASE WHEN (ativa IS NOT NULL AND ativa>0) OR (reativa IS NOT NULL AND reativa>0) OR (ponta IS NOT NULL AND ponta>0) OR (agua IS NOT NULL AND agua>0) THEN 1 ELSE 0 END) AS dias_preenchidos,
                   MAX(data) AS ultima_data,
                   SUM(COALESCE(agua,0)) AS agua_total,
                   SUM(COALESCE(diferenca,0)) AS soma_diferencas
            FROM leituras_mensais
            GROUP BY local, mes, ano
            ORDER BY ano DESC, mes DESC, local ASC
        ''').fetchall()
    locais = [r[1] for r in get_locais()] if 'get_locais' in globals() else []
    conn.close()
    return render_template('leituras_mensal_arquivo.html', rows=rows, local=local, locais=locais)


def _draw_right(c, x, y, txt, font='Helvetica', size=8):
    c.setFont(font, size); c.drawRightString(x, y, str(txt))


def _fmt_pdf(v, nd=2):
    try:
        return _fmt_mil(v, nd)
    except Exception:
        return str(v)


@app.route('/leituras_mensal/fatura_edm_pdf')
def leituras_mensal_fatura_edm_pdf():
    local = request.args.get('local','').strip()
    mes = (request.args.get('mes') or datetime.now().strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or datetime.now().year)
    if not local:
        return redirect(url_for('leituras_mensal'))
    ctx = _montar_contexto_fatura_mensal(local, mes, ano)
    invoice_id = _arquivar_fatura_mensal_snapshot(ctx)

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    W, H = landscape(A4)
    margin = 28
    logo_path = os.path.join(BASE_DIR, 'static', 'adrmm_logo.png')

    if os.path.exists(logo_path):
        try:
            c.saveState(); c.setFillAlpha(0.08)
            c.drawImage(logo_path, W/2-145, H/2-120, width=290, height=240, preserveAspectRatio=True, mask='auto')
            c.restoreState()
        except Exception:
            pass
    c.setStrokeColor(colors.HexColor('#073b78')); c.setLineWidth(1.2)
    c.line(margin, H-76, W-margin, H-76)
    if os.path.exists(logo_path):
        try: c.drawImage(logo_path, margin, H-68, width=46, height=46, preserveAspectRatio=True, mask='auto')
        except Exception: pass
    c.setFillColor(colors.HexColor('#073b78'))
    c.setFont('Helvetica-Bold', 17); c.drawString(margin+58, H-44, 'Fatura de Energia - EDM (Modelo Interno)')
    c.setFont('Helvetica', 8.5); c.setFillColor(colors.HexColor('#244f78'))
    c.drawString(margin+58, H-58, 'Sistema de Gestão de Energia - Monitorização, análise e controlo energético')
    c.setFillColor(colors.black)
    _draw_right(c, W-margin, H-36, f"Local: {local}", 'Helvetica-Bold', 9)
    _draw_right(c, W-margin, H-50, f"Período: {ctx['periodo']}", 'Helvetica-Bold', 9)
    _draw_right(c, W-margin, H-64, f"Registo SGE: #{invoice_id or '-'}", 'Helvetica', 8)

    y = H-94
    c.setFont('Helvetica-Bold', 9); c.setFillColor(colors.HexColor('#073b78'))
    c.drawString(margin, y, 'Parâmetros:')
    c.setFillColor(colors.black); c.setFont('Helvetica', 8)
    c.drawString(margin+75, y, f"Pot. contratada: {_fmt_pdf(ctx['pot_contratada'],2)} kVA")
    c.drawString(margin+230, y, f"Tarifa ativa: {_fmt_pdf(ctx['tarifa_ativa'],4)} MT/kWh")
    c.drawString(margin+385, y, f"Tarifa reativa: {_fmt_pdf(ctx['tarifa_reativa'],4)} MT/kVArh")
    c.drawString(margin+555, y, f"Tarifa ponta: {_fmt_pdf(ctx['tarifa_ponta'],4)} MT/kW")
    y -= 18
    c.setFillColor(colors.HexColor('#f4faff')); c.rect(margin, y-18, W-2*margin, 28, fill=1, stroke=0)
    c.setFillColor(colors.HexColor('#102b43')); c.setFont('Helvetica', 7.6)
    c.drawString(margin+8, y-2, f"Base: Ativa = {_fmt_pdf(ctx['leitura_final_ativa'],2)} - {_fmt_pdf(ctx['leitura_base_ativa'],2)} = {_fmt_pdf(ctx['kwh_ativa'],2)} kWh | Reativa = {_fmt_pdf(ctx['leitura_final_reativa'],2)} - {_fmt_pdf(ctx['leitura_base_reativa'],2)} = {_fmt_pdf(ctx['kvarh_reativa'],2)} kVArh | Excedente = máx(Reativa - 0,75 × Ativa, 0)")

    def table_header(x, y, w, title):
        c.setFillColor(colors.HexColor('#eaf4ff')); c.rect(x, y, w, 18, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#073b78')); c.setFont('Helvetica-Bold', 9); c.drawString(x+7, y+5, title)
        c.setStrokeColor(colors.HexColor('#cfddeb')); c.rect(x, y-120, w, 138, fill=0, stroke=1)
    left_x = margin; right_x = W/2+6; box_w = W/2-margin-12; top = y-46
    table_header(left_x, top, box_w, 'Resumo de energia')
    table_header(right_x, top, box_w, 'Taxas, IVA e total')
    def row(x, y, name, qty, tarifa, valor, bold=False, dark=False):
        if dark:
            c.setFillColor(colors.HexColor('#0f2337')); c.rect(x, y-2, box_w, 17, fill=1, stroke=0); c.setFillColor(colors.white)
        else:
            c.setFillColor(colors.black)
        c.setFont('Helvetica-Bold' if bold else 'Helvetica', 8)
        c.drawString(x+7, y+3, name)
        if qty is not None: _draw_right(c, x+box_w-170, y+3, qty, 'Helvetica-Bold' if bold else 'Helvetica', 8)
        if tarifa is not None: _draw_right(c, x+box_w-88, y+3, tarifa, 'Helvetica-Bold' if bold else 'Helvetica', 8)
        _draw_right(c, x+box_w-8, y+3, valor, 'Helvetica-Bold' if bold else 'Helvetica', 8)
    yy = top-20
    row(left_x, yy, 'Energia ativa', f"{_fmt_pdf(ctx['kwh_ativa'],2)} kWh", _fmt_pdf(ctx['tarifa_ativa'],4), _fmt_pdf(ctx['valor_ativa'],2)); yy-=19
    row(left_x, yy, 'Reativa excedente', f"{_fmt_pdf(ctx['kvarh_excedente'],2)} kVArh", _fmt_pdf(ctx['tarifa_reativa'],4), _fmt_pdf(ctx['valor_reativa'],2)); yy-=19
    row(left_x, yy, 'Demanda de ponta', f"{_fmt_pdf(ctx['demanda_ponta_kw'],2)} kW", _fmt_pdf(ctx['tarifa_ponta'],4), _fmt_pdf(ctx['valor_ponta'],2)); yy-=22
    row(left_x, yy, 'Subtotal energia', None, None, _fmt_pdf(ctx['subtotal_energia'],2), True); yy-=22
    c.setFont('Helvetica', 7.2); c.setFillColor(colors.HexColor('#4e6982'))
    c.drawString(left_x+7, yy+5, f"Ponta máxima considerada: {_fmt_pdf(ctx['kw_ponta_lida'],2)} kW | Limite reativa: {_fmt_pdf(ctx['limite_reativa'],2)} kVArh")

    yy = top-20
    row(right_x, yy, 'Taxa fixa', None, None, _fmt_pdf(ctx['taxa_fixa'],2)); yy-=17
    row(right_x, yy, 'Taxa rádio', None, None, _fmt_pdf(ctx['taxa_radio'],2)); yy-=17
    row(right_x, yy, 'Taxa lixo', None, None, _fmt_pdf(ctx['taxa_lixo'],2)); yy-=18
    row(right_x, yy, 'Subtotal taxas', None, None, _fmt_pdf(ctx['subtotal_taxas'],2), True); yy-=18
    row(right_x, yy, 'Subtotal', None, None, _fmt_pdf(ctx['subtotal'],2)); yy-=17
    row(right_x, yy, f"Base IVA ({_fmt_pdf(ctx['base_iva_percent'],0)}%)", None, None, _fmt_pdf(ctx['base_iva'],2)); yy-=17
    row(right_x, yy, f"IVA {_fmt_pdf(ctx['iva_percent'],0)}%", None, None, _fmt_pdf(ctx['valor_iva'],2)); yy-=20
    row(right_x, yy, 'TOTAL A PAGAR', None, None, _fmt_pdf(ctx['total'],2), True, True)

    y2 = top-158
    c.setFillColor(colors.HexColor('#f8fbff')); c.rect(margin, y2-38, W-2*margin, 42, fill=1, stroke=1)
    c.setFillColor(colors.HexColor('#073b78')); c.setFont('Helvetica-Bold', 8.5); c.drawString(margin+8, y2-10, 'Valor a pagar por extenso:')
    c.setFillColor(colors.black); c.setFont('Helvetica', 8)
    ext = ctx.get('total_extenso') or _mzn_extenso(ctx['total'])
    max_chars = 145
    lines = [ext[i:i+max_chars] for i in range(0, len(ext), max_chars)]
    for i, line in enumerate(lines[:2]): c.drawString(margin+8, y2-24-(i*10), line)

    y3 = y2-58
    c.setFont('Helvetica-Bold', 8.5); c.setFillColor(colors.HexColor('#073b78')); c.drawString(margin, y3, 'Indicadores adicionais')
    c.setFillColor(colors.black); c.setFont('Helvetica', 8)
    c.drawString(margin, y3-14, f"Água total: {_fmt_pdf(ctx['agua_total'],2)} m³")
    ce = '-' if ctx['consumo_especifico_medio'] is None else f"{_fmt_pdf(ctx['consumo_especifico_medio'],3)} kWh/m³"
    c.drawString(margin+150, y3-14, f"Consumo específico médio: {ce}")
    c.drawRightString(W-margin, y3-14, 'Documento gerado pelo SGE / Equipa de Eficiência Energética')
    c.setStrokeColor(colors.HexColor('#cfddeb')); c.line(margin, 34, W-margin, 34)
    c.setFont('Helvetica', 7); c.setFillColor(colors.HexColor('#607d9d'))
    c.drawCentredString(W/2, 22, 'Documento interno de apoio à conferência da fatura EDM. Valores sujeitos à validação da fatura oficial e parâmetros tarifários vigentes.')
    c.showPage(); c.save(); buffer.seek(0)
    filename = f"Fatura_EDM_{local.replace(' ','_')}_{mes}_{ano}.pdf"
    return Response(buffer.getvalue(), mimetype='application/pdf', headers={'Content-Disposition': f'attachment; filename="{filename}"'})

def _mt_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def _mt_exec(conn, sql, args=()):
    cur = conn.cursor()
    cur.execute(sql, args)
    conn.commit()
    return cur

# --- Migração MT ---
def _mt_init_db():
    conn = _mt_conn(); c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS mt_config (
            id INTEGER PRIMARY KEY CHECK (id=1),
            alfa_reativa REAL DEFAULT 0.50,
            iva_taxa REAL DEFAULT 0.16,
            iva_base REAL DEFAULT 0.62,
            tarifa_ativa REAL DEFAULT 4.780,
            tarifa_reativa REAL DEFAULT 1.430,
            tarifa_potencia REAL DEFAULT 497.000
        )
    """)
    c.execute("INSERT OR IGNORE INTO mt_config (id) VALUES (1)")

    c.execute("""
        CREATE TABLE IF NOT EXISTS mt_leituras_raw (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            local_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            hora TEXT NOT NULL,
            ea_leitura REAL NOT NULL,
            er_leitura REAL NOT NULL,
            demanda_lida REAL NOT NULL,
            obs TEXT,
            UNIQUE(local_id, data, hora),
            FOREIGN KEY(local_id) REFERENCES locais(id) ON DELETE CASCADE
        )
    """)
    conn.commit(); conn.close()

_mt_init_db()

def _mt_cfg():
    conn = _mt_conn()
    row = _mt_exec(conn, "SELECT * FROM mt_config WHERE id=1").fetchone(); conn.close()
    return row

def _mt_get_local_cfg(local_id: int):
    """Obtém FM, PC e tarifas do local a partir de locais_cfg, com defaults de segurança."""
    conn = _mt_conn()
    row = _mt_exec(conn, """
        SELECT COALESCE(fator_mult,1.0) AS fm,
               COALESCE(pot_contratada,0.0) AS pc,
               COALESCE(tarifa_ativa,4.780) AS t_ativa,
               COALESCE(tarifa_reativa,1.430) AS t_reat,
               COALESCE(tarifa_ponta,497.000) AS t_pot
          FROM locais_cfg WHERE local_id=?
    """, (local_id,)).fetchone()
    conn.close()
    if not row:
        return (1.0, 0.0, 4.780, 1.430, 497.000)
    return (float(row["fm"]), float(row["pc"]), float(row["t_ativa"]), float(row["t_reat"]), float(row["t_pot"]))

def _mt_month_bounds(ano: int, mes: int):
    import calendar
    from datetime import date
    first = date(ano, mes, 1)
    last = date(ano, mes, calendar.monthrange(ano, mes)[1])
    return first.isoformat(), last.isoformat()

# --------- Rotas: Configuração MT ---------
@app.route("/mt/config", methods=["GET","POST"])
def mt_config():
    if request.method == "POST":
        alfa = float(request.form.get("alfa_reativa", 0.50) or 0.50)
        iva_taxa = float(request.form.get("iva_taxa", 0.16) or 0.16)
        iva_base = float(request.form.get("iva_base", 0.62) or 0.62)
        t_ativa = float(request.form.get("tarifa_ativa", 4.780) or 4.780)
        t_reat = float(request.form.get("tarifa_reativa", 1.430) or 1.430)
        t_pot = float(request.form.get("tarifa_potencia", 497.000) or 497.000)
        conn = _mt_conn()
        _mt_exec(conn, """UPDATE mt_config SET alfa_reativa=?, iva_taxa=?, iva_base=?, 
                          tarifa_ativa=?, tarifa_reativa=?, tarifa_potencia=? WHERE id=1""",
                 (alfa, iva_taxa, iva_base, t_ativa, t_reat, t_pot))
        conn.close()
        flash("Configuração MT atualizada.", "success")
        return redirect(url_for("mt_config"))
    cfg = _mt_cfg()
    return render_template_string("""
    {% extends "base.html" %}
    {% block content %}
    <h3>Configuração MT (Global)</h3>
    <form method="post" class="row g-3">
      <div class="col-md-2">
        <label class="form-label">α Reativa</label>
        <input name="alfa_reativa" type="number" step="0.01" class="form-control" value="{{ '%.2f'|format(cfg['alfa_reativa']) }}">
      </div>
      <div class="col-md-2">
        <label class="form-label">IVA (taxa)</label>
        <input name="iva_taxa" type="number" step="0.01" class="form-control" value="{{ '%.2f'|format(cfg['iva_taxa']) }}">
      </div>
      <div class="col-md-2">
        <label class="form-label">Base IVA (fator)</label>
        <input name="iva_base" type="number" step="0.01" class="form-control" value="{{ '%.2f'|format(cfg['iva_base']) }}">
      </div>
      <div class="col-md-2">
        <label class="form-label">Tarifa ativa (MZN/kWh)</label>
        <input name="tarifa_ativa" type="number" step="0.001" class="form-control" value="{{ '%.3f'|format(cfg['tarifa_ativa']) }}">
      </div>
      <div class="col-md-2">
        <label class="form-label">Tarifa reativa (MZN/kVArh)</label>
        <input name="tarifa_reativa" type="number" step="0.001" class="form-control" value="{{ '%.3f'|format(cfg['tarifa_reativa']) }}">
      </div>
      <div class="col-md-2">
        <label class="form-label">Tarifa potência (MZN/kVA)</label>
        <input name="tarifa_potencia" type="number" step="0.001" class="form-control" value="{{ '%.3f'|format(cfg['tarifa_potencia']) }}">
      </div>
      <div class="col-12">
        <button class="btn btn-primary mt-2">Guardar</button>
      </div>
    </form>
    <p class="mt-3 text-muted">IVA efetivo = taxa × base. Ex.: 0,16 × 0,62 = 0,0992 (9,92%).</p>
    {% endblock %}
    """, cfg=cfg)

# --------- Rotas: Leituras MT ---------
@app.route("/mt/<int:local_id>/leituras")
def mt_leituras(local_id):
    from datetime import date
    hoje = date.today()
    ano = int(request.args.get("ano", hoje.year))
    mes = int(request.args.get("mes", hoje.month))
    first, last = _mt_month_bounds(ano, mes)

    conn = _mt_conn()
    local = _mt_exec(conn, "SELECT id, nome FROM locais WHERE id=?", (local_id,)).fetchone()
    if not local:
        conn.close()
        flash("Local não encontrado.", "danger")
        return redirect(url_for("index"))

    rows = _mt_exec(conn, """
        SELECT * FROM mt_leituras_raw
        WHERE local_id=? AND date(data) BETWEEN ? AND ?
        ORDER BY date(data), time(hora)
    """, (local_id, first, last)).fetchall()
    cfg = _mt_cfg()
    fm, Pc, t_ativa_l, t_reat_l, t_pot_l = _mt_get_local_cfg(local_id)
    conn.close()

    # última leitura do dia
    by_day = {}
    for r in rows:
        d = r["data"]
        if d not in by_day or r["hora"] > by_day[d]["hora"]:
            by_day[d] = dict(r)

    dias = sorted(by_day.keys())
    tabela = []
    EA_total = ER_total = 0.0
    Dmax = 0.0
    prev_ea = prev_er = None

    for d in dias:
        r = by_day[d]
        ea_aj = float(r["ea_leitura"]) * fm
        er_aj = float(r["er_leitura"]) * fm
        demanda_aj = float(r["demanda_lida"]) * fm

        if prev_ea is None:
            dea = 0.0; der = 0.0
        else:
            dea = max(0.0, ea_aj - prev_ea)
            der = max(0.0, er_aj - prev_er)

        custo_ativo_d = dea * float(cfg["tarifa_ativa"])

        # indicadores simples
        P_d = (dea/24.0) if dea>0 else 0.0
        Q_d = (der/24.0) if der>0 else 0.0
        S_d = math.sqrt(P_d**2 + Q_d**2) if (P_d or Q_d) else 0.0
        FP_d = (P_d/S_d) if S_d>0 else None

        tabela.append({
            "data": d, "hora": r["hora"],
            "ea_leitura": ea_aj, "er_leitura": er_aj,
            "dea": dea, "der": der,
            "custo_ativo_d": custo_ativo_d,
            "demanda": demanda_aj,
            "FP_d": FP_d
        })

        EA_total += dea; ER_total += der
        Dmax = max(Dmax, demanda_aj)
        prev_ea, prev_er = ea_aj, er_aj

    alfa = float(cfg["alfa_reativa"])
    ER_exced = max(0.0, ER_total - alfa * EA_total)

    C_ativo = EA_total * float(cfg["tarifa_ativa"])
    C_reativa = ER_exced * float(cfg["tarifa_reativa"])
    P_fat = 0.2 * Pc + 0.8 * Dmax
    C_pot = P_fat * float(cfg["tarifa_potencia"])

    subtotal = C_ativo + C_reativa + C_pot
    iva = float(cfg["iva_taxa"]) * float(cfg["iva_base"]) * subtotal
    total = subtotal + iva

    resumo = {
        "EA_total": EA_total, "ER_total": ER_total,
        "ER_exced": ER_exced, "C_ativo": C_ativo, "C_reativa": C_reativa,
        "Dmax": Dmax, "Pc": Pc, "P_fat": P_fat, "C_pot": C_pot,
        "subtotal": subtotal, "iva": iva, "total": total,
        "alfa": alfa,
        "iva_taxa": float(cfg["iva_taxa"]), "iva_base": float(cfg["iva_base"]),
        "tarifa_ativa": float(cfg["tarifa_ativa"]),
        "tarifa_reativa": float(cfg["tarifa_reativa"]),
        "tarifa_potencia": float(cfg["tarifa_potencia"]),
        "ano": ano, "mes": mes
    }

    return render_template_string("""
    {% extends "base.html" %}
    {% block content %}
    <div class="d-flex justify-content-between align-items-center">
      <h3>Leituras Mensais (MT) — {{ local['nome'] }}</h3>
      <div><a class="btn btn-sm btn-primary" href="{{ url_for('mt_nova_leitura', local_id=local['id']) }}">+ Nova leitura</a></div>
    </div>

    <form method="get" class="row g-2 my-2">
      <div class="col-auto">
        <label class="form-label">Mês</label>
        <input class="form-control" type="number" name="mes" value="{{ mes }}" min="1" max="12">
      </div>
      <div class="col-auto">
        <label class="form-label">Ano</label>
        <input class="form-control" type="number" name="ano" value="{{ ano }}" min="2000" max="2100">
      </div>
      <div class="col-auto align-self-end">
        <button class="btn btn-outline-secondary">Ir</button>
      </div>
    </form>

    <div class="table-responsive">
    <table class="table table-striped table-sm align-middle">
      <thead>
        <tr>
          <th>Data</th><th>Hora</th>
          <th>Leitura Ativa (aj.)</th>
          <th>Leitura Reativa (aj.)</th>
          <th>Δ Ativa (kWh)</th>
          <th>Δ Reativa (kVArh)</th>
          <th>Custo ativo (MZN)</th>
          <th>Demanda lida (kVA)</th>
          <th>FP (aprox.)</th>
        </tr>
      </thead>
      <tbody>
        {% for r in tabela %}
          <tr>
            <td>{{ r.data }}</td>
            <td>{{ r.hora }}</td>
            <td>{{ '%.3f'|format(r.ea_leitura) }}</td>
            <td>{{ '%.3f'|format(r.er_leitura) }}</td>
            <td>{{ '%.3f'|format(r.dea) }}</td>
            <td>{{ '%.3f'|format(r.der) }}</td>
            <td>{{ '%.2f'|format(r.custo_ativo_d) }}</td>
            <td>{{ '%.3f'|format(r.demanda) }}</td>
            <td>
              {% if r.FP_d is not none %}
                <span class="badge bg-{{ 'success' if r.FP_d>=0.92 else 'warning' }}">{{ '%.3f'|format(r.FP_d) }}</span>
              {% else %} — {% endif %}
            </td>
          </tr>
        {% else %}
          <tr><td colspan="9" class="text-center text-muted">Sem leituras para o mês.</td></tr>
        {% endfor %}
      </tbody>
    </table>
    </div>

    <hr>
    <h5>Resumo do mês {{ "%02d"|format(mes) }}/{{ ano }}</h5>
    <div class="row g-3">
      <div class="col-md-3">
        <div class="card"><div class="card-body">
          <div class="small text-muted">Energia Ativa total</div>
          <div class="h5">{{ '%.3f'|format(resumo.EA_total) }} kWh</div>
          <div class="small text-muted">Tarifa ativa: {{ '%.3f'|format(resumo.tarifa_ativa) }} MZN/kWh</div>
          <div class="small">Custo: <strong>{{ '%.2f'|format(resumo.C_ativo) }} MZN</strong></div>
        </div></div>
      </div>
      <div class="col-md-3">
        <div class="card"><div class="card-body">
          <div class="small text-muted">Energia Reativa total</div>
          <div class="h5">{{ '%.3f'|format(resumo.ER_total) }} kVArh</div>
          <div class="small text-muted">Excedente (α={{ '%.2f'|format(resumo.alfa) }}): <strong>{{ '%.3f'|format(resumo.ER_exced) }} kVArh</strong></div>
          <div class="small">Custo: <strong>{{ '%.2f'|format(resumo.C_reativa) }} MZN</strong></div>
        </div></div>
      </div>
      <div class="col-md-3">
        <div class="card"><div class="card-body">
          <div class="small text-muted">Demanda máxima (Dmax)</div>
          <div class="h5">{{ '%.3f'|format(resumo.Dmax) }} kVA</div>
          <div class="small text-muted">Ponta faturável = 0,2·PC + 0,8·Dmax</div>
          <div class="small">PC: {{ '%.2f'|format(resumo.Pc) }} kVA</div>
          <div class="small">P_fat: <strong>{{ '%.3f'|format(resumo.P_fat) }} kVA</strong></div>
          <div class="small">Custo potência: <strong>{{ '%.2f'|format(resumo.C_pot) }} MZN</strong></div>
        </div></div>
      </div>
      <div class="col-md-3">
        <div class="card"><div class="card-body">
          <div class="small text-muted">Totais</div>
          <div class="small">Subtotal: <strong>{{ '%.2f'|format(resumo.subtotal) }} MZN</strong></div>
          <div class="small">IVA ({{ '%.2f'|format(resumo.iva_taxa*100) }}% de {{ '%.0f'|format(resumo.iva_base*100) }}%):
            <strong>{{ '%.2f'|format(resumo.iva) }} MZN</strong></div>
          <div class="h5">Total: {{ '%.2f'|format(resumo.total) }} MZN</div>
        </div></div>
      </div>
    </div>

    <div class="mt-3">
      <a class="btn btn-secondary" href="{{ url_for('index') }}">Início</a>
      <a class="btn btn-outline-primary" href="{{ url_for('mt_nova_leitura', local_id=local['id']) }}">Lançar outra leitura</a>
    </div>
    {% endblock %}
    """, local=local, tabela=tabela, resumo=resumo, ano=ano, mes=mes)

@app.route("/mt/<int:local_id>/leituras/novo", methods=["GET","POST"])
def mt_nova_leitura(local_id):
    conn = _mt_conn()
    local = _mt_exec(conn, "SELECT id, nome FROM locais WHERE id=?", (local_id,)).fetchone()
    if not local:
        conn.close()
        flash("Local não encontrado.", "danger")
        return redirect(url_for("index"))
    if request.method == "POST":
        data = (request.form.get("data") or "").strip()
        hora = (request.form.get("hora") or "").strip()
        try:
            _ = data and hora  # formatos validados pelo HTML5; backend tolerante
        except:
            pass
        ea = float(request.form.get("ea_leitura", 0) or 0)
        er = float(request.form.get("er_leitura", 0) or 0)
        demanda = float(request.form.get("demanda_lida", 0) or 0)
        obs = request.form.get("obs")
        try:
            _mt_exec(conn, """INSERT INTO mt_leituras_raw (local_id, data, hora, ea_leitura, er_leitura, demanda_lida, obs)
                              VALUES (?,?,?,?,?,?,?)""",
                     (local_id, data, hora, ea, er, demanda, obs))
            conn.close()
            flash("Leitura registada (MT).", "success")
        except sqlite3.IntegrityError:
            conn.close()
            flash("Já existe leitura para este local nesta data/hora.", "warning")
        return redirect(url_for("mt_leituras", local_id=local_id))
    conn.close()
    return render_template_string("""
    {% extends "base.html" %}
    {% block content %}
    <h3>Nova leitura (MT) — {{ local['nome'] }}</h3>
    <form method="post" class="row g-3">
      <div class="col-md-3">
        <label class="form-label">Data</label>
        <input name="data" type="date" class="form-control" required>
      </div>
      <div class="col-md-2">
        <label class="form-label">Hora</label>
        <input name="hora" type="time" class="form-control" required>
      </div>
      <div class="col-md-3">
        <label class="form-label">Leitura Ativa (kWh)</label>
        <input name="ea_leitura" type="number" step="0.001" class="form-control" required>
      </div>
      <div class="col-md-3">
        <label class="form-label">Leitura Reativa (kVArh)</label>
        <input name="er_leitura" type="number" step="0.001" class="form-control" required>
      </div>
      <div class="col-md-3">
        <label class="form-label">Demanda lida (kVA máx. no mês)</label>
        <input name="demanda_lida" type="number" step="0.001" class="form-control" required>
      </div>
      <div class="col-12">
        <label class="form-label">Observações</label>
        <textarea name="obs" class="form-control" rows="2"></textarea>
      </div>
      <div class="col-12">
        <button class="btn btn-primary">Registar</button>
        <a class="btn btn-secondary" href="{{ url_for('mt_leituras', local_id=local['id']) }}">Voltar</a>
      </div>
    </form>
    {% endblock %}
    """, local=local)


# === API: Configurações do Local por ID ===
@app.route("/api/local_cfg_by_id/<int:local_id>")
def api_local_cfg_by_id(local_id):
    from flask import jsonify, g
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(
        "SELECT l.id, l.nome, "
        "COALESCE(lc.pot_contratada,0.0), COALESCE(lc.pot_instalada,0.0), COALESCE(lc.fator_mult,1.0), "
        "COALESCE(lc.tarifa_ativa,0.0), COALESCE(lc.tarifa_reativa,0.0), COALESCE(lc.tarifa_ponta,0.0), COALESCE(lc.tarifa_perdas,0.0), "
        "COALESCE(lc.taxa_fixa,0.0), COALESCE(lc.taxa_radio,0.0), COALESCE(lc.taxa_lixo,0.0), COALESCE(lc.iva,0.0) "
        "FROM locais l LEFT JOIN locais_cfg lc ON l.id = lc.local_id WHERE l.id = ?",
        (local_id,)
    )
    row = c.fetchone(); conn.close()
    if not row:
        return jsonify({"erro":"Local não encontrado","id":local_id}), 404
    keys = ["id","nome","pot_contratada","pot_instalada","fator_mult",
            "tarifa_ativa","tarifa_reativa","tarifa_ponta","tarifa_perdas",
            "taxa_fixa","taxa_radio","taxa_lixo","iva"]
    return jsonify(dict(zip(keys,row)))


# === API: Calcular Fatura (Leituras Mensais) ===

# === API: Configurações do Local por NOME ===
@app.route("/api/local_cfg_by_name/<path:local_name>")
def api_local_cfg_by_name(local_name):
    from flask import jsonify, g
    name = (local_name or "").strip()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute(
        "SELECT l.id, l.nome, "
        "COALESCE(lc.pot_contratada,0.0), COALESCE(lc.pot_instalada,0.0), COALESCE(lc.fator_mult,1.0), "
        "COALESCE(lc.tarifa_ativa,0.0), COALESCE(lc.tarifa_reativa,0.0), COALESCE(lc.tarifa_ponta,0.0), COALESCE(lc.tarifa_perdas,0.0), "
        "COALESCE(lc.taxa_fixa,0.0), COALESCE(lc.taxa_radio,0.0), COALESCE(lc.taxa_lixo,0.0), COALESCE(lc.iva,0.0) "
        "FROM locais l LEFT JOIN locais_cfg lc ON l.id = lc.local_id WHERE l.nome = ?",
        (name,)
    )
    row = c.fetchone(); conn.close()
    if not row:
        return jsonify({"erro":"Local não encontrado","nome":name}), 404
    keys = ["id","nome","pot_contratada","pot_instalada","fator_mult",
            "tarifa_ativa","tarifa_reativa","tarifa_ponta","tarifa_perdas",
            "taxa_fixa","taxa_radio","taxa_lixo","iva"]
    return jsonify(dict(zip(keys,row)))

@app.route('/api/leituras_mensal/calcular', methods=['POST'])
def api_calc_fatura_leituras():
    from flask import request, jsonify, g
    data = request.get_json(silent=True) or {}
    local = (data.get('local') or '').strip()
    mes   = str(data.get('mes') or '').zfill(2)
    ano   = int(data.get('ano') or 0)
    if not local or not mes or not ano:
        return jsonify({'erro':'Parâmetros inválidos'}), 400

    lid = None
    for (lid_, nome) in get_locais():
        if nome == local:
            lid = lid_
            break
    cfg = get_local_cfg_full(lid) if lid is not None else {}
    t_ativa   = float(cfg.get('tarifa_ativa', 0) or 0)
    t_reativa = float(cfg.get('tarifa_reativa', 0) or 0)
    t_ponta   = float(cfg.get('tarifa_ponta', 0) or 0)
    t_perdas  = float(cfg.get('tarifa_perdas', 0) or 0)
    taxa_fixa = float(cfg.get('taxa_fixa', 0) or 0)
    taxa_radio= float(cfg.get('taxa_radio', 0) or 0)
    taxa_lixo = float(cfg.get('taxa_lixo', 0) or 0)
    iva       = float(cfg.get('iva', 0) or 0)

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute(
        "SELECT IFNULL(ativa,0), IFNULL(reativa,0), IFNULL(ponta,0) "
        "FROM leituras_mensais WHERE local=? AND mes=? AND ano=?",
        (local, mes, ano)
    ).fetchall()
    conn.close()
    if not rows:
        return jsonify({'erro':'Sem dados de leituras para o período selecionado.'}), 404

    kwh_total   = sum((r[0] or 0) for r in rows)
    kvarh_total = sum((r[1] or 0) for r in rows)
    demanda_max = max((r[2] or 0) for r in rows)

    sub_ativa   = kwh_total   * t_ativa
    sub_reativa = kvarh_total * t_reativa
    sub_ponta   = demanda_max * t_ponta
    sub_perdas  = kwh_total   * t_perdas
    sub_energia = sub_ativa + sub_reativa + sub_ponta + sub_perdas
    sub_taxas   = taxa_fixa + taxa_radio + taxa_lixo
    total_siva  = sub_energia + sub_taxas
    total_civa  = total_siva * (1 + (iva/100.0))

    return jsonify({
        'local': local, 'mes': mes, 'ano': ano,
        'totais': {'kwh': kwh_total, 'kvarh': kvarh_total, 'demanda_kw': demanda_max},
        'tarifas': {'ativa': t_ativa, 'reativa': t_reativa, 'ponta': t_ponta, 'perdas': t_perdas},
        'taxas': {'fixa': taxa_fixa, 'radio': taxa_radio, 'lixo': taxa_lixo, 'iva_percent': iva},
        'subtotal': {'ativa': sub_ativa, 'reativa': sub_reativa, 'ponta': sub_ponta, 'perdas': sub_perdas,
                     'energia': sub_energia, 'taxas': sub_taxas},
        'total': {'sem_iva': total_siva, 'com_iva': total_civa}
    })



# ----------------- API AUX: leitura anterior (para auto-preenchimento) -----------------
from flask import jsonify, request
import sqlite3

@app.get('/api/leituras/prev')
def api_prev_leitura():
    """Retorna a última leitura atual (leit_atual/leitura_atual) anterior à 'data' para o 'local' dado.
    Params: local (int), data (YYYY-MM-DD)"""
    try:
        local = request.args.get('local', type=int)
        data = request.args.get('data', type=str)
        if not local or not data:
            return jsonify(success=False, error="missing_params"), 400
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            c.execute("""
                SELECT COALESCE(leit_atual, leitura_atual, 0)
                FROM leituras_mensais
                WHERE local=? AND date(data) < date(?)
                ORDER BY date(data) DESC, rowid DESC
                LIMIT 1
            """, (local, data))
        except Exception:
            c.execute("""
                SELECT COALESCE(leit_atual, leitura_atual, 0)
                FROM leituras_mensais
                WHERE local=? AND data < ?
                ORDER BY data DESC, rowid DESC
                LIMIT 1
            """, (local, data))
        row = c.fetchone()
        conn.close()
        prev = float(row[0]) if row and row[0] is not None else 0.0
        return jsonify(success=True, prev=prev)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500
# --- Compat: rota alias entre /leituras_mensal e /leituras_mensais ---
try:
    from flask import redirect, url_for, request, g, render_template

    # 1) /leituras_mensal → rota principal do módulo mensal
    #    Se já existir uma função 'leituras_mensal' noutro sítio, este bloco NÃO cria outra.
    if not any(r.endpoint == 'leituras_mensal' for r in app.url_map.iter_rules()):
        @app.route('/leituras_mensal', methods=['GET'], endpoint='leituras_mensal')
        def leituras_mensal():
            # Abre o template mensal (sem redirecionar para export, nem procurar outras rotas)
            return render_template('leituras_mensal.html')

    # 2) /leituras_mensais → alias simples para /leituras_mensal
    #    Não usa mais procura por substring 'mensais' ou 'mensal', evitando cair em export.
    if not any(r.endpoint == 'leituras_mensais' for r in app.url_map.iter_rules()):
        @app.route('/leituras_mensais', methods=['GET'], endpoint='leituras_mensais')
        def _alias_leituras_mensais():
            return redirect(url_for('leituras_mensal'))

except Exception as _e:
    # Não falhar a app por causa de alias
    print("Alias leituras_mensal/mensais falhou:", _e)


# --- Safe single-shot DB init after all defs ---
try:
    init_db()
except Exception as _e:
    print("init_db falhou:", _e)


# === RESPOSTAS JSON PADRÃO PARA ERROS EM AJAX/JSON ===
from flask import jsonify, g
def _wants_json():
    if request.is_json:
        return True
    hx = request.headers.get('X-Requested-With','').lower()
    return 'xmlhttprequest' in hx or request.path.startswith('/api/')

@app.errorhandler(400)
def _err_400(e):
    if _wants_json():
        return jsonify(success=False, error="bad_request", message=str(e)), 400
    return e

@app.errorhandler(404)
def _err_404(e):
    if _wants_json():
        return jsonify(success=False, error="not_found", message="Recurso não encontrado"), 404
    return e

@app.errorhandler(500)
def _err_500(e):
    if _wants_json():
        return jsonify(success=False, error="server_error", message="Erro interno no servidor"), 500
    return e


# === VALIDAÇÕES POR LOCAL PARA LEITURAS_MENSAIS ===
from flask import abort, g

def _get_validacao_local(local):
    try:
        return get_validacao_local(local)
    except Exception:
        return {'fp_min': 0.85, 'kwh_dia_max': None, 'permitir_regressivo': 0}

def _calc_fp(ativa, reativa):
    try:
        ativa = float(str(ativa).replace(',','.'))
        reativa = float(str(reativa).replace(',','.'))
    except Exception:
        return None
    try:
        import math
        aparente = (ativa**2 + reativa**2)**0.5
        if aparente <= 0: 
            return None
        return round(ativa/aparente, 4)
    except Exception:
        return None

@app.before_request
def _leituras_mensais_guard():
    # Intercepta apenas POST aos módulos de leituras_mensais
    if request.method != 'POST':
        return
    p = request.path.lower()
    if 'leituras_mensal' not in p:  # cobre 'mensal' e 'mensais' pelo contains
        return
    try:
        local = request.form.get('local') or request.json.get('local') if request.is_json else None
        ativa = request.form.get('ativa') or request.json.get('ativa') if request.is_json else None
        reativa = request.form.get('reativa') or request.json.get('reativa') if request.is_json else None
        anterior = request.form.get('anterior') or request.json.get('anterior') if request.is_json else None
        atual = request.form.get('atual') or request.json.get('atual') if request.is_json else None

        if not local:
            return  # deixa a rota tratar campos obrigatórios

        rules = _get_validacao_local(local)
        # FP mínimo
        if ativa is not None and reativa is not None and rules.get('fp_min'):
            fp = _calc_fp(ativa, reativa)
            if fp is not None and fp < float(rules['fp_min']):
                msg = f"Fator de potência ({fp}) abaixo do mínimo ({rules['fp_min']}) definido para o local {local}."
                if _wants_json():
                    return jsonify(success=False, error="fp_min", message=msg), 400
                abort(400, description=msg)

        # Regressivo
        if rules.get('permitir_regressivo') in (0, '0', None):
            try:
                a0 = float(str(anterior).replace(',','.')) if anterior is not None else None
                a1 = float(str(atual).replace(',','.')) if atual is not None else None
                if a0 is not None and a1 is not None and a1 < a0:
                    msg = "Leitura atual inferior à anterior (regressivo não permitido para este local)."
                    if _wants_json():
                        return jsonify(success=False, error="regressivo", message=msg), 400
                    abort(400, description=msg)
            except Exception:
                pass

        # Limite diário de kWh (se front enviar um 'delta' já calculado)
        kwh_dia = request.form.get('kwh_dia') or (request.json.get('kwh_dia') if request.is_json else None)
        if rules.get('kwh_dia_max') and kwh_dia is not None:
            try:
                kd = float(str(kwh_dia).replace(',','.'))
                if kd > float(rules['kwh_dia_max']):
                    msg = f"Consumo diário ({kd} kWh) excede o limite ({rules['kwh_dia_max']} kWh) deste local."
                    if _wants_json():
                        return jsonify(success=False, error="kwh_dia_max", message=msg), 400
                    abort(400, description=msg)
            except Exception:
                pass
    except Exception as _e:
        # Não bloquear request em caso de exceção de validação
        if _logger:
            _logger.warning("Validação leituras_mensais falhou: %s", _e)
        return


def _ensure_indices():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    try:
        c.execute("CREATE TABLE IF NOT EXISTS saved_filters (id INTEGER PRIMARY KEY AUTOINCREMENT, user TEXT, modulo TEXT, nome TEXT, query_json TEXT, created_at TEXT DEFAULT (datetime('now','localtime')) )")
        cols = {r[1] for r in c.execute('PRAGMA table_info(saved_filters)').fetchall()}
        if 'nome' not in cols:
            c.execute("ALTER TABLE saved_filters ADD COLUMN nome TEXT")
            if 'name' in cols:
                c.execute("UPDATE saved_filters SET nome = COALESCE(nome, name)")
        if 'name' not in cols:
            c.execute("ALTER TABLE saved_filters ADD COLUMN name TEXT")
            c.execute("UPDATE saved_filters SET name = COALESCE(name, nome)")
        if 'query_json' not in cols:
            c.execute("ALTER TABLE saved_filters ADD COLUMN query_json TEXT")
    except Exception:
        pass
    try:
        c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_leit_mensal_unique ON leituras_mensais(local, data, mes, ano)")
    except Exception:
        pass
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_leit_mensal_periodo ON leituras_mensais(mes, ano, local)")
    except Exception:
        pass
    try:
        c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_locais_nome ON locais(nome)")
    except Exception:
        pass
    conn.commit(); conn.close()

# Garante criação após migrações
try:
    _ensure_indices()
except Exception as _e:
    print("Falha ao garantir índices:", _e)

# --- Endpoint aliases para compatibilidade com templates antigos ---
from flask import redirect, url_for, request

def _alias_to(endpoint, **kwargs):
    try:
        return redirect(url_for(endpoint, **kwargs))
    except Exception:
        # fallback: redireciona para home
        return redirect(url_for('index'))

# /leituras  (endpoint antigo 'leituras' aponta para função 'leituras_list')
try:
    if not any(r.endpoint == 'leituras' for r in app.url_map.iter_rules()):
        @app.route('/legacy/leituras', methods=['GET'], endpoint='leituras')
        def leituras():
            return _alias_to('leituras_list', **request.args.to_dict())
except Exception:
    pass

# 'locais' (endpoint antigo) -> 'listar_locais'
try:
    if not any(r.endpoint == 'locais' for r in app.url_map.iter_rules()):
        @app.route('/legacy/locais', methods=['GET'], endpoint='locais')
        def locais():
            return _alias_to('listar_locais', **request.args.to_dict())
except Exception:
    pass

# 'config_mt' -> 'mt_config'
try:
    if not any(r.endpoint == 'config_mt' for r in app.url_map.iter_rules()):
        @app.route('/legacy/mt/config', methods=['GET','POST'], endpoint='config_mt')
        def config_mt():
            return _alias_to('mt_config', **request.args.to_dict())
except Exception:
    pass

# 'leituras_diarias_*' -> mapeamentos aproximados
try:
    if not any(r.endpoint == 'leituras_diarias_view' for r in app.url_map.iter_rules()):
        @app.route('/leituras_diarias/view', methods=['GET'], endpoint='leituras_diarias_view')
        def leituras_diarias_view():
            return _alias_to('leituras_list', **request.args.to_dict())
except Exception:
    pass

try:
    if not any(r.endpoint == 'leituras_diarias_add' for r in app.url_map.iter_rules()):
        @app.route('/leituras_diarias/add', methods=['GET','POST'], endpoint='leituras_diarias_add')
        def leituras_diarias_add():
            return _alias_to('leituras_mensal', **request.args.to_dict())
except Exception:
    pass

try:
    if not any(r.endpoint == 'leituras_diarias_export' for r in app.url_map.iter_rules()):
        @app.route('/leituras_diarias/export', methods=['GET'], endpoint='leituras_diarias_export')
        def leituras_diarias_export():
            # se existir export específico mensal, prioriza
            for r in app.url_map.iter_rules():
                if r.endpoint == 'leituras_mensal_export':
                    return _alias_to('leituras_mensal_export', **request.args.to_dict())
            return _alias_to('leituras_export_csv', **request.args.to_dict())
except Exception:
    pass


# === ERROS: Render de template para navegação normal ===
from flask import render_template

def _render_error(status_code:int, message:str):
    try:
        # Tenta usar um template dedicado se existir
        return render_template('error.html', status_code=status_code, message=message), status_code
    except Exception:
        # Fallback simples em HTML
        return f"<h2>Erro {status_code}</h2><p>{message}</p>", status_code

@app.errorhandler(400)
def _err_400_page(e):
    if _wants_json():
        from flask import jsonify
        return jsonify(success=False, error="bad_request", message=str(getattr(e, 'description', e))), 400
    return _render_error(400, str(getattr(e, 'description', e)))

@app.errorhandler(404)
def _err_404_page(e):
    if _wants_json():
        from flask import jsonify
        return jsonify(success=False, error="not_found", message="Recurso não encontrado"), 404
    return _render_error(404, "Recurso não encontrado")

@app.errorhandler(500)
def _err_500_page(e):
    if _wants_json():
        from flask import jsonify
        return jsonify(success=False, error="server_error", message="Erro interno no servidor"), 500
    return _render_error(500, "Erro interno no servidor")


# === Normalização de endpoints (nomes padronizados) ===
# Mantém compatibilidade com os existentes e os legados.
try:
    # lista padronizada
    if not any(r.endpoint == 'leituras_mensais_list' for r in app.url_map.iter_rules()):
        @app.route('/legacy/leituras_mensais', methods=['GET'], endpoint='leituras_mensais_list')
        def leituras_mensais_list():
            return _alias_to('leituras_list', **request.args.to_dict())

    if not any(r.endpoint == 'leituras_mensais_add' for r in app.url_map.iter_rules()):
        @app.route('/leituras_mensais/add', methods=['GET','POST'], endpoint='leituras_mensais_add')
        def leituras_mensais_add():
            return _alias_to('leituras_mensal', **request.args.to_dict())

    if not any(r.endpoint == 'leituras_mensais_export' for r in app.url_map.iter_rules()):
        @app.route('/leituras_mensais/export', methods=['GET'], endpoint='leituras_mensais_export')
        def leituras_mensais_export():
            for r in app.url_map.iter_rules():
                if r.endpoint == 'leituras_mensal_export':
                    return _alias_to('leituras_mensal_export', **request.args.to_dict())
            return _alias_to('leituras_export_csv', **request.args.to_dict())
except Exception as _e:
    pass


# === Fase 7B: Dashboard auditado, financeiro e executivo ===
def _dashboard_get_cfg_map():
    """Devolve configurações dos locais usadas no dashboard sem quebrar bases antigas."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    cfg = {}
    try:
        rows = c.execute("""
            SELECT l.id, l.nome,
                   COALESCE(lc.fator_mult, l.fator_multiplicativo, 1.0) AS fator_mult,
                   COALESCE(lc.pot_contratada, l.potencia_contratada, l.potencia_contratada_kva, 0.0) AS pot_contratada,
                   COALESCE(lc.tarifa_ativa, 4.78) AS tarifa_ativa,
                   COALESCE(lc.tarifa_reativa, 0) AS tarifa_reativa,
                   COALESCE(lc.tarifa_ponta, 497.0) AS tarifa_ponta,
                   COALESCE(lc.taxa_fixa, 0) AS taxa_fixa,
                   COALESCE(lc.taxa_radio, 0) AS taxa_radio,
                   COALESCE(lc.taxa_lixo, 0) AS taxa_lixo,
                   COALESCE(lc.iva, 16) AS iva
            FROM locais l
            LEFT JOIN locais_cfg lc ON lc.local_id = l.id
        """).fetchall()
        for r in rows:
            d = dict(r)
            if not d.get('tarifa_reativa') and d.get('tarifa_ativa'):
                d['tarifa_reativa'] = float(d.get('tarifa_ativa') or 0) * 0.30
            cfg[d['nome']] = d
    except Exception:
        pass
    finally:
        conn.close()
    return cfg


def _dashboard_month_finance(local_nome, mes, ano, cfg):
    """Calcula valores do dashboard usando a mesma filosofia da fatura EDM mensal."""
    try:
        q = _quantidades_fatura_mensal(local_nome, str(mes).zfill(2), int(ano))
    except Exception:
        q = {'kwh_ativa':0.0,'kvarh_reativa':0.0,'kvarh_excedente':0.0,'kw_ponta_lida':0.0,'agua_total':0.0,'consumo_especifico':None,'avisos':[]}
    tarifa_ativa = float(cfg.get('tarifa_ativa') or 0)
    tarifa_reativa = float(cfg.get('tarifa_reativa') or (tarifa_ativa * 0.30 if tarifa_ativa else 0))
    tarifa_ponta = float(cfg.get('tarifa_ponta') or 0)
    pot_contratada = float(cfg.get('pot_contratada') or 0)
    taxa_fixa = float(cfg.get('taxa_fixa') or 0)
    taxa_radio = float(cfg.get('taxa_radio') or 0)
    taxa_lixo = float(cfg.get('taxa_lixo') or 0)
    ponta_faturavel = _ponta_faturavel_edm(pot_contratada, float(q.get('kw_ponta_lida') or 0))
    valor_ativa = float(q.get('kwh_ativa') or 0) * tarifa_ativa
    valor_reativa = float(q.get('kvarh_excedente') or 0) * tarifa_reativa
    valor_ponta = ponta_faturavel * tarifa_ponta
    subtotal = valor_ativa + valor_reativa + valor_ponta + taxa_fixa + taxa_radio + taxa_lixo
    valor_iva = subtotal * 0.62 * 0.16
    total = subtotal + valor_iva
    return {
        'qfat': q,
        'ponta_faturavel': ponta_faturavel,
        'valor_ativa': valor_ativa,
        'valor_reativa': valor_reativa,
        'valor_ponta': valor_ponta,
        'subtotal': subtotal,
        'valor_iva': valor_iva,
        'total': total,
        'tarifa_ativa': tarifa_ativa,
        'tarifa_reativa': tarifa_reativa,
        'tarifa_ponta': tarifa_ponta,
    }


def _agg_dashboard(mes, ano, local_id=None):
    """
    Dashboard 7B: consolida leituras com a lógica da fatura mensal.
    Mantém a mesma assinatura usada pela rota /dashboard.
    """
    _ensure_idx_dashboard()
    cfg_map = _dashboard_get_cfg_map()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        params = []
        where = "WHERE COALESCE(l.ativo,1)=1"
        if local_id:
            where += " AND l.id=?"
            params.append(int(local_id))
        locais_rows = c.execute(f"SELECT id, nome FROM locais l {where} ORDER BY nome", params).fetchall()
        dias_mes = _dias_no_mes(mes, int(ano))
        cards = []
        energia_total = reativa_exc_total = fatura_total = ponta_faturavel_global = agua_total = 0.0
        locais_fp_baixo = locais_cobertura_baixa = locais_sem_base = 0
        fp_vals = []
        for lr in locais_rows:
            lid, lname = lr['id'], lr['nome']
            cfg = cfg_map.get(lname, {})
            fin = _dashboard_month_finance(lname, mes, ano, cfg)
            q = fin['qfat']
            kwh = float(q.get('kwh_ativa') or 0)
            rexc = float(q.get('kvarh_excedente') or 0)
            agua = float(q.get('agua_total') or 0)
            stat = c.execute("""
                SELECT COUNT(DISTINCT data) AS dias,
                       SUM(CASE WHEN fp IS NOT NULL AND fp>0 AND fp<0.80 THEN 1 ELSE 0 END) AS fp_baixo,
                       AVG(CASE WHEN fp IS NOT NULL AND fp>0 THEN fp ELSE NULL END) AS fp_medio,
                       MAX(CASE WHEN ponta IS NOT NULL THEN ponta ELSE 0 END) AS ponta_max
                FROM leituras_mensais
                WHERE local=? AND mes=? AND ano=?
            """, (lname, str(mes).zfill(2), int(ano))).fetchone()
            dias_com_dados = int((stat['dias'] if stat else 0) or 0)
            fp_baixo = int((stat['fp_baixo'] if stat else 0) or 0)
            fp_medio = float((stat['fp_medio'] if stat else 0) or 0)
            ponta_max = float(q.get('kw_ponta_lida') or (stat['ponta_max'] if stat else 0) or 0)
            if fp_medio:
                fp_vals.append(fp_medio)
            try:
                horas = c.execute("""
                    SELECT ROUND(COALESCE(SUM(r.duracao_min),0)/60.0,2)
                    FROM motor_runs r JOIN equipamentos e ON e.id=r.equipamento_id
                    WHERE e.local_id=? AND strftime('%m', r.start_time)=? AND strftime('%Y', r.start_time)=?
                """, (lid, str(mes).zfill(2), str(ano))).fetchone()[0] or 0
            except Exception:
                horas = 0
            cobertura_pct = round((dias_com_dados * 100.0 / dias_mes) if dias_mes else 0, 1)
            if cobertura_pct < 80:
                locais_cobertura_baixa += 1
            if fp_baixo > 0:
                locais_fp_baixo += 1
            if not q.get('tem_base_mes_anterior_ativa') and kwh > 0:
                locais_sem_base += 1
            status = 'Normal'
            if fp_baixo > 0 or rexc > 0 or cobertura_pct < 50:
                status = 'Crítico' if (fp_baixo > 3 or cobertura_pct < 40) else 'Atenção'
            cards.append({
                'local_id': lid,
                'local': lname,
                'energia_mes': round(kwh,2),
                'reativa_excedente': round(rexc,2),
                'ponta_max': round(ponta_max,2),
                'ponta_faturavel': round(fin['ponta_faturavel'],2),
                'fatura_estimativa': round(fin['total'],2),
                'valor_ativa': round(fin['valor_ativa'],2),
                'valor_reativa': round(fin['valor_reativa'],2),
                'valor_ponta': round(fin['valor_ponta'],2),
                'agua_total': round(agua,2),
                'consumo_especifico': q.get('consumo_especifico'),
                'fp_baixo': fp_baixo,
                'fp_medio': round(fp_medio,3) if fp_medio else None,
                'horas_motores': float(horas or 0),
                'dias_com_dados': dias_com_dados,
                'dias_mes': dias_mes,
                'cobertura_pct': cobertura_pct,
                'tarifa_kwh': fin['tarifa_ativa'],
                'custo_estimado': round(fin['total'],2),
                'status': status,
                'avisos': q.get('avisos', []),
            })
            energia_total += kwh
            reativa_exc_total += rexc
            fatura_total += fin['total']
            ponta_faturavel_global = max(ponta_faturavel_global, fin['ponta_faturavel'])
            agua_total += agua
        cards = sorted(cards, key=lambda x: x['energia_mes'], reverse=True)
        top = cards[:8]
        rank = {'labels':[x['local'] for x in top], 'data':[x['energia_mes'] for x in top]}
        params_daily = [str(mes).zfill(2), int(ano)]
        where_daily = ''
        if local_id:
            where_daily = ' AND l.id=?'
            params_daily.append(int(local_id))
        daily = c.execute(f"""
            SELECT m.data, ROUND(COALESCE(SUM(CASE WHEN m.diferenca>0 THEN m.diferenca ELSE 0 END),0),2) AS kwh
            FROM leituras_mensais m JOIN locais l ON l.nome=m.local
            WHERE m.mes=? AND m.ano=? {where_daily}
            GROUP BY m.data ORDER BY m.data
        """, params_daily).fetchall()
        trend = {'labels':[r['data'] for r in daily], 'data':[float(r['kwh'] or 0) for r in daily]}
        pm, py = _prev_month(str(mes).zfill(2), int(ano))
        prev_total = 0.0
        for lr in locais_rows:
            cfg = cfg_map.get(lr['nome'], {})
            try:
                prev_total += _dashboard_month_finance(lr['nome'], pm, py, cfg)['qfat'].get('kwh_ativa',0) or 0
            except Exception:
                pass
        def _pct_delta(curr, prev):
            try:
                curr=float(curr or 0); prev=float(prev or 0)
                if prev == 0:
                    return None
                return round(((curr-prev)/prev)*100, 1)
            except Exception:
                return None
        fp_medio_global = round(sum(fp_vals)/len(fp_vals),3) if fp_vals else None
        kpis = {
            'energia_total': round(energia_total,2),
            'reativa_excedente_total': round(reativa_exc_total,2),
            'ponta_max_global': round(ponta_faturavel_global,2),
            'ponta_faturavel_global': round(ponta_faturavel_global,2),
            'locais_fp_baixo': locais_fp_baixo,
            'horas_motores_total': round(sum(x['horas_motores'] for x in cards),2),
            'locais_cobertura_baixa': int(locais_cobertura_baixa),
            'custo_total': round(fatura_total,2),
            'fatura_total': round(fatura_total,2),
            'custos_habilitados': True,
            'delta_energia_pct': _pct_delta(energia_total, prev_total),
            'delta_custo_pct': None,
            'fp_medio_global': fp_medio_global,
            'agua_total': round(agua_total,2),
            'consumo_especifico_global': round(energia_total/agua_total,4) if agua_total else None,
            'locais_sem_base': locais_sem_base,
            'estado_geral': 'Crítico' if (locais_fp_baixo>0 or reativa_exc_total>0) else ('Atenção' if locais_cobertura_baixa>0 else 'Normal'),
        }
        return cards, kpis, rank, trend
    finally:
        conn.close()


@app.route('/dashboard/api')
def dashboard_api():
    hoje = datetime.now()
    mes = (request.args.get('mes') or hoje.strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or hoje.year)
    local_id = request.args.get('local_id')
    cards, kpis, rank, trend = _agg_dashboard(mes, ano, local_id=local_id)
    return jsonify({'mes': mes, 'ano': ano, 'local_id': local_id, 'kpis': kpis, 'cards': cards, 'rank': rank, 'trend': trend})


@app.route('/dashboard/relatorio')
def dashboard_relatorio():
    hoje = datetime.now()
    mes = (request.args.get('mes') or hoje.strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or hoje.year)
    local_id = request.args.get('local_id')
    cards, kpis, rank, trend = _agg_dashboard(mes, ano, local_id=local_id)
    locais = get_locais()
    return render_template('dashboard_relatorio.html', cards=cards, kpis=kpis, rank=rank, trend=trend, locais=locais, mes=mes, ano=ano, local_id=local_id)


# === Fase 7C: Dashboard final - score executivo, ações prioritárias e exportação estruturada ===
def _dashboard_score_e_acoes(cards, kpis):
    """Cria score executivo e plano de ações a partir dos dados já agregados do dashboard."""
    score = 100
    acoes = []
    def add(nivel, titulo, origem, impacto, acao, link, local=None, score_penalty=0):
        nonlocal score
        score -= score_penalty
        acoes.append({
            'nivel': nivel,
            'titulo': titulo,
            'origem': origem,
            'local': local or 'Geral',
            'impacto': impacto,
            'acao': acao,
            'link': link,
            'prioridade': {'Crítico': 100, 'Atenção': 70, 'Informativo': 35}.get(nivel, 50) + score_penalty
        })
    if (kpis.get('locais_sem_base') or 0) > 0:
        add('Atenção','Validar base do mês anterior','Leituras Mensais','Pode distorcer energia faturável e comparação mensal.','Confirmar última leitura válida do mês anterior e recalcular o período.','/leituras_mensal/arquivo', score_penalty=10)
    if (kpis.get('locais_cobertura_baixa') or 0) > 0:
        add('Atenção','Completar leituras em falta','Qualidade de Dados','A baixa cobertura reduz a confiança dos indicadores executivos.','Preencher dias em falta ou justificar ausência de dados.','/leituras_mensal/arquivo', score_penalty=12)
    if (kpis.get('reativa_excedente_total') or 0) > 0:
        add('Crítico','Reativa excedente no período','Faturação / FP','Aumenta o custo da fatura e indica baixo fator de potência.','Avaliar compensação reativa e regime de operação dos motores.','/alertas', score_penalty=15)
    if (kpis.get('locais_fp_baixo') or 0) > 0:
        add('Crítico','Locais com fator de potência baixo','Qualidade de Energia','Pode gerar penalização por reativa e perda de eficiência.','Abrir ação correctiva e analisar banco de capacitores/motores.','/alertas', score_penalty=15)
    delta = kpis.get('delta_energia_pct')
    if delta is not None and abs(float(delta or 0)) >= 20:
        add('Atenção','Variação anormal de consumo','Consumo Mensal','A energia mudou mais de 20% face ao mês anterior.','Comparar operação, bombas em serviço, caudais e leituras base.','/dashboard/relatorio', score_penalty=8)
    # ações por local
    for c in cards[:12]:
        loc = c.get('local')
        if (c.get('reativa_excedente') or 0) > 0 or (c.get('fp_baixo') or 0) > 0:
            add('Crítico','Corrigir FP / reativa excedente','Local crítico','Há impacto técnico e financeiro por baixo FP ou reativa excedente.','Inspecionar cargas indutivas, motores e compensação reativa.', c.get('link') or '/alertas', local=loc, score_penalty=5)
        if (c.get('cobertura_pct') or 0) < 80:
            add('Atenção','Melhorar cobertura de leituras','Dados do local','Dados incompletos reduzem a precisão da fatura e dos KPIs.', 'Completar planilha mensal e validar dados do operador.', c.get('link') or '/leituras_mensal', local=loc, score_penalty=4)
        if (c.get('status') or '').lower().startswith('cr'):
            add('Crítico','Local em estado crítico','Dashboard', 'O local concentra anomalias relevantes no período.', 'Abrir investigação técnica e definir responsável.', '/alertas', local=loc, score_penalty=5)
    if not acoes:
        add('Informativo','Manter rotina de acompanhamento','Gestão','Não foram detectados desvios críticos no período.','Manter preenchimento diário, validar faturas e monitorar tendências.','/monitoria', score_penalty=0)
    score = max(0, min(100, int(round(score))))
    acoes = sorted(acoes, key=lambda x: x.get('prioridade',0), reverse=True)
    return score, acoes[:30]

# Reforça a função de agregação existente sem alterar a assinatura usada pelas rotas anteriores.
_dashboard_agg_base_7c = _agg_dashboard
def _agg_dashboard(mes, ano, local_id=None):
    cards, kpis, rank, trend = _dashboard_agg_base_7c(mes, ano, local_id=local_id)
    score, acoes = _dashboard_score_e_acoes(cards, kpis)
    kpis['score_executivo'] = score
    kpis['acoes_prioritarias'] = acoes
    kpis['total_acoes_prioritarias'] = len(acoes)
    kpis['estado_geral'] = 'Crítico' if score < 60 else ('Atenção' if score < 82 else (kpis.get('estado_geral') or 'Normal'))
    return cards, kpis, rank, trend

@app.route('/dashboard/acoes')
def dashboard_acoes():
    hoje = datetime.now()
    mes = (request.args.get('mes') or hoje.strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or hoje.year)
    local_id = request.args.get('local_id')
    cards, kpis, rank, trend = _agg_dashboard(mes, ano, local_id=local_id)
    locais = get_locais()
    return render_template('dashboard_acoes.html', cards=cards, kpis=kpis, rank=rank, trend=trend, locais=locais, mes=mes, ano=ano, local_id=local_id, acoes=kpis.get('acoes_prioritarias', []))

@app.route('/dashboard/export.json')
def dashboard_export_json():
    hoje = datetime.now()
    mes = (request.args.get('mes') or hoje.strftime('%m')).zfill(2)
    ano = int(request.args.get('ano') or hoje.year)
    local_id = request.args.get('local_id')
    cards, kpis, rank, trend = _agg_dashboard(mes, ano, local_id=local_id)
    payload = {'periodo': {'mes': mes, 'ano': ano, 'local_id': local_id}, 'kpis': kpis, 'locais': cards, 'ranking': rank, 'tendencia': trend}
    return Response(json.dumps(payload, ensure_ascii=False, indent=2), mimetype='application/json', headers={'Content-Disposition': f'attachment; filename=dashboard_{ano}_{mes}.json'})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', '5000'))
    debug = os.environ.get('FLASK_DEBUG', '0').lower() in ('1','true','yes')
    app.run(host='0.0.0.0', port=port, debug=debug, use_reloader=debug)

