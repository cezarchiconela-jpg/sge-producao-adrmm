"""Leitura controlada de PIGI e planilhas operacionais normalizadas."""
from __future__ import annotations
import calendar, hashlib, io, re, tempfile, unicodedata
from datetime import date, datetime
from typing import Any

class OperationalImportError(ValueError): pass

ALIASES={
 "boane":"CD Boane","cd de boane":"CD Boane","matola rio":"CD Matola Rio",
 "matola":"CD Matola","matola zona alta":"CD Matola","machava":"CD Machava",
 "tsalala":"CD Tsalala","chamanculo":"CD Chamanculo","maxaquene":"CD Maxaquene",
 "alto mae":"CD Alto Maé","katembe":"CD Ka Tembe","ka tembe":"CD Ka Tembe",
 "laulane":"CD Laulane","intaka":"CD Intaka","vila olimpica":"CD Vila Olímpica",
 "mathemele":"CD Matlhemele","matlhemele":"CD Matlhemele","guava":"CD Guava",
 "hulene esquadra":"Hulene Esquadrea","hulene esquadrea":"Hulene Esquadrea",
 "3 de fevereiro":"3 de Fevereiro","eta umbeluzi":"ETA Umbeluzi",
 "eta umbeluze":"ETA Umbeluzi","eta corumana":"ETA Corumana",
}

def _norm(v):
    s=unicodedata.normalize("NFKD",str(v or "")).encode("ascii","ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+"," ",s).strip()

def _number(v):
    if v in (None,"") or isinstance(v,bool): return None
    try:
        n=float(str(v).replace(" ","").replace(",",".")); return n if n==n else None
    except (TypeError,ValueError): return None

def file_hash(content): return hashlib.sha256(content).hexdigest()

def normalise_site(value):
    raw=str(value or "").strip(); key=_norm(raw)
    if key in ALIASES: return ALIASES[key]
    key=re.sub(r"^(cds?|ao|delegacao de|delegacao da)\s+","",key).strip()
    return ALIASES.get(key,raw)

def match_local(conn, source_name):
    target=_norm(normalise_site(source_name)); compact=target.removeprefix("cd ").removeprefix("eta ")
    for row in conn.execute("SELECT id,nome FROM locais WHERE COALESCE(ativo,1)=1"):
        name=_norm(row[1])
        if name==target or name.removeprefix("cd ").removeprefix("eta ")==compact: return int(row[0])
    return None

def _rows(content,filename,sheet):
    if filename.lower().endswith(".xlsb"):
        try: from pyxlsb import open_workbook
        except ImportError as exc: raise OperationalImportError("Instale as dependências atualizadas para ler .xlsb.") from exc
        with tempfile.NamedTemporaryFile(suffix=".xlsb") as f:
            f.write(content); f.flush()
            with open_workbook(f.name) as wb:
                if sheet not in wb.sheets: return []
                with wb.get_sheet(sheet) as ws: return [[c.v for c in row] for row in ws.rows()]
    from openpyxl import load_workbook
    wb=load_workbook(io.BytesIO(content),read_only=True,data_only=True)
    if sheet not in wb.sheetnames: return []
    return [list(r) for r in wb[sheet].iter_rows(values_only=True)]

def _period(filename):
    months={"janeiro":1,"fevereiro":2,"marco":3,"abril":4,"maio":5,"junho":6,"julho":7,"agosto":8,"setembro":9,"outubro":10,"novembro":11,"dezembro":12}
    text=_norm(filename); year=re.search(r"\b(20\d{2})\b",text)
    for name,number in months.items():
        if name in text and year: return int(year.group(1)),number
    return None

def parse_pigi(content,filename):
    period=_period(filename)
    if not period: raise OperationalImportError("Inclua mês e ano no nome do ficheiro PIGI.")
    year,month=period; days=calendar.monthrange(year,month)[1]; rows=_rows(content,filename,"Energia kw")
    if not rows: raise OperationalImportError("A folha 'Energia kw' não foi encontrada.")
    # A tabela com coeficiente é a fonte energética bruta. As linhas de
    # "Específico" podem conter fórmulas antigas e não são importadas.
    coefficient_started=False; energy_lines={}
    for row in rows:
        heading=_norm(" ".join(str(x) for x in row[:3] if x is not None))
        if "consumos de energia com coeficiente" in heading: coefficient_started=True; continue
        if coefficient_started and "consumo especifico de energia" in heading: break
        if not coefficient_started or len(row)<3: continue
        label=_norm(row[1] if len(row)>1 else "")
        if not label or label.startswith("total ") or label in ("consumo de energia","consumo por bombagem"): continue
        values=[_number(row[d+2] if d+2<len(row) else None) for d in range(days)]
        if any(v is not None for v in values): energy_lines[label]=values
    source_labels={
      "CD Boane":["boane"],"CD Matola Rio":["matola rio"],"CD Matola":["matola zona alta"],
      "CD Machava":["machava"],"transferencia para tsalala":["machava para tsalala"],
      "CD Tsalala":["tsalala"],"transferencia para coca cola de matola gare":["tsalala para coca cola"],
      "CD Chamanculo":["chamanculo"],"CD Maxaquene":["maxaquene"],
      "CD Alto Maé":["alto mae"],"CD Ka Tembe":["katembe"],"CD Laulane":["laulane"],
      "3 de Fevereiro":["3 de fevereiro"],"Hulene Esquadrea":["hulene esquadra"],
      "hulene mercado":["hulene mercado"],"CD Intaka":["intaka"],
      "CD Vila Olímpica":["vila olimpica"],"CD Matlhemele":["mathemele","matlhemele"],"CD Guava":["guava"],
    }
    volume_rows=_rows(content,filename,"Volumes"); volume_lines={}; any_volume_lines={}; in_summary=False; section=""
    for row in volume_rows:
        heading=_norm(" ".join(str(x) for x in row[:3] if x is not None))
        if "tabela resumo dos volumes" in heading: in_summary=True; continue
        if len(row)<3: continue
        a=_norm(row[0] if len(row)>0 else ""); b=_norm(row[1] if len(row)>1 else "")
        if a and not b: section=a
        if b:
            values=[_number(row[d+2] if d+2<len(row) else None) for d in range(days)]
            if any(v is not None for v in values):
                any_volume_lines[b]=values
                if in_summary: volume_lines[(section,b)]=values
    volume_labels={
      "CD Boane":[("boane","cd")],"CD Matola Rio":[("matola rio","cd m rio dn300")],
      "CD Matola":[("matola","cd matola za")],"CD Machava":[("machava","cd machava")],
      "CD Tsalala":[("tsalala","cd tsalala")],"transferencia para coca cola de matola gare":[("tsalala","coca cola m gare")],
      "CD Maxaquene":[("maxaquene","cd maxaquene za")],"CD Alto Maé":[("chamanculo","cd a mae za")],
      "CD Ka Tembe":[("maxaquene","katembe torre dn160"),("maxaquene","katembe cd dn90")],
      "CD Laulane":[("laulane","cd laulane za")],"3 de Fevereiro":[("laulane","3 de fevereiro")],
      "Hulene Esquadrea":[("laulane","hulene esquadra")],"hulene mercado":[("laulane","hulene mercado")],
      "CD Intaka":[("", "volume total do intaka")],
      "CD Matlhemele":[("tsalala","cd malhemele"),("tsalala","cd matlhemele")],"CD Guava":[("laulane","cd guava")],
    }
    records=[]; current=None
    for i,row in enumerate(rows):
        label=str(row[1] if len(row)>1 and row[1] is not None else "").strip()
        heading=_norm(" ".join(str(x) for x in row[:3] if x is not None))
        found=re.search(r"consumo de energia kw m3 (.+)",heading)
        if found: current=normalise_site(found.group(1)); continue
        if current and _norm(label)=="volume":
            erow=rows[i+1] if i+1<len(rows) else []
            if _norm(erow[1] if len(erow)>1 else "")!="energia": current=None; continue
            block_series=[_number(erow[d+2] if d+2<len(erow) else None) for d in range(days)]
            coefficient_series=[]
            for day_index in range(days):
                values=[energy_lines[x][day_index] for x in source_labels.get(current,[]) if x in energy_lines and energy_lines[x][day_index] is not None]
                coefficient_series.append(sum(values) if values else None)
            block_total=sum(x or 0 for x in block_series); coefficient_total=sum(x or 0 for x in coefficient_series)
            source_warning=None
            if block_total>0 and coefficient_total>0 and abs(block_total-coefficient_total)/block_total>0.02:
                source_warning=f"divergência interna PIGI: tabela de energia {coefficient_total:.2f} vs bloco específico {block_total:.2f} kWh"
            for d in range(days):
                col=d+2; volume=_number(row[col] if col<len(row) else None); energy=_number(erow[col] if col<len(erow) else None)
                labels=source_labels.get(current,[]); source_values=[]
                for source_label in labels:
                    if source_label in energy_lines and energy_lines[source_label][d] is not None:
                        source_values.append(energy_lines[source_label][d])
                if energy is None and source_values: energy=source_values[0]
                volume_values=[]
                for source_key in volume_labels.get(current,[]):
                    exact=any_volume_lines.get(source_key[1]) if not source_key[0] else volume_lines.get(source_key)
                    if exact and exact[d] is not None: volume_values.append(exact[d]); continue
                    # tolera pequenas alterações DN90/DN110 e acentuação nos títulos.
                    for (sec,label_key),vals in volume_lines.items():
                        if (not source_key[0] or sec==source_key[0]) and source_key[1] in label_key and vals[d] is not None:
                            volume_values.append(vals[d]); break
                if volume_values: volume=sum(volume_values)
                warnings=[]
                if source_warning: warnings.append(source_warning)
                if volume is None or volume<=0: warnings.append("volume ausente ou zero")
                if energy is None or energy<0: warnings.append("energia ausente ou inválida")
                records.append({"sheet":"Energia kw","source_row":i+1,"site":current,"date":date(year,month,d+1).isoformat(),"energy_kwh":energy,"volume_m3":volume,"hours":None,"data_type":"medido","quality":"provisoria" if warnings else "validavel","warnings":warnings})
            current=None
    occurrences=[]
    for i,row in enumerate(_rows(content,filename,"Ocorrências")):
        when=row[1] if len(row)>1 else None; desc=str(row[2] if len(row)>2 and row[2] is not None else "").strip()
        if not desc or "ocorrencias do sistema" in _norm(desc): continue
        if isinstance(when,datetime): when=when.date().isoformat()
        elif isinstance(when,date): when=when.isoformat()
        else: when=date(year,month,min(i+1,days)).isoformat()
        occurrences.append({"date":str(when)[:10],"description":desc,"source_row":i+1})
    if not records: raise OperationalImportError("Nenhum par diário de energia e volume foi encontrado no PIGI.")
    return {"format":"PIGI","period":f"{year:04d}-{month:02d}","records":records,"occurrences":occurrences}

HEADERS={"local":"site","instalacao":"site","data":"date","energia kwh":"energy_kwh","consumo kwh":"energy_kwh","leitura energia kwh":"meter_reading_kwh","agua m3":"volume_m3","volume m3":"volume_m3","agua elevada m3":"volume_m3","volume produzido m3":"volume_m3","horas":"hours","horas operacao":"hours","tipo dado":"data_type","observacoes":"notes"}

def parse_standard(content,filename):
    rows=_rows(content,filename,"Dados") or _rows(content,filename,"Leituras")
    if not rows: raise OperationalImportError("Use uma folha chamada 'Dados' ou 'Leituras'.")
    hi=next((i for i,r in enumerate(rows[:20]) if any(_norm(x) in HEADERS for x in r)),None)
    if hi is None: raise OperationalImportError("Cabeçalhos reconhecíveis não encontrados.")
    headers=[HEADERS.get(_norm(x),_norm(x)) for x in rows[hi]]; records=[]; previous={}
    for i,row in enumerate(rows[hi+1:],hi+2):
        item={headers[j]:v for j,v in enumerate(row) if j<len(headers) and headers[j]}
        if not item.get("site") or not item.get("date"): continue
        when=item["date"]
        if isinstance(when,datetime): when=when.date().isoformat()
        elif isinstance(when,date): when=when.isoformat()
        else:
            try: when=datetime.fromisoformat(str(when).strip()).date().isoformat()
            except ValueError: continue
        site=normalise_site(item["site"]); energy=_number(item.get("energy_kwh")); meter=_number(item.get("meter_reading_kwh")); warnings=[]
        if energy is None and meter is not None:
            energy=meter-previous[site] if site in previous else None; previous[site]=meter
            if energy is None: warnings.append("primeira leitura acumulada usada como referência")
            elif energy<0: warnings.append("leitura acumulada regressiva"); energy=None
        volume=_number(item.get("volume_m3"))
        if energy is None and volume is None: continue
        records.append({"sheet":"Dados","source_row":i,"site":site,"date":when,"energy_kwh":energy,"volume_m3":volume,"hours":_number(item.get("hours")),"data_type":_norm(item.get("data_type")) or "medido","quality":"provisoria" if warnings else "validavel","warnings":warnings,"notes":str(item.get("notes") or "")[:500]})
    if not records: raise OperationalImportError("Nenhuma linha válida foi encontrada.")
    periods=sorted({x["date"][:7] for x in records})
    return {"format":"EDM/OPERACIONAL","period":periods[0] if len(periods)==1 else None,"records":records,"occurrences":[]}

def parse_workbook(content,filename,kind="auto"):
    if not filename.lower().endswith((".xlsx",".xlsb")): raise OperationalImportError("Utilize .xlsx ou .xlsb.")
    return parse_pigi(content,filename) if kind=="pigi" or (kind=="auto" and "pigi" in _norm(filename)) else parse_standard(content,filename)
