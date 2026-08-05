"""Serviço canónico do cadastro mestre de locais e activos do SGE.

O módulo não depende de Flask. Faz leitura, normalização, pré-visualização,
reconciliação e importação transaccional de cadastros Excel/CSV, preservando
os dados manuais e os identificadores já existentes no SGE.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import sqlite3
import unicodedata
from collections import Counter, defaultdict, deque
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable


BUNDLED_REGISTRY_MARKER = 2026080503
BUNDLED_REGISTRY_FILENAME = "activos_dima_todos_20260804_154738.xlsx"
MAX_REGISTRY_ROWS = 20_000


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def fold_text(value: Any) -> str:
    text = unicodedata.normalize("NFD", clean_text(value))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def header_key(value: Any) -> str:
    return fold_text(value).lower().replace(" ", "_")


HEADER_ALIASES = {
    "n": "source_record_no", "no": "source_record_no", "numero": "source_record_no",
    "id": "source_record_no", "n_registo": "source_record_no",
    "equipamento": "nome", "nome": "nome", "nome_equipamento": "nome", "activo": "nome", "ativo": "nome",
    "sistema": "sistema", "processo": "sistema",
    "instalacao": "instalacao", "subinstalacao": "instalacao", "area": "instalacao", "sub_area": "instalacao",
    "estado": "estado_operacional", "estado_operacional": "estado_operacional", "condicao": "estado_operacional",
    "criticidade": "criticidade_fonte", "criticalidade": "criticidade_fonte",
    "sector": "sector_operacional", "setor": "sector_operacional", "sector_operacional": "sector_operacional",
    "marca": "fabricante", "fabricante": "fabricante", "marca_fabricante": "fabricante",
    "modelo": "modelo", "modelo_especificacao": "modelo",
    "periodicidade": "periodicidade_manutencao", "periodicidade_manutencao": "periodicidade_manutencao",
    "frequencia_manutencao": "periodicidade_manutencao",
    "local": "local", "local_principal": "local",
    "codigo_activo": "referencia_externa", "codigo_ativo": "referencia_externa",
    "referencia_externa": "referencia_externa", "asset_code": "referencia_externa",
    "codigo_do_activo": "referencia_externa", "codigo_do_ativo": "referencia_externa",
    "tag": "tag", "etiqueta": "tag",
    "especificacao": "especificacao", "descricao_tecnica": "especificacao",
    "ano": "ano_instalacao", "ano_instalacao": "ano_instalacao",
    "quantidade": "quantidade", "qtd": "quantidade",
    "categoria": "categoria", "tipo_equipamento": "categoria",
    "numero_serie": "numero_serie", "n_serie": "numero_serie", "serie": "numero_serie",
    "custo": "custo_aquisicao", "custo_aquisicao": "custo_aquisicao",
    "vida_util": "vida_util_anos", "vida_util_anos": "vida_util_anos",
    "activo_sge": "ativo", "ativo_sge": "ativo", "activo_no_sge": "ativo", "ativo_no_sge": "ativo",
    "n_na_fonte": "source_record_no", "no_na_fonte": "source_record_no",
    "potencia_kw": "potencia_kw", "potencia": "potencia_kw",
    "tensao_v": "tensao_v", "tensao": "tensao_v",
    "corrente_a": "corrente_a", "corrente": "corrente_a",
    "fornecedor": "fornecedor", "contrato": "contrato_num", "contrato_num": "contrato_num",
    "garantia_fim": "garantia_fim", "observacoes": "observacoes",
}


def _canonical_header(value: Any) -> str:
    key = header_key(value)
    return HEADER_ALIASES.get(key, key)


def _has_supported_header(mapped: Iterable[str]) -> bool:
    keys = set(mapped)
    return "nome" in keys and ("local" in keys or {"sector_operacional", "instalacao"}.issubset(keys))


def _source_bytes(file_or_path: Any) -> bytes:
    if isinstance(file_or_path, (str, os.PathLike, Path)):
        return Path(file_or_path).read_bytes()
    if hasattr(file_or_path, "read"):
        data = file_or_path.read()
        try:
            file_or_path.seek(0)
        except Exception:
            pass
        return bytes(data)
    raise TypeError("Fonte de cadastro não suportada.")


def _xlsx_rows(data: bytes) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    sheets_used: list[str] = []
    warnings: list[str] = []
    for worksheet in workbook.worksheets:
        header_row = None
        raw_headers: list[str] = []
        mapped_headers: list[str] = []
        probe = list(worksheet.iter_rows(min_row=1, max_row=min(30, worksheet.max_row), values_only=True))
        for row_number, values in enumerate(probe, start=1):
            raw = [clean_text(value) for value in values]
            mapped = [_canonical_header(value) for value in raw]
            if _has_supported_header(mapped):
                header_row = row_number
                raw_headers = raw
                mapped_headers = mapped
                break
        if header_row is None:
            if any(any(clean_text(value) for value in row) for row in probe):
                warnings.append(f'Folha "{worksheet.title}" ignorada: cabeçalho de cadastro não reconhecido.')
            continue
        sheets_used.append(worksheet.title)
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            obj: dict[str, Any] = {"_sheet": worksheet.title, "_row": row_number}
            for index, value in enumerate(values):
                if index >= len(mapped_headers):
                    continue
                key = mapped_headers[index]
                if key and key not in obj:
                    obj[key] = value
            if not clean_text(obj.get("nome")):
                continue
            rows.append(obj)
            if len(rows) > MAX_REGISTRY_ROWS:
                raise ValueError(f"O cadastro excede o limite de {MAX_REGISTRY_ROWS:,} activos.")
    workbook.close()
    if not rows:
        raise ValueError("Nenhuma folha com um cadastro de equipamentos válido foi encontrada.")
    return rows, sheets_used, warnings


def _csv_rows(data: bytes) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:10_000]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t,").delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    matrix = list(reader)
    header_index = None
    mapped_headers: list[str] = []
    for index, values in enumerate(matrix[:30]):
        mapped = [_canonical_header(value) for value in values]
        if _has_supported_header(mapped):
            header_index = index
            mapped_headers = mapped
            break
    if header_index is None:
        raise ValueError("Cabeçalho do cadastro CSV não reconhecido.")
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(matrix[header_index + 1 :], start=header_index + 2):
        obj: dict[str, Any] = {"_sheet": "CSV", "_row": row_number}
        for index, value in enumerate(values):
            if index < len(mapped_headers) and mapped_headers[index] and mapped_headers[index] not in obj:
                obj[mapped_headers[index]] = value
        if clean_text(obj.get("nome")):
            rows.append(obj)
        if len(rows) > MAX_REGISTRY_ROWS:
            raise ValueError(f"O cadastro excede o limite de {MAX_REGISTRY_ROWS:,} activos.")
    return rows, ["CSV"], []


def _to_number(value: Any) -> float | None:
    text = clean_text(value).replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        number = float(text)
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def _to_bool(value: Any) -> int | None:
    folded = fold_text(value)
    if folded in {"1", "SIM", "S", "YES", "TRUE", "ACTIVO", "ATIVO"}:
        return 1
    if folded in {"0", "NAO", "N", "NO", "FALSE", "INACTIVO", "INATIVO"}:
        return 0
    return None


def _extract_number(text: str, unit_pattern: str) -> float | None:
    match = re.search(rf"(?<![A-Z0-9])([0-9]+(?:[.,][0-9]+)?)\s*{unit_pattern}\b", text, re.I)
    return _to_number(match.group(1)) if match else None


def extract_technical_values(model: str, specification: str = "") -> dict[str, float | None]:
    text = f"{clean_text(model)} {clean_text(specification)}"
    return {
        "potencia_kw": _extract_number(text, r"kW"),
        "tensao_v": _extract_number(text, r"V"),
        "corrente_a": _extract_number(text, r"A"),
    }


def normalize_state(value: Any) -> str:
    folded = fold_text(value)
    if folded in {"BOM", "OPERACIONAL", "EM SERVICO"}:
        return "Operacional"
    if "AVARI" in folded:
        return "Avariado"
    if folded in {"FORA DE SERVICO", "INOPERACIONAL", "DESACTIVADO", "DESATIVADO"}:
        return "Fora de serviço"
    if folded in {"", ".", "0", "NA", "N A"}:
        return "Não informado"
    return clean_text(value)


def normalize_criticality(value: Any) -> str:
    folded = fold_text(value)
    if folded in {"CRITICO", "ALTA", "MUITO ALTA"}:
        return "Alta"
    if folded in {"NORMAL", "MEDIA", "MEDIO"}:
        return "Média"
    if folded in {"REDUNDANCIA", "BAIXA"}:
        return "Baixa"
    return ""


def normalize_periodicity(value: Any) -> str:
    folded = fold_text(value)
    if folded in {"TRIMESTRAL", "TRIMENSTRAL", "TRIMETRAL"}:
        return "Trimestral"
    if folded == "SEMESTRAL":
        return "Semestral"
    if folded == "MENSAL":
        return "Mensal"
    if folded == "ANUAL":
        return "Anual"
    return clean_text(value)


def infer_equipment_category(name: str) -> str:
    folded = fold_text(name)
    rules = (
        ("Motor eléctrico", ("MOTOR",)),
        ("Bomba", ("BOMBA", "ELECTROBOMBA", "ELETROBOMBA")),
        ("Válvula", ("VALVULA",)),
        ("Quadro eléctrico", ("QUADRO", "PAINEL ELECTRICO", "PAINEL ELETRICO", "ARMARIO ELECTRICO")),
        ("Transformador", ("TRANSFORMADOR", "TRAFO")),
        ("Comporta", ("COMPORTA",)),
        ("Instrumentação", ("MEDIDOR", "TRANSMISSOR", "SENSOR", "MANOMETRO", "CAUDALIMETRO", "NIVEL")),
        ("Protecção eléctrica", ("DISJUNTOR", "RELE", "FUSIVEL", "PARA RAIO", "SECCIONADOR")),
        ("Iluminação", ("LUMINARIA", "PROJECTOR", "PROJETOR", "LAMPADA")),
        ("Compressores e sopradores", ("COMPRESSOR", "SOPRADOR", "SUPRESSOR")),
        ("Gerador", ("GERADOR", "GRUPO GERADOR")),
        ("Equipamento de elevação", ("PONTE ROLANTE", "GUINDASTE", "TALHA")),
        ("Canalização", ("TUBAGEM", "CANALIZACAO", "JUNTA", "VENTOSA")),
        ("Infraestrutura", ("PORTA", "JANELA", "GRADE", "ESCADA", "VEDACAO")),
    )
    for category, keys in rules:
        if any(key in folded for key in keys):
            return category
    return "Outro equipamento"


SMALL_SYSTEMS = {
    "3 DE FEVEREIRO", "HULENE", "JONASSE", "MUSSUMBULUKU", "RADIO MARCONI",
}


def infer_local_name(row: dict[str, Any]) -> str:
    explicit = clean_text(row.get("local"))
    if explicit:
        return explicit
    sector = fold_text(row.get("sector_operacional"))
    installation = clean_text(row.get("instalacao"))
    installation_folded = fold_text(installation)
    if sector == "UMBELUZI":
        return "ETA Umbeluzi"
    if sector == "SABIE":
        return "ETA Sabié"
    if sector in {"CDS", "CD"}:
        return installation or "Centros de Distribuição"
    if sector in {"ADUCAO", "ADUCAO E TRANSFERENCIA"}:
        if "UMBELUZI" in installation_folded:
            return "ETA Umbeluzi"
        if "SABIE" in installation_folded:
            return "ETA Sabié"
        return "Rede de Adução e Transferência"
    return installation or clean_text(row.get("sector_operacional")) or "Local por classificar"


def infer_local_profile(name: str, sector: str = "") -> dict[str, Any]:
    folded = fold_text(name)
    if "UMBELUZI" in folded and "ETA" in folded:
        return {"codigo": "ETA-UMB", "tipo_local": "ETA", "categoria_operacional": "Produção e tratamento", "prioridade": "Alta"}
    if "SABIE" in folded and "ETA" in folded:
        return {"codigo": "ETA-SAB", "tipo_local": "ETA", "categoria_operacional": "Produção e tratamento", "prioridade": "Alta"}
    if folded == "REDE DE ADUCAO E TRANSFERENCIA":
        return {"codigo": "ADU-ASM", "tipo_local": "Sistema de Adução", "categoria_operacional": "Adução e transferência", "prioridade": "Alta"}
    if "CAMPO DE FURO" in folded:
        local_type = "Campo de Furos"
        category = "Captação subterrânea"
    elif "TRANSFERENCIA" in folded or folded.startswith("EB ") or "ESTACAO DE BOMBAGEM" in folded:
        local_type = "Estação de Bombagem"
        category = "Bombagem e transferência"
    elif folded in SMALL_SYSTEMS:
        local_type = "Pequeno Sistema"
        category = "Abastecimento local"
    else:
        local_type = "Centro de Distribuição"
        category = "Distribuição de água"
    code_seed = re.sub(r"[^A-Z0-9]", "", folded)[:12] or "LOCAL"
    return {"codigo": f"LOC-{code_seed}", "tipo_local": local_type, "categoria_operacional": category, "prioridade": "Média"}


LOCAL_ALIASES = {
    "ETA UMBELUZI": {"ETA DE UMBELUZI", "ETA UMBELUZI"},
    "ETA SABIE": {"ETA DE SABIE", "ETA SABIE"},
    "CHAMANCULO": {"CHAMANCULO", "CD CHAMANCULO", "CD DE CHAMANCULO"},
    "VILA OLIMPICA": {"VILA OLIMPICA", "VILA OLIMPICA CD"},
    "RADIO MARCONI": {"RADIO MARCONI", "RADIO MARCONI CD"},
}


def _local_aliases(name: str) -> set[str]:
    folded = fold_text(name)
    aliases = set(LOCAL_ALIASES.get(folded, {folded}))
    aliases.add(folded)
    return aliases


def _float_or_source(value: Any, fallback: float | None) -> float | None:
    parsed = _to_number(value)
    return parsed if parsed is not None else fallback


def prepare_registry_rows(raw_rows: list[dict[str, Any]], source_name: str, source_hash: str) -> list[dict[str, Any]]:
    occurrence: Counter[str] = Counter()
    prepared: list[dict[str, Any]] = []
    for raw in raw_rows:
        nome = clean_text(raw.get("nome"))
        if not nome:
            continue
        sector = clean_text(raw.get("sector_operacional"))
        installation = clean_text(raw.get("instalacao"))
        system = clean_text(raw.get("sistema"))
        local_name = infer_local_name(raw)
        base_identity = "|".join(map(fold_text, (local_name, sector, installation, system, nome)))
        occurrence[base_identity] += 1
        external = clean_text(raw.get("referencia_externa"))
        if external:
            source_key = external
        else:
            digest = hashlib.sha256(base_identity.encode("utf-8")).hexdigest()[:20]
            source_key = f"DIMA-{digest}-{occurrence[base_identity]:03d}"
        model = clean_text(raw.get("modelo"))
        specification = clean_text(raw.get("especificacao"))
        technical = extract_technical_values(model, specification)
        category = clean_text(raw.get("categoria")) or infer_equipment_category(nome)
        state = normalize_state(raw.get("estado_operacional"))
        criticality = normalize_criticality(raw.get("criticidade_fonte"))
        periodicity = normalize_periodicity(raw.get("periodicidade_manutencao"))
        row = {
            "source_key": source_key,
            "source_record_no": clean_text(raw.get("source_record_no")) or str(raw.get("_row") or ""),
            "source_sheet": clean_text(raw.get("_sheet")),
            "source_name": source_name,
            "source_file_hash": source_hash,
            "local_name": local_name,
            "nome": nome,
            "sector_operacional": sector,
            "instalacao": installation,
            "sistema": system,
            "estado_operacional": state,
            "criticidade": criticality,
            "criticidade_fonte": clean_text(raw.get("criticidade_fonte")),
            "fabricante": clean_text(raw.get("fabricante")),
            "modelo": model,
            "periodicidade_manutencao": periodicity,
            "tag": clean_text(raw.get("tag")),
            "especificacao": specification,
            "ano_instalacao": clean_text(raw.get("ano_instalacao")),
            "quantidade": max(1, int(_to_number(raw.get("quantidade")) or 1)),
            "ativo": _to_bool(raw.get("ativo")),
            "categoria": category,
            "numero_serie": clean_text(raw.get("numero_serie")),
            "custo_aquisicao": _to_number(raw.get("custo_aquisicao")),
            "vida_util_anos": int(_to_number(raw.get("vida_util_anos")) or 0) or None,
            "potencia_kw": _float_or_source(raw.get("potencia_kw"), technical["potencia_kw"]),
            "tensao_v": _float_or_source(raw.get("tensao_v"), technical["tensao_v"]),
            "corrente_a": _float_or_source(raw.get("corrente_a"), technical["corrente_a"]),
            "fornecedor": clean_text(raw.get("fornecedor")),
            "contrato_num": clean_text(raw.get("contrato_num")),
            "garantia_fim": clean_text(raw.get("garantia_fim")),
            "observacoes": clean_text(raw.get("observacoes")),
        }
        provenance = {
            "source_hash", "source_name", "source_file_hash", "source_sheet",
            "source_record_no", "source_key", "criticidade_fonte",
        }
        managed = {key: value for key, value in row.items() if key not in provenance}
        row["source_hash"] = hashlib.sha256(
            json.dumps(managed, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()
        prepared.append(row)
    return prepared


def parse_registry_file(file_or_path: Any, filename: str | None = None) -> dict[str, Any]:
    data = _source_bytes(file_or_path)
    name = filename or (Path(file_or_path).name if isinstance(file_or_path, (str, os.PathLike, Path)) else "cadastro.xlsx")
    extension = Path(name).suffix.lower()
    if extension == ".xlsx":
        raw_rows, sheets, warnings = _xlsx_rows(data)
    elif extension in {".csv", ".txt"}:
        raw_rows, sheets, warnings = _csv_rows(data)
    elif extension == ".xls":
        raise ValueError("O formato .xls antigo não é aceite. Abra o ficheiro e guarde-o como .xlsx.")
    else:
        raise ValueError("Formato não suportado. Utilize .xlsx ou .csv.")
    digest = hashlib.sha256(data).hexdigest()
    rows = prepare_registry_rows(raw_rows, clean_text(name), digest)
    if not rows:
        raise ValueError("O ficheiro não contém equipamentos válidos.")
    return {
        "source_name": clean_text(name),
        "source_hash": digest,
        "sheets": sheets,
        "warnings": warnings,
        "rows": rows,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }


def _existing_location_map(conn: sqlite3.Connection) -> tuple[dict[str, int], dict[int, str]]:
    rows = conn.execute("SELECT id, nome FROM locais").fetchall()
    by_alias: dict[str, int] = {}
    names: dict[int, str] = {}
    for local_id, name in rows:
        names[int(local_id)] = clean_text(name)
        for alias in _local_aliases(clean_text(name)):
            by_alias.setdefault(alias, int(local_id))
    return by_alias, names


def _find_local_id(by_alias: dict[str, int], name: str) -> int | None:
    for alias in _local_aliases(name):
        if alias in by_alias:
            return by_alias[alias]
    return None


def preview_registry(db_path: str, parsed: dict[str, Any]) -> dict[str, Any]:
    rows = parsed.get("rows") or []
    conn = sqlite3.connect(db_path)
    try:
        location_aliases, _ = _existing_location_map(conn)
        existing_refs = {
            row[1]: (int(row[0]), clean_text(row[2]))
            for row in conn.execute(
                "SELECT id, referencia_externa, source_hash FROM equipamentos WHERE COALESCE(referencia_externa,'')<>''"
            ).fetchall()
        }
        legacy = Counter()
        for local_id, name in conn.execute(
            "SELECT local_id, nome FROM equipamentos WHERE COALESCE(referencia_externa,'')='' AND COALESCE(deleted_at,'')=''"
        ).fetchall():
            if local_id is not None:
                legacy[(int(local_id), fold_text(name))] += 1
        legacy_used: Counter[tuple[int, str]] = Counter()
        actions = Counter()
        locations_new: set[str] = set()
        location_counts = Counter()
        sector_counts = Counter()
        system_counts = Counter()
        quality = Counter()
        for row in rows:
            local_name = row["local_name"]
            location_counts[local_name] += 1
            sector_counts[row["sector_operacional"] or "Não informado"] += 1
            system_counts[row["sistema"] or "Não informado"] += 1
            if not row["fabricante"]:
                quality["sem_fabricante"] += 1
            if not row["modelo"]:
                quality["sem_modelo"] += 1
            if row["estado_operacional"] == "Não informado":
                quality["sem_estado"] += 1
            if not row["criticidade"]:
                quality["sem_criticidade"] += 1
            reference = existing_refs.get(row["source_key"])
            if reference:
                actions["sem_alteracao" if reference[1] == row["source_hash"] else "actualizar"] += 1
                continue
            local_id = _find_local_id(location_aliases, local_name)
            if local_id is None:
                locations_new.add(local_name)
                actions["inserir"] += 1
                continue
            legacy_key = (local_id, fold_text(row["nome"]))
            if legacy_used[legacy_key] < legacy[legacy_key]:
                legacy_used[legacy_key] += 1
                actions["reconciliar"] += 1
            else:
                actions["inserir"] += 1
        duplicate_keys = len(rows) - len({row["source_key"] for row in rows})
        return {
            "total": len(rows),
            "actions": dict(actions),
            "new_locations": sorted(locations_new, key=fold_text),
            "new_locations_count": len(locations_new),
            "locations_count": len(location_counts),
            "duplicate_keys": duplicate_keys,
            "quality": dict(quality),
            "locations": location_counts.most_common(),
            "sectors": sector_counts.most_common(),
            "systems_top": system_counts.most_common(15),
            "warnings": list(parsed.get("warnings") or []),
            "source_name": parsed.get("source_name"),
            "source_hash": parsed.get("source_hash"),
            "sheets": parsed.get("sheets") or [],
        }
    finally:
        conn.close()


def _upsert_location(
    conn: sqlite3.Connection,
    by_alias: dict[str, int],
    names: dict[int, str],
    local_name: str,
    sector: str,
    actor: str,
) -> tuple[int, bool]:
    local_id = _find_local_id(by_alias, local_name)
    profile = infer_local_profile(local_name, sector)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if local_id is not None:
        conn.execute(
            """
            UPDATE locais SET
                codigo=COALESCE(NULLIF(TRIM(codigo),''), ?),
                tipo_local=COALESCE(NULLIF(TRIM(tipo_local),''), ?),
                categoria_operacional=COALESCE(NULLIF(TRIM(categoria_operacional),''), ?),
                prioridade=COALESCE(NULLIF(TRIM(prioridade),''), ?),
                sector_operacional=CASE WHEN ?<>'' THEN ? ELSE sector_operacional END,
                fonte_cadastro='Cadastro DIMA 2026', referencia_externa=COALESCE(NULLIF(referencia_externa,''), ?),
                ultima_sincronizacao=?, classificacao_confirmada=COALESCE(classificacao_confirmada,0)
            WHERE id=?
            """,
            (
                profile["codigo"], profile["tipo_local"], profile["categoria_operacional"], profile["prioridade"],
                sector, sector, f"LOCAL-{hashlib.sha256(fold_text(local_name).encode()).hexdigest()[:16]}", now, local_id,
            ),
        )
        return local_id, False
    cursor = conn.execute(
        """
        INSERT INTO locais(
            nome, codigo, ativo, tipo_local, categoria_operacional, estado_tecnico, prioridade,
            sector_operacional, fonte_cadastro, referencia_externa, ultima_sincronizacao,
            classificacao_confirmada
        ) VALUES(?,?,1,?,?, 'Normal', ?, ?, 'Cadastro DIMA 2026', ?, ?, 0)
        """,
        (
            local_name, profile["codigo"], profile["tipo_local"], profile["categoria_operacional"],
            profile["prioridade"], sector,
            f"LOCAL-{hashlib.sha256(fold_text(local_name).encode()).hexdigest()[:16]}", now,
        ),
    )
    local_id = int(cursor.lastrowid)
    conn.execute("INSERT OR IGNORE INTO locais_cfg(local_id) VALUES(?)", (local_id,))
    conn.execute(
        "INSERT INTO locais_history(local_id, evento, detalhe, actor) VALUES(?,?,?,?)",
        (local_id, "Local criado por cadastro mestre", f"Origem: {sector or 'não informada'}", actor),
    )
    names[local_id] = local_name
    for alias in _local_aliases(local_name):
        by_alias[alias] = local_id
    return local_id, True


def _existing_equipment_state(conn: sqlite3.Connection) -> tuple[dict[str, dict[str, Any]], dict[tuple[int, str], deque[int]]]:
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT id, local_id, nome, referencia_externa, source_hash, deleted_at
        FROM equipamentos
        ORDER BY id
        """
    ).fetchall()
    by_reference: dict[str, dict[str, Any]] = {}
    legacy: dict[tuple[int, str], deque[int]] = defaultdict(deque)
    for row in rows:
        reference = clean_text(row["referencia_externa"])
        if reference:
            by_reference[reference] = dict(row)
        elif row["local_id"] is not None and not clean_text(row["deleted_at"]):
            legacy[(int(row["local_id"]), fold_text(row["nome"]))].append(int(row["id"]))
    conn.row_factory = None
    return by_reference, legacy


def import_registry(db_path: str, parsed: dict[str, Any], actor: str = "sge") -> dict[str, Any]:
    rows = parsed.get("rows") or []
    if not rows:
        raise ValueError("Não há registos preparados para importar.")
    duplicate_count = len(rows) - len({row["source_key"] for row in rows})
    if duplicate_count:
        raise ValueError(
            f"O ficheiro contém {duplicate_count} código(s) de activo duplicado(s). "
            "Corrija-os antes de confirmar a actualização."
        )
    conn = sqlite3.connect(db_path, timeout=30)
    conn.execute("PRAGMA busy_timeout=30000")
    batch_id = None
    try:
        conn.execute("BEGIN IMMEDIATE")
        cursor = conn.execute(
            """
            INSERT INTO asset_import_batches(source_name, source_hash, actor, status, total_rows)
            VALUES(?,?,?,?,?)
            """,
            (parsed.get("source_name"), parsed.get("source_hash"), actor, "Em curso", len(rows)),
        )
        batch_id = int(cursor.lastrowid)
        location_aliases, location_names = _existing_location_map(conn)
        local_cache: dict[str, int] = {}
        created_locations = 0
        for row in rows:
            key = fold_text(row["local_name"])
            if key in local_cache:
                continue
            local_id, created = _upsert_location(
                conn, location_aliases, location_names, row["local_name"], row["sector_operacional"], actor
            )
            local_cache[key] = local_id
            created_locations += int(created)

        by_reference, legacy = _existing_equipment_state(conn)
        counters = Counter()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row in rows:
            local_id = local_cache[fold_text(row["local_name"])]
            existing = by_reference.get(row["source_key"])
            equipment_id = int(existing["id"]) if existing else None
            reconciled = False
            if equipment_id is None:
                candidates = legacy.get((local_id, fold_text(row["nome"])))
                if candidates:
                    equipment_id = candidates.popleft()
                    reconciled = True
            if equipment_id is not None:
                old_hash = clean_text(existing.get("source_hash")) if existing else ""
                if existing and old_hash == row["source_hash"] and not clean_text(existing.get("deleted_at")):
                    counters["sem_alteracao"] += 1
                    continue
                conn.execute(
                    """
                    UPDATE equipamentos SET
                        local_id=?, nome=?,
                        tag=CASE WHEN ?<>'' THEN ? ELSE tag END,
                        especificacao=CASE WHEN COALESCE(TRIM(especificacao),'')='' THEN ? ELSE especificacao END,
                        ano_instalacao=CASE WHEN ?<>'' THEN ? ELSE ano_instalacao END,
                        quantidade=CASE WHEN ? > 0 THEN ? ELSE COALESCE(quantidade,1) END,
                        ativo=CASE WHEN ? IS NOT NULL THEN ? ELSE ativo END,
                        categoria=?,
                        fabricante=CASE WHEN ?<>'' THEN ? ELSE fabricante END,
                        modelo=CASE WHEN ?<>'' THEN ? ELSE modelo END,
                        numero_serie=CASE WHEN ?<>'' THEN ? ELSE numero_serie END,
                        custo_aquisicao=COALESCE(custo_aquisicao, ?),
                        vida_util_anos=COALESCE(vida_util_anos, ?),
                        criticidade=CASE WHEN ?<>'' THEN ? ELSE criticidade END,
                        potencia_kw=COALESCE(potencia_kw, ?), tensao_v=COALESCE(tensao_v, ?), corrente_a=COALESCE(corrente_a, ?),
                        fornecedor=CASE WHEN ?<>'' THEN ? ELSE fornecedor END,
                        contrato_num=CASE WHEN ?<>'' THEN ? ELSE contrato_num END,
                        garantia_fim=CASE WHEN ?<>'' THEN ? ELSE garantia_fim END,
                        sistema=?, instalacao=?, estado_operacional=?, periodicidade_manutencao=?,
                        sector_operacional=?, referencia_externa=?, fonte_cadastro=?, source_record_no=?,
                        source_hash=?, ultima_sincronizacao=?, updated_at=?, deleted_at=NULL
                    WHERE id=?
                    """,
                    (
                        local_id, row["nome"], row["tag"], row["tag"], row["especificacao"] or row["modelo"],
                        row["ano_instalacao"], row["ano_instalacao"], row["quantidade"], row["quantidade"],
                        row["ativo"], row["ativo"], row["categoria"],
                        row["fabricante"], row["fabricante"], row["modelo"], row["modelo"], row["numero_serie"], row["numero_serie"],
                        row["custo_aquisicao"], row["vida_util_anos"],
                        row["criticidade"], row["criticidade"], row["potencia_kw"], row["tensao_v"], row["corrente_a"],
                        row["fornecedor"], row["fornecedor"], row["contrato_num"], row["contrato_num"],
                        row["garantia_fim"], row["garantia_fim"], row["sistema"], row["instalacao"], row["estado_operacional"],
                        row["periodicidade_manutencao"], row["sector_operacional"], row["source_key"], parsed.get("source_name"),
                        row["source_record_no"], row["source_hash"], now, now, equipment_id,
                    ),
                )
                if reconciled:
                    counters["reconciliados"] += 1
                    action = "reconciliar_cadastro"
                else:
                    counters["actualizados"] += 1
                    action = "actualizar_cadastro"
                if action:
                    conn.execute(
                        "INSERT INTO equipamentos_audit(equipamento_id, acao, detalhes, actor) VALUES(?,?,?,?)",
                        (equipment_id, action, f"Lote #{batch_id}; origem {parsed.get('source_name')}", actor),
                    )
                by_reference[row["source_key"]] = {"id": equipment_id, "source_hash": row["source_hash"]}
                continue

            cursor = conn.execute(
                """
                INSERT INTO equipamentos(
                    nome, local_id, tag, especificacao, ano_instalacao, quantidade, ativo,
                    created_at, updated_at, categoria, fabricante, modelo, numero_serie,
                    custo_aquisicao, vida_util_anos, criticidade, potencia_kw, tensao_v, corrente_a, fornecedor, contrato_num,
                    garantia_fim, sistema, instalacao, estado_operacional, periodicidade_manutencao,
                    sector_operacional, referencia_externa, fonte_cadastro, source_record_no,
                    source_hash, ultima_sincronizacao
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    row["nome"], local_id, row["tag"], row["especificacao"] or row["modelo"], row["ano_instalacao"] or None,
                    row["quantidade"], 1 if row["ativo"] is None else row["ativo"], now, now,
                    row["categoria"], row["fabricante"], row["modelo"], row["numero_serie"],
                    row["custo_aquisicao"], row["vida_util_anos"],
                    row["criticidade"], row["potencia_kw"], row["tensao_v"], row["corrente_a"], row["fornecedor"],
                    row["contrato_num"], row["garantia_fim"] or None, row["sistema"], row["instalacao"],
                    row["estado_operacional"], row["periodicidade_manutencao"], row["sector_operacional"], row["source_key"],
                    parsed.get("source_name"), row["source_record_no"], row["source_hash"], now,
                ),
            )
            equipment_id = int(cursor.lastrowid)
            counters["inseridos"] += 1
            conn.execute(
                "INSERT INTO equipamentos_audit(equipamento_id, acao, detalhes, actor) VALUES(?,?,?,?)",
                (equipment_id, "importar_cadastro", f"Lote #{batch_id}; origem {parsed.get('source_name')}", actor),
            )

        result = {
            "batch_id": batch_id,
            "total": len(rows),
            "locais_criados": created_locations,
            **{key: int(value) for key, value in counters.items()},
        }
        conn.execute(
            """
            UPDATE asset_import_batches SET status='Concluído', inserted_rows=?, updated_rows=?,
                reconciled_rows=?, unchanged_rows=?, new_locations=?, summary_json=?, completed_at=datetime('now','localtime')
            WHERE id=?
            """,
            (
                counters["inseridos"], counters["actualizados"], counters["reconciliados"],
                counters["sem_alteracao"], created_locations, json.dumps(result, ensure_ascii=False), batch_id,
            ),
        )
        conn.commit()
        return result
    except Exception as exc:
        conn.rollback()
        if batch_id:
            try:
                conn.execute(
                    "UPDATE asset_import_batches SET status='Falhou', error_message=?, completed_at=datetime('now','localtime') WHERE id=?",
                    (clean_text(exc)[:1000], batch_id),
                )
                conn.commit()
            except Exception:
                conn.rollback()
        raise
    finally:
        conn.close()


def registry_dashboard(db_path: str) -> dict[str, Any]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        total = conn.execute("SELECT COUNT(*) FROM equipamentos WHERE COALESCE(deleted_at,'')=''").fetchone()[0]
        source_total = conn.execute(
            "SELECT COUNT(*) FROM equipamentos WHERE COALESCE(referencia_externa,'')<>'' AND COALESCE(deleted_at,'')=''"
        ).fetchone()[0]
        locations = conn.execute("SELECT COUNT(*) FROM locais WHERE COALESCE(ativo,1)=1").fetchone()[0]
        etas = conn.execute("SELECT COUNT(*) FROM locais WHERE COALESCE(ativo,1)=1 AND tipo_local='ETA'").fetchone()[0]
        missing = conn.execute(
            """
            SELECT
              SUM(CASE WHEN COALESCE(TRIM(fabricante),'')='' THEN 1 ELSE 0 END),
              SUM(CASE WHEN COALESCE(TRIM(modelo),'')='' THEN 1 ELSE 0 END),
              SUM(CASE WHEN COALESCE(TRIM(estado_operacional),'')='' OR estado_operacional='Não informado' THEN 1 ELSE 0 END),
              SUM(CASE WHEN COALESCE(TRIM(criticidade),'')='' THEN 1 ELSE 0 END)
            FROM equipamentos WHERE COALESCE(deleted_at,'')=''
            """
        ).fetchone()
        by_sector = [dict(row) for row in conn.execute(
            "SELECT COALESCE(NULLIF(sector_operacional,''),'Não informado') AS nome, COUNT(*) AS total FROM equipamentos WHERE COALESCE(deleted_at,'')='' GROUP BY 1 ORDER BY total DESC"
        )]
        by_type = [dict(row) for row in conn.execute(
            "SELECT COALESCE(NULLIF(tipo_local,''),'Outro') AS nome, COUNT(*) AS total FROM locais WHERE COALESCE(ativo,1)=1 GROUP BY 1 ORDER BY total DESC"
        )]
        batches = [dict(row) for row in conn.execute(
            "SELECT id, source_name, actor, status, total_rows, inserted_rows, updated_rows, reconciled_rows, unchanged_rows, new_locations, started_at, completed_at FROM asset_import_batches ORDER BY id DESC LIMIT 10"
        )]
        return {
            "total": int(total or 0), "source_total": int(source_total or 0), "locations": int(locations or 0), "etas": int(etas or 0),
            "missing_manufacturer": int(missing[0] or 0), "missing_model": int(missing[1] or 0),
            "missing_state": int(missing[2] or 0), "missing_criticality": int(missing[3] or 0),
            "by_sector": by_sector, "by_type": by_type, "batches": batches,
        }
    finally:
        conn.close()


def bootstrap_bundled_registry(db_path: str, base_dir: str) -> dict[str, Any] | None:
    conn = sqlite3.connect(db_path)
    try:
        if conn.execute("SELECT 1 FROM schema_migrations WHERE version=?", (BUNDLED_REGISTRY_MARKER,)).fetchone():
            return None
    finally:
        conn.close()
    source = Path(base_dir) / "data" / BUNDLED_REGISTRY_FILENAME
    if not source.is_file():
        return None
    parsed = parse_registry_file(source, source.name)
    result = import_registry(db_path, parsed, actor="migração_cadastro_dima")
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            "INSERT OR IGNORE INTO schema_migrations(version, description) VALUES(?,?)",
            (BUNDLED_REGISTRY_MARKER, "Cadastro mestre DIMA de locais e 3.147 activos"),
        )
        conn.commit()
    finally:
        conn.close()
    return result
